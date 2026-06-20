# -*- coding: utf-8 -*-
"""Εστία — evaluations.py (Module Αξιολόγησης Προσωπικού, Φ1).
Plug-in: import από το ΤΕΛΟΣ του app.py (πριν το init_db ώστε το create_all να πιάσει τα μοντέλα).
HR/admin-only. Reference: CONDIAN Employee Evaluation Form (25 σταθμισμένα κριτήρια, κλίμακα 1-10).
"""
from datetime import datetime, date
from flask import request, redirect, url_for, render_template, session, jsonify, Response
from app import (app, db, current_user, is_admin, log_activity, role_rank, ROLE_RANK,
                 User, Hotel, allowed_hotels)

# ── Μοντέλα ───────────────────────────────────────────────────────────────────
class EvalTemplate(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    name       = db.Column(db.String(80), nullable=False)
    scope      = db.Column(db.String(40), default='general')   # 'general' | dept-key (Φ2 variants)
    scale_max  = db.Column(db.Integer, default=10)
    is_active  = db.Column(db.Boolean, default=True)
    hotel_id   = db.Column(db.Integer, db.ForeignKey('hotel.id'))   # null = όλα τα ξενοδοχεία
    version    = db.Column(db.Integer, default=1)
    created_at = db.Column(db.DateTime, default=datetime.now)
    hotel      = db.relationship('Hotel')
    criteria   = db.relationship('EvalCriterion', backref='template',
                                 order_by='EvalCriterion.sort', cascade='all, delete-orphan')

class EvalCriterion(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    template_id = db.Column(db.Integer, db.ForeignKey('eval_template.id'))
    grp         = db.Column(db.String(40))      # ομάδα (Βασικά / Ξένες γλώσσες)
    label       = db.Column(db.String(200))
    weight      = db.Column(db.Float, default=0)
    sort        = db.Column(db.Integer, default=0)
    max_score   = db.Column(db.Integer, default=10)

class Evaluation(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    code          = db.Column(db.String(32), index=True)  # κοινό σε όλες τις versions
    employee_id   = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    evaluator_id  = db.Column(db.Integer, db.ForeignKey('user.id'))
    hotel_id      = db.Column(db.Integer, db.ForeignKey('hotel.id'))
    department_id = db.Column(db.Integer)
    template_id   = db.Column(db.Integer, db.ForeignKey('eval_template.id'))
    period_label  = db.Column(db.String(40))    # σαιζόν/περίοδος
    year          = db.Column(db.Integer)
    eval_date     = db.Column(db.Date, default=date.today)
    status        = db.Column(db.String(12), default='draft')   # draft | finalized
    score_pct     = db.Column(db.Float)
    band          = db.Column(db.String(24))
    prev_eval_id  = db.Column(db.Integer)       # auto trend
    general_comment = db.Column(db.Text)
    version       = db.Column(db.Integer, default=1)
    supersedes_id = db.Column(db.Integer)
    created_at    = db.Column(db.DateTime, default=datetime.now)
    updated_at    = db.Column(db.DateTime)
    employee  = db.relationship('User', foreign_keys=[employee_id])
    evaluator = db.relationship('User', foreign_keys=[evaluator_id])
    hotel     = db.relationship('Hotel')
    template  = db.relationship('EvalTemplate')
    scores    = db.relationship('EvalScore', backref='evaluation', cascade='all, delete-orphan')
    goals     = db.relationship('EvalGoal', backref='evaluation', cascade='all, delete-orphan')

class EvalScore(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    evaluation_id = db.Column(db.Integer, db.ForeignKey('evaluation.id'))
    criterion_id  = db.Column(db.Integer, db.ForeignKey('eval_criterion.id'))
    score         = db.Column(db.Float)
    comment       = db.Column(db.Text)

class EvalGoal(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    evaluation_id = db.Column(db.Integer, db.ForeignKey('evaluation.id'))
    text          = db.Column(db.String(300))
    done          = db.Column(db.Boolean, default=False)
    created_at    = db.Column(db.DateTime, default=datetime.now)

class EvalPeriod(db.Model):
    """v12.100 — Περίοδοι αξιολόγησης ορισμένες από HR (π.χ. «1η αξιολόγηση» 15/05/2026)."""
    id        = db.Column(db.Integer, primary_key=True)
    name      = db.Column(db.String(80))
    ref_date  = db.Column(db.Date)
    is_active = db.Column(db.Boolean, default=True)
    sort      = db.Column(db.Integer, default=0)
    created_at= db.Column(db.DateTime, default=datetime.now)
    @property
    def label(self):
        return ('%s - %s' % (self.name, self.ref_date.strftime('%d/%m/%Y'))) if self.ref_date else (self.name or '')

# ── Seed προτύπου (reference CONDIAN) ─────────────────────────────────────────
_CORE = [
    ('Γνώση & Κατανόηση της δουλειάς', 0.05), ('Συνέπεια στο ωράριο', 0.06),
    ('Εργάζεται οργανωμένα / ποιοτικά / παραγωγικά, με σωστή διαχείριση χρόνου', 0.05),
    ('Ακολουθεί διαδικασίες και ποιοτικές προδιαγραφές', 0.06),
    ('Φροντίζει τον εξοπλισμό / μηχανήματα και την ασφάλειά τους', 0.04),
    ('Υπευθυνότητα χωρίς συνεχή επίβλεψη', 0.04),
    ('Ευελιξία και προσαρμοστικότητα σε αλλαγές', 0.05),
    ('Αναπτύσσει πρωτοβουλίες & προβλέπει τις ανάγκες δουλειάς και τμήματος', 0.05),
    ('Επίδειξη ικανοτήτων εξέλιξης & προαγωγής', 0.05), ('Γνώση Η/Υ', 0.04),
    ('Προσωπική φροντίδα και υγιεινή', 0.06), ('Φροντίζει την στολή του / Φοράει την κονκάρδα', 0.05),
    ('Αποδοχή συμβουλών και οδηγιών', 0.03), ('Προθυμία για εκπαίδευση και για νέες γνώσεις', 0.05),
    ('Εθελοντισμός & Ευαισθησία', 0.05), ('Προσφορά βοήθειας, προθυμία, εξυπηρέτηση συναδέλφων', 0.03),
    ('Ευγένεια και φιλικότητα', 0.05), ('Εξυπηρέτηση και προθυμία προς τους πελάτες', 0.05),
    ('Εργασία κάτω από πίεση', 0.04),
]
_LANG = [('Γερμανικά', 0.02), ('Αγγλικά', 0.025), ('Γαλλικά', 0.025),
         ('Ιταλικά', 0.01), ('Ρώσικα', 0.01), ('Άλλες ξένες γλώσσες', 0.01)]

def _seed_menu_evals():
    """Μία φορά: πρόσθεσε 'evals' στο αποθηκευμένο menu_vis (manager on) ώστε να εμφανίζεται & να ρυθμίζεται."""
    try:
      with app.app_context():
        from app import Setting
        if Setting.query.get('menu_vis_evals_seeded'):
            return
        row = Setting.query.get('menu_vis')
        if row and row.value:
            import json as _j
            cfg = _j.loads(row.value)
            cfg.setdefault('manager', [])
            if 'evals' not in cfg['manager']:
                cfg['manager'].append('evals')
            row.value = _j.dumps(cfg, ensure_ascii=False)
        db.session.add(Setting(key='menu_vis_evals_seeded', value='1'))
        db.session.commit()
    except Exception as e:
        db.session.rollback(); print('[evaluations] menu seed skipped:', e)

def ensure_eval_setup():
    """Seed του κοινού προτύπου CONDIAN (idempotent)."""
    try:
      with app.app_context():
        db.create_all()
        if EvalTemplate.query.first():
            return
        t = EvalTemplate(name='Πρότυπο CONDIAN', scope='general', scale_max=10, is_active=True)
        db.session.add(t); db.session.flush()
        i = 0
        for lab, w in _CORE:
            db.session.add(EvalCriterion(template_id=t.id, grp='Βασικά κριτήρια', label=lab, weight=w, sort=i, max_score=10)); i += 1
        for lab, w in _LANG:
            db.session.add(EvalCriterion(template_id=t.id, grp='Ξένες γλώσσες', label=lab, weight=w, sort=i, max_score=10)); i += 1
        db.session.commit()
        print('[evaluations] seeded πρότυπο CONDIAN (%d κριτήρια)' % (len(_CORE)+len(_LANG)))
    except Exception as e:
        db.session.rollback(); print('[evaluations] seed skipped:', e)

# ── Seed προτύπων τμημάτων F&B (Kitchen / Service) από reference 2026 ──────────
_TPL_KITCHEN = [
    ('Γνώση & Κατανόηση της δουλειάς', 0.05, 'Βασικά κριτήρια'),
    ('Συνέπεια στο ωράριο', 0.05, 'Βασικά κριτήρια'),
    ('Εργάζεται οργανωμένα / ποιοτικά / παραγωγικά, με σωστή διαχείριση χρόνου', 0.05, 'Βασικά κριτήρια'),
    ('Ακολουθεί διαδικασίες και ποιοτικές προδιαγραφές', 0.05, 'Βασικά κριτήρια'),
    ('Φροντίζει τον εξοπλισμό / μηχανήματα και την ασφάλειά τους', 0.04, 'Βασικά κριτήρια'),
    ('Υπευθυνότητα χωρίς συνεχή επίβλεψη', 0.05, 'Βασικά κριτήρια'),
    ('Ευελιξία και προσαρμοστικότητα σε αλλαγές', 0.05, 'Βασικά κριτήρια'),
    ('Αναπτύσσει πρωτοβουλίες & προβλέπει τις ανάγκες δουλειάς και τμήματος', 0.05, 'Βασικά κριτήρια'),
    ('Επίδειξη ικανοτήτων εξέλιξης & προαγωγής', 0.05, 'Βασικά κριτήρια'),
    ('Γνώση και χρήση εξοπλισμού', 0.04, 'Βασικά κριτήρια'),
    ('Προσωπική φροντίδα και υγιεινή', 0.04, 'Βασικά κριτήρια'),
    ('Φροντίζει την στολή του / Φοράει την κονκάρδα', 0.03, 'Βασικά κριτήρια'),
    ('Αποδοχή συμβουλών και οδηγιών', 0.04, 'Βασικά κριτήρια'),
    ('Προθυμία για εκπαίδευση και για νέες γνώσεις', 0.04, 'Βασικά κριτήρια'),
    ('Εθελοντισμός & Ευαισθησία', 0.04, 'Βασικά κριτήρια'),
    ('Προσφορά βοήθειας, προθυμία, εξυπηρέτηση συναδέλφων', 0.03, 'Βασικά κριτήρια'),
    ('Ευγένεια και φιλικότητα', 0.03, 'Βασικά κριτήρια'),
    ('Εξυπηρέτηση και προθυμία προς τους πελάτες', 0.03, 'Βασικά κριτήρια'),
    ('Εργασία κάτω από πίεση', 0.04, 'Βασικά κριτήρια'),
    ('Γνώσεις HACCP & ISO', 0.04, 'Τεχνικά τμήματος'),
    ('Λειτουργία του συστήματος FIFO-LIFO', 0.04, 'Τεχνικά τμήματος'),
    ('Χρήση σωστών χρωματικών κωδικών στην κουζίνα', 0.025, 'Τεχνικά τμήματος'),
    ('Γνώση θερμοκρασιών και αναθέρμανσης', 0.04, 'Τεχνικά τμήματος'),
    ('Χρήση του συστήματος Clean as you go', 0.03, 'Τεχνικά τμήματος'),
    ('Αγγλικά', 0.025, 'Ξένες γλώσσες'),
]
_TPL_SERVICE = [
    ('Γνώση & Κατανόηση της δουλειάς', 0.04, 'Βασικά κριτήρια'),
    ('Συνέπεια στο ωράριο', 0.04, 'Βασικά κριτήρια'),
    ('Εργάζεται οργανωμένα / ποιοτικά / παραγωγικά, με σωστή διαχείριση χρόνου', 0.04, 'Βασικά κριτήρια'),
    ('Ακολουθεί διαδικασίες και ποιοτικές προδιαγραφές', 0.04, 'Βασικά κριτήρια'),
    ('Φροντίζει τον εξοπλισμό / μηχανήματα και την ασφάλειά τους', 0.04, 'Βασικά κριτήρια'),
    ('Υπευθυνότητα χωρίς συνεχή επίβλεψη', 0.05, 'Βασικά κριτήρια'),
    ('Ευελιξία και προσαρμοστικότητα σε αλλαγές', 0.05, 'Βασικά κριτήρια'),
    ('Αναπτύσσει πρωτοβουλίες & προβλέπει τις ανάγκες δουλειάς και τμήματος', 0.05, 'Βασικά κριτήρια'),
    ('Επίδειξη ικανοτήτων εξέλιξης & προαγωγής', 0.05, 'Βασικά κριτήρια'),
    ('Γνώση και χρήση εξοπλισμού', 0.03, 'Βασικά κριτήρια'),
    ('Προσωπική φροντίδα και υγιεινή', 0.04, 'Βασικά κριτήρια'),
    ('Φροντίζει την στολή του / Φοράει την κονκάρδα', 0.03, 'Βασικά κριτήρια'),
    ('Αποδοχή συμβουλών και οδηγιών', 0.03, 'Βασικά κριτήρια'),
    ('Προθυμία για εκπαίδευση και για νέες γνώσεις', 0.04, 'Βασικά κριτήρια'),
    ('Εθελοντισμός & Ευαισθησία', 0.04, 'Βασικά κριτήρια'),
    ('Προσφορά βοήθειας, προθυμία, εξυπηρέτηση συναδέλφων', 0.03, 'Βασικά κριτήρια'),
    ('Ευγένεια και φιλικότητα', 0.03, 'Βασικά κριτήρια'),
    ('Εξυπηρέτηση και προθυμία προς τους πελάτες', 0.03, 'Βασικά κριτήρια'),
    ('Εργασία κάτω από πίεση', 0.04, 'Βασικά κριτήρια'),
    ('Γνώσεις αλλεργιογόνων', 0.03, 'Τεχνικά τμήματος'),
    ('Λειτουργία του συστήματος FIFO-LIFO', 0.03, 'Τεχνικά τμήματος'),
    ('Χρήση σωστών χρωματικών κωδικών στον χώρο του εστιατορίου', 0.02, 'Τεχνικά τμήματος'),
    ('Γνώση θερμοκρασιών και αναθέρμανσης', 0.02, 'Τεχνικά τμήματος'),
    ('Χρήση του συστήματος Clean as you go', 0.03, 'Τεχνικά τμήματος'),
    ('Upselling', 0.03, 'Τεχνικά τμήματος'),
    ('Γνώσεις ποτών και κρασιών', 0.03, 'Τεχνικά τμήματος'),
    ('Αγγλικά', 0.02, 'Ξένες γλώσσες'),
    ('Γερμανικά', 0.01, 'Ξένες γλώσσες'),
    ('Γνώση και χρήση PDA/PC', 0.02, 'Τεχνικά τμήματος'),
    ('Τήρηση των χρόνων διαλειμμάτων', 0.02, 'Τεχνικά τμήματος'),
]

def _seed_one_tpl(name, scope, rows):
    if EvalTemplate.query.filter_by(name=name).first():
        return False
    t = EvalTemplate(name=name, scope=scope, scale_max=10, is_active=True)
    db.session.add(t); db.session.flush()
    for i, (lab, w, g) in enumerate(rows):
        db.session.add(EvalCriterion(template_id=t.id, grp=g, label=lab, weight=w, sort=i, max_score=10))
    return True

def ensure_eval_dept_templates():
    """Seed προτύπων τμημάτων F&B (Kitchen/Service) — self-healing:
    1) καθαρίζει ΑΧΡΗΣΙΜΟΠΟΙΗΤΑ διπλά (race 2 gunicorn workers), κρατά το παλαιότερο·
    2) δημιουργεί ΜΟΝΟ αν δεν υπάρχει κανένα. Idempotent & ανθεκτικό σε re-boot."""
    try:
      with app.app_context():
        db.create_all()
        made = []; cleaned = 0
        for name, scope, rows in [('F&B — Κουζίνα (Kitchen)', 'Kitchen', _TPL_KITCHEN),
                                  ('F&B — Σέρβις (Service)', 'Service', _TPL_SERVICE)]:
            existing = EvalTemplate.query.filter_by(name=name).order_by(EvalTemplate.id).all()
            if len(existing) > 1:
                for dup in existing[1:]:
                    if Evaluation.query.filter_by(template_id=dup.id).count() == 0:
                        db.session.delete(dup); cleaned += 1
                db.session.commit()
            elif not existing:
                if _seed_one_tpl(name, scope, rows):
                    db.session.commit(); made.append(scope)
        if made:
            print('[evaluations] seeded πρότυπα τμημάτων: %s' % ', '.join(made))
        if cleaned:
            print('[evaluations] καθαρίστηκαν %d διπλά πρότυπα (race)' % cleaned)
    except Exception as e:
        db.session.rollback(); print('[evaluations] dept templates seed skipped:', e)

def ensure_eval_columns():
    """v12.141 — auto-migration: eval_template.hotel_id (ανάθεση σε ξενοδοχείο)."""
    try:
      with app.app_context():
        from app import _add_col
        _add_col('eval_template', 'hotel_id', 'hotel_id INTEGER')
    except Exception as e:
        print('[evaluations] ensure_eval_columns skipped:', e)


# ── Helpers ───────────────────────────────────────────────────────────────────
def _active_periods():
    return EvalPeriod.query.filter_by(is_active=True).order_by(EvalPeriod.sort, EvalPeriod.ref_date).all()

def band_for(pct):
    if pct is None: return '—'
    if pct >= 80: return 'Εξαιρετική'
    if pct >= 60: return 'Καλή'
    return 'Χρειάζεται προσοχή'

BAND_COLOR = {'Εξαιρετική': ('#dcfce7', '#16a34a'), 'Καλή': ('#fef3c7', '#b45309'),
              'Χρειάζεται προσοχή': ('#fee2e2', '#dc2626'), '—': ('#f1f5f9', '#64748b')}
APPROVED_STATES = ('approved', 'finalized')   # finalized = legacy
ST_LABEL = {'draft': 'Πρόχειρο', 'submitted': 'Προς έγκριση', 'approved': 'Εγκρίθηκε',
            'finalized': 'Εγκρίθηκε', 'returned': 'Επιστράφηκε'}
ST_COLOR = {'draft': ('#f1f5f9', '#64748b'), 'submitted': ('#fef3c7', '#b45309'),
            'approved': ('#dcfce7', '#16a34a'), 'finalized': ('#dcfce7', '#16a34a'),
            'returned': ('#fee2e2', '#dc2626')}

def compute_pct(scores_map, criteria):
    """scores_map: {criterion_id: score}. Σταθμισμένο % με βάση weight & max_score."""
    num = den = 0.0
    for c in criteria:
        v = scores_map.get(c.id)
        if v is None or v == '':
            continue
        try: v = float(v)
        except (TypeError, ValueError): continue
        num += v * (c.weight or 0)
        den += (c.weight or 0) * (c.max_score or 10)
    if den <= 0: return None
    return round(num / den * 100, 1)

def _hotel_code(hid):
    h = Hotel.query.get(hid) if hid else None
    if not h:
        return 'XXX'
    try:
        from schedule import _hotel_short
        c = _hotel_short(h.name)
        if c:
            return c
    except Exception:
        pass
    import re
    s = re.sub(r'[^A-Z0-9]', '', (h.name or '').upper())
    return s[:3] or 'XXX'

def _gen_code(ev):
    """Μορφή κωδικού: ΕΤΟΣ-HOTELCODE-DEADLINE(ddmmyyyy)-SUBMISSIONID(0000)."""
    yr = ev.year or date.today().year
    hc = _hotel_code(ev.hotel_id)
    dl = ev.eval_date.strftime('%d%m%Y') if ev.eval_date else ''
    sid = '%04d' % (ev.id or 0)
    return '-'.join([p for p in [str(yr), hc, dl, sid] if p])

def _dept_name(did):
    if not did: return ''
    try:
        from schedule import Department
        d = Department.query.get(did)
        return d.name if d else ''
    except Exception:
        return ''

def _employees():
    """Ενεργοί εργαζόμενοι από Μητρώο (για επιλογή + autofill)."""
    out = []
    try:
        from schedule import Department
        deps = {d.id: d.name for d in Department.query.all()}
    except Exception:
        deps = {}
    hotels = {h.id: h.name for h in Hotel.query.all()}
    spec = {}; ecode = {}; eafm = {}
    try:
        from payroll import EmployeePII
        for p in EmployeePII.query.all():
            spec[p.user_id] = getattr(p, 'ergani_specialty', None)
            ecode[p.user_id] = getattr(p, 'emp_code', None)
            eafm[p.user_id] = getattr(p, 'afm', None)
    except Exception:
        pass
    hids = _scope_hids()
    q = User.query.filter(User.is_active == True)
    for u in q.order_by(User.full_name).all():
        if getattr(u, 'employment_active', None) is False:
            continue
        if hids is not None and getattr(u, 'home_hotel_id', None) not in hids:
            continue
        out.append({'id': u.id, 'name': u.full_name,
                    'department_id': getattr(u, 'department_id', None),
                    'department': deps.get(getattr(u, 'department_id', None), ''),
                    'hotel_id': getattr(u, 'home_hotel_id', None),
                    'hotel': hotels.get(getattr(u, 'home_hotel_id', None), ''),
                    'specialty': spec.get(u.id, '') or '',
                    'emp_code': ecode.get(u.id, '') or '',
                    'afm': eafm.get(u.id, '') or ''})
    return out

def _auth():
    return is_admin()

def _auth_eval():
    """manager+ (admin/master always). Οι managers με scope ξενοδοχείων."""
    u = current_user()
    return u is not None and role_rank(u.role) >= ROLE_RANK['manager']

def _scope_hids():
    """None = όλα (admin). Αλλιώς set ξενοδοχείων του manager (allowed_hotels)."""
    if is_admin():
        return None
    u = current_user()
    try:
        return {h.id for h in allowed_hotels(u)} if u else set()
    except Exception:
        return set()

def _hid_ok(hid):
    hs = _scope_hids()
    return hs is None or (hid in hs)

# ── ROUTES ────────────────────────────────────────────────────────────────────
@app.route('/dashboard/evaluations')
def evaluations_list():
    if not _auth_eval():
        return redirect(url_for('login'))
    f_hotel = request.args.get('hotel_id', type=int)
    f_band  = request.args.get('band', '')
    f_period= request.args.get('period', '')
    q = (request.args.get('q') or '').strip()
    qry = Evaluation.query
    if f_hotel: qry = qry.filter(Evaluation.hotel_id == f_hotel)
    if f_band:  qry = qry.filter(Evaluation.band == f_band)
    if f_period:qry = qry.filter(Evaluation.period_label == f_period)
    items = qry.order_by(Evaluation.eval_date.desc(), Evaluation.id.desc()).limit(500).all()
    if q:
        ql = q.lower()
        items = [e for e in items if e.employee and ql in (e.employee.full_name or '').lower()]
    _hs = _scope_hids()
    if _hs is not None:
        items = [e for e in items if e.hotel_id in _hs]
    periods = [p[0] for p in db.session.query(Evaluation.period_label).distinct().all() if p[0]]
    hotels = Hotel.query.filter_by(is_active=True).order_by(Hotel.name).all()
    submitted_n = Evaluation.query.filter_by(status='submitted').count()
    return render_template('evaluations_list.html', items=items, hotels=hotels,
                           periods=periods, f_hotel=f_hotel, f_band=f_band, f_period=f_period, q=q,
                           band_color=BAND_COLOR, dept_name=_dept_name, st_label=ST_LABEL, st_color=ST_COLOR,
                           submitted_n=submitted_n, is_admin=is_admin())

@app.route('/dashboard/evaluations/new', methods=['GET', 'POST'])
def evaluation_new():
    if not _auth_eval():
        return redirect(url_for('login'))
    tid = request.args.get('template_id', type=int)
    tmpl = (EvalTemplate.query.filter_by(id=tid, is_active=True).first() if tid else None) \
           or EvalTemplate.query.filter_by(is_active=True, scope='general').order_by(EvalTemplate.id).first() \
           or EvalTemplate.query.filter_by(is_active=True).order_by(EvalTemplate.id).first()
    if request.method == 'POST' and tmpl:
        return _save_evaluation(None, tmpl)
    templates = EvalTemplate.query.filter_by(is_active=True).order_by(EvalTemplate.scope, EvalTemplate.name).all()
    _u = current_user(); _hs = allowed_hotels(_u) if _u else []
    return render_template('evaluation_form.html', ev=None, tmpl=tmpl, templates=templates,
                           criteria=tmpl.criteria if tmpl else [], employees=_employees(),
                           scores={}, goals=[], periods=_active_periods(), today=date.today(),
                           form_hotels=_hs)

@app.route('/dashboard/evaluations/<int:eid>')
def evaluation_view(eid):
    if not _auth_eval():
        return redirect(url_for('login'))
    ev = Evaluation.query.get_or_404(eid)
    if not _hid_ok(ev.hotel_id):
        return redirect(url_for('evaluations_list') + '?embed=1')
    smap = {s.criterion_id: s for s in ev.scores}
    prev = Evaluation.query.get(ev.prev_eval_id) if ev.prev_eval_id else None
    return render_template('evaluation_view.html', ev=ev, smap=smap, prev=prev,
                           band_color=BAND_COLOR, dept_name=_dept_name, st_label=ST_LABEL, st_color=ST_COLOR,
                           goals=sorted(ev.goals, key=lambda g: g.id), is_admin=is_admin(),
                           criteria=ev.template.criteria if ev.template else [])

@app.route('/dashboard/evaluations/<int:eid>/edit', methods=['GET', 'POST'])
def evaluation_edit(eid):
    if not _auth_eval():
        return redirect(url_for('login'))
    ev = Evaluation.query.get_or_404(eid)
    if not _hid_ok(ev.hotel_id):
        return redirect(url_for('evaluations_list') + '?embed=1')
    tmpl = ev.template or EvalTemplate.query.filter_by(is_active=True).first()
    if request.method == 'POST':
        return _save_evaluation(ev, tmpl)
    scores = {s.criterion_id: s for s in ev.scores}
    _u = current_user(); _hs = allowed_hotels(_u) if _u else []
    return render_template('evaluation_form.html', ev=ev, tmpl=tmpl, templates=None,
                           criteria=tmpl.criteria if tmpl else [], employees=_employees(),
                           scores=scores, goals=sorted(ev.goals, key=lambda g: g.id),
                           periods=_active_periods(), today=date.today(), form_hotels=_hs)

@app.route('/dashboard/evaluations/<int:eid>/delete', methods=['POST'])
def evaluation_delete(eid):
    if not _auth():
        return redirect(url_for('login'))
    ev = Evaluation.query.get_or_404(eid)
    db.session.delete(ev); db.session.commit()
    log_activity('evaluation_delete', '#%d' % eid)
    return redirect(url_for('evaluations_list') + '?embed=1')

def _save_evaluation(ev, tmpl):
    fm = request.form
    action = fm.get('action', 'save')
    emp_id = fm.get('employee_id', type=int)
    if not emp_id or not tmpl:
        return redirect(url_for('evaluation_new') + '?embed=1')
    user = current_user()
    criteria = tmpl.criteria
    scores_map = {}
    for c in criteria:
        raw = fm.get('score_%d' % c.id, '')
        if raw != '':
            try: scores_map[c.id] = float(raw)
            except ValueError: pass
    pct = compute_pct(scores_map, criteria)
    yr = fm.get('year', type=int) or date.today().year
    # ΟΛΑ τα queries ΠΡΙΝ δημιουργήσουμε το νέο record (αποφυγή autoflush half-built)
    emp = User.query.get(emp_id)
    if not _hid_ok(getattr(emp, 'home_hotel_id', None)):
        return redirect(url_for('evaluations_list') + '?embed=1')
    exclude_code = ev.code if ev else None
    _pq = Evaluation.query.filter(Evaluation.employee_id == emp_id, Evaluation.status.in_(APPROVED_STATES))
    if exclude_code:
        _pq = _pq.filter(Evaluation.code != exclude_code)
    prev = _pq.order_by(Evaluation.eval_date.desc()).first()
    # versioning: αν επεξεργαζόμαστε ΟΡΙΣΤΙΚΟΠΟΙΗΜΕΝΗ → νέα version
    if ev and ev.status in APPROVED_STATES:
        old = ev
        ev = Evaluation(supersedes_id=old.id, version=(old.version or 1) + 1, code=old.code)
        db.session.add(ev)
    if ev is None:
        ev = Evaluation(); db.session.add(ev)
    ev.employee_id = emp_id
    ev.evaluator_id = user.id if user else None
    ev.hotel_id = getattr(emp, 'home_hotel_id', None)
    ev.department_id = getattr(emp, 'department_id', None)
    ev.template_id = tmpl.id
    pid = fm.get('period_id', type=int)
    per = EvalPeriod.query.get(pid) if pid else None
    if per:
        ev.period_label = per.label[:60]
        ev.year = per.ref_date.year if per.ref_date else yr
        ev.eval_date = per.ref_date or date.today()
    else:
        ev.period_label = (fm.get('period_label') or '').strip()[:60]
        ev.year = yr
        try:
            ev.eval_date = datetime.strptime(fm.get('eval_date', ''), '%Y-%m-%d').date()
        except ValueError:
            ev.eval_date = date.today()
    ev.score_pct = pct
    ev.band = band_for(pct)
    ev.prev_eval_id = prev.id if prev else None
    ev.general_comment = (fm.get('general_comment') or '')[:4000]
    if action == 'submit':
        ev.status = 'submitted'
    elif action == 'approve':
        ev.status = 'approved'
    elif action == 'finalize':
        ev.status = 'approved'
    else:
        ev.status = 'draft'
    ev.updated_at = datetime.now()
    db.session.flush()
    if not ev.code:
        ev.code = _gen_code(ev)
    # scores
    for s in list(ev.scores):
        db.session.delete(s)
    for c in criteria:
        if c.id in scores_map:
            db.session.add(EvalScore(evaluation_id=ev.id, criterion_id=c.id,
                                     score=scores_map[c.id],
                                     comment=(fm.get('comment_%d' % c.id) or '')[:1000]))
    for g in list(ev.goals):
        db.session.delete(g)
    for gtext in fm.getlist('goal_text'):
        gtext = (gtext or '').strip()
        if gtext:
            db.session.add(EvalGoal(evaluation_id=ev.id, text=gtext[:300], done=False))
    db.session.commit()
    log_activity('evaluation_save', '%s %s' % (ev.code, ev.status))
    return redirect(url_for('evaluation_view', eid=ev.id) + '?embed=1')

@app.route('/dashboard/evaluations/<int:eid>/export.xlsx')
def evaluation_export(eid):
    if not _auth_eval():
        return redirect(url_for('login'))
    ev = Evaluation.query.get_or_404(eid)
    if not _hid_ok(ev.hotel_id):
        return redirect(url_for('evaluations_list') + '?embed=1')
    import io, openpyxl
    from openpyxl.styles import Font, Alignment
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Αξιολόγηση'
    smap = {s.criterion_id: s for s in ev.scores}
    emp = ev.employee
    ws['A1'] = 'CONDIAN HOTELS'; ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = 'Employee Evaluation Form – Αξιολόγηση Εργαζομένου'; ws['A2'].font = Font(bold=True)
    meta = [('ΞΕΝΟΔΟΧΕΙΑΚΗ ΜΟΝΑΔΑ', ev.hotel.name if ev.hotel else ''),
            ('ΟΝΟΜΑΤΕΠΩΝΥΜΟ', emp.full_name if emp else ''),
            ('ΤΜΗΜΑ', _dept_name(ev.department_id)),
            ('ΠΕΡΙΟΔΟΣ', '%s %s' % (ev.period_label or '', ev.year or '')),
            ('ΗΜΕΡΟΜΗΝΙΑ ΑΞΙΟΛΟΓΗΣΗΣ', ev.eval_date.strftime('%d/%m/%Y') if ev.eval_date else ''),
            ('ΑΞΙΟΛΟΓΗΤΗΣ', ev.evaluator.full_name if ev.evaluator else ''),
            ('ΚΩΔΙΚΟΣ', ev.code)]
    r = 4
    for k, v in meta:
        ws.cell(r, 2, k); ws.cell(r, 3, v); r += 1
    r += 1
    hdr = ['Α/Α', 'ΚΡΙΤΗΡΙΑ', 'ΒΑΘΜΟΛΟΓΙΑ', 'ΣΥΝΤΕΛΕΣΤΗΣ', 'ΣΧΟΛΙΟ']
    for j, h in enumerate(hdr): c = ws.cell(r, j + 1, h); c.font = Font(bold=True)
    r += 1; i = 1
    for crit in (ev.template.criteria if ev.template else []):
        s = smap.get(crit.id)
        ws.cell(r, 1, i); ws.cell(r, 2, crit.label)
        ws.cell(r, 3, (s.score if s else None)); ws.cell(r, 4, crit.weight)
        ws.cell(r, 5, (s.comment if s else '')); r += 1; i += 1
    r += 1
    ws.cell(r, 2, 'ΒΑΘΜΟΛΟΓΙΑ %'); c = ws.cell(r, 3, (ev.score_pct or 0)); c.font = Font(bold=True)
    ws.cell(r + 1, 2, 'ΑΞΙΟΛΟΓΗΣΗ'); ws.cell(r + 1, 3, ev.band or '')
    ws.cell(r + 2, 2, 'ΓΕΝΙΚΟ ΣΧΟΛΙΟ'); ws.cell(r + 2, 3, ev.general_comment or '')
    ws.column_dimensions['B'].width = 55
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    log_activity('evaluation_export', ev.code)
    fname = 'evaluation-%s.xlsx' % (ev.code or ev.id)
    return Response(buf.read(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment; filename=' + fname})


@app.route('/dashboard/evaluations/<int:eid>/export.pdf')
def evaluation_export_pdf(eid):
    if not _auth_eval():
        return redirect(url_for('login'))
    ev = Evaluation.query.get_or_404(eid)
    if not _hid_ok(ev.hotel_id):
        return redirect(url_for('evaluations_list') + '?embed=1')
    import os
    from app import BASE_DIR
    from fpdf import FPDF
    NAVY = (25, 56, 71); GREY = (100, 116, 139)
    smap = {s.criterion_id: s for s in ev.scores}
    emp = ev.employee
    pdf = FPDF(orientation='P', unit='mm', format='A4'); pdf.set_auto_page_break(True, margin=14)
    pdf.add_page()
    pdf.add_font('dv', '', os.path.join(BASE_DIR, 'assets', 'fonts', 'DejaVuSans.ttf'))
    pdf.add_font('dv', 'B', os.path.join(BASE_DIR, 'assets', 'fonts', 'DejaVuSans-Bold.ttf'))
    pdf.set_font('dv', 'B', 15); pdf.set_text_color(*NAVY); pdf.cell(0, 9, 'CONDIAN HOTELS', ln=1)
    pdf.set_font('dv', 'B', 11); pdf.set_text_color(*NAVY); pdf.cell(0, 6, 'Έντυπο Αξιολόγησης Εργαζομένου', ln=1)
    pdf.ln(2)
    meta = [('ΞΕΝΟΔΟΧΕΙΟ', ev.hotel.name if ev.hotel else '—'),
            ('ΟΝΟΜΑΤΕΠΩΝΥΜΟ', emp.full_name if emp else '—'),
            ('ΤΜΗΜΑ', _dept_name(ev.department_id) or '—'),
            ('ΠΕΡΙΟΔΟΣ', ('%s %s' % (ev.period_label or '', ev.year or '')).strip()),
            ('ΗΜ/ΝΙΑ', ev.eval_date.strftime('%d/%m/%Y') if ev.eval_date else '—'),
            ('ΑΞΙΟΛΟΓΗΤΗΣ', ev.evaluator.full_name if ev.evaluator else '—'),
            ('ΚΩΔΙΚΟΣ', ev.code or '—')]
    for k, v in meta:
        pdf.set_font('dv', '', 10); pdf.set_text_color(40, 40, 40)
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(0, 6, '%s:  %s' % (k, v))
    pdf.ln(2)
    pdf.set_font('dv', 'B', 10); pdf.set_text_color(*NAVY); pdf.cell(0, 7, 'Κριτήρια', ln=1)
    i = 1
    for crit in (ev.template.criteria if ev.template else []):
        sc = smap.get(crit.id)
        val = ('%g' % sc.score) if (sc and sc.score is not None) else '–'
        pdf.set_font('dv', '', 9); pdf.set_text_color(20, 20, 20)
        line = '%d. %s   [Βαθμός: %s/%d · Συντ. %d%%]' % (i, crit.label, val, crit.max_score or 10, round((crit.weight or 0) * 100))
        if sc and sc.comment:
            line += '  — ' + sc.comment
        pdf.set_x(pdf.l_margin); pdf.multi_cell(0, 5.5, line)
        i += 1
    pdf.ln(3)
    pdf.set_font('dv', 'B', 12); pdf.set_text_color(*NAVY)
    pdf.cell(0, 8, 'Βαθμολογία: %s%%    ·    %s' % (
        ('%.1f' % ev.score_pct) if ev.score_pct is not None else '—', ev.band or '—'), ln=1)
    if ev.goals:
        pdf.ln(1); pdf.set_font('dv', 'B', 10); pdf.set_text_color(*NAVY); pdf.cell(0, 6, 'Στόχοι:', ln=1)
        pdf.set_font('dv', '', 9); pdf.set_text_color(40, 40, 40)
        for g in sorted(ev.goals, key=lambda x: x.id):
            pdf.set_x(pdf.l_margin); pdf.multi_cell(0, 5.5, ('☑ ' if g.done else '☐ ') + g.text)
    if ev.general_comment:
        pdf.ln(1); pdf.set_font('dv', 'B', 10); pdf.set_text_color(*NAVY); pdf.cell(0, 6, 'Γενικό σχόλιο:', ln=1)
        pdf.set_font('dv', '', 9); pdf.set_text_color(40, 40, 40); pdf.set_x(pdf.l_margin); pdf.multi_cell(0, 5.5, ev.general_comment)
    log_activity('evaluation_export_pdf', ev.code)
    out = pdf.output()
    return Response(bytes(out), mimetype='application/pdf',
                    headers={'Content-Disposition': 'attachment; filename=evaluation-%s.pdf' % (ev.code or ev.id)})

# ── Φ2: Στατιστικά ────────────────────────────────────────────────────────────
def latest_finalized():
    """Μία εγγραφή ανά code: η τελευταία ΟΡΙΣΤΙΚΟΠΟΙΗΜΕΝΗ version (αγνοεί drafts & παλιές versions)."""
    best = {}
    for e in Evaluation.query.filter(Evaluation.status.in_(APPROVED_STATES)).all():
        k = e.code or ('id%d' % e.id)
        if k not in best or (e.version or 1) > (best[k].version or 1):
            best[k] = e
    return list(best.values())

def _maps():
    hotels = {h.id: h.name for h in Hotel.query.all()}
    deps = {}
    try:
        from schedule import Department
        deps = {d.id: d.name for d in Department.query.all()}
    except Exception:
        pass
    return hotels, deps

def _grp(e, mode, deps, hotels):
    if mode == 'current' and e.employee:
        did = getattr(e.employee, 'department_id', None); hid = getattr(e.employee, 'home_hotel_id', None)
    else:
        did = e.department_id; hid = e.hotel_id
    return (deps.get(did, '—'), hotels.get(hid, '—'), hid)

def _avg(vals):
    vals = [v for v in vals if v is not None]
    return round(sum(vals) / len(vals), 1) if vals else None

@app.route('/dashboard/evaluations/stats')
def evaluations_stats():
    if not _auth():
        return redirect(url_for('login'))
    mode = request.args.get('mode', 'snapshot')
    if mode not in ('snapshot', 'current'): mode = 'snapshot'
    f_year = request.args.get('year', type=int)
    f_period = (request.args.get('period') or '').strip()
    f_hotel = request.args.get('hotel_id', type=int)
    hotels_m, deps_m = _maps()
    evs = latest_finalized()
    if f_year:   evs = [e for e in evs if e.year == f_year]
    if f_period: evs = [e for e in evs if e.period_label == f_period]
    if f_hotel:
        evs = [e for e in evs if _grp(e, mode, deps_m, hotels_m)[2] == f_hotel]
    # KPIs
    pcts = [e.score_pct for e in evs if e.score_pct is not None]
    kpi = {'n': len(evs), 'avg': _avg(pcts),
           'att': sum(1 for e in evs if e.score_pct is not None and e.score_pct < 60),
           'top': sum(1 for e in evs if e.score_pct is not None and e.score_pct >= 80)}
    # roll-up ανά τμήμα & ανά property
    bd, bh = {}, {}
    for e in evs:
        dn, hn, hid = _grp(e, mode, deps_m, hotels_m)
        bd.setdefault(dn, []).append(e.score_pct)
        bh.setdefault(hn, []).append(e.score_pct)
    by_dept = sorted([{'name': k, 'count': len(v), 'avg': _avg(v),
                       'att': sum(1 for x in v if x is not None and x < 60)} for k, v in bd.items()],
                     key=lambda r: (r['avg'] is None, -(r['avg'] or 0)))
    by_hotel = sorted([{'name': k, 'count': len(v), 'avg': _avg(v)} for k, v in bh.items()],
                      key=lambda r: (r['avg'] is None, -(r['avg'] or 0)))
    def _row(e):
        dn, hn, _ = _grp(e, mode, deps_m, hotels_m)
        return {'id': e.id, 'emp_id': e.employee_id, 'name': e.employee.full_name if e.employee else '—',
                'dept': dn, 'hotel': hn, 'pct': e.score_pct, 'band': e.band,
                'period': '%s %s' % (e.period_label or '', e.year or '')}
    ranked = sorted([e for e in evs if e.score_pct is not None], key=lambda e: e.score_pct, reverse=True)
    top = [_row(e) for e in ranked[:8]]
    attention = [_row(e) for e in ranked if e.score_pct < 60]
    years = sorted({e.year for e in latest_finalized() if e.year}, reverse=True)
    periods = sorted({e.period_label for e in latest_finalized() if e.period_label})
    hotels = Hotel.query.filter_by(is_active=True).order_by(Hotel.name).all()
    return render_template('eval_stats.html', kpi=kpi, by_dept=by_dept, by_hotel=by_hotel,
                           top=top, attention=attention, mode=mode, years=years, periods=periods,
                           hotels=hotels, f_year=f_year, f_period=f_period, f_hotel=f_hotel,
                           band_color=BAND_COLOR)

@app.route('/dashboard/evaluations/employee/<int:uid>')
def evaluations_employee(uid):
    if not _auth_eval():
        return redirect(url_for('login'))
    emp = User.query.get_or_404(uid)
    if not _hid_ok(getattr(emp, 'home_hotel_id', None)):
        return redirect(url_for('evaluations_list') + '?embed=1')
    evs = [e for e in latest_finalized() if e.employee_id == uid]
    evs.sort(key=lambda e: (e.year or 0, e.eval_date or date.min))
    # trend
    trend = [{'period': '%s %s' % (e.period_label or '', e.year or ''), 'pct': e.score_pct,
              'band': e.band, 'id': e.id} for e in evs]
    # per-criterion average (σε όλες τις αξιολογήσεις του)
    crit_sum, crit_cnt, crit_lab = {}, {}, {}
    for e in evs:
        for sc in e.scores:
            if sc.score is None: continue
            crit_sum[sc.criterion_id] = crit_sum.get(sc.criterion_id, 0) + sc.score
            crit_cnt[sc.criterion_id] = crit_cnt.get(sc.criterion_id, 0) + 1
    if evs and evs[-1].template:
        for c in evs[-1].template.criteria:
            crit_lab[c.id] = (c.label, c.grp)
    crit_avg = []
    for cid, tot in crit_sum.items():
        lab, grp = crit_lab.get(cid, ('(κριτήριο)', ''))
        crit_avg.append({'label': lab, 'grp': grp, 'avg': round(tot / crit_cnt[cid], 1)})
    crit_avg.sort(key=lambda r: r['avg'])
    hotels_m, deps_m = _maps()
    return render_template('eval_employee.html', emp=emp, trend=trend, crit_avg=crit_avg,
                           dept=deps_m.get(getattr(emp, 'department_id', None), '—'),
                           hotel=hotels_m.get(getattr(emp, 'home_hotel_id', None), '—'),
                           band_color=BAND_COLOR, avg=_avg([t['pct'] for t in trend]))

# ── Φ2: Διαχείριση / variants προτύπων ───────────────────────────────────────
def _criterion_used(cid):
    return EvalScore.query.filter_by(criterion_id=cid).count() > 0

@app.route('/dashboard/evaluations/templates')
def eval_templates():
    if not _auth():
        return redirect(url_for('login'))
    tmpls = EvalTemplate.query.order_by(EvalTemplate.scope, EvalTemplate.name).all()
    rows = []
    for t in tmpls:
        wsum = round(sum(c.weight or 0 for c in t.criteria), 3)
        used = Evaluation.query.filter_by(template_id=t.id).count()
        rows.append({'t': t, 'n': len(t.criteria), 'wsum': wsum, 'used': used})
    hotels = Hotel.query.filter_by(is_active=True).order_by(Hotel.name).all()
    return render_template('eval_templates.html', rows=rows, hotels=hotels,
                           ok=request.args.get('ok'), err=request.args.get('err'))

@app.route('/dashboard/evaluations/templates/<int:tid>', methods=['GET', 'POST'])
def eval_template_edit(tid):
    if not _auth():
        return redirect(url_for('login'))
    t = EvalTemplate.query.get_or_404(tid)
    if request.method == 'POST':
        fm = request.form
        t.name = (fm.get('name') or t.name).strip()[:80]
        t.scope = (fm.get('scope') or 'general').strip()[:40]
        _hid = fm.get('hotel_id', type=int); t.hotel_id = _hid or None
        try: t.scale_max = max(1, int(fm.get('scale_max') or 10))
        except ValueError: pass
        # update existing criteria
        for c in list(t.criteria):
            if fm.get('del_%d' % c.id) and not _criterion_used(c.id):
                db.session.delete(c); continue
            c.label = (fm.get('label_%d' % c.id) or c.label).strip()[:200]
            c.grp = (fm.get('grp_%d' % c.id) or c.grp or '').strip()[:40]
            try: c.weight = float(fm.get('weight_%d' % c.id) or 0)
            except ValueError: pass
            try: c.sort = int(fm.get('sort_%d' % c.id) or c.sort)
            except ValueError: pass
        # add new criteria (παράλληλες λίστες)
        labels = fm.getlist('new_label'); grps = fm.getlist('new_grp'); weights = fm.getlist('new_weight')
        base = (max([c.sort for c in t.criteria], default=-1)) + 1
        for i, lab in enumerate(labels):
            lab = (lab or '').strip()
            if not lab: continue
            try: w = float(weights[i]) if i < len(weights) and weights[i] else 0
            except ValueError: w = 0
            g = (grps[i].strip() if i < len(grps) and grps[i] else 'Βασικά κριτήρια')
            db.session.add(EvalCriterion(template_id=t.id, grp=g[:40], label=lab[:200], weight=w, sort=base + i, max_score=t.scale_max))
        db.session.commit()
        log_activity('eval_template_save', t.name)
        return redirect(url_for('eval_template_edit', tid=t.id) + '?embed=1&ok=1')
    wsum = round(sum(c.weight or 0 for c in t.criteria), 3)
    used_ids = {c.id for c in t.criteria if _criterion_used(c.id)}
    hotels = Hotel.query.filter_by(is_active=True).order_by(Hotel.name).all()
    return render_template('eval_template_edit.html', t=t, criteria=t.criteria, wsum=wsum,
                           used_ids=used_ids, ok=request.args.get('ok'), hotels=hotels)

@app.route('/dashboard/evaluations/templates/<int:tid>/clone', methods=['POST'])
def eval_template_clone(tid):
    if not _auth():
        return redirect(url_for('login'))
    src = EvalTemplate.query.get_or_404(tid)
    name = (request.form.get('name') or (src.name + ' (αντίγραφο)')).strip()[:80]
    scope = (request.form.get('scope') or 'general').strip()[:40]
    _hid = request.form.get('hotel_id', type=int)
    nt = EvalTemplate(name=name, scope=scope, scale_max=src.scale_max, is_active=True,
                      hotel_id=(_hid if _hid else src.hotel_id))
    db.session.add(nt); db.session.flush()
    for c in src.criteria:
        db.session.add(EvalCriterion(template_id=nt.id, grp=c.grp, label=c.label, weight=c.weight, sort=c.sort, max_score=c.max_score))
    db.session.commit()
    log_activity('eval_template_clone', '%s -> %s' % (src.name, name))
    return redirect(url_for('eval_template_edit', tid=nt.id) + '?embed=1')

@app.route('/dashboard/evaluations/templates/<int:tid>/toggle', methods=['POST'])
def eval_template_toggle(tid):
    if not _auth():
        return redirect(url_for('login'))
    t = EvalTemplate.query.get_or_404(tid)
    t.is_active = not bool(t.is_active); db.session.commit()
    return redirect(url_for('eval_templates') + '?embed=1')

@app.route('/dashboard/evaluations/templates/<int:tid>/delete', methods=['POST'])
def eval_template_delete(tid):
    if not _auth():
        return redirect(url_for('login'))
    t = EvalTemplate.query.get_or_404(tid)
    used = Evaluation.query.filter_by(template_id=tid).count()
    if used:
        return redirect(url_for('eval_templates') + '?embed=1&err=used')
    name = t.name
    db.session.delete(t); db.session.commit()
    log_activity('eval_template_delete', name)
    return redirect(url_for('eval_templates') + '?embed=1&ok=del')

# ── Φ2: Group roll-up matrix (κριτήρια × υπάλληλοι ανά τμήμα/περίοδο) ──────────
def _group_data(dept_id, year, period, mode):
    hotels_m, deps_m = _maps()
    evs = latest_finalized()
    if year:   evs = [e for e in evs if e.year == year]
    if period: evs = [e for e in evs if e.period_label == period]
    def did_of(e):
        return (getattr(e.employee, 'department_id', None) if (mode == 'current' and e.employee) else e.department_id)
    if dept_id:
        evs = [e for e in evs if did_of(e) == dept_id]
    evs.sort(key=lambda e: (e.employee.full_name if e.employee else ''))
    # γραμμές κριτηρίων: από το πρότυπο του 1ου eval (ή general), match με LABEL
    tmpl = (evs[0].template if evs and evs[0].template else
            EvalTemplate.query.filter_by(scope='general', is_active=True).first() or EvalTemplate.query.first())
    rows = [{'label': c.label, 'grp': c.grp, 'weight': c.weight} for c in (tmpl.criteria if tmpl else [])]
    cols = []
    for e in evs:
        by_lab = {}
        for sc in e.scores:
            cr = EvalCriterion.query.get(sc.criterion_id)
            if cr: by_lab[cr.label] = sc.score
        cols.append({'name': e.employee.full_name if e.employee else '—', 'id': e.id,
                     'emp_id': e.employee_id, 'pct': e.score_pct, 'band': e.band, 'by_lab': by_lab})
    # dept avg ανά κριτήριο
    for r in rows:
        vals = [c['by_lab'].get(r['label']) for c in cols]
        vals = [v for v in vals if v is not None]
        r['avg'] = round(sum(vals) / len(vals), 1) if vals else None
    return rows, cols, deps_m

@app.route('/dashboard/evaluations/group')
def evaluations_group():
    if not _auth():
        return redirect(url_for('login'))
    mode = request.args.get('mode', 'snapshot')
    if mode not in ('snapshot', 'current'): mode = 'snapshot'
    dept_id = request.args.get('department_id', type=int)
    year = request.args.get('year', type=int)
    period = (request.args.get('period') or '').strip()
    try:
        from schedule import Department
        depts = Department.query.order_by(Department.name).all()
    except Exception:
        depts = []
    if not dept_id and depts:
        # default: το πρώτο τμήμα που έχει αξιολογήσεις
        used = {e.department_id for e in latest_finalized() if e.department_id}
        for d in depts:
            if d.id in used: dept_id = d.id; break
    rows, cols, deps_m = _group_data(dept_id, year, period, mode)
    years = sorted({e.year for e in latest_finalized() if e.year}, reverse=True)
    periods = sorted({e.period_label for e in latest_finalized() if e.period_label})
    return render_template('eval_group.html', rows=rows, cols=cols, depts=depts, deps_m=deps_m,
                           dept_id=dept_id, year=year, period=period, mode=mode,
                           years=years, periods=periods, band_color=BAND_COLOR)

@app.route('/dashboard/evaluations/group/export.xlsx')
def evaluations_group_export():
    if not _auth():
        return redirect(url_for('login'))
    mode = request.args.get('mode', 'snapshot')
    dept_id = request.args.get('department_id', type=int)
    year = request.args.get('year', type=int)
    period = (request.args.get('period') or '').strip()
    rows, cols, deps_m = _group_data(dept_id, year, period, mode)
    import io as _io, openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Group'
    ws['A1'] = 'CONDIAN HOTELS — GROUP EVALUATION'; ws['A1'].font = Font(bold=True, size=13)
    ws['A2'] = 'Τμήμα: %s · %s %s' % (deps_m.get(dept_id, '—'), period or '', year or '')
    hr = 4
    ws.cell(hr, 1, 'ΚΡΙΤΗΡΙΟ').font = Font(bold=True)
    ws.cell(hr, 2, 'ΒΑΡΟΣ').font = Font(bold=True)
    for j, c in enumerate(cols):
        ws.cell(hr, 3 + j, c['name']).font = Font(bold=True)
    ws.cell(hr, 3 + len(cols), 'ΜΟ ΤΜΗΜΑΤΟΣ').font = Font(bold=True)
    r = hr + 1
    for row in rows:
        ws.cell(r, 1, row['label']); ws.cell(r, 2, row['weight'])
        for j, c in enumerate(cols):
            ws.cell(r, 3 + j, c['by_lab'].get(row['label']))
        ws.cell(r, 3 + len(cols), row['avg']); r += 1
    ws.cell(r + 1, 1, 'ΣΥΝΟΛΟ %').font = Font(bold=True)
    for j, c in enumerate(cols):
        ws.cell(r + 1, 3 + j, c['pct'])
    ws.column_dimensions['A'].width = 50
    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    log_activity('eval_group_export', '%s %s' % (deps_m.get(dept_id, ''), period))
    return Response(buf.read(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment; filename=group-evaluation.xlsx'})

# ── Φ3: Workflow (έγκριση/επιστροφή), goals, εκκρεμείς ────────────────────────
@app.route('/dashboard/evaluations/<int:eid>/approve', methods=['POST'])
def evaluation_approve(eid):
    if not _auth():
        return redirect(url_for('login'))
    ev = Evaluation.query.get_or_404(eid)
    ev.status = 'approved'; ev.updated_at = datetime.now(); db.session.commit()
    log_activity('evaluation_approve', ev.code or str(eid))
    return redirect(url_for('evaluation_view', eid=eid) + '?embed=1')

@app.route('/dashboard/evaluations/<int:eid>/return', methods=['POST'])
def evaluation_return(eid):
    if not _auth():
        return redirect(url_for('login'))
    ev = Evaluation.query.get_or_404(eid)
    ev.status = 'returned'; ev.updated_at = datetime.now(); db.session.commit()
    log_activity('evaluation_return', ev.code or str(eid))
    return redirect(url_for('evaluation_view', eid=eid) + '?embed=1')

@app.route('/dashboard/evaluations/goal/<int:gid>/toggle', methods=['POST'])
def evaluation_goal_toggle(gid):
    if not _auth_eval():
        return redirect(url_for('login'))
    g = EvalGoal.query.get_or_404(gid)
    _ev = Evaluation.query.get(g.evaluation_id)
    if _ev and not _hid_ok(_ev.hotel_id):
        return redirect(url_for('evaluations_list') + '?embed=1')
    g.done = not bool(g.done); db.session.commit()
    return redirect(url_for('evaluation_view', eid=g.evaluation_id) + '?embed=1')

@app.route('/dashboard/evaluations/pending')
def evaluations_pending():
    """Εκκρεμείς: ενεργοί εργαζόμενοι ΧΩΡΙΣ εγκεκριμένη αξιολόγηση για το επιλεγμένο έτος."""
    if not _auth_eval():
        return redirect(url_for('login'))
    year = request.args.get('year', type=int) or date.today().year
    f_period = (request.args.get('period') or '').strip()
    done_ids = set()
    for e in latest_finalized():
        if e.year == year and (not f_period or e.period_label == f_period):
            done_ids.add(e.employee_id)
    pend = [emp for emp in _employees() if emp['id'] not in done_ids]
    pend.sort(key=lambda x: (x['hotel'], x['department'], x['name']))
    periods = sorted({e.period_label for e in latest_finalized() if e.period_label})
    years = sorted({e.year for e in latest_finalized() if e.year} | {date.today().year}, reverse=True)
    return render_template('eval_pending.html', pend=pend, year=year, f_period=f_period,
                           periods=periods, years=years, done_n=len(done_ids))

# ── Περίοδοι αξιολόγησης (HR) ─────────────────────────────────────────────────
@app.route('/dashboard/evaluations/periods', methods=['GET', 'POST'])
def eval_periods():
    if not _auth():
        return redirect(url_for('login'))
    if request.method == 'POST':
        name = (request.form.get('name') or '').strip()
        ds = request.form.get('ref_date', '')
        if name:
            try:
                rd = datetime.strptime(ds, '%Y-%m-%d').date()
            except ValueError:
                rd = None
            mx = db.session.query(db.func.max(EvalPeriod.sort)).scalar() or 0
            db.session.add(EvalPeriod(name=name[:80], ref_date=rd, sort=mx + 1, is_active=True))
            db.session.commit(); log_activity('eval_period_add', name)
        return redirect(url_for('eval_periods') + '?embed=1')
    periods = EvalPeriod.query.order_by(EvalPeriod.sort, EvalPeriod.ref_date).all()
    used = {}
    for pr in periods:
        used[pr.id] = Evaluation.query.filter_by(period_label=pr.label).count()
    return render_template('eval_periods.html', periods=periods, used=used, today=date.today())

@app.route('/dashboard/evaluations/periods/<int:pid>/delete', methods=['POST'])
def eval_period_delete(pid):
    if not _auth():
        return redirect(url_for('login'))
    pr = EvalPeriod.query.get_or_404(pid)
    db.session.delete(pr); db.session.commit()
    return redirect(url_for('eval_periods') + '?embed=1')

@app.route('/dashboard/evaluations/periods/<int:pid>/toggle', methods=['POST'])
def eval_period_toggle(pid):
    if not _auth():
        return redirect(url_for('login'))
    pr = EvalPeriod.query.get_or_404(pid)
    pr.is_active = not bool(pr.is_active); db.session.commit()
    return redirect(url_for('eval_periods') + '?embed=1')

print('[evaluations] module loaded (Αξιολόγηση Προσωπικού Φ1)')
