# -*- coding: utf-8 -*-
"""
Εστία — Module «Μισθοδοσία» (Payroll) — Φάση 1 (θεμέλιο & μητρώο)
================================================================
Plug-in: `import payroll` από το ΤΕΛΟΣ του app.py (αφού οριστούν app/db/helpers,
ΠΡΙΝ το init_db() ώστε το create_all να πιάσει τους νέους πίνακες).
Spec: 02_MODULES_ESTIA/ΜΙΣΘΟΔΟΣΙΑ/00_SPEC.md

Φ1: Company, Agreement, EmployeePII, PayrollRates (seed 2026) · Hotel.company_id ·
admin-only (P-08) · όψεις μητρώο/καρτέλα/εταιρείες/συντελεστές · διαβάζει schedule.py.

Αποφάσεις Giannis (14/06): Λογιστήριο όψη = Epsilon import (αλήθεια)· Management = τι
πληρώνεται· ωρομίσθιο = ημερομίσθιο÷8 (Α-02)· μενού νέα ομάδα «Οικονομικά» (Α-08).
ΔΕΝ αλλάζει layout (D-09). Μηχανή/Run/Line/Forecast/Outputs = Φ2+.
"""
import os, re, json, unicodedata
from datetime import datetime, date
from flask import request, redirect, url_for, render_template, session
from app import (app, db, current_user, is_admin, log_activity,
                 Hotel, User, Setting, role_rank, ROLE_RANK,
                 notify, notify_admins)

try:
    from schedule import EmploymentProfile, monthly_settlement  # L1/L2 source (S-13)
except Exception:
    EmploymentProfile = None
    monthly_settlement = None


# ── Αυθεντικός χάρτης νομικών οντοτήτων (Giannis 14/06) ───────────────────────
# Εταιρεία (ΑΦΜ) -> ξενοδοχεία, με Κωδικό Υποκαταστήματος (ΥΠΟΚ) ανά ξενοδοχείο.
COMPANY_INFO = {   # company_code -> (legal_name, ΑΦΜ)
    'GIAN': ('Γ. ΓΙΑΝΝΟΥΛΑΚΗΣ - Α. ΓΙΑΝΝΟΥΛΑΚΗ Α.Ε.', '094082480'),
    'SERG': ('ΑΦΟΙ ΣΕΡΓΙΟΥ Α.Ε.',                      '094084694'),
    'PISK': ('ΠΙΣΚΟΠΙΑΝΟ Α.Ε.',                        '094121123'),
}
HOTEL_INFO = {   # hotel_code -> (name, company_code, ypok_code, ypok_desc)
    'AST': ('Asterias Village Resort',   'GIAN', '0002', 'ΥΠΟΚ/ΜΑ 2 ΑΣΤΕΡΙΑΣ'),
    'CNT': ('Central Hersonissos Hotel', 'GIAN', '0001', 'ΥΠΟΚ/ΜΑ 1 CENTRAL'),
    'IRO': ('Iro Hotel',                 'GIAN', '0000', 'Κεντρικό'),
    'SRG': ('Sergios Hotel',             'SERG', '0000', 'Κεντρικό'),
    'PSV': ('Piskopiano Village Resort', 'PISK', '0000', 'Κεντρικό'),
    'MINI':('Mini Market',               'GIAN', '0003', 'Mini Market (μη λειτουργικό)'),
}
COMPANY_CODE = {   # filename/hotel prefix -> εταιρεία (CND/HQ ασφαλίζεται στη Γιαννουλάκη)
    'AST': 'GIAN', 'CNT': 'GIAN', 'IRO': 'GIAN', 'CND': 'GIAN',
    'SRG': 'SERG', 'PSV': 'PISK',
}
HOTEL_NAME_CODE = {   # όνομα ξενοδοχείου -> hotel code
    'asterias': 'AST', 'central': 'CNT', 'iro': 'IRO', 'ηρω': 'IRO',
    'sergios': 'SRG', 'piskopiano': 'PSV', 'minimarket': 'MINI', 'mini': 'MINI',
}
HOTEL_YPOK = {hc: info[2] for hc, info in HOTEL_INFO.items()}
YPOK_TO_HOTEL = {(info[1], info[2]): hc for hc, info in HOTEL_INFO.items()}  # (company,ypok)->hotel
# Managerial κέντρο κόστους ανά (εταιρεία,ΥΠΟΚ) — «Κεντρικό» Γιαννουλάκη -> Headquarter (CND)
YPOK_TO_CC = {('GIAN','0002'):'AST', ('GIAN','0001'):'CNT', ('GIAN','0000'):'CND', ('GIAN','0003'):'CND',
              ('SERG','0000'):'SRG', ('PISK','0000'):'PSV'}

# Managerial «κέντρο κόστους / μονάδα» (εσωτερικό· μπορεί να διαφέρει από το νομικό ξεν.)
# π.χ. κάποιος ασφαλισμένος στο Iro αλλά ανήκει στα Κεντρικά Γραφεία (CND).
COST_CENTERS = {
    'AST': 'Asterias Village Resort', 'CNT': 'Central Hersonissos Hotel', 'IRO': 'Iro Hotel',
    'SRG': 'Sergios Hotel', 'PSV': 'Piskopiano Village Resort', 'CND': 'Headquarter',
}


def _acc(s):
    return ''.join(c for c in unicodedata.normalize('NFD', str(s)) if unicodedata.category(c) != 'Mn')

def _norm(s):
    return re.sub(r'[^a-zα-ω0-9]', '', _acc(s).strip().lower()) if s else ''

def hotel_code(hotel):
    if not hotel:
        return None
    n = _norm(hotel.name)
    for key, code in HOTEL_NAME_CODE.items():
        if _norm(key) in n:
            return code
    return None


# ── ΜΟΝΤΕΛΑ ───────────────────────────────────────────────────────────────────
class Company(db.Model):
    id           = db.Column(db.Integer, primary_key=True)
    code         = db.Column(db.String(12), unique=True, index=True)
    legal_name   = db.Column(db.String(160), nullable=False)
    vat          = db.Column(db.String(20))
    subunit_code = db.Column(db.String(20))
    active       = db.Column(db.Boolean, default=True)


class Agreement(db.Model):
    """Συμφωνία με ιστορικό ισχύος (SPEC §4.2). Φ1: δομή + προαιρετική καταγραφή."""
    id                 = db.Column(db.Integer, primary_key=True)
    user_id            = db.Column(db.Integer, db.ForeignKey('user.id'), index=True, nullable=False)
    effective_from     = db.Column(db.Date, default=date.today)
    effective_to       = db.Column(db.Date)
    agreement_type     = db.Column(db.String(20), default='Μηνιαίος')
    agreed_amount      = db.Column(db.Float)
    folder_fixed       = db.Column(db.Float, default=0.0)
    hour_wage_override = db.Column(db.Float)
    channels_json      = db.Column(db.Text)
    note               = db.Column(db.String(200))
    created_at         = db.Column(db.DateTime, default=datetime.now)
    created_by         = db.Column(db.Integer, db.ForeignKey('user.id'))


class EmployeePII(db.Model):
    """Ευαίσθητα στοιχεία — ξεχωριστός πίνακας, admin-gated (P-07/GDPR)."""
    id               = db.Column(db.Integer, primary_key=True)
    user_id          = db.Column(db.Integer, db.ForeignKey('user.id'), unique=True, nullable=False)
    afm              = db.Column(db.String(12))
    amka             = db.Column(db.String(12))
    ika_am           = db.Column(db.String(15))
    last_name        = db.Column(db.String(80))
    first_name       = db.Column(db.String(80))
    father_name      = db.Column(db.String(80))
    ergani_specialty = db.Column(db.String(120))
    contract_type    = db.Column(db.String(40))
    employment_kind  = db.Column(db.String(40))
    bank_name        = db.Column(db.String(60))
    bank_iban        = db.Column(db.String(34))
    hired_at         = db.Column(db.Date)
    left_at          = db.Column(db.Date)
    locked           = db.Column(db.Boolean, default=False)   # v12.56: κλειδωμένο μητρώο (πηγή=Epsilon)
    cost_center      = db.Column(db.String(8))   # v12.59: managerial μονάδα (AST/CNT/IRO/SRG/PSV/CND)
    emp_code         = db.Column(db.String(12), index=True)  # v12.60: Κωδ. Εργαζομένου (master Excel)
    updated_at       = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    updated_by       = db.Column(db.Integer, db.ForeignKey('user.id'))


class PayrollRates(db.Model):
    """Συντελεστές ανά έτος (admin-editable). Seed 2026 ΓΙΑ ΕΚΤΙΜΗΣΗ/ΠΡΟΒΛΕΨΗ.
       Η αλήθεια των νόμιμων καθαρών έρχεται από Epsilon import."""
    id                    = db.Column(db.Integer, primary_key=True)
    year                  = db.Column(db.Integer, unique=True, index=True, nullable=False)
    efka_employee_pct     = db.Column(db.Float, default=13.87)
    efka_employer_pct     = db.Column(db.Float, default=22.29)
    efka_aux_employee_pct = db.Column(db.Float, default=3.25)
    efka_aux_employer_pct = db.Column(db.Float, default=3.25)
    tax_free_threshold    = db.Column(db.Float, default=8636.0)
    tax_brackets_json     = db.Column(db.Text)
    digital_fee           = db.Column(db.Float, default=0.0)
    note                  = db.Column(db.String(200))
    valid_from            = db.Column(db.Date)
    valid_to              = db.Column(db.Date)


TAX_BRACKETS_2026 = [[10000, 9], [20000, 22], [30000, 28], [40000, 36], [None, 44]]


# ── AUTH (admin-only — P-08) ──────────────────────────────────────────────────
def _padmin():
    return ('user_id' in session) and is_admin()


# ── MIGRATION (μη καταστροφικό) ───────────────────────────────────────────────
def ensure_payroll_columns():
    with app.app_context():
        try:
            from app import _add_col
            _add_col('hotel', 'company_id', 'company_id INTEGER')
            _add_col('hotel', 'ypok_code', "ypok_code VARCHAR(8)")
            _add_col('legal_net_import', 'period_kind', "period_kind VARCHAR(24)")
            _add_col('payroll_line', 'extra_legal_net', 'extra_legal_net FLOAT')
            _add_col('payroll_line', 'extra_employer_cost', 'extra_employer_cost FLOAT')
            _add_col('payroll_line', 'extras_json', 'extras_json TEXT')
            _add_col('payroll_line', 'paid', 'paid FLOAT')
            _add_col('payroll_run', 'approved_by', 'approved_by INTEGER')
            _add_col('payroll_run', 'approved_at', 'approved_at DATETIME')
            _add_col('employee_pii', 'locked', 'locked BOOLEAN')
            _add_col('employee_pii', 'cost_center', "cost_center VARCHAR(8)")
            _add_col('employee_pii', 'emp_code', "emp_code VARCHAR(12)")
            _add_col('employee_pii', 'last_name', "last_name VARCHAR(80)")
            _add_col('employee_pii', 'first_name', "first_name VARCHAR(80)")
        except Exception as e:
            print('ensure_payroll_columns skipped:', e)


# ── SEED (idempotent) ─────────────────────────────────────────────────────────
def seed_payroll():
    with app.app_context():
        try:
            # 1) Εταιρείες (αυθεντικός χάρτης + ΑΦΜ)
            for ccode, (legal, vat) in COMPANY_INFO.items():
                c = Company.query.filter_by(code=ccode).first()
                if not c:
                    c = Company(code=ccode, legal_name=legal, vat=vat); db.session.add(c)
                else:
                    c.legal_name = legal
                    if not c.vat:
                        c.vat = vat
            db.session.commit()
            # καταργημένες εταιρείες (PLM/ΦΥΤΩΡΙΑ, CND) -> ανενεργές
            for old_code in ('FYTO', 'CND'):
                oc = Company.query.filter_by(code=old_code).first()
                if oc:
                    oc.active = False
            db.session.commit()

            # Mini Market (ΥΠΟΚ 0003) — μη λειτουργικό (is_active=False), μόνο μισθοδοσία/νομικά
            if not Hotel.query.filter(Hotel.name == 'Mini Market').first():
                db.session.add(Hotel(name='Mini Market', is_active=False)); db.session.commit()
            # 2) Ξενοδοχεία -> εταιρεία + Κωδικός ΥΠΟΚ (αυθεντικά)
            comp_by_code = {c.code: c for c in Company.query.all()}
            for h in Hotel.query.all():
                hc = hotel_code(h)
                info = HOTEL_INFO.get(hc) if hc else None
                if not info:
                    continue
                cc = info[1]
                if cc in comp_by_code:
                    h.company_id = comp_by_code[cc].id
                if hasattr(h, 'ypok_code'):
                    h.ypok_code = info[2]
            db.session.commit()

            if not PayrollRates.query.filter_by(year=2026).first():
                db.session.add(PayrollRates(
                    year=2026,
                    tax_brackets_json=json.dumps(TAX_BRACKETS_2026, ensure_ascii=False),
                    note='Seed εκτίμησης (η αλήθεια από Epsilon import). Admin-editable.',
                    valid_from=date(2026, 1, 1), valid_to=date(2026, 12, 31)))
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print('seed_payroll skipped:', e)


# ── HELPERS ───────────────────────────────────────────────────────────────────
def _company_for_hotel(hotel_id):
    h = Hotel.query.get(hotel_id) if hotel_id else None
    if h and getattr(h, 'company_id', None):
        return Company.query.get(h.company_id)
    return None

def _years_available():
    yrs = [r[0] for r in db.session.query(LegalNetImport.year).distinct().all() if r[0]]
    return sorted(set(int(y) for y in yrs))

def _legal_net_map():
    """{afm: {year: {'reg','tot','months','extra'}}} από LegalNetImport (Epsilon).
    reg = τακτικοί μήνες (period_kind=='monthly')· tot = + δώρα/άδεια· extra = δώρα/άδεια."""
    out = {}
    for li in LegalNetImport.query.all():
        if not li.afm or not li.year:
            continue
        d = out.setdefault(str(li.afm), {}).setdefault(int(li.year),
                {'reg': 0.0, 'tot': 0.0, 'months': 0, 'extra': 0.0})
        net = li.net_legal or 0.0
        d['tot'] += net
        if (li.period_kind or 'monthly') == 'monthly':
            d['reg'] += net; d['months'] += 1
        else:
            d['extra'] += net
    return out

def _employees(status='active'):
    q = User.query
    if status == 'inactive':
        q = q.filter(User.employment_active == False)
    elif status == 'all':
        pass
    else:
        q = q.filter((User.employment_active == True) | (User.employment_active.is_(None)))
    users = q.all()
    out = []
    for u in users:
        prof = EmploymentProfile.query.filter_by(user_id=u.id).first() if EmploymentProfile else None
        pii = EmployeePII.query.filter_by(user_id=u.id).first()
        has_hr = bool(prof or pii or getattr(u, 'home_hotel_id', None) or getattr(u, 'department_id', None))
        if not has_hr:
            continue
        hid = getattr(u, 'home_hotel_id', None)
        comp = _company_for_hotel(hid)
        hotel = Hotel.query.get(hid) if hid else None
        cc = pii.cost_center if (pii and pii.cost_center) else None
        ma = current_assignment(u.id)
        out.append({'user': u, 'profile': prof, 'pii': pii, 'company': comp,
                    'hotel_name': (hotel.name if hotel else ''), 'hotel_id': hid,
                    'dept_id': getattr(u, 'department_id', None), 'mgmt': ma,
                    'mgmt_dept': (ma.department if ma else None), 'mgmt_pos': (ma.position if ma else None),
                    'mgmt_unit': (ma.unit if ma else None),
                    'cost_center': cc, 'cost_center_label': (COST_CENTERS.get(cc, cc) if cc else '')})
    out.sort(key=lambda r: (r['company'].legal_name if r['company'] else 'ω', r['user'].full_name or ''))
    return out


# ── ROUTES ────────────────────────────────────────────────────────────────────
@app.route('/dashboard/payroll')
def payroll_home():
    if not _padmin():
        return redirect(url_for('login'))
    status = request.args.get('status') or 'active'
    if status not in ('active', 'inactive', 'all'):
        status = 'active'
    allrows = _employees(status)
    companies = Company.query.filter_by(active=True).order_by(Company.legal_name).all()
    # πλήθη ανά κατάσταση (για τα κουμπιά)
    n_active = len(_employees('active')); n_inactive = len(_employees('inactive'))
    n_all = n_active + n_inactive
    # KPIs στο σύνολο της τρέχουσας κατάστασης
    n_emp = len(allrows)
    n_agree = sum(1 for r in allrows if r['profile'] and r['profile'].agreement_amount)
    n_pii = sum(1 for r in allrows if r['pii'] and r['pii'].afm)
    # Χωρίς κλειδί: λείπει ΑΦΜ Ή Κωδ. Εργαζομένου (να φαίνονται για διόρθωση)
    def _nokey(r):
        pii = r['pii']
        return (not (pii and pii.afm)) or (not (pii and pii.emp_code))
    n_nokey = sum(1 for r in allrows if _nokey(r))
    # έτος + ποσά (από Epsilon/LegalNetImport)
    net_map = _legal_net_map()
    years = _years_available()
    year = request.args.get('year', type=int)
    if year is None:
        year = (years[-1] if years else 0)
    if year != 0 and year not in years:
        year = (years[-1] if years else 0)
    # φίλτρα
    company_id = request.args.get('company_id', type=int)
    hotel_id = request.args.get('hotel_id', type=int)
    cc = request.args.get('cc')
    flt = request.args.get('filter')
    rows = allrows
    if company_id: rows = [r for r in rows if r['company'] and r['company'].id == company_id]
    if hotel_id: rows = [r for r in rows if r.get('hotel_id') == hotel_id]
    if cc: rows = [r for r in rows if r.get('cost_center') == cc]
    if flt == 'agree':   rows = [r for r in rows if r['profile'] and r['profile'].agreement_amount]
    elif flt == 'noagree': rows = [r for r in rows if not (r['profile'] and r['profile'].agreement_amount)]
    elif flt == 'pii':   rows = [r for r in rows if r['pii'] and r['pii'].afm]
    elif flt == 'nopii': rows = [r for r in rows if not (r['pii'] and r['pii'].afm)]
    elif flt == 'nokey':  rows = [r for r in rows if _nokey(r)]
    # ποσά ανά γραμμή για το επιλεγμένο έτος (0 = όλα τα έτη)
    for r in rows:
        afm = r['pii'].afm if (r['pii'] and r['pii'].afm) else None
        info = net_map.get(str(afm), {}) if afm else {}
        if year == 0:
            reg = sum(v['reg'] for v in info.values()); tot = sum(v['tot'] for v in info.values())
            mon = sum(v['months'] for v in info.values())
        else:
            d = info.get(year)
            reg = d['reg'] if d else 0.0; tot = d['tot'] if d else 0.0; mon = d['months'] if d else 0
        r['net_reg'] = reg; r['net_tot'] = tot; r['net_months'] = mon
    hotels = Hotel.query.order_by(Hotel.name).all()
    rates = PayrollRates.query.filter_by(year=2026).first()
    log_activity('payroll_view', 'μητρώο')
    return render_template('payroll_home.html',
        rows=rows, companies=companies, hotels=hotels, n_emp=n_emp, n_agree=n_agree, n_pii=n_pii,
        company_id=company_id, hotel_id=hotel_id, cc=cc, cost_centers=COST_CENTERS,
        flt=flt, total_shown=len(rows), n_nokey=n_nokey, rates=rates, is_admin=is_admin(),
        status=status, n_active=n_active, n_inactive=n_inactive, n_all=n_all,
        years=years, year=year)


@app.route('/dashboard/payroll/companies', methods=['GET', 'POST'])
def payroll_companies():
    if not _padmin():
        return redirect(url_for('login'))
    if request.method == 'POST':
        cid = request.form.get('id', type=int)
        c = Company.query.get(cid) if cid else None
        if c:
            c.legal_name   = (request.form.get('legal_name') or c.legal_name).strip()
            c.vat          = (request.form.get('vat') or '').strip() or None
            c.subunit_code = (request.form.get('subunit_code') or '').strip() or None
            c.active       = request.form.get('active') == '1'
            db.session.commit()
            log_activity('payroll_company_edit', c.legal_name)
        return redirect(url_for('payroll_companies'))
    companies = Company.query.order_by(Company.legal_name).all()
    hcount = {}
    for h in Hotel.query.all():
        cid = getattr(h, 'company_id', None)
        if cid:
            hcount.setdefault(cid, []).append(h.name)
    return render_template('payroll_companies.html', companies=companies, hcount=hcount, is_admin=is_admin())


@app.route('/dashboard/payroll/rates', methods=['GET', 'POST'])
def payroll_rates():
    if not _padmin():
        return redirect(url_for('login'))
    r = PayrollRates.query.filter_by(year=2026).first()
    if not r:
        r = PayrollRates(year=2026, tax_brackets_json=json.dumps(TAX_BRACKETS_2026, ensure_ascii=False))
        db.session.add(r); db.session.commit()
    if request.method == 'POST':
        def _f(name, cur):
            v = request.form.get(name)
            try:
                return float(v) if v not in (None, '') else cur
            except Exception:
                return cur
        r.efka_employee_pct     = _f('efka_employee_pct', r.efka_employee_pct)
        r.efka_employer_pct     = _f('efka_employer_pct', r.efka_employer_pct)
        r.efka_aux_employee_pct = _f('efka_aux_employee_pct', r.efka_aux_employee_pct)
        r.efka_aux_employer_pct = _f('efka_aux_employer_pct', r.efka_aux_employer_pct)
        r.tax_free_threshold    = _f('tax_free_threshold', r.tax_free_threshold)
        r.digital_fee           = _f('digital_fee', r.digital_fee)
        r.note                  = (request.form.get('note') or '').strip() or r.note
        db.session.commit()
        log_activity('payroll_rates_edit', '2026')
        return redirect(url_for('payroll_rates'))
    try:
        brackets = json.loads(r.tax_brackets_json or '[]')
    except Exception:
        brackets = []
    return render_template('payroll_rates.html', r=r, brackets=brackets, is_admin=is_admin())


@app.route('/dashboard/payroll/employee/<int:uid>', methods=['GET', 'POST'])
def payroll_employee(uid):
    if not _padmin():
        return redirect(url_for('login'))
    u = User.query.get_or_404(uid)
    prof = EmploymentProfile.query.filter_by(user_id=uid).first() if EmploymentProfile else None
    pii = EmployeePII.query.filter_by(user_id=uid).first()
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'pii':
            if not pii:
                pii = EmployeePII(user_id=uid); db.session.add(pii)
            for f in ('afm', 'amka', 'ika_am', 'father_name', 'ergani_specialty',
                      'contract_type', 'employment_kind', 'bank_name', 'bank_iban'):
                setattr(pii, f, (request.form.get(f) or '').strip() or None)
            for f in ('hired_at', 'left_at'):
                v = request.form.get(f)
                try:
                    setattr(pii, f, datetime.strptime(v, '%Y-%m-%d').date() if v else None)
                except Exception:
                    pass
            cu = current_user()
            pii.updated_by = cu.id if cu else None
            db.session.commit()
            log_activity('payroll_pii_edit', u.full_name or u.username)
        elif action == 'agreement' and prof is not None:
            amt = request.form.get('agreement_amount')
            try:
                prof.agreement_amount = float(amt) if amt not in (None, '') else prof.agreement_amount
            except Exception:
                pass
            prof.agreement_type = request.form.get('agreement_type') or prof.agreement_type
            db.session.commit()
            log_activity('payroll_agreement_edit', u.full_name or u.username)
        return redirect(url_for('payroll_employee', uid=uid))
    comp = _company_for_hotel(getattr(u, 'home_hotel_id', None))
    hotel = Hotel.query.get(u.home_hotel_id) if getattr(u, 'home_hotel_id', None) else None
    agreements = (Agreement.query.filter_by(user_id=uid)
                  .order_by(Agreement.effective_from.desc()).all())
    # Οικονομικά (Λογιστήριο = Epsilon): καθαρά ανά έτος, τακτικά vs δώρα/άδεια
    fin = []
    afm = pii.afm if (pii and pii.afm) else None
    if afm:
        nm = _legal_net_map().get(str(afm), {})
        for y in sorted(nm.keys()):
            d = nm[y]
            avg = (d['reg'] / d['months']) if d['months'] else 0.0
            fin.append({'year': y, 'months': d['months'], 'reg': d['reg'],
                        'extra': d['extra'], 'tot': d['tot'], 'avg': avg})
    fin_tot = {'reg': sum(f['reg'] for f in fin), 'extra': sum(f['extra'] for f in fin),
               'tot': sum(f['tot'] for f in fin), 'months': sum(f['months'] for f in fin)}
    fin_months = {}
    if afm:
        for li in LegalNetImport.query.filter_by(afm=str(afm)).all():
            if (li.period_kind or 'monthly') == 'monthly' and li.year and li.month:
                fin_months.setdefault(int(li.year), {})
                fin_months[int(li.year)][int(li.month)] = fin_months[int(li.year)].get(int(li.month), 0.0) + (li.net_legal or 0.0)
    import people as PPL
    assignments = assignments_for(uid)
    events = PPL.events_for(uid)
    flags = PPL.open_flags_for(uid)
    ev_labels = PPL.EVENT_LABELS; flag_labels = PPL.FLAG_LABELS
    merge_cands = []
    if pii and pii.last_name:
        for p2 in EmployeePII.query.filter(EmployeePII.last_name == pii.last_name).all():
            if p2.user_id == uid: continue
            if PPL.is_dismissed(uid, p2.user_id): continue
            u2 = User.query.get(p2.user_id)
            if u2 and (u2.employment_active is not False):
                merge_cands.append(u2)
    return render_template('payroll_employee.html',
        u=u, prof=prof, pii=pii, comp=comp, hotel=hotel, agreements=agreements,
        fin=fin, fin_tot=fin_tot, fin_months=fin_months, month_gr=_MONTH_GR,
        assignments=assignments, events=events, flags=flags,
        ev_labels=ev_labels, flag_labels=flag_labels, merge_cands=merge_cands,
        is_admin=is_admin())


# ══════════════════════════════════════════════════════════════════════════════
# ΦΑΣΗ 2 — Μηχανή υπολογισμού + δύο όψεις + Epsilon import (Λογιστήριο = αλήθεια)
# ══════════════════════════════════════════════════════════════════════════════
import hashlib, io, zipfile
try:
    import openpyxl
except Exception:
    openpyxl = None

def _safe_load_xlsx(raw):
    """Ανθεκτικό άνοιγμα .xlsx — διορθώνει Epsilon exports με cellStyle χωρίς όνομα
    (openpyxl TypeError: name should be str but value is None)."""
    try:
        return openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except TypeError:
        zin = zipfile.ZipFile(io.BytesIO(raw)); buf = io.BytesIO()
        zout = zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED)
        for it in zin.namelist():
            data = zin.read(it)
            if it == 'xl/styles.xml':
                txt = data.decode('utf-8', 'replace')
                def _fix(m):
                    tag = m.group(0)
                    return tag if 'name=' in tag else tag[:-2] + ' name="Normal_x"/>'
                txt = re.sub(r'<cellStyle [^>]*/>', _fix, txt)
                data = txt.encode('utf-8')
            zout.writestr(it, data)
        zout.close(); buf.seek(0)
        return openpyxl.load_workbook(buf, data_only=True)

MONTHS_EL2 = ['', 'Ιανουάριος','Φεβρουάριος','Μάρτιος','Απρίλιος','Μάιος','Ιούνιος',
              'Ιούλιος','Αύγουστος','Σεπτέμβριος','Οκτώβριος','Νοέμβριος','Δεκέμβριος']
_EPSILON_MONTH = {}
for _i, _n in enumerate(MONTHS_EL2):
    if _i: _EPSILON_MONTH[_norm(_n)] = _i
_EPSILON_MONTH.update({_norm(k): v for k, v in {
    'ΙΑΝ':1,'ΦΕΒ':2,'ΜΑΡ':3,'ΑΠΡ':4,'ΜΑΙ':5,'ΜΑΪΟΣ':5,'ΙΟΥΝ':6,'ΙΟΥΛ':7,'ΑΥΓ':8,
    'ΣΕΠ':9,'ΟΚΤ':10,'ΝΟΕ':11,'ΔΕΚ':12,'JANUARY':1,'FEBRUARY':2,'MARCH':3,'APRIL':4,
    'MAY':5,'JUNE':6,'JULY':7,'AUGUST':8,'SEPTEMBER':9,'OCTOBER':10,'NOVEMBER':11,'DECEMBER':12,
}.items()})

RUN_STATUS = ('draft', 'calculated', 'verified', 'locked', 'paid')
RUN_LABELS = {'draft':'Πρόχειρο','calculated':'Υπολογίστηκε','verified':'Εγκρίθηκε (ΓΔ)',
              'locked':'Κλειδωμένο','paid':'Πληρωμένο'}


class PayrollRun(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    company_id    = db.Column(db.Integer, db.ForeignKey('company.id'), index=True, nullable=False)
    year          = db.Column(db.Integer, nullable=False)
    month         = db.Column(db.Integer, nullable=False)
    status        = db.Column(db.String(12), default='draft')
    rates_version = db.Column(db.Integer)
    note          = db.Column(db.String(200))
    created_by    = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_at    = db.Column(db.DateTime, default=datetime.now)
    locked_at     = db.Column(db.DateTime)
    approved_by   = db.Column(db.Integer, db.ForeignKey('user.id'))
    approved_at   = db.Column(db.DateTime)
    __table_args__ = (db.UniqueConstraint('company_id', 'year', 'month', name='uq_payrollrun'),)


class PayrollLine(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    run_id        = db.Column(db.Integer, db.ForeignKey('payroll_run.id'), index=True, nullable=False)
    user_id       = db.Column(db.Integer, db.ForeignKey('user.id'), index=True, nullable=False)
    # ώρες/ημέρες (από monthly_settlement)
    work_days     = db.Column(db.Integer, default=0)
    sundays       = db.Column(db.Integer, default=0)
    holidays_worked = db.Column(db.Integer, default=0)
    extra_hours   = db.Column(db.Float, default=0.0)
    repo          = db.Column(db.Integer, default=0)
    total_days    = db.Column(db.Integer, default=0)
    elsewhere_days= db.Column(db.Integer, default=0)
    # Management όψη
    gross_agreement = db.Column(db.Float, default=0.0)
    extra_pay     = db.Column(db.Float, default=0.0)
    gross_total   = db.Column(db.Float, default=0.0)
    bank_target   = db.Column(db.Float)
    bonus         = db.Column(db.Float, default=0.0)
    bank_total    = db.Column(db.Float, default=0.0)
    folder_total  = db.Column(db.Float, default=0.0)
    in_hand       = db.Column(db.Float, default=0.0)
    # Λογιστήριο όψη (εκτίμηση) + αλήθεια Epsilon
    efka_employee = db.Column(db.Float, default=0.0)
    fmy           = db.Column(db.Float, default=0.0)
    net_calc      = db.Column(db.Float, default=0.0)
    employer_cost = db.Column(db.Float, default=0.0)
    net_legal     = db.Column(db.Float)        # από Epsilon
    employer_cost_legal = db.Column(db.Float)
    net_diff      = db.Column(db.Float)
    extra_legal_net = db.Column(db.Float, default=0.0)      # δώρα/άδεια καθαρά (Epsilon)
    extra_employer_cost = db.Column(db.Float, default=0.0)  # κόστος εργοδ. δώρων/άδειας
    extras_json   = db.Column(db.Text)                      # JSON [{kind,net,cost}]
    paid          = db.Column(db.Float, default=0.0)        # πληρωμένο (παρακολούθηση)
    note          = db.Column(db.String(200))


class LegalNetImport(db.Model):
    """Νόμιμα καθαρά από Epsilon (η αλήθεια του Λογιστηρίου). Idempotent ανά εταιρεία/μήνα/ΑΦΜ."""
    id            = db.Column(db.Integer, primary_key=True)
    company_id    = db.Column(db.Integer, db.ForeignKey('company.id'), index=True)
    user_id       = db.Column(db.Integer, db.ForeignKey('user.id'), index=True)
    year          = db.Column(db.Integer, index=True)
    month         = db.Column(db.Integer, index=True)
    afm           = db.Column(db.String(12), index=True)
    emp_name      = db.Column(db.String(120))
    gross_legal   = db.Column(db.Float)
    efka_employee_legal = db.Column(db.Float)
    fmy_legal     = db.Column(db.Float)
    net_legal     = db.Column(db.Float)
    employer_cost_legal = db.Column(db.Float)
    period_kind   = db.Column(db.String(24), default='monthly')  # monthly/Δώρο Πάσχα/Επίδομα Αδείας...
    import_hash   = db.Column(db.String(40), unique=True, index=True)
    source_file   = db.Column(db.String(160))
    created_at    = db.Column(db.DateTime, default=datetime.now)


# ── EPSILON PARSER ────────────────────────────────────────────────────────────
def _epsilon_headers(ws):
    """Χάρτης normalized-header -> column index από την 1η γραμμή."""
    h = {}
    for c in range(1, ws.max_column + 1):
        h[_norm(ws.cell(1, c).value)] = c
    return h

def _ypok4(v):
    """Κωδικός ΥΠΟΚ -> 4ψήφιο string (π.χ. 2 -> '0002')."""
    if v in (None, ''):
        return None
    sv = str(v).strip()
    try:
        return '%04d' % int(float(sv))
    except Exception:
        return sv

def parse_epsilon(wb):
    """Επιστρέφει λίστα dict ανά εργαζόμενο από workbook Epsilon (44 στήλες)."""
    ws = wb.worksheets[0]
    H = _epsilon_headers(ws)
    def col(*names, default=None):
        for n in names:
            c = H.get(_norm(n))
            if c: return c
        return default
    c_epon=col('Επώνυμο'); c_onoma=col('Όνομα'); c_afm=col('ΑΦΜ')
    c_amka=col('ΑΜΚΑ'); c_ika=col('Α.Μ. ΙΚΑ','ΑΜ ΙΚΑ'); c_pat=col('Όνομα Πατρός')
    c_spec=col('Ειδικότητα'); c_kind=col('Είδος Εργάζ','Είδος Εργαζ'); c_contract=col('Διάρκεια Σύμβασης')
    c_sub=col('Περιγραφή Υποκαταστήματος'); c_dept=col('Περιγραφή Τμήματος'); c_comp=col('Εταιρεία')
    c_subcode=col('Κωδικός Υποκαταστήματος')
    c_bank=col('Τράπεζα'); c_period=col('Περίοδος'); c_year=col('Έτος')
    c_gross=col('Συν.Αποδ.','Συν.Αποδ'); c_emain=col('Εισφ. Εργάζ. Κύριου Ταμείου')
    c_eaux=col('Εισφ. Εργάζ. Επικ. Ταμείου'); c_fmy=col('Φ.Μ.Υ','ΦΜΥ')
    c_net=col('Καθαρές Αποδοχές'); c_cost=col('Συνολικό Κόστος')
    rows=[]
    for r in range(2, ws.max_row + 1):
        epon = ws.cell(r, c_epon).value if c_epon else None
        if not epon: continue
        onoma = ws.cell(r, c_onoma).value if c_onoma else ''
        per = ws.cell(r, c_period).value if c_period else None
        yr  = ws.cell(r, c_year).value if c_year else None
        month = _EPSILON_MONTH.get(_norm(per)) if per else None
        period_raw = str(per).strip() if per else ''
        try: yr = int(yr) if yr else None
        except Exception: yr = None
        def num(c):
            v = ws.cell(r, c).value if c else None
            try: return float(v) if v not in (None,'') else 0.0
            except Exception: return 0.0
        emain=num(c_emain); eaux=num(c_eaux)
        rows.append({
            'epon': str(epon).strip(), 'onoma': str(onoma).strip() if onoma else '',
            'afm': str(ws.cell(r,c_afm).value).strip() if c_afm and ws.cell(r,c_afm).value else None,
            'amka': str(ws.cell(r,c_amka).value).strip() if c_amka and ws.cell(r,c_amka).value else None,
            'ika': str(ws.cell(r,c_ika).value).strip() if c_ika and ws.cell(r,c_ika).value else None,
            'father': (ws.cell(r,c_pat).value if c_pat else None),
            'specialty': (ws.cell(r,c_spec).value if c_spec else None),
            'kind': (ws.cell(r,c_kind).value if c_kind else None),
            'contract': (ws.cell(r,c_contract).value if c_contract else None),
            'subunit_desc': (ws.cell(r,c_sub).value if c_sub else None),
            'dept_desc': (ws.cell(r,c_dept).value if c_dept else None),
            'subunit_code': _ypok4(ws.cell(r,c_subcode).value if c_subcode else None),
            'company_name': (ws.cell(r,c_comp).value if c_comp else None),
            'bank': (ws.cell(r,c_bank).value if c_bank else None),
            'year': yr, 'month': month, 'period_raw': period_raw,
            'gross': num(c_gross), 'efka_employee': round(emain+eaux,2),
            'fmy': num(c_fmy), 'net': num(c_net), 'employer_cost': num(c_cost),
        })
    return rows

def _match_user_by_afm_or_name(afm, epon, onoma):
    if afm:
        pii = EmployeePII.query.filter_by(afm=str(afm)).first()
        if pii: return User.query.get(pii.user_id)
    full = _norm((str(epon) + str(onoma or '')))
    for u in User.query.all():
        if _norm(u.full_name) == full:
            return u
    # χαλαρό: επώνυμο match μοναδικό
    cand = [u for u in User.query.all() if _norm(u.full_name).startswith(_norm(epon))]
    return cand[0] if len(cand) == 1 else None

_KIND_MAP = {
    'δωροπασχα':'Δώρο Πάσχα', 'δωροχριστουγεννων':'Δώρο Χριστουγέννων',
    'επιδομααδειας':'Επίδομα Αδείας', 'αποζημιωσηαδειας':'Αποζημίωση Αδείας',
    'αποδοχεςαδειας':'Αποδοχές Αδείας',
    'επιδομααδειαςπροηγουμενουετους':'Επίδ.Αδείας π.έ.',
    'αποζημιωσηαδειαςπροηγουμενουετους':'Αποζ.Αδείας π.έ.',
    'αποδοχεςαδειαςπροηγουμενουετους':'Αποδ.Αδείας π.έ.',
    'δωροπασχαπροηγουμενουετους':'Δώρο Πάσχα π.έ.',
    'δωροχριστουγεννωνπροηγουμενουετους':'Δώρο Χριστ. π.έ.',
}
def _period_kind(period_raw):
    n = _norm(period_raw)
    if n in _EPSILON_MONTH: return 'monthly'
    lbl = _KIND_MAP.get(n)
    if lbl: return lbl
    # άγνωστη ταμπέλα: κράτα ≤24 χαρ. (όριο στήλης)
    return (str(period_raw).strip() or 'extra')[:24]

_HOTEL_KW = {
    'AST': ['asterias', 'αστεριας'], 'CNT': ['central', 'χερσονησ', 'hersoniss'],  # ΟΧΙ 'κεντρικο' (=έδρα)
    'IRO': ['iro', 'ηρω'], 'SRG': ['sergios', 'σεργιος', 'σεργιου'],
    'PSV': ['piskopiano', 'πισκοπιανο'], 'PLM': ['palm', 'παλμ'], 'CND': ['condian', 'κονντιαν'],
}
def _hotel_from_epsilon(row, comp=None):
    """Ξενοδοχείο ΝΤΕΤΕΡΜΙΝΙΣΤΙΚΑ από Κωδικό ΥΠΟΚ + εταιρεία (αυθεντικός χάρτης).
    π.χ. Γιαννουλάκης: 0002->AST, 0001->CNT, 0000->IRO. Μονοξενοδοχειακή -> το ξεν. της."""
    if comp:
        ypok = row.get('subunit_code')
        hc = YPOK_TO_HOTEL.get((comp.code, ypok)) if ypok else None
        if hc:
            for h in Hotel.query.filter_by(company_id=comp.id).all():
                if hotel_code(h) == hc:
                    return h
        hs = Hotel.query.filter_by(company_id=comp.id).all()
        if len(hs) == 1:
            return hs[0]
        # εφεδρεία: περιγραφή ΥΠΟΚ
        blob = _norm(str(row.get('subunit_desc') or '') + ' ' + str(row.get('dept_desc') or ''))
        for h in hs:
            if any(_norm(k) in blob for k in _HOTEL_KW.get(hotel_code(h), [])):
                return h
        return None
    return None

def _create_locked_employee(row):
    """Δημιουργεί χρήστη-εργαζόμενο (login-off) από Epsilon (master)."""
    from werkzeug.security import generate_password_hash
    full = (str(row['epon']).strip() + ' ' + (str(row['onoma']).strip() if row['onoma'] else '')).strip()
    base = re.sub(r'[^a-z0-9.]', '', _acc(full).lower().replace(' ', '.')) or ('emp' + os.urandom(3).hex())
    uname = base[:40]
    _guard = 0
    while User.query.filter_by(username=uname).first():
        _guard += 1
        uname = (base[:34] + '.' + os.urandom(3).hex())[:46]
        if _guard > 50:
            uname = 'emp.' + os.urandom(6).hex()
            break
    u = User(username=uname, password=generate_password_hash(os.urandom(8).hex()),
             full_name=full[:100], role='staff', approved=True, is_active=True)
    for attr, val in [('login_enabled', False), ('employment_active', True)]:
        if hasattr(u, attr):
            setattr(u, attr, val)
    db.session.add(u); db.session.flush()
    return u

def _apply_epsilon_identity(u, row, comp=None, cost_center=None):
    """Ταυτότητα από Epsilon (αλήθεια): ξενοδοχείο + PII + κλείδωμα."""
    hotel = _hotel_from_epsilon(row, comp)
    if hotel:
        u.home_hotel_id = hotel.id
    pii = EmployeePII.query.filter_by(user_id=u.id).first()
    if not pii:
        pii = EmployeePII(user_id=u.id); db.session.add(pii)
    filled = False
    for fld, val in [('afm',row['afm']),('amka',row['amka']),('ika_am',row['ika']),
                     ('father_name',row['father']),('ergani_specialty',row['specialty']),
                     ('employment_kind',row['kind']),('contract_type',row['contract']),
                     ('bank_name',row['bank'])]:
        if val and not getattr(pii, fld, None):
            setattr(pii, fld, str(val)[:120]); filled = True
    # όνομα/επώνυμο χωριστά
    if row.get('epon') and not pii.last_name: pii.last_name = str(row['epon']).strip()[:80]
    if row.get('onoma') and not pii.first_name: pii.first_name = str(row['onoma']).strip()[:80]
    pii.locked = True
    # managerial κέντρο κόστους: file-level αν δόθηκε, αλλιώς default = νομικό ξενοδοχείο
    if cost_center:
        pii.cost_center = cost_center
    elif not pii.cost_center:
        ypk = row.get('subunit_code')
        cc_def = YPOK_TO_CC.get((comp.code if comp else None, ypk)) if comp else None
        if not cc_def and hotel:
            cc_def = hotel_code(hotel)
        if cc_def:
            pii.cost_center = cc_def
    db.session.flush()
    return filled

def _company_from_name(name):
    """Ταιριάζει το «Εταιρεία» (C1) του Epsilon σε Company (normalized)."""
    n = _norm(name)
    if not n:
        return None
    for c in Company.query.all():
        cn = _norm(c.legal_name)
        if cn and (cn in n or n in cn or cn[:18] == n[:18]):
            return c
    return None

def import_epsilon_bytes(raw, filename='', company_id=None):
    """Εισάγει Epsilon workbook -> LegalNetImport (κανονικός + δώρα/άδεια). Idempotent."""
    if openpyxl is None:
        return {'error': 'openpyxl μη διαθέσιμο'}
    if filename and filename.lower().endswith('.xls'):
        return {'error': 'Παλιό format .xls — άνοιξέ το και αποθήκευσέ το ως .xlsx και ξαναδοκίμασε.'}
    wb = _safe_load_xlsx(raw)
    rows = parse_epsilon(wb); wb.close()
    comp = Company.query.get(company_id) if company_id else None
    if not comp and rows:
        comp = _company_from_name(rows[0].get('company_name'))   # από περιεχόμενο (C1)
    if not comp and filename:
        pref = re.split(r'[ _]', filename)[0].upper()
        cc = COMPANY_CODE.get(pref)
        if cc: comp = Company.query.filter_by(code=cc).first()
    # κυρίαρχος μήνας αρχείου (από τις γραμμές που ΕΧΟΥΝ μήνα)
    from collections import Counter
    mc = Counter((r['year'], r['month']) for r in rows if r['year'] and r['month'])
    dom = mc.most_common(1)[0][0] if mc else (None, None)
    # managerial κέντρο κόστους από το ΟΝΟΜΑ αρχείου (AST/CNT/IRO/SRG/PSV/CND)
    _ccpref = (re.split(r'[ _]', filename)[0].upper() if filename else None)
    cost_center = _ccpref if _ccpref in COST_CENTERS else None
    added = updated = matched = pii_filled = extras = created = 0
    period = set()
    for row in rows:
        kind = _period_kind(row['period_raw'])
        if kind == 'monthly' and row['year'] and row['month']:
            yr, mo = row['year'], row['month']
        else:
            yr, mo = dom            # δώρα/άδεια -> κυρίαρχος μήνας αρχείου
            if kind != 'monthly': extras += 1
        if not yr or not mo: continue
        period.add((yr, mo))
        # ΑΦΜ-first αυστηρά (μην ενώνεις λάθος με όνομα)
        afm0 = row['afm']
        if afm0:
            pp0 = EmployeePII.query.filter_by(afm=afm0).first()
            u = User.query.get(pp0.user_id) if pp0 else None
        else:
            u = _match_user_by_afm_or_name(None, row['epon'], row['onoma'])
        if u is None and (afm0 or row['epon']):   # δημιούργησε από ΟΠΟΙΑΔΗΠΟΤΕ γραμμή (& δώρου)
            u = _create_locked_employee(row)
            if afm0:
                pp0 = EmployeePII.query.filter_by(user_id=u.id).first()
                if not pp0:
                    pp0 = EmployeePII(user_id=u.id); db.session.add(pp0)
                pp0.afm = afm0; db.session.flush()
            created += 1
        if u is not None:
            matched += 1
            if _apply_epsilon_identity(u, row, comp, cost_center):  # cost_center None -> default=νομικό ξεν.
                pii_filled += 1
        key = '%s|%s|%s|%s|%s' % (comp.id if comp else 'x', yr, mo, kind, row['afm'] or (row['epon']+row['onoma']))
        h = hashlib.sha1(key.encode('utf-8')).hexdigest()[:40]
        rec = LegalNetImport.query.filter_by(import_hash=h).first()
        if not rec:
            rec = LegalNetImport(import_hash=h); db.session.add(rec); added += 1
        else:
            updated += 1
        rec.company_id = comp.id if comp else None
        rec.user_id = u.id if u else None
        rec.year = yr; rec.month = mo; rec.afm = row['afm']; rec.period_kind = (kind or 'monthly')[:24]
        rec.emp_name = (row['epon'] + ' ' + (row['onoma'] or '')).strip()
        rec.gross_legal = row['gross']; rec.efka_employee_legal = row['efka_employee']
        rec.fmy_legal = row['fmy']; rec.net_legal = row['net']; rec.employer_cost_legal = row['employer_cost']
        rec.source_file = filename[:160]
    db.session.commit()
    return {'added': added, 'updated': updated, 'matched': matched, 'rows': len(rows),
            'created': created, 'pii_filled': pii_filled, 'extras': extras,
            'company': comp.legal_name if comp else None, 'periods': sorted(period)}


# ── ΜΗΧΑΝΗ: build_run ─────────────────────────────────────────────────────────
def _hotels_of_company(company_id):
    return [h.id for h in Hotel.query.filter_by(company_id=company_id).all()]

def build_run(company_id, year, month, created_by=None):
    """Δημιουργεί/ξαναϋπολογίζει PayrollRun + PayrollLines για εταιρεία×μήνα."""
    if monthly_settlement is None:
        return {'error': 'schedule module μη διαθέσιμο'}
    run = PayrollRun.query.filter_by(company_id=company_id, year=year, month=month).first()
    if run and run.status in ('verified', 'locked', 'paid'):
        return {'run_id': run.id, 'lines': PayrollLine.query.filter_by(run_id=run.id).count(),
                'locked': True}   # εγκεκριμένη — δεν ξαναϋπολογίζεται
    if not run:
        run = PayrollRun(company_id=company_id, year=year, month=month, created_by=created_by)
        db.session.add(run); db.session.flush()
    rates = PayrollRates.query.filter_by(year=year).first()
    run.rates_version = rates.id if rates else None
    prev_paid = {l.user_id: (l.paid or 0.0) for l in PayrollLine.query.filter_by(run_id=run.id).all()}
    PayrollLine.query.filter_by(run_id=run.id).delete()
    hotel_ids = _hotels_of_company(company_id)
    seen = set()
    legal_by_user = {}; extras_by_user = {}
    for li in LegalNetImport.query.filter_by(company_id=company_id, year=year, month=month).all():
        if not li.user_id: continue
        if (li.period_kind or 'monthly') == 'monthly':
            legal_by_user[li.user_id] = li
        else:
            extras_by_user.setdefault(li.user_id, []).append(li)
    n = 0
    for hid in hotel_ids:
        for row in monthly_settlement(year, month, hid):
            u = row['user']; agg = row['agg']; prof = row['profile']
            if u.id in seen: continue
            seen.add(u.id)
            day_wage = prof.day_wage if prof else 0.0
            hour_wage = (prof.hour_wage if prof else 0.0)
            if prof and getattr(prof, 'agreement_type', '') == 'Management' and prof.agreement_amount:
                gross_agreement = round(prof.agreement_amount, 2)
            else:
                gross_agreement = round(day_wage * agg['total_days'], 2)
            extra_pay = round(hour_wage * agg['extra_hours'], 2)
            gross_total = round(gross_agreement + extra_pay, 2)
            # Λογιστήριο: αλήθεια Epsilon αν υπάρχει, αλλιώς εκτίμηση
            li = legal_by_user.get(u.id)
            if li and li.net_legal:
                net_ref = li.net_legal; efka = li.efka_employee_legal or 0.0
                fmy = li.fmy_legal or 0.0; emp_cost = li.employer_cost_legal or 0.0
                net_legal = li.net_legal; net_calc = net_ref
            else:
                erate = ((rates.efka_employee_pct or 0) + (rates.efka_aux_employee_pct or 0)) / 100.0 if rates else 0.1387
                efka = round(gross_total * erate, 2)
                fmy = 0.0
                net_calc = round(gross_total - efka - fmy, 2)
                emp_rate = ((rates.efka_employer_pct or 0) + (rates.efka_aux_employer_pct or 0)) / 100.0 if rates else 0.2229
                emp_cost = round(gross_total * (1 + emp_rate), 2)
                net_ref = net_calc; net_legal = None
            ag = Agreement.query.filter_by(user_id=u.id, effective_to=None).first()
            # bank_target: στόχος τράπεζας από συμφωνία (Φ2: αν δηλωθεί στο channels_json)· αλλιώς None
            bank_target = None
            if ag and ag.channels_json:
                try:
                    bank_target = (json.loads(ag.channels_json) or {}).get('bank_target')
                except Exception:
                    bank_target = None
            bonus = round(max(0.0, (bank_target - net_ref)), 2) if bank_target else 0.0
            bank_total = round(net_ref + bonus, 2)
            folder_fixed = (ag.folder_fixed if ag else 0.0) or 0.0
            folder_total = round(folder_fixed + extra_pay, 2)
            in_hand = round(bank_total + folder_total, 2)
            ex_list = extras_by_user.get(u.id, [])
            extra_legal_net = round(sum((e.net_legal or 0) for e in ex_list), 2)
            extra_emp_cost = round(sum((e.employer_cost_legal or 0) for e in ex_list), 2)
            extras_json = json.dumps([{'kind': e.period_kind, 'net': round(e.net_legal or 0, 2),
                                       'cost': round(e.employer_cost_legal or 0, 2)} for e in ex_list],
                                     ensure_ascii=False) if ex_list else None
            line = PayrollLine(run_id=run.id, user_id=u.id,
                extra_legal_net=extra_legal_net, extra_employer_cost=extra_emp_cost, extras_json=extras_json,
                work_days=agg['work_days'], sundays=agg['sundays'],
                holidays_worked=agg['holidays_worked'], extra_hours=agg['extra_hours'],
                repo=agg['repo'], total_days=agg['total_days'], elsewhere_days=agg['elsewhere_days'],
                gross_agreement=gross_agreement, extra_pay=extra_pay, gross_total=gross_total,
                bank_target=bank_target, bonus=bonus, bank_total=bank_total,
                folder_total=folder_total, in_hand=in_hand,
                efka_employee=efka, fmy=fmy, net_calc=net_calc, employer_cost=emp_cost,
                net_legal=net_legal,
                employer_cost_legal=(li.employer_cost_legal if li else None),
                net_diff=(round((net_legal - net_calc), 2) if net_legal is not None else None))
            line.paid = prev_paid.get(u.id, 0.0)
            db.session.add(line); n += 1
    run.status = 'calculated'
    db.session.commit()
    return {'run_id': run.id, 'lines': n, 'legal_matched': len(legal_by_user)}

def run_totals(run_id):
    lines = PayrollLine.query.filter_by(run_id=run_id).all()
    t = {'gross_total':0.0,'extra_pay':0.0,'bank_total':0.0,'folder_total':0.0,'in_hand':0.0,
         'bonus':0.0,'net_legal':0.0,'employer_cost':0.0,'extra_legal':0.0,
         'in_hand_total':0.0,'employer_cost_total':0.0,'n':len(lines),'n_legal':0,'n_extras':0}
    for l in lines:
        t['gross_total']+=l.gross_total or 0; t['extra_pay']+=l.extra_pay or 0
        t['bank_total']+=l.bank_total or 0; t['folder_total']+=l.folder_total or 0
        t['in_hand']+=l.in_hand or 0; t['bonus']+=l.bonus or 0
        ec = (l.employer_cost_legal or l.employer_cost or 0)
        t['employer_cost']+= ec
        eln = l.extra_legal_net or 0; eec = l.extra_employer_cost or 0
        t['extra_legal']+= eln
        t['in_hand_total']+= (l.in_hand or 0) + eln
        t['employer_cost_total']+= ec + eec
        if eln: t['n_extras']+=1
        if l.net_legal is not None: t['net_legal']+=l.net_legal; t['n_legal']+=1
    for k in t:
        if isinstance(t[k], float): t[k]=round(t[k],2)
    return t


# ── ROUTES Φ2 ─────────────────────────────────────────────────────────────────
@app.route('/dashboard/payroll/runs', methods=['GET', 'POST'])
def payroll_runs():
    if not _padmin():
        return redirect(url_for('login'))
    if request.method == 'POST':
        cid = request.form.get('company_id', type=int)
        yr = request.form.get('year', type=int); mo = request.form.get('month', type=int)
        if cid and yr and mo:
            res = build_run(cid, yr, mo, created_by=(current_user().id if current_user() else None))
            log_activity('payroll_run_build', '%s/%s/%s -> %s' % (cid, yr, mo, res))
            run = PayrollRun.query.filter_by(company_id=cid, year=yr, month=mo).first()
            if run: return redirect(url_for('payroll_run_view', rid=run.id))
        return redirect(url_for('payroll_runs'))
    runs = PayrollRun.query.order_by(PayrollRun.year.desc(), PayrollRun.month.desc()).all()
    companies = Company.query.filter_by(active=True).order_by(Company.legal_name).all()
    comp_by_id = {c.id: c for c in Company.query.all()}
    tot_by_run = {r.id: run_totals(r.id) for r in runs}
    return render_template('payroll_runs.html', runs=runs, companies=companies,
        comp_by_id=comp_by_id, tot_by_run=tot_by_run, months=MONTHS_EL2,
        run_labels=RUN_LABELS, is_admin=is_admin())

@app.route('/dashboard/payroll/run/<int:rid>')
def payroll_run_view(rid):
    if not _padmin():
        return redirect(url_for('login'))
    run = PayrollRun.query.get_or_404(rid)
    comp = Company.query.get(run.company_id)
    lines = PayrollLine.query.filter_by(run_id=rid).all()
    umap = {u.id: u for u in User.query.all()}
    lines.sort(key=lambda l: (umap.get(l.user_id).full_name if umap.get(l.user_id) else ''))
    view = request.args.get('view', 'management')
    return render_template('payroll_run.html', run=run, comp=comp, lines=lines, umap=umap,
        totals=run_totals(rid), months=MONTHS_EL2, run_labels=RUN_LABELS, view=view, is_admin=is_admin())

@app.route('/dashboard/payroll/import', methods=['GET', 'POST'])
def payroll_import():
    if not _padmin():
        return redirect(url_for('login'))
    results = []
    if request.method == 'POST':
        files = request.files.getlist('file')
        cid = request.form.get('company_id', type=int)
        for f in files:
            if not (f and f.filename):
                continue
            try:
                r = import_epsilon_bytes(f.read(), filename=f.filename, company_id=cid)
            except Exception as e:
                r = {'error': str(e)}
            r['_file'] = f.filename
            results.append(r)
        log_activity('payroll_epsilon_import', '%d αρχεία' % len(results))
    companies = Company.query.filter_by(active=True).order_by(Company.legal_name).all()
    n_legal = LegalNetImport.query.count()
    return render_template('payroll_import.html', companies=companies, results=results,
        n_legal=n_legal, is_admin=is_admin())

# ══════════════════════════════════════════════════════════════════════════════
# ΦΑΣΗ 2.2 — Σελίδα «Έλεγχος & Έγκριση» (πίνακας Γενικού Διευθυντή)
# ══════════════════════════════════════════════════════════════════════════════
try:
    from schedule import Department as _Dept
except Exception:
    _Dept = None

def _control_rows(year, month, company_id=None, hotel_id=None, dept_id=None, cc=None):
    runs_q = PayrollRun.query.filter_by(year=year, month=month)
    if company_id:
        runs_q = runs_q.filter_by(company_id=company_id)
    runs = runs_q.all()
    comp_by_id = {c.id: c for c in Company.query.all()}
    umap = {u.id: u for u in User.query.all()}
    dmap = {d.id: d.name for d in _Dept.query.all()} if _Dept else {}
    rows = []
    tot = {'work_days':0,'repo':0,'extra_hours':0.0,'mgmt':0.0,'legal':0.0,
           'ektos':0.0,'payable':0.0,'paid':0.0,'remaining':0.0}
    for run in runs:
        comp = comp_by_id.get(run.company_id)
        for l in PayrollLine.query.filter_by(run_id=run.id).all():
            u = umap.get(l.user_id)
            if not u:
                continue
            hid = getattr(u, 'home_hotel_id', None)
            if hotel_id and hid != hotel_id:
                continue
            did = getattr(u, 'department_id', None)
            if dept_id and did != dept_id:
                continue
            pii_u = EmployeePII.query.filter_by(user_id=u.id).first()
            ccu = pii_u.cost_center if (pii_u and pii_u.cost_center) else None
            if cc and ccu != cc:
                continue
            hotel = Hotel.query.get(hid) if hid else None
            mgmt = round((l.in_hand or 0) + (l.extra_legal_net or 0), 2)         # στο χέρι σύνολο
            legal = (round((l.net_legal or 0), 2) if l.net_legal is not None else None)
            ektos = round(mgmt - (l.net_legal or 0), 2)                          # εκτός λογιστηρίου
            paid = round(l.paid or 0.0, 2)
            remaining = round(mgmt - paid, 2)
            diff = (round((l.net_legal or 0) - (l.net_calc or 0), 2) if l.net_legal is not None else None)
            flags = []
            if not (l.gross_agreement or 0): flags.append('χωρίς συμφωνία')
            if (l.repo or 0) == 0 and (l.work_days or 0) > 0: flags.append('0 ρεπό')
            if (l.extra_hours or 0) > 40: flags.append('πολλές έξτρα')
            if legal is None and (l.work_days or 0) > 0: flags.append('χωρίς Epsilon')
            rows.append({'user': u, 'company': comp, 'run': run,
                'hotel': hotel.name if hotel else '', 'dept': dmap.get(did, ''), 'cc': ccu or '',
                'work_days': l.work_days or 0, 'repo': l.repo or 0,
                'extra_hours': l.extra_hours or 0, 'mgmt': mgmt, 'legal': legal,
                'diff': diff, 'ektos': ektos, 'payable': mgmt, 'paid': paid,
                'remaining': remaining, 'flags': flags})
            tot['work_days'] += l.work_days or 0; tot['repo'] += l.repo or 0
            tot['extra_hours'] += l.extra_hours or 0; tot['mgmt'] += mgmt
            tot['legal'] += (l.net_legal or 0); tot['ektos'] += ektos
            tot['payable'] += mgmt; tot['paid'] += paid; tot['remaining'] += remaining
    for k in tot:
        if isinstance(tot[k], float): tot[k] = round(tot[k], 2)
    rows.sort(key=lambda r: (r['company'].legal_name if r['company'] else '', r['hotel'], r['user'].full_name or ''))
    return rows, runs, tot


@app.route('/dashboard/payroll/control', methods=['GET', 'POST'])
def payroll_control():
    if not _padmin():
        return redirect(url_for('login'))
    year = request.values.get('year', type=int) or 2026
    month = request.values.get('month', type=int) or 5
    if request.method == 'POST' and request.form.get('action') == 'calc':
        # υπολόγισε όλες τις εταιρείες για τον μήνα
        for c in Company.query.filter_by(active=True).all():
            build_run(c.id, year, month, created_by=(current_user().id if current_user() else None))
        log_activity('payroll_calc_month', '%s/%s' % (year, month))
        return redirect(url_for('payroll_control', year=year, month=month))
    company_id = request.values.get('company_id', type=int)
    hotel_id = request.values.get('hotel_id', type=int)
    dept_id = request.values.get('dept_id', type=int)
    cc = request.values.get('cc')
    rows, runs, tot = _control_rows(year, month, company_id, hotel_id, dept_id, cc)
    companies = Company.query.filter_by(active=True).order_by(Company.legal_name).all()
    hotels = Hotel.query.order_by(Hotel.name).all()
    depts = _Dept.query.order_by(_Dept.name).all() if _Dept else []
    # per-company run status (για κουμπιά έγκρισης)
    run_map = {r.company_id: r for r in runs}
    return render_template('payroll_control.html', rows=rows, tot=tot, runs=runs, run_map=run_map,
        companies=companies, hotels=hotels, depts=depts, months=MONTHS_EL2,
        year=year, month=month, company_id=company_id, hotel_id=hotel_id, dept_id=dept_id,
        cc=cc, cost_centers=COST_CENTERS, run_labels=RUN_LABELS, is_admin=is_admin())


@app.route('/dashboard/payroll/approve', methods=['POST'])
def payroll_approve():
    if not _padmin():
        return ('', 403)
    rid = request.form.get('run_id', type=int)
    run = PayrollRun.query.get(rid) if rid else None
    if run and run.status not in ('verified', 'locked', 'paid'):
        cu = current_user()
        run.status = 'verified'
        run.approved_by = cu.id if cu else None
        run.approved_at = datetime.now()
        run.locked_at = datetime.now()
        db.session.commit()
        comp = Company.query.get(run.company_id)
        period = '%s %s' % (MONTHS_EL2[run.month], run.year)
        link = '/dashboard/payroll/run/%d?embed=1' % run.id
        msg = 'Εγκρίθηκε μισθοδοσία: %s — %s (προς πληρωμή/λογιστήριο)' % (comp.legal_name if comp else '', period)
        # ειδοποίηση: προς το παρόν ΜΟΝΟ masteradmin (πλατφόρμα + email) — απόφαση Giannis
        masters = User.query.filter_by(role='masteradmin').all()
        for u in masters:
            notify(u.id, msg, link)
        emails = [u.email for u in masters if u.email]
        try:
            from app import send_email
            if emails:
                send_email('Εστία — Έγκριση μισθοδοσίας: %s %s' % (comp.legal_name if comp else '', period),
                           '<p>%s</p><p>Πίνακας ανά ξενοδοχείο: εξαγωγή από τη σελίδα εκτέλεσης.</p>' % msg,
                           emails)
        except Exception as e:
            print('approve email skipped:', e)
        log_activity('payroll_approve', msg)
    return redirect(request.referrer or url_for('payroll_control'))


@app.route('/dashboard/payroll/run/<int:rid>/markpaid', methods=['POST'])
def payroll_markpaid(rid):
    if not _padmin():
        return ('', 403)
    run = PayrollRun.query.get_or_404(rid)
    for l in PayrollLine.query.filter_by(run_id=rid).all():
        l.paid = round((l.in_hand or 0) + (l.extra_legal_net or 0), 2)
    run.status = 'paid'
    db.session.commit()
    log_activity('payroll_markpaid', str(rid))
    return redirect(request.referrer or url_for('payroll_run_view', rid=rid))


@app.route('/dashboard/payroll/run/<int:rid>/export.xlsx')
def payroll_run_export(rid):
    if not _padmin():
        return redirect(url_for('login'))
    if openpyxl is None:
        return ('openpyxl μη διαθέσιμο', 500)
    run = PayrollRun.query.get_or_404(rid)
    comp = Company.query.get(run.company_id)
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = 'Μισθοδοσία'
    umap = {u.id: u for u in User.query.all()}
    hmap = {h.id: h.name for h in Hotel.query.all()}
    hdr = ['Ξενοδοχείο', 'Εργαζόμενος', 'Εργάσιμες', 'Ρεπό', 'Έξτρα ώρες',
           'Management (στο χέρι)', 'Λογιστήριο (καθαρά)', 'Δώρα/άδεια',
           'Εκτός λογιστηρίου', 'Πληρωτέο', 'Πληρωμένο', 'Υπόλοιπο']
    ws.append(hdr)
    lines = PayrollLine.query.filter_by(run_id=rid).all()
    def hof(u): return hmap.get(getattr(u, 'home_hotel_id', None), '')
    lines.sort(key=lambda l: (hof(umap.get(l.user_id)), (umap.get(l.user_id).full_name if umap.get(l.user_id) else '')))
    for l in lines:
        u = umap.get(l.user_id)
        mgmt = round((l.in_hand or 0) + (l.extra_legal_net or 0), 2)
        ws.append([hof(u), (u.full_name if u else l.user_id), l.work_days or 0, l.repo or 0,
                   round(l.extra_hours or 0, 2), mgmt, (l.net_legal if l.net_legal is not None else ''),
                   round(l.extra_legal_net or 0, 2), round(mgmt - (l.net_legal or 0), 2),
                   mgmt, round(l.paid or 0, 2), round(mgmt - (l.paid or 0), 2)])
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    from flask import send_file
    fn = 'misthodosia_%s_%s_%s.xlsx' % ((comp.code if comp else 'X'), run.year, run.month)
    return send_file(bio, as_attachment=True, download_name=fn,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ══════════════════════════════════════════════════════════════════════════════
# v12.60 — SYNC MASTER FILE (Excel -> Βάση) · on-demand + νυχτερινό SharePoint
# ══════════════════════════════════════════════════════════════════════════════
import threading, urllib.parse, urllib.request

def _hmap(ws):
    """normalized header -> column index (γραμμή 1)."""
    h = {}
    for c in range(1, ws.max_column + 1):
        h[_norm(ws.cell(1, c).value)] = c
    return h

def _col(hmap, *keys, exact=None):
    if exact:
        for e in exact:
            if _norm(e) in hmap: return hmap[_norm(e)]
    for k in keys:
        kk = _norm(k)
        for hn, ci in hmap.items():
            if kk in hn: return ci
    return None

def _pdate(v):
    if not v: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    sv = str(v).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d.%m.%Y'):
        try: return datetime.strptime(sv, fmt).date()
        except Exception: pass
    return None

def _hotel_by_code(code):
    if not code: return None
    code = str(code).strip().upper()
    for h in Hotel.query.all():
        if hotel_code(h) == code: return h
    return None

def sync_master_workbook(wb):
    """Διαβάζει φύλλα «Μητρώο» + «Συμφωνίες» -> upsert στη βάση (κλειδί ΑΦΜ/Κωδ.)."""
    res = {'reg_new': 0, 'reg_upd': 0, 'agr_upd': 0, 'skipped': 0}
    comp_by_code = {c.code: c for c in Company.query.all()}
    # --- Μητρώο ---
    if 'Μητρώο' in wb.sheetnames:
        ws = wb['Μητρώο']; H = _hmap(ws)
        c = {
            'afm': _col(H, exact=['ΑΦΜ'], ), 'code': _col(H, 'κωδ'),
            'epon': _col(H, exact=['Επώνυμο']), 'onoma': _col(H, exact=['Όνομα']),
            'father': _col(H, 'πατρ'), 'ika': _col(H, exact=['Α.Μ. ΙΚΑ'], ) or _col(H, 'μικα'),
            'amka': _col(H, exact=['ΑΜΚΑ']), 'company': _col(H, 'εταιρει'),
            'hotel': _col(H, 'νομικοξεν') or _col(H, 'νομικο'), 'cc': _col(H, 'κεντροκοστ') or _col(H, 'κεντρο'),
            'spec': _col(H, 'ειδικοτητα'), 'kind': _col(H, 'ειδος'), 'contract': _col(H, 'συμβαση'),
            'bank': _col(H, 'τραπεζα'), 'iban': _col(H, 'iban'),
            'hired': _col(H, 'προσληψ'), 'left': _col(H, 'αποχωρ'),
            'status': _col(H, 'κατασταση'),
        }
        def g(r, key):
            ci = c.get(key); v = ws.cell(r, ci).value if ci else None
            return v
        for r in range(2, ws.max_row + 1):
            afm = g(r, 'afm'); afm = str(afm).strip() if afm not in (None, '') else None
            epon = g(r, 'epon')
            if not afm and not epon:
                continue
            row = {'afm': afm, 'epon': str(epon).strip() if epon else '',
                   'onoma': str(g(r,'onoma')).strip() if g(r,'onoma') else ''}
            # ΑΦΜ-first αυστηρά (idempotent): αν υπάρχει ΑΦΜ, ΜΟΝΟ με ΑΦΜ
            u = None
            if afm:
                pp = EmployeePII.query.filter_by(afm=afm).first()
                u = User.query.get(pp.user_id) if pp else None
            else:
                u = _match_user_by_afm_or_name(None, row['epon'], row['onoma'])
            if not u:
                if not (row['epon'] or afm):
                    res['skipped'] += 1; continue
                u = _create_locked_employee(row)
                if afm:   # κλείδωσε το ΑΦΜ ΑΜΕΣΩΣ ώστε το re-run να ταιριάζει
                    pp = EmployeePII.query.filter_by(user_id=u.id).first()
                    if not pp:
                        pp = EmployeePII(user_id=u.id); db.session.add(pp)
                    pp.afm = afm; db.session.flush()
                res['reg_new'] += 1
            else:
                res['reg_upd'] += 1
            # User full name + hotel + company
            full = (row['epon'] + ' ' + row['onoma']).strip()
            if full: u.full_name = full[:100]
            h = _hotel_by_code(g(r, 'hotel'))
            if h:
                u.home_hotel_id = h.id
            sv = g(r, 'status')
            if sv and hasattr(u, 'employment_active'):
                u.employment_active = ('ανενεργ' not in _norm(sv))
            # PII
            pii = EmployeePII.query.filter_by(user_id=u.id).first()
            if not pii:
                pii = EmployeePII(user_id=u.id); db.session.add(pii)
            def sset(field, key):
                v = g(r, key)
                if v not in (None, ''):
                    setattr(pii, field, str(v).strip()[:120])
            sset('afm','afm'); sset('amka','amka'); sset('ika_am','ika'); sset('father_name','father')
            sset('ergani_specialty','spec'); sset('employment_kind','kind'); sset('contract_type','contract')
            sset('bank_name','bank'); sset('bank_iban','iban')
            if row['epon']: pii.last_name = row['epon'][:80]
            if row['onoma']: pii.first_name = row['onoma'][:80]
            cc = g(r,'cc')
            if cc: pii.cost_center = str(cc).strip()[:8]
            code = g(r,'code')
            if code: pii.emp_code = str(code).strip()[:12]
            hv = _pdate(g(r,'hired'));  lv = _pdate(g(r,'left'))
            if hv: pii.hired_at = hv
            if lv: pii.left_at = lv
            pii.locked = True
            db.session.flush()
    # --- Συμφωνίες ---
    if 'Συμφωνίες' in wb.sheetnames:
        ws = wb['Συμφωνίες']; H = _hmap(ws)
        c = {'afm': _col(H, exact=['ΑΦΜ']), 'type': _col(H, 'τυπος'),
             'amount': _col(H, 'συμφωνημ'), 'bank': _col(H, 'στοχοςτραπ') or _col(H, 'στοχος'),
             'folder': _col(H, 'φακελο'), 'hour': _col(H, 'ωρομισθ')}
        for r in range(2, ws.max_row + 1):
            afm = ws.cell(r, c['afm']).value if c['afm'] else None
            if not afm: continue
            amt = ws.cell(r, c['amount']).value if c['amount'] else None
            if not isinstance(amt, (int, float)) or amt <= 0:
                continue
            pii = EmployeePII.query.filter_by(afm=str(afm).strip()).first()
            if not pii: continue
            prof = EmploymentProfile.query.filter_by(user_id=pii.user_id).first() if EmploymentProfile else None
            if EmploymentProfile is None: break
            if not prof:
                prof = EmploymentProfile(user_id=pii.user_id); db.session.add(prof)
            prof.agreement_amount = round(float(amt), 2)
            tv = ws.cell(r, c['type']).value if c['type'] else None
            if tv: prof.agreement_type = str(tv).strip()
            hw = ws.cell(r, c['hour']).value if c['hour'] else None
            res['agr_upd'] += 1
            db.session.flush()
    db.session.commit()
    return res

def sync_master_bytes(raw):
    if openpyxl is None:
        return {'error': 'openpyxl μη διαθέσιμο'}
    wb = _safe_load_xlsx(raw)
    out = sync_master_workbook(wb); wb.close()
    Setting.query.filter_by(key='master_last_sync')  # touch
    st = Setting.query.get('master_last_sync') if hasattr(Setting, 'query') else None
    try:
        row = Setting.query.get('master_last_sync')
        if not row:
            row = Setting(key='master_last_sync'); db.session.add(row)
        row.value = datetime.now().isoformat(timespec='minutes')
        db.session.commit()
    except Exception:
        db.session.rollback()
    return out

def sp_download_master():
    """Κατεβάζει το master xlsx από SharePoint (env SP_MASTER_FILE = πλήρης διαδρομή)."""
    path = os.environ.get('SP_MASTER_FILE', '')
    try:
        import backup as _B
        if not (path and _B.GRAPH_CLIENT_ID and _B.SP_HOST and _B.SP_SITE_PATH):
            return None
        from app import _graph_token
        token = _graph_token(); sid = _B._site_id(token)
        url = ('https://graph.microsoft.com/v1.0/sites/%s/drive/root:/%s:/content'
               % (sid, urllib.parse.quote(path)))
        req = urllib.request.Request(url, headers={'Authorization': 'Bearer ' + token})
        return urllib.request.urlopen(req, timeout=60).read()
    except Exception as e:
        print('sp_download_master:', e); return None


@app.route('/dashboard/payroll/sync', methods=['GET', 'POST'])
def payroll_sync():
    if not _padmin():
        return redirect(url_for('login'))
    result = None
    if request.method == 'POST':
        act = request.form.get('action')
        if act == 'sharepoint':
            raw = sp_download_master()
            result = sync_master_bytes(raw) if raw else {'error': 'Δεν βρέθηκε αρχείο στο SharePoint (env SP_MASTER_FILE).'}
        else:
            f = request.files.get('file')
            if f and f.filename:
                try:
                    result = sync_master_bytes(f.read())
                except Exception as e:
                    result = {'error': str(e)}
        log_activity('payroll_master_sync', str(result))
    last = Setting.query.get('master_last_sync')
    return render_template('payroll_sync.html', result=result,
        last_sync=(last.value if last else None),
        sp_ready=bool(os.environ.get('SP_MASTER_FILE')), is_admin=is_admin())


# ── Νυχτερινό auto-sync από SharePoint ────────────────────────────────────────
def _master_sync_tick():
    with app.app_context():
        try:
            hour = int(os.environ.get('MASTER_SYNC_HOUR', '4'))
        except Exception:
            hour = 4
        now = datetime.now()
        if now.hour != hour:
            return
        slot = now.strftime('%Y-%m-%d')
        row = Setting.query.get('master_sync_slot')
        if row and row.value == slot:
            return
        raw = sp_download_master()
        if raw:
            try:
                sync_master_bytes(raw)
                print('[master-sync] νυχτερινός συγχρονισμός OK')
            except Exception as e:
                print('[master-sync] σφάλμα:', e)
        if not row:
            row = Setting(key='master_sync_slot'); db.session.add(row)
        row.value = slot; db.session.commit()

def _master_sync_loop():
    import time as _t
    while True:
        try:
            _master_sync_tick()
        except Exception as e:
            print('[master-sync] loop:', e)
        _t.sleep(900)  # 15'

def start_master_sync_scheduler():
    if os.environ.get('SP_MASTER_FILE'):
        threading.Thread(target=_master_sync_loop, daemon=True).start()
        print('[master-sync] scheduler started')

@app.route('/dashboard/payroll/reset', methods=['POST'])
def payroll_reset():
    """Πλήρης καθαρισμός ΟΛΟΥ του εισαγμένου προσωπικού (locked + unlocked) + σχετικών."""
    cu = current_user()
    if not (cu and cu.role == 'masteradmin'):
        return ('', 403)
    if (request.form.get('confirm') or '').strip() != 'ΔΙΑΓΡΑΦΗ':
        return redirect(request.referrer or url_for('payroll_sync'))
    # χρήστες-προσωπικό = login-off (κρατάμε πραγματικούς λογαριασμούς)
    uids = [u.id for u in User.query.filter(User.login_enabled == False).all()]
    n = len(uids)
    try:
        PayrollLine.query.delete(synchronize_session=False)
        PayrollRun.query.delete(synchronize_session=False)
        LegalNetImport.query.delete(synchronize_session=False)
        try:
            MgmtAssignment.query.delete(synchronize_session=False)
            import people as _PPL
            _PPL.ProfileEvent.query.delete(synchronize_session=False)
            _PPL.AttentionFlag.query.delete(synchronize_session=False)
            _PPL.NotDuplicate.query.delete(synchronize_session=False)
        except Exception: pass
        if uids:
            Agreement.query.filter(Agreement.user_id.in_(uids)).delete(synchronize_session=False)
            EmployeePII.query.filter(EmployeePII.user_id.in_(uids)).delete(synchronize_session=False)
            try:
                from schedule import EmploymentProfile as _EP, ShiftAssignment as _SA
                _SA.query.filter(_SA.user_id.in_(uids)).delete(synchronize_session=False)
                _EP.query.filter(_EP.user_id.in_(uids)).delete(synchronize_session=False)
            except Exception:
                pass
            try:
                import faults as _flt
                _flt.UserSpecialty.query.filter(_flt.UserSpecialty.user_id.in_(uids)).delete(synchronize_session=False)
            except Exception:
                pass
            for u in User.query.filter(User.id.in_(uids)).all():
                db.session.delete(u)
        db.session.commit()
        log_activity('payroll_reset', '%d εργαζόμενοι διαγράφηκαν' % n)
    except Exception as e:
        db.session.rollback()
        log_activity('payroll_reset_error', str(e))
    return redirect(url_for('payroll_sync'))


# ── Φόρτωση ΜΗΤΡΩΟΥ από το κεντρικό Excel (νέα πηγή αλήθειας) ─────────────────
_COMP_SHORT = {'γιαννουλακης': 'GIAN', 'σεργιου': 'SERG', 'πισκοπιανο': 'PISK'}

def _wipe_personnel():
    """Καθαρισμός ΟΛΟΥ του εισαγμένου προσωπικού (login-off) + σχετικών. Επιστρέφει πλήθος."""
    uids = [u.id for u in User.query.filter(User.login_enabled == False).all()]
    n = len(uids)
    PayrollLine.query.delete(synchronize_session=False)
    PayrollRun.query.delete(synchronize_session=False)
    LegalNetImport.query.delete(synchronize_session=False)
    if uids:
        Agreement.query.filter(Agreement.user_id.in_(uids)).delete(synchronize_session=False)
        EmployeePII.query.filter(EmployeePII.user_id.in_(uids)).delete(synchronize_session=False)
        try:
            from schedule import EmploymentProfile as _EP, ShiftAssignment as _SA
            _SA.query.filter(_SA.user_id.in_(uids)).delete(synchronize_session=False)
            _EP.query.filter(_EP.user_id.in_(uids)).delete(synchronize_session=False)
        except Exception:
            pass
        try:
            import faults as _flt
            _flt.UserSpecialty.query.filter(_flt.UserSpecialty.user_id.in_(uids)).delete(synchronize_session=False)
        except Exception:
            pass
        for u in User.query.filter(User.id.in_(uids)).all():
            db.session.delete(u)
    db.session.flush()
    return n

def import_mitroo_central(raw, wipe=True):
    """Φορτώνει το ΜΗΤΡΩΟ εργαζομένων από το φύλλο «ΛΟΓΙΣΤΗΡΙΟ (σύνοψη)» του κεντρικού Excel.
    Δημιουργεί κλειδωμένες ταυτότητες (πηγή=Λογιστήριο). Δεν φορτώνει μηνιαία μισθοδοσία."""
    if openpyxl is None:
        return {'error': 'openpyxl μη διαθέσιμο'}
    wb = _safe_load_xlsx(raw)
    ws = None
    for s in wb.worksheets:
        if _norm(s.title).startswith(_norm('ΛΟΓΙΣΤΗΡΙΟ')):
            ws = s; break
    if ws is None:
        wb.close(); return {'error': 'Δεν βρέθηκε φύλλο «ΛΟΓΙΣΤΗΡΙΟ (σύνοψη)».'}
    hdr_r = None
    for r in range(1, 7):
        vals = [_norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if _norm('ΑΦΜ') in vals and _norm('Επώνυμο') in vals:
            hdr_r = r; break
    if not hdr_r:
        wb.close(); return {'error': 'Δεν βρέθηκαν επικεφαλίδες (ΑΦΜ/Επώνυμο) στο φύλλο.'}
    H = {_norm(ws.cell(hdr_r, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(hdr_r, c).value}
    def col(*names):
        for nm in names:
            c = H.get(_norm(nm))
            if c: return c
        return None
    def gv(r, *names):
        c = col(*names)
        return ws.cell(r, c).value if c else None
    wiped = _wipe_personnel() if wipe else 0
    created = 0; active_n = 0; comp_cache = {}; seen_afm = set()
    for r in range(hdr_r + 1, ws.max_row + 1):
        epon = gv(r, 'Επώνυμο')
        if not epon or not str(epon).strip():
            continue
        afm = gv(r, 'ΑΦΜ'); afm = str(afm).strip() if afm not in (None, '') else None
        if afm and afm in seen_afm:
            continue
        if afm: seen_afm.add(afm)
        compname = str(gv(r, 'Εταιρεία (τελ.)', 'Εταιρεία') or '')
        code = _COMP_SHORT.get(_norm(compname))
        comp = None
        if code:
            if code not in comp_cache:
                comp_cache[code] = Company.query.filter_by(code=code).first()
            comp = comp_cache[code]
        ypok = gv(r, 'Κωδ.ΥΠΟΚ', 'ΥΠΟΚ'); ypok = _ypok4(ypok) if ypok not in (None, '') else None
        ccv = gv(r, 'Κέντρο κόστους', 'Κ.κόστους'); ccv = str(ccv).strip() if ccv not in (None, '') else None
        row = {'epon': epon, 'onoma': gv(r, 'Όνομα'), 'afm': afm, 'amka': gv(r, 'ΑΜΚΑ', 'Α.Μ.Κ.Α'),
               'ika': gv(r, 'Α.Μ.ΙΚΑ', 'ΑΜ ΙΚΑ'), 'father': gv(r, 'Πατρώνυμο'),
               'specialty': gv(r, 'Ειδικότητα'), 'kind': gv(r, 'Είδος'),
               'contract': gv(r, 'Σύμβαση', 'Τύπος Σύμβασης'), 'bank': gv(r, 'Τράπεζα'),
               'subunit_code': ypok, 'subunit_desc': None, 'dept_desc': gv(r, 'Τμήμα'),
               'company_name': compname}
        u = _create_locked_employee(row)
        _apply_epsilon_identity(u, row, comp, cost_center=ccv)
        pii = EmployeePII.query.filter_by(user_id=u.id).first()
        ecode = gv(r, 'Κωδ.Εργ', 'Κωδ. Εργ', 'Κωδ Εργ')
        if pii and ecode not in (None, ''):
            pii.emp_code = str(ecode).strip()[:12]
        if pii:
            ib = gv(r, 'IBAN')
            if ib not in (None, '') and not pii.bank_iban: pii.bank_iban = str(ib).strip()[:40]
            from datetime import datetime as _dt2
            for _fld, _nm in [('hired_at', 'Ημ.Πρόσληψης'), ('left_at', 'Ημ.Αποχώρησης')]:
                _dv = gv(r, _nm)
                if _dv not in (None, ''):
                    try:
                        setattr(pii, _fld, _dv.date() if hasattr(_dv, 'date') else _dt2.strptime(str(_dv)[:10], '%Y-%m-%d').date())
                    except Exception: pass
        st = _norm(gv(r, 'Κατάσταση'))
        active = (st != _norm('Ανενεργός'))
        if hasattr(u, 'employment_active'):
            u.employment_active = active
        if active: active_n += 1
        created += 1
    db.session.commit()
    # v12.82 — ποσά από «Μισθοδοσία ανά μήνα» (αν υπάρχει) -> LegalNetImport (lossless round-trip)
    net_loaded = 0
    msheet = None
    for s2 in wb.worksheets:
        if _norm(s2.title).startswith(_norm('Μισθοδοσία αν')) or _norm(s2.title).startswith(_norm('Μισθοδοσια αν')):
            msheet = s2; break
    if msheet is not None:
        mhr = None
        for r in range(1, 6):
            vals = [_norm(msheet.cell(r, c).value) for c in range(1, msheet.max_column + 1)]
            if _norm('ΑΦΜ') in vals and _norm('Έτος') in vals:
                mhr = r; break
        if mhr:
            MH = {_norm(msheet.cell(mhr, c).value): c for c in range(1, msheet.max_column + 1) if msheet.cell(mhr, c).value}
            def mcol(*names):
                for nm in names:
                    c = MH.get(_norm(nm))
                    if c: return c
            def mgv(r, *names):
                c = mcol(*names); return msheet.cell(r, c).value if c else None
            import hashlib as _hl
            pmap = {p.afm: p for p in EmployeePII.query.all() if p.afm}
            def _ff(v):
                try: return float(v) if v not in (None, '') else None
                except Exception: return None
            def _ii(v):
                try: return int(float(v)) if v not in (None, '') else None
                except Exception: return None
            for r in range(mhr + 1, msheet.max_row + 1):
                afm2 = mgv(r, 'ΑΦΜ'); afm2 = str(afm2).strip() if afm2 not in (None, '') else None
                yr = _ii(mgv(r, 'Έτος'))
                if not afm2 or not yr: continue
                mo = _ii(mgv(r, 'Μήνας'))
                kind = (str(mgv(r, 'Είδος', 'period_kind') or 'monthly').strip())[:24]
                h = _hl.md5(('%s|%s|%s|%s' % (afm2, yr, mo, kind)).encode('utf-8')).hexdigest()[:40]
                rec = LegalNetImport.query.filter_by(import_hash=h).first()
                if not rec:
                    rec = LegalNetImport(import_hash=h); db.session.add(rec); net_loaded += 1
                p = pmap.get(afm2)
                rec.afm = afm2; rec.year = yr; rec.month = mo; rec.period_kind = kind
                rec.user_id = (p.user_id if p else None)
                rec.gross_legal = _ff(mgv(r, 'Μικτά', 'Συν.Αποδ.'))
                rec.efka_employee_legal = _ff(mgv(r, 'ΕΦΚΑ εργαζ', 'Εισφ. Εργάζ.'))
                rec.fmy_legal = _ff(mgv(r, 'ΦΜΥ', 'Φ.Μ.Υ'))
                rec.net_legal = _ff(mgv(r, 'Καθαρά', 'Καθαρές Αποδοχές'))
                rec.employer_cost_legal = _ff(mgv(r, 'Κόστος εργοδότη', 'Συνολικό Κόστος'))
            db.session.commit()
    wb.close()
    return {'ok': True, 'created': created, 'active': active_n,
            'inactive': created - active_n, 'wiped': wiped, 'net_rows': net_loaded}

@app.route('/dashboard/payroll/mitroo', methods=['GET', 'POST'])
def payroll_mitroo():
    if not _padmin():
        return redirect(url_for('login'))
    cu = current_user(); result = None
    if request.method == 'POST':
        if not (cu and cu.role == 'masteradmin'):
            result = {'error': 'Μόνο ο masteradmin μπορεί να φορτώσει/καθαρίσει το μητρώο.'}
        elif (request.form.get('confirm') or '').strip() != 'ΦΟΡΤΩΣΗ':
            result = {'error': 'Για επιβεβαίωση γράψε ΦΟΡΤΩΣΗ (καθαρίζει & ξαναφορτώνει το προσωπικό).'}
        else:
            f = request.files.get('file')
            if f and f.filename:
                try:
                    result = import_mitroo_central(f.read(), wipe=True)
                except Exception as e:
                    db.session.rollback(); result = {'error': str(e)}
            else:
                result = {'error': 'Δεν δόθηκε αρχείο .xlsx.'}
        log_activity('payroll_mitroo_import', str(result))
    cnt = EmployeePII.query.count()
    locked = EmployeePII.query.filter_by(locked=True).count()
    return render_template('payroll_mitroo.html', result=result, count=cnt, locked=locked, is_admin=is_admin())



# ══════════════════════════════════════════════════════════════════════════════
# v12.71 — MANAGEMENT LAYER: αναθέσεις (ιστορικό) + import Management + merge
# ══════════════════════════════════════════════════════════════════════════════
class MgmtAssignment(db.Model):
    """Ανάθεση Management ανά εργαζόμενο — ιστορικό τμήματος/θέσης/συμφωνίας.
    valid_to NULL = τρέχουσα. Πολλές αναθέσεις/εργαζόμενο (αλλαγή ρόλου ή cross-hotel)."""
    id            = db.Column(db.Integer, primary_key=True)
    user_id       = db.Column(db.Integer, db.ForeignKey('user.id'), index=True)
    unit          = db.Column(db.String(40))
    department    = db.Column(db.String(60))
    position      = db.Column(db.String(80))
    agreement_amount = db.Column(db.Float)
    days_per_month   = db.Column(db.Float)
    hours_per_day    = db.Column(db.Float)
    day_wage      = db.Column(db.Float)
    hour_wage     = db.Column(db.Float)
    accommodation = db.Column(db.String(10))
    phone         = db.Column(db.String(40))
    email         = db.Column(db.String(120))
    notes         = db.Column(db.Text)
    valid_from    = db.Column(db.Date)
    valid_to      = db.Column(db.Date)
    needs_date    = db.Column(db.Boolean, default=False)
    source        = db.Column(db.String(40), default='mgmt2026')
    created_at    = db.Column(db.DateTime, default=datetime.now)


def _hotel_by_unit(unit):
    if not unit: return None
    u = _norm(unit)
    for h in Hotel.query.all():
        if u and u in _norm(h.name):
            return h
    return None


def assignments_for(uid):
    return (MgmtAssignment.query.filter_by(user_id=uid)
            .order_by(MgmtAssignment.valid_to.is_(None).desc(),
                      MgmtAssignment.valid_from.is_(None).desc(),
                      MgmtAssignment.id.desc()).all())


def current_assignment(uid):
    a = (MgmtAssignment.query.filter_by(user_id=uid, valid_to=None)
         .order_by(MgmtAssignment.id.desc()).first())
    if a: return a
    return (MgmtAssignment.query.filter_by(user_id=uid)
            .order_by(MgmtAssignment.id.desc()).first())


import re as _re2
def _hidx(row, *names):
    norm = [_norm(c) for c in row]
    for n in names:
        t = _norm(n)
        for i, h in enumerate(norm):
            if h == t: return i
    return None

def _asdate(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None

def import_management(raw):
    import people as PPL
    if openpyxl is None:
        return {'error': 'openpyxl μη διαθέσιμο'}
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    if 'Μητρώο Management' not in wb.sheetnames:
        return {'error': 'Λείπει το φύλλο «Μητρώο Management».'}
    M = list(wb['Μητρώο Management'].iter_rows(values_only=True))
    mh = M[0]
    c_code=_hidx(mh,'Κωδ.Εργαζομένου'); c_unit=_hidx(mh,'Μονάδα'); c_act=_hidx(mh,'Ενεργός')
    c_dept=_hidx(mh,'Τμήμα'); c_pos=_hidx(mh,'Θέση'); c_last=_hidx(mh,'Επώνυμο'); c_first=_hidx(mh,'Όνομα')
    c_agr=_hidx(mh,'Συμφωνία €','Συμφωνία'); c_dpm=_hidx(mh,'Ημέρες/μήνα'); c_hpd=_hidx(mh,'Ώρες/μέρα')
    c_day=_hidx(mh,'Ημερομίσθιο'); c_hour=_hidx(mh,'Ωρομίσθιο')
    c_hire=_hidx(mh,'Ημ. Πρόσληψης'); c_left=_hidx(mh,'Ημ. Αποχώρησης')
    c_stay=_hidx(mh,'Διαμονή'); c_tel=_hidx(mh,'Τηλέφωνο'); c_em=_hidx(mh,'Email'); c_note=_hidx(mh,'Σχόλια')
    kmap = {}
    if 'ΠΡΟΣ ΑΝΤΙΣΤΟΙΧΙΣΗ' in wb.sheetnames:
        T = list(wb['ΠΡΟΣ ΑΝΤΙΣΤΟΙΧΙΣΗ'].iter_rows(values_only=True))
        th = T[0]
        t_unit=_hidx(th,'Μονάδα'); t_last=_hidx(th,'Επώνυμο'); t_first=_hidx(th,'Όνομα')
        t_kl=[i for i,h in enumerate(th) if h and 'Σημείωση' in str(h)]
        t_k = t_kl[0] if t_kl else None
        for r in T[1:]:
            if not r or t_last is None or not r[t_last]: continue
            kv = str(r[t_k]).strip() if (t_k is not None and r[t_k]) else ''
            kmap[(_norm(r[t_unit]), _norm(r[t_last]), _norm(r[t_first]))] = kv
    stats = {'rows':0,'assignments':0,'matched':0,'orphan':0,'created_users':0,
             'multi':0,'dup_rows':0,'errors':[]}
    seen = {}; by_user = {}
    code_re = _re2.compile(r'(E0?\d{2,})', _re2.I)
    for r in M[1:]:
        if not r or c_last is None or not r[c_last]: continue
        stats['rows'] += 1
        unit=r[c_unit]; last=r[c_last]; first=r[c_first]
        code = str(r[c_code]).strip() if (c_code is not None and r[c_code]) else ''
        if not code:
            kv = kmap.get((_norm(unit),_norm(last),_norm(first)), '')
            m = code_re.search(kv or '')
            code = m.group(1).upper() if m else ''
        user = None
        if code:
            pii = EmployeePII.query.filter_by(emp_code=code).first()
            if pii: user = User.query.get(pii.user_id)
        if user is None:
            full = ('%s %s' % (str(last).strip(), str(first or '').strip())).strip()
            u = None
            for cand in User.query.filter_by(login_enabled=False).all():
                if _norm(cand.full_name) == _norm(full):
                    u = cand; break
            if u is None:
                u = User(username=('mgmt_%s' % _norm(full).replace(' ','_'))[:60] or 'mgmt_x',
                         password='!', full_name=full, role='staff',
                         login_enabled=False, employment_active=True)
                h = _hotel_by_unit(unit)
                if h and hasattr(u,'home_hotel_id'): u.home_hotel_id = h.id
                db.session.add(u); db.session.flush()
                pii2 = EmployeePII(user_id=u.id)
                if hasattr(pii2,'last_name'): pii2.last_name = str(last).strip()
                if hasattr(pii2,'first_name'): pii2.first_name = str(first or '').strip()
                db.session.add(pii2)
                PPL.log_event(u.id,'created','Ορφανό προφίλ από Management (χωρίς κλειδί Λογιστηρίου)')
                stats['created_users'] += 1
            user = u
            PPL.add_flag(user.id,'orphan','warn','Εκτός αρχείων Λογιστηρίου')
            PPL.add_flag(user.id,'no_afm','warn','Χωρίς ΑΦΜ (ορφανό)')
            PPL.add_flag(user.id,'no_code','warn','Χωρίς Κωδ. Εργαζομένου')
            stats['orphan'] += 1
        else:
            stats['matched'] += 1
        k=(user.id, _norm(unit), _norm(r[c_dept]))
        a = seen.get(k)
        num=lambda c: (float(r[c]) if (c is not None and isinstance(r[c],(int,float))) else None)
        if a is None:
            a = (MgmtAssignment.query
                 .filter_by(user_id=user.id, unit=str(unit or ''), department=str(r[c_dept] or ''), source='mgmt2026')
                 .first())
            if a is None:
                a = MgmtAssignment(user_id=user.id, source='mgmt2026'); db.session.add(a)
                stats['assignments'] += 1
            seen[k]=a
        else:
            stats['dup_rows'] += 1
            PPL.add_flag(user.id,'possible_dup','info',
                         'Η πηγή Management είχε 2 ίδιες γραμμές (%s/%s)' % (unit, r[c_dept]))
        a.unit=str(unit or ''); a.department=str(r[c_dept] or ''); a.position=str(r[c_pos] or '') or None
        a.agreement_amount=num(c_agr); a.days_per_month=num(c_dpm); a.hours_per_day=num(c_hpd)
        a.day_wage=num(c_day); a.hour_wage=num(c_hour)
        a.accommodation=str(r[c_stay]).strip() if (c_stay is not None and r[c_stay]) else None
        a.phone=str(r[c_tel]).strip() if (c_tel is not None and r[c_tel]) else None
        a.email=str(r[c_em]).strip() if (c_em is not None and r[c_em]) else None
        a.notes=str(r[c_note]).strip() if (c_note is not None and r[c_note]) else None
        if a.valid_from is None and c_hire is not None:
            a.valid_from=_asdate(r[c_hire])
        if a.valid_to is None and c_left is not None:
            a.valid_to=_asdate(r[c_left])
        by_user.setdefault(user.id, []).append(a)
        PPL.log_event(user.id,'import','Management: %s / %s%s' %
                      (unit, r[c_dept], (' - ' + str(r[c_pos]) if (c_pos is not None and r[c_pos]) else '')))
        if code:
            PPL.clear_flags(user.id,'no_code'); PPL.clear_flags(user.id,'orphan')
    db.session.flush()
    for uid, lst in by_user.items():
        if len(lst) > 1:
            stats['multi'] += 1
            need = [a for a in lst if a.valid_from is None and a.valid_to is None]
            if len(need) >= 2:
                for a in need: a.needs_date = True
                PPL.add_flag(uid,'assignment_no_date','warn',
                             '%d αναθέσεις χωρίς ημερομηνία - όρισε περιόδους' % len(need))
    db.session.commit()
    try:
        scan_completeness()
    except Exception:
        pass
    return stats


@app.route('/dashboard/payroll/mgmt-import', methods=['GET','POST'])
def payroll_mgmt_import():
    if not _padmin():
        return redirect(url_for('login'))
    result = None
    if request.method == 'POST':
        f = request.files.get('file')
        if f and f.filename:
            try:
                result = import_management(f.read())
            except Exception as e:
                db.session.rollback(); result = {'error': str(e)}
        else:
            result = {'error': 'Δεν δόθηκε αρχείο .xlsx.'}
        log_activity('payroll_mgmt_import', str(result))
    n_assign = MgmtAssignment.query.count()
    return render_template('payroll_mgmt_import.html', result=result, n_assign=n_assign, is_admin=is_admin())


@app.route('/dashboard/payroll/employee/<int:uid>/assign', methods=['POST'])
def payroll_assign(uid):
    if not _padmin():
        return redirect(url_for('login'))
    import people as PPL
    from datetime import timedelta
    u = User.query.get_or_404(uid)
    act = request.form.get('action')
    def _d(name):
        v=request.form.get(name)
        try: return datetime.strptime(v,'%Y-%m-%d').date() if v else None
        except Exception: return None
    def _f(name):
        v=request.form.get(name)
        try: return float(v) if v not in (None,'') else None
        except Exception: return None
    if act == 'assign_new':
        vf = _d('valid_from')
        if vf:
            for a in MgmtAssignment.query.filter_by(user_id=uid, valid_to=None).all():
                a.valid_to = vf - timedelta(days=1); a.needs_date=False
        a = MgmtAssignment(user_id=uid, source='manual',
                           unit=(request.form.get('unit') or '').strip() or None,
                           department=(request.form.get('department') or '').strip() or None,
                           position=(request.form.get('position') or '').strip() or None,
                           agreement_amount=_f('agreement_amount'), valid_from=vf)
        db.session.add(a)
        PPL.log_event(uid,'assignment','Νέα ανάθεση: %s / %s (από %s)' % (a.unit, a.department, vf or '-'))
        PPL.clear_flags(uid,'assignment_no_date')
    elif act == 'assign_edit':
        a = MgmtAssignment.query.get(request.form.get('aid', type=int))
        if a and a.user_id == uid:
            a.valid_from=_d('valid_from'); a.valid_to=_d('valid_to')
            if request.form.get('department'): a.department=request.form.get('department').strip()
            if request.form.get('position') is not None: a.position=(request.form.get('position') or '').strip() or None
            if _f('agreement_amount') is not None: a.agreement_amount=_f('agreement_amount')
            if a.valid_from: a.needs_date=False
            PPL.log_event(uid,'assignment','Επεξεργασία ανάθεσης %s/%s' % (a.unit,a.department))
            if not MgmtAssignment.query.filter_by(user_id=uid, needs_date=True).count():
                PPL.clear_flags(uid,'assignment_no_date')
    db.session.commit()
    return redirect(url_for('payroll_employee', uid=uid) + '?embed=1')


@app.route('/dashboard/payroll/employee/<int:uid>/merge', methods=['POST'])
def payroll_merge(uid):
    if not _padmin():
        return redirect(url_for('login'))
    import people as PPL
    keep = User.query.get_or_404(uid)
    other_id = request.form.get('other_id', type=int)
    other = User.query.get(other_id) if other_id else None
    if not other or other.id == keep.id:
        return redirect(url_for('payroll_employee', uid=uid) + '?embed=1')
    for a in MgmtAssignment.query.filter_by(user_id=other.id).all():
        a.user_id = keep.id
    for li in LegalNetImport.query.filter_by(user_id=other.id).all():
        li.user_id = keep.id
    pk = EmployeePII.query.filter_by(user_id=keep.id).first()
    po = EmployeePII.query.filter_by(user_id=other.id).first()
    if pk and po:
        for fld in ('afm','amka','ika_am','father_name','bank_name','bank_iban','emp_code'):
            if not getattr(pk, fld, None) and getattr(po, fld, None):
                setattr(pk, fld, getattr(po, fld))
    other.employment_active = False; other.login_enabled = False
    PPL.clear_flags(other.id)
    PPL.clear_flags(keep.id, 'possible_dup')
    PPL.log_event(keep.id,'merge','Συγχώνευση: «%s» (#%d) -> σε αυτό το προφίλ' % (other.full_name or other.username, other.id))
    PPL.log_event(other.id,'merge','Συγχωνεύτηκε στο «%s» (#%d) - αρχειοθετήθηκε' % (keep.full_name or keep.username, keep.id))
    db.session.commit()
    log_activity('payroll_merge', '%s -> %s' % (other.id, keep.id))
    return redirect(url_for('payroll_employee', uid=uid) + '?embed=1')



# ── v12.72 — Σάρωμα πληρότητας προφίλ (flags για κάθε κενό) ───────────────────
def scan_completeness():
    import people as PPL
    from collections import defaultdict
    AF = PPL.AttentionFlag
    SCAN_TYPES = {'no_afm','no_code','missing_amka','missing_ika','missing_father',
                  'missing_iban','missing_bank','missing_hire','missing_specialty',
                  'missing_phone','missing_email','missing_cc','missing_assignment','missing_agreement','missing_position'}
    by = defaultdict(dict)
    for f in AF.query.filter_by(entity_type='employee', resolved=False).all():
        by[f.entity_id][f.flag_type] = f
    uids = set()
    for (uid,) in db.session.query(EmployeePII.user_id).all(): uids.add(uid)
    for (uid,) in db.session.query(MgmtAssignment.user_id).all(): uids.add(uid)
    if EmploymentProfile:
        for (uid,) in db.session.query(EmploymentProfile.user_id).all(): uids.add(uid)
    uids.discard(None)
    def has(v): return bool(v and str(v).strip())
    n_scanned=n_added=n_cleared=0
    for uid in uids:
        u = User.query.get(uid)
        if not u: continue
        pii = EmployeePII.query.filter_by(user_id=uid).first()
        prof = EmploymentProfile.query.filter_by(user_id=uid).first() if EmploymentProfile else None
        ma = current_assignment(uid)
        miss=set()
        if not (pii and has(pii.afm)): miss.add('no_afm')
        if not (pii and has(pii.emp_code)): miss.add('no_code')
        if not (pii and has(pii.amka)): miss.add('missing_amka')
        if not (pii and has(pii.ika_am)): miss.add('missing_ika')
        if not (pii and has(pii.father_name)): miss.add('missing_father')
        if not (pii and has(pii.bank_iban)): miss.add('missing_iban')
        if not (pii and has(pii.bank_name)): miss.add('missing_bank')
        if not (pii and pii.hired_at): miss.add('missing_hire')
        if not (pii and has(pii.ergani_specialty)): miss.add('missing_specialty')
        if not (has(getattr(u,'phone',None)) or (ma and has(ma.phone))): miss.add('missing_phone')
        if not (has(getattr(u,'email',None)) or (ma and has(ma.email))): miss.add('missing_email')
        if not (pii and has(pii.cost_center)): miss.add('missing_cc')
        if ma is None: miss.add('missing_assignment')
        else:
            _allpos = [a.position for a in MgmtAssignment.query.filter_by(user_id=uid).all()]
            if not any((x or '').strip() for x in _allpos): miss.add('missing_position')
        if not ((ma and ma.agreement_amount) or (prof and prof.agreement_amount)): miss.add('missing_agreement')
        existing = {t for t in by.get(uid, {}) if t in SCAN_TYPES}
        for t in (miss - existing):
            db.session.add(AF(entity_type='employee', entity_id=uid, flag_type=t,
                              severity='high', detail=PPL.FLAG_LABELS.get(t, t))); n_added+=1
        for t in (existing - miss):
            f = by[uid][t]; f.resolved=True
            from datetime import datetime as _dt; f.resolved_at=_dt.utcnow(); n_cleared+=1
        n_scanned+=1
    db.session.commit()
    return {'scanned': n_scanned, 'added': n_added, 'cleared': n_cleared}


@app.route('/dashboard/payroll/scan', methods=['POST'])
def payroll_scan():
    if not _padmin():
        return redirect(url_for('login'))
    res = scan_completeness()
    log_activity('payroll_scan', str(res))
    return redirect(url_for('attention_center', entity='employee') + '&embed=1')



# ── v12.77 — Πιθανές διπλοεγγραφές (κεντρικά) + dismiss + μήνες ───────────────
_MONTH_GR = ['', 'Ιαν', 'Φεβ', 'Μαρ', 'Απρ', 'Μάι', 'Ιουν', 'Ιουλ', 'Αυγ', 'Σεπ', 'Οκτ', 'Νοε', 'Δεκ']

def _dup_pairs():
    """Ζεύγη πιθανών διπλών: ίδιο επώνυμο (ή ίδιο ΑΦΜ), ενεργά, μη-απορριφθέντα."""
    import people as PPL
    from collections import defaultdict
    by_last = defaultdict(list); by_afm = defaultdict(list)
    for pii in EmployeePII.query.all():
        u = User.query.get(pii.user_id)
        if not u or u.employment_active is False:
            continue
        if pii.last_name:
            by_last[_norm(pii.last_name)].append((u, pii))
        if pii.afm:
            by_afm[str(pii.afm)].append((u, pii))
    pairs = []
    seen = set()
    def add(a, pa, b, pb, reason):
        k = (min(a.id, b.id), max(a.id, b.id))
        if k in seen: return
        if PPL.is_dismissed(a.id, b.id): return
        seen.add(k); pairs.append((a, pa, b, pb, reason))
    for afm, lst in by_afm.items():
        for i in range(len(lst)):
            for j in range(i+1, len(lst)):
                add(lst[i][0], lst[i][1], lst[j][0], lst[j][1], 'Ίδιο ΑΦΜ')
    for last, lst in by_last.items():
        if len(lst) < 2: continue
        for i in range(len(lst)):
            for j in range(i+1, len(lst)):
                add(lst[i][0], lst[i][1], lst[j][0], lst[j][1], 'Ίδιο επώνυμο')
    return pairs

def _net_summary(uid):
    afm = None
    pii = EmployeePII.query.filter_by(user_id=uid).first()
    if pii and pii.afm: afm = str(pii.afm)
    nm = _legal_net_map().get(afm, {}) if afm else {}
    yrs = ', '.join('%s:%.0f' % (y, nm[y]['tot']) for y in sorted(nm)) if nm else '—'
    na = MgmtAssignment.query.filter_by(user_id=uid).count()
    return {'afm': (afm or '—'), 'code': (pii.emp_code if pii else None) or '—',
            'years': yrs, 'assignments': na}

@app.route('/dashboard/payroll/duplicates')
def payroll_duplicates():
    if not _padmin():
        return redirect(url_for('login'))
    rows = []
    for a, pa, b, pb, reason in _dup_pairs():
        rows.append({'a': a, 'b': b, 'reason': reason,
                     'a_sum': _net_summary(a.id), 'b_sum': _net_summary(b.id)})
    log_activity('payroll_duplicates', '%d ζεύγη' % len(rows))
    return render_template('payroll_duplicates.html', rows=rows, total=len(rows), is_admin=is_admin())

@app.route('/dashboard/payroll/dup/dismiss', methods=['POST'])
def payroll_dup_dismiss():
    if not _padmin():
        return redirect(url_for('login'))
    import people as PPL
    a = request.form.get('a_id', type=int); b = request.form.get('b_id', type=int)
    if a and b:
        PPL.dismiss_pair(a, b)
        PPL.clear_flags(a, 'possible_dup'); PPL.clear_flags(b, 'possible_dup')
        PPL.log_event(a, 'merge', 'Δηλώθηκε ΟΤΙ ΔΕΝ είναι διπλό με #%d' % b)
        PPL.log_event(b, 'merge', 'Δηλώθηκε ΟΤΙ ΔΕΝ είναι διπλό με #%d' % a)
        db.session.commit()
        log_activity('payroll_dup_dismiss', '%s / %s' % (a, b))
    back = request.form.get('back') or url_for('payroll_duplicates')
    return redirect(back + ('?embed=1' if '?' not in back else '&embed=1'))


# ── v12.79 — Διαγραφή ανάθεσης + Μαζική επεξεργασία (grid) ────────────────────
@app.route('/dashboard/payroll/assignment/<int:aid>/delete', methods=['POST'])
def payroll_assignment_delete(aid):
    if not _padmin():
        return redirect(url_for('login'))
    import people as PPL
    a = MgmtAssignment.query.get_or_404(aid)
    uid = a.user_id
    PPL.log_event(uid, 'assignment', 'Διαγραφή ανάθεσης: %s / %s' % (a.unit, a.department))
    db.session.delete(a); db.session.commit()
    log_activity('payroll_assignment_delete', '%s' % aid)
    return redirect(url_for('payroll_employee', uid=uid) + '?embed=1')


@app.route('/dashboard/payroll/grid', methods=['GET', 'POST'])
def payroll_grid():
    if not _padmin():
        return redirect(url_for('login'))
    import people as PPL
    if request.method == 'POST':
        ids = [int(x) for x in (request.form.get('ids') or '').split(',') if x.strip().isdigit()]
        changed = 0
        def g(name, uid):
            v = request.form.get('%s_%d' % (name, uid))
            return v.strip() if v is not None else None
        for uid in ids:
            u = User.query.get(uid)
            if not u: continue
            pii = EmployeePII.query.filter_by(user_id=uid).first()
            a = current_assignment(uid)
            ch = False
            ph = g('f_phone', uid); em = g('f_email', uid); cc = g('f_cc', uid)
            act = g('f_active', uid)
            if ph is not None and (u.phone or '') != ph: u.phone = ph or None; ch = True
            if em is not None and (u.email or '') != em: u.email = em or None; ch = True
            if cc is not None and pii and (pii.cost_center or '') != cc: pii.cost_center = cc or None; ch = True
            if act in ('1', '0'):
                want = (act == '1')
                if u.employment_active != want: u.employment_active = want; ch = True
            if a:
                un = g('f_unit', uid); dp = g('f_dept', uid); ps = g('f_pos', uid); ag = g('f_agr', uid)
                if un is not None and (a.unit or '') != un: a.unit = un or None; ch = True
                if dp is not None and (a.department or '') != dp: a.department = dp or None; ch = True
                if ps is not None and (a.position or '') != ps: a.position = ps or None; ch = True
                if ag is not None:
                    try: agv = float(ag) if ag != '' else None
                    except Exception: agv = a.agreement_amount
                    if a.agreement_amount != agv: a.agreement_amount = agv; ch = True
            if ch:
                changed += 1
                PPL.log_event(uid, 'edit', 'Μαζική επεξεργασία (grid)')
        db.session.commit()
        log_activity('payroll_grid_save', '%d αλλαγές' % changed)
        return redirect(url_for('payroll_grid') + '?embed=1&status=%s&saved=%d' % (request.form.get('status') or 'active', changed))
    status = request.args.get('status') or 'active'
    if status not in ('active', 'inactive', 'all'): status = 'active'
    rows = _employees(status)
    saved = request.args.get('saved', type=int)
    n_active = len(_employees('active')); n_inactive = len(_employees('inactive'))
    log_activity('payroll_grid_view', status)
    return render_template('payroll_grid.html', rows=rows, status=status, saved=saved,
                           n_active=n_active, n_inactive=n_inactive, n_all=n_active + n_inactive,
                           ids=','.join(str(r['user'].id) for r in rows), is_admin=is_admin())



# ── v12.80 — Export 2 μητρώων ΑΠΟ ΤΗ ΒΑΣΗ (καθρέφτης live) ────────────────────
def _xlsx_response(wb, fname):
    from flask import send_file
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/dashboard/payroll/export/logistirio')
def export_logistirio():
    if not _padmin(): return redirect(url_for('login'))
    if openpyxl is None: return 'openpyxl μη διαθέσιμο', 500
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    HF = PatternFill('solid', fgColor='1F4E5F'); HFONT = Font(bold=True, color='FFFFFF', size=10)
    # Σύνοψη
    s = wb.active; s.title = 'ΛΟΓΙΣΤΗΡΙΟ (σύνοψη)'
    cols = ['Κωδ.Εργ','ΑΦΜ','Επώνυμο','Όνομα','Πατρώνυμο','Α.Μ.ΙΚΑ','ΑΜΚΑ','Ειδικότητα','Είδος','Σύμβαση','Τράπεζα','IBAN','Εταιρεία (τελ.)','Κωδ.ΥΠΟΚ','Ξενοδοχείο','Τμήμα','Κέντρο κόστους','Ημ.Πρόσληψης','Ημ.Αποχώρησης','Κατάσταση']
    s.append(cols)
    for u in User.query.all():
        pii = EmployeePII.query.filter_by(user_id=u.id).first()
        if not pii: continue
        h = Hotel.query.get(u.home_hotel_id) if getattr(u, 'home_hotel_id', None) else None
        comp = _company_for_hotel(u.home_hotel_id) if getattr(u, 'home_hotel_id', None) else None
        s.append([pii.emp_code, pii.afm, pii.last_name, pii.first_name, pii.father_name, pii.ika_am, pii.amka,
                  pii.ergani_specialty, pii.employment_kind, pii.contract_type, pii.bank_name, pii.bank_iban,
                  (comp.legal_name if comp else ''), (getattr(h, 'ypok_code', None) if h else ''),
                  (h.name if h else ''), '', pii.cost_center, pii.hired_at, pii.left_at,
                  ('Ενεργός' if u.employment_active is not False else 'Ανενεργός')])
    s2 = wb.create_sheet('Μισθοδοσία ανά μήνα')
    s2.append(['ΑΦΜ','Επώνυμο','Όνομα','Έτος','Μήνας','Είδος','Μικτά','ΕΦΚΑ εργαζ','ΦΜΥ','Καθαρά','Κόστος εργοδότη','Εταιρεία'])
    for li in LegalNetImport.query.order_by(LegalNetImport.afm, LegalNetImport.year, LegalNetImport.month).all():
        pii = EmployeePII.query.filter_by(user_id=li.user_id).first() if li.user_id else None
        s2.append([li.afm, (pii.last_name if pii else (li.emp_name or '')), (pii.first_name if pii else ''),
                   li.year, li.month, li.period_kind, li.gross_legal, li.efka_employee_legal, li.fmy_legal,
                   li.net_legal, li.employer_cost_legal, ''])
    for ws in (s, s2):
        for c in range(1, ws.max_column + 1):
            ws.cell(1, c).fill = HF; ws.cell(1, c).font = HFONT
        ws.freeze_panes = 'A2'
    log_activity('export_logistirio', '%d' % (s.max_row - 1))
    return _xlsx_response(wb, 'ΜΗΤΡΩΟ ΜΙΣΘΟΔΟΣΙΑΣ ΛΟΓΙΣΤΗΡΙΟΥ (live).xlsx')

@app.route('/dashboard/payroll/export/management')
def export_management():
    if not _padmin(): return redirect(url_for('login'))
    if openpyxl is None: return 'openpyxl μη διαθέσιμο', 500
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    HF = PatternFill('solid', fgColor='1F4E5F'); HFONT = Font(bold=True, color='FFFFFF', size=10)
    # Σύνοψη ανά άτομο (readable)
    s = wb.active; s.title = 'Εργαζόμενοι (σύνοψη)'
    s.append(['Κωδ.Εργ', 'ΑΦΜ', 'Επώνυμο', 'Όνομα', 'Κατάσταση', 'Πλήθος αναθέσεων', 'Μονάδες', 'Τρέχον Τμήμα', 'Τρέχουσα Θέση', 'Συμφωνία €', 'Τηλέφωνο', 'Email'])
    # Μητρώο Management (importable, 1 γραμμή/ανάθεση)
    s2 = wb.create_sheet('Μητρώο Management')
    s2.append(['Α/Α','Κωδ.Εργαζομένου','ΑΦΜ','Κατάσταση ταύτισης','Μονάδα','Ενεργός','Τμήμα','Θέση','Επώνυμο','Όνομα','Συμφωνία €','Ημέρες/μήνα','Ώρες/μέρα','Ημερομίσθιο','Ωρομίσθιο','Ημ. Πρόσληψης','Ημ. Αποχώρησης','Διαμονή','Τηλέφωνο','Email','Σχόλια'])
    aa = 0
    for u in User.query.all():
        asg = MgmtAssignment.query.filter_by(user_id=u.id).order_by(MgmtAssignment.id).all()
        if not asg: continue
        pii = EmployeePII.query.filter_by(user_id=u.id).first()
        cur = current_assignment(u.id)
        s.append([(pii.emp_code if pii else ''), (pii.afm if pii else ''),
                  (pii.last_name if pii else u.full_name), (pii.first_name if pii else ''),
                  ('Ενεργός' if u.employment_active is not False else 'Ανενεργός'),
                  len(asg), ' / '.join(sorted({a.unit or '' for a in asg})),
                  (cur.department if cur else ''), (cur.position if cur else ''),
                  (cur.agreement_amount if cur else ''), u.phone, u.email])
        for a in asg:
            aa += 1
            s2.append([aa, (pii.emp_code if pii else ''), (pii.afm if pii else ''), '',
                       a.unit, ('ΝΑΙ' if u.employment_active is not False else 'ΟΧΙ'),
                       a.department, a.position, (pii.last_name if pii else u.full_name), (pii.first_name if pii else ''),
                       a.agreement_amount, a.days_per_month, a.hours_per_day, a.day_wage, a.hour_wage,
                       a.valid_from, a.valid_to, a.accommodation, a.phone, a.email, a.notes])
    for ws in (s, s2):
        for c in range(1, ws.max_column + 1):
            ws.cell(1, c).fill = HF; ws.cell(1, c).font = HFONT
        ws.freeze_panes = 'A2'
    log_activity('export_management', '%d' % (s.max_row - 1))
    return _xlsx_response(wb, 'ΜΗΤΡΩΟ ΜΙΣΘΟΔΟΣΙΑΣ MANAGEMENT (live).xlsx')


print('payroll module loaded (Φ2.4 μητρώο + v12.71 Management + v12.72 scan + v12.77 dups + v12.79 grid + v12.80 export)')
