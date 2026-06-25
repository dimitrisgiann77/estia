# -*- coding: utf-8 -*-
"""
Εστία — measurements.py — Ενοποίηση μετρήσεων Συντήρησης (Φ1→Φ3b-2).
Plug-in: import από το ΤΕΛΟΣ του app.py (ΠΡΙΝ το init_db: create_all να πιάσει το MonitorPeriod).

Περιεχόμενα:
  Φ1  — MonitorPeriod + seed templates «pool»/«znx» + default περίοδοι (seed_measurement_engine, boot)
  Φ2  — σημεία από Pool/WaterSystem + ΑΝΤΙΓΡΑΦΗ legacy records → Reading (idempotent)
  Φ3a — περίοδοι CRUD
  Φ3b — generic φόρμα καταχώρησης (Reading) + προτεινόμενες ενέργειες
  Φ3b-2 — granular σημεία ανά περιοχή (Option B) + ΕΝΙΑΙΑ κονσόλα ρυθμίσεων (tabs)
Καθαρά προσθετικό· οι legacy φόρμες/κονσόλα παραμένουν.
"""
from datetime import date, timedelta
from flask import request, redirect, url_for, render_template, session, Response, jsonify
from app import (app, db, current_user, is_admin, can_log, scoped_hotel_ids, log_activity, area_actions,
                 MonitorTemplate, MonitorParam, Hotel, Pool, WaterSystem,
                 PoolRecord, WaterRecord, Area, Reading, FREQ_LABEL)
import json as _json


# ── Μοντέλο: περίοδοι/βάρδιες ανά template (ορίζονται από admin) ──────────────
class MonitorPeriod(db.Model):
    id           = db.Column(db.Integer, primary_key=True)
    template_key = db.Column(db.String(30), db.ForeignKey('monitor_template.key'), nullable=False)
    key          = db.Column(db.String(20), nullable=False)
    label        = db.Column(db.String(40), nullable=False)
    time         = db.Column(db.String(5))
    sort         = db.Column(db.Integer, default=0)


# ── Φ1 (SPEC βιβλιοθήκης): ανάθεση μετρήσεων ανά σημείο (drag&drop) ───────────
# Κενό για ένα σημείο = πέφτει στις παραμέτρους του template του (συμβατότητα).
class AreaParam(db.Model):
    id      = db.Column(db.Integer, primary_key=True)
    area_id = db.Column(db.Integer, db.ForeignKey('area.id'), nullable=False, index=True)
    pkey    = db.Column(db.String(40), nullable=False)
    sort    = db.Column(db.Integer, default=0)

# ── Φ-Α: Δομή δικτύων ως ιεραρχία-δέντρο (κόμβος με γονέα, απεριόριστο βάθος) ──
# Καθολικός κατάλογος-αναφορά. Τα σημεία (Area.node_id) κρέμονται από οποιονδήποτε
# κόμβο. Area.node_id = null → σημερινή συμπεριφορά (καμία ορατή αλλαγή).
class MonitorNode(db.Model):
    id        = db.Column(db.Integer, primary_key=True)
    key       = db.Column(db.String(40), unique=True)
    parent_id = db.Column(db.Integer, db.ForeignKey('monitor_node.id'))
    name      = db.Column(db.String(80), nullable=False)
    node_kind = db.Column(db.String(20))                 # 'group' | 'subgroup' (ενημερωτικό)
    icon      = db.Column(db.String(40), default='')
    sort      = db.Column(db.Integer, default=0)
    is_active = db.Column(db.Boolean, default=True)


_CLO2_LOW  = 'ClO2 {n} <1 ppm: αύξησε τη δοσομέτρηση ClO2· έλεγξε δοσομετρική αντλία/απόθεμα.'
_CLO2_HIGH = 'ClO2 {n} >2 ppm: μείωσε τη δοσομέτρηση ClO2.'

# (pkey, label, unit, min_v, max_v, action_low, action_high)
POOL_PARAMS = [
    ('free_chlorine', 'Ελεύθερο χλώριο', 'mg/L', 0.4, 1.5,
     'Χαμηλό χλώριο — κάνε χλωρίωση και επανέλεγξε σε 30΄.',
     'Υψηλό χλώριο — σταμάτα τη δοσομέτρηση/άσε να πέσει· απόφυγε χρήση μέχρι <1.5 mg/L.'),
    ('combined_chlorine', 'Συνδεδεμένο χλώριο', 'mg/L', None, 0.5, None,
     'Υψηλό δεσμευμένο χλώριο — υπερχλωρίωση (shock) + αερισμός· έλεγξε ανανέωση νερού.'),
    ('ph', 'pH', '', 7.2, 7.8,
     'Χαμηλό pH — πρόσθεσε pH plus (ανθρακική σόδα).',
     'Υψηλό pH — πρόσθεσε pH minus (οξύ), σταδιακά.'),
    ('temp', 'Θερμοκρασία', '°C', None, 32.0, None,
     'Υψηλή θερμοκρασία — έλεγξε/μείωσε θέρμανση· παρακολούθησε χλώριο.'),
    ('turbidity', 'Θολότητα', 'NTU', None, 1.0, None,
     'Θολό νερό — backwash φίλτρου, έλεγξε διήθηση/κυκλοφορία, εξέτασε κροκίδωση.'),
    ('cyanuric_acid', 'Κυανουρικό οξύ', 'mg/L', None, 75.0, None,
     'Υψηλό κυανουρικό οξύ — μερική ανανέωση νερού (αραίωση)· μείωσε σταθεροποιητή.'),
    ('total_alkalinity', 'Ολική αλκαλικότητα', 'mg/L', 80.0, 120.0,
     'Χαμηλή αλκαλικότητα — πρόσθεσε alkalinity up (ανθρακική σόδα).',
     'Υψηλή αλκαλικότητα — πρόσθεσε οξύ σταδιακά.'),
    ('orp', 'ORP', 'mV', 650.0, None,
     'Χαμηλό ORP — ανέβασε ελεύθερο χλώριο και ρύθμισε pH στο 7.2–7.6.', None),
    ('backwash_done', 'Backwash έγινε', '', None, None, None, None),
]

ZNX_PARAMS = [
    ('clo2_tank', 'ClO2 Δεξαμενή', 'ppm', 1.0, 2.0, _CLO2_LOW.format(n='Δεξαμενή'), _CLO2_HIGH.format(n='Δεξαμενή')),
    ('clo2_kitchen', 'ClO2 Κουζίνα', 'ppm', 1.0, 2.0, _CLO2_LOW.format(n='Κουζίνα'), _CLO2_HIGH.format(n='Κουζίνα')),
    ('clo2_remote', 'ClO2 Απομακρυσμένο', 'ppm', 1.0, 2.0, _CLO2_LOW.format(n='Απομακρυσμένο'), _CLO2_HIGH.format(n='Απομακρυσμένο')),
    ('clo2_dhw_out', 'ClO2 Αναχώρηση ΖΝΧ', 'ppm', 1.0, 2.0, _CLO2_LOW.format(n='Αναχώρηση ΖΝΧ'), _CLO2_HIGH.format(n='Αναχώρηση ΖΝΧ')),
    ('clo2_dhw_return', 'ClO2 Επιστροφή ΖΝΧ', 'ppm', 1.0, 2.0, _CLO2_LOW.format(n='Επιστροφή ΖΝΧ'), _CLO2_HIGH.format(n='Επιστροφή ΖΝΧ')),
    ('clo2_ro', 'ClO2 Αντ. Όσμωση', 'ppm', 1.0, 2.0, _CLO2_LOW.format(n='Αντ. Όσμωση'), _CLO2_HIGH.format(n='Αντ. Όσμωση')),
    ('temp_dhw_out', 'Κολεκτέρ ΖΝΧ (Αναχ.)', '°C', 60.0, None,
     'Κολεκτέρ ΖΝΧ <60°C: ανέβασε θερμοκρασία αποθήκευσης ≥60°C (κίνδυνος legionella)· έλεγξε λέβητα/εναλλάκτη/θερμοστάτη.', None),
    ('temp_dhw_return', 'Επιστροφή ανακυκλ.', '°C', 50.0, None,
     'Επιστροφή ανακυκλοφορίας <50°C: ανεπαρκής ανακυκλοφορία· έλεγξε αντλία & βάνες· εξέτασε θερμική απολύμανση/flushing.', None),
    ('temp_kitchen_hot', 'Κουζίνα Ζεστό', '°C', 50.0, None,
     'Ζεστό Κουζίνας <50°C: flushing του σημείου· έλεγξε ανακυκλοφορία/μόνωση γραμμής.', None),
    ('temp_remote_hot', 'Απομακρυσμένο Ζεστό', '°C', 50.0, None,
     'Ζεστό Απομακρυσμένου <50°C: flushing· έλεγξε ανακυκλοφορία (κρίσιμο τελευταίο σημείο δικτύου).', None),
    ('temp_tank', 'Δεξαμενή (κρύο)', '°C', None, 20.0, None,
     'Δεξαμενή (κρύο) >20°C: εξέτασε ψύξη/μόνωση/ανανέωση νερού· κίνδυνος ανάπτυξης μικροβίων.'),
    ('temp_kitchen_cold', 'Κουζίνα Κρύο', '°C', None, None, None, None),
    ('temp_remote_cold', 'Απομακρυσμένο Κρύο', '°C', None, None, None, None),
    ('temp_ro', 'Αντ. Όσμωση (θερμ.)', '°C', None, None, None, None),
    ('ph_tank', 'pH Δεξαμενής', '', None, None, None, None),
]

# Granular templates ΖΝΧ ανά περιοχή (Option B): (key, label, [pkeys])
ZNX_LOCATIONS = [
    ('znx_tank',    'ΖΝΧ — Δεξαμενή / Μηχανοστάσιο', ['clo2_tank', 'temp_tank', 'ph_tank']),
    ('znx_kitchen', 'ΖΝΧ — Κουζίνα',                 ['clo2_kitchen', 'temp_kitchen_hot', 'temp_kitchen_cold', 'location_kitchen']),
    ('znx_remote',  'ΖΝΧ — Απομακρυσμένο',           ['clo2_remote', 'temp_remote_hot', 'temp_remote_cold', 'location_remote']),
    ('znx_dhw',     'ΖΝΧ — Αναχώρηση / Επιστροφή',   ['clo2_dhw_out', 'clo2_dhw_return', 'temp_dhw_out', 'temp_dhw_return']),
    ('znx_ro',      'ΖΝΧ — Αντίστροφη Όσμωση',       ['clo2_ro', 'temp_ro']),
]
# location_* params (text) δεν είναι στο ZNX_PARAMS με όρια — ορισμός εδώ:
_TEXT_PARAMS = {'location_kitchen': 'Σημείο Κουζίνας', 'location_remote': 'Σημείο Απομακρ.'}

DEFAULT_PERIODS = [('morning', 'Πρωί', '08:00', 1), ('afternoon', 'Απόγευμα', '17:00', 2)]


# ── helpers seed ─────────────────────────────────────────────────────────────
def _seed_template(key, name, icon, params):
    if MonitorTemplate.query.get(key):
        return False
    db.session.add(MonitorTemplate(key=key, name=name, icon=icon, frequency='twice', sort=0, is_active=True))
    db.session.flush()
    for i, (pkey, label, unit, mn, mx, low, high) in enumerate(params, start=1):
        db.session.add(MonitorParam(template_key=key, pkey=pkey, label=label, unit=unit or '',
                                    min_v=mn, max_v=mx, action_low=low, action_high=high, sort=i))
    return True


def _seed_periods(key):
    if MonitorPeriod.query.filter_by(template_key=key).first():
        return
    for pk, label, t, s in DEFAULT_PERIODS:
        db.session.add(MonitorPeriod(template_key=key, key=pk, label=label, time=t, sort=s))


def seed_measurement_engine():
    """boot (module-level) → χρειάζεται app context (όπως schedule/payroll)."""
    with app.app_context():
        try:
            created = False
            created = _seed_template('pool', 'Πισίνα', 'ti-pool', POOL_PARAMS) or created
            created = _seed_template('znx', 'ΖΝΧ / Δίκτυο νερού', 'ti-droplet', ZNX_PARAMS) or created
            created = _seed_template('generic', 'Σημείο μέτρησης', 'ti-map-pin', []) or created
            db.session.commit()
            for key in ('pool', 'znx', 'generic'):
                _seed_periods(key)
            db.session.commit()
            if created:
                print('[measurements] Φ1 seed: templates pool/znx + periods OK')
        except Exception as e:
            db.session.rollback()
            print(f'[measurements] seed skipped: {e}')


# ── Φ-Α: seed καταλόγου-αναφοράς δέντρου (ομάδα/υποομάδα) ─────────────────────
# (key, parent_key, name, node_kind, icon, sort) — γονείς ΠΡΙΝ τα παιδιά.
NODE_CATALOG = [
    ('water',           None,              'Δίκτυο Νερού (πόσιμο/οικιακό)',       'group',    'ti-droplet',  1),
    ('water_source',    'water',           'Πηγή / τροφοδοσία',                  'subgroup', '',            1),
    ('water_treatment', 'water',           'Επεξεργασία',                        'subgroup', '',            2),
    ('water_ro',        'water_treatment', 'Όσμωση (RO)',                        'subgroup', '',            1),
    ('water_storage',   'water',           'Αποθήκευση (δεξαμενές/calorifiers)', 'subgroup', '',            3),
    ('water_cold',      'water',           'Κρύο νερό (ΨΝΧ)',                    'subgroup', '',            4),
    ('water_hot',       'water',           'Ζεστό νερό (ΖΝΧ)',                   'subgroup', '',            5),
    ('aerosol',         None,              'Αερόλυμα / υψηλού κινδύνου',         'group',    'ti-wind',     2),
    ('aer_spa',         'aerosol',         'Spa / jacuzzi',                      'subgroup', '',            1),
    ('aer_cooling',     'aerosol',         'Πύργοι ψύξης',                       'subgroup', '',            2),
    ('aer_misting',     'aerosol',         'Misting / δροσισμός',                'subgroup', '',            3),
    ('aer_fountain',    'aerosol',         'Σιντριβάνια',                        'subgroup', '',            4),
    ('aer_ice',         'aerosol',         'Παγομηχανές',                        'subgroup', '',            5),
    ('pools',           None,              'Πισίνες',                            'group',    'ti-pool',     3),
    ('irrigation',      None,              'Άρδευση',                            'group',    'ti-plant-2',  4),
    ('sewage',          None,              'Λύματα / βιολογικός',                'group',    'ti-recycle',  5),
]


def seed_node_catalog():
    """Φ-Α (boot, idempotent): σπέρνει τον κατάλογο-αναφορά κόμβων (ομάδα/υποομάδα).
    ΔΕΝ αναθέτει σημεία — καμία ορατή αλλαγή. Γονείς πριν τα παιδιά."""
    with app.app_context():
        try:
            created = 0
            for key, pkey, name, kind, icon, sort in NODE_CATALOG:
                if MonitorNode.query.filter_by(key=key).first():
                    continue
                parent_id = None
                if pkey:
                    par = MonitorNode.query.filter_by(key=pkey).first()
                    parent_id = par.id if par else None
                db.session.add(MonitorNode(key=key, parent_id=parent_id, name=name,
                                           node_kind=kind, icon=icon or '', sort=sort, is_active=True))
                db.session.flush()
                created += 1
            if created:
                db.session.commit()
                print(f'[measurements] Φ-Α seed: {created} κόμβοι δικτύων OK')
        except Exception as e:
            db.session.rollback()
            print(f'[measurements] node catalog seed skipped: {e}')


# ── Φ2: σημεία (coarse) + αντιγραφή legacy ───────────────────────────────────
_POOL_KEYS = ['free_chlorine', 'combined_chlorine', 'ph', 'temp', 'turbidity',
              'cyanuric_acid', 'total_alkalinity', 'orp', 'backwash_done']
_ZNX_KEYS  = ['clo2_tank', 'clo2_kitchen', 'clo2_remote', 'clo2_dhw_out', 'clo2_dhw_return',
              'clo2_ro', 'temp_dhw_out', 'temp_dhw_return', 'temp_kitchen_hot', 'temp_remote_hot',
              'temp_tank', 'temp_kitchen_cold', 'temp_remote_cold', 'temp_ro', 'ph_tank',
              'location_kitchen', 'location_remote']


def _values_from(rec, keys):
    out = {}
    for k in keys:
        v = getattr(rec, k, None)
        if v is not None and v != '':
            out[k] = v
    return out


def ensure_measurement_points():
    """coarse: ΕΝΑ Area ανά Pool (template 'pool') & WaterSystem (template 'znx'). idempotent."""
    made = 0
    for p in Pool.query.all():
        if not Area.query.filter_by(legacy_kind='pool', legacy_id=p.id, template_key='pool').first():
            db.session.add(Area(hotel_id=p.hotel_id, template_key='pool', name=p.name, location=p.location,
                                is_active=True, engine_only=True, legacy_kind='pool', legacy_id=p.id))
            made += 1
    for w in WaterSystem.query.all():
        if not Area.query.filter_by(legacy_kind='water', legacy_id=w.id, template_key='znx').first():
            db.session.add(Area(hotel_id=w.hotel_id, template_key='znx', name=w.name, location=w.location,
                                is_active=True, engine_only=True, legacy_kind='water', legacy_id=w.id))
            made += 1
    db.session.commit()
    return made


def _point_map():
    m = {}
    for a in Area.query.filter(Area.engine_only.is_(True)).all():
        if a.legacy_kind and a.legacy_id and a.template_key in ('pool', 'znx'):
            m[(a.legacy_kind, a.legacy_id)] = a.id
    return m


def migrate_legacy_records():
    ensure_measurement_points()
    pm = _point_map()
    res = {'pool': 0, 'water': 0, 'pool_skip': 0, 'water_skip': 0, 'orphan': 0}
    n = 0
    for r in PoolRecord.query.all():
        if Reading.query.filter_by(source_kind='pool', source_id=r.id).first():
            res['pool_skip'] += 1; continue
        aid = pm.get(('pool', r.pool_id))
        if not aid:
            res['orphan'] += 1; continue
        db.session.add(Reading(area_id=aid, template_key='pool', user_id=r.user_id,
                               record_date=r.record_date, period=r.period, recorded_at=r.recorded_at,
                               updated_at=r.updated_at, updated_by=r.updated_by,
                               values=_json.dumps(_values_from(r, _POOL_KEYS)), notes=r.notes,
                               source_kind='pool', source_id=r.id))
        res['pool'] += 1; n += 1
        if n % 500 == 0:
            db.session.commit()
    for r in WaterRecord.query.all():
        if Reading.query.filter_by(source_kind='water', source_id=r.id).first():
            res['water_skip'] += 1; continue
        aid = pm.get(('water', r.water_system_id))
        if not aid:
            res['orphan'] += 1; continue
        db.session.add(Reading(area_id=aid, template_key='znx', user_id=r.user_id,
                               record_date=r.record_date, period=r.period, recorded_at=r.recorded_at,
                               updated_at=r.updated_at, updated_by=r.updated_by,
                               values=_json.dumps(_values_from(r, _ZNX_KEYS)), notes=r.notes,
                               source_kind='water', source_id=r.id))
        res['water'] += 1; n += 1
        if n % 500 == 0:
            db.session.commit()
    db.session.commit()
    return res


def migration_status():
    return {
        'pool_records': PoolRecord.query.count(), 'water_records': WaterRecord.query.count(),
        'pool_migrated': Reading.query.filter_by(source_kind='pool').count(),
        'water_migrated': Reading.query.filter_by(source_kind='water').count(),
        'pools': Pool.query.count(), 'systems': WaterSystem.query.count(),
    }


# ── Φ3b-2: granular σημεία ανά περιοχή ───────────────────────────────────────
def _znx_param(pkey):
    for tup in ZNX_PARAMS:
        if tup[0] == pkey:
            return tup
    if pkey in _TEXT_PARAMS:
        return (pkey, _TEXT_PARAMS[pkey], '', None, None, None, None)
    return None


def autocreate_granular_points():
    """Σπάει το ΖΝΧ σε σημεία ανά περιοχή (sub-templates + Area/δίκτυο). idempotent.
    Πισίνες: ensure ένα σημείο/πισίνα. Coarse 'znx' σημεία → ανενεργά (μένουν ως ιστορικό)."""
    made_t = made_p = 0
    for key, label, pkeys in ZNX_LOCATIONS:
        if not MonitorTemplate.query.get(key):
            db.session.add(MonitorTemplate(key=key, name=label, icon='ti-droplet', frequency='twice', sort=5, is_active=True))
            db.session.flush()
            for i, pk in enumerate(pkeys, start=1):
                tup = _znx_param(pk)
                if tup:
                    _, lab, unit, mn, mx, low, high = tup
                    db.session.add(MonitorParam(template_key=key, pkey=pk, label=lab, unit=unit or '',
                                                min_v=mn, max_v=mx, action_low=low, action_high=high, sort=i))
            made_t += 1
        _seed_periods(key)
    db.session.commit()
    for w in WaterSystem.query.all():
        for key, label, _pk in ZNX_LOCATIONS:
            if not Area.query.filter_by(legacy_kind='water', legacy_id=w.id, template_key=key).first():
                db.session.add(Area(hotel_id=w.hotel_id, template_key=key, name=label, location=w.name,
                                    is_active=True, engine_only=True, legacy_kind='water', legacy_id=w.id))
                made_p += 1
    made_p += ensure_measurement_points()  # πισίνες (+coarse znx)
    # coarse znx σημεία → ανενεργά (ιστορικό), για να μη μπαίνουν στην καταχώρηση
    for a in Area.query.filter_by(template_key='znx', engine_only=True).all():
        a.is_active = False
    db.session.commit()
    return made_t, made_p


# ── helpers UI ───────────────────────────────────────────────────────────────
def _param_input_kind(p):
    """Είδος εισόδου: προτεραιότητα στο αποθηκευμένο MonitorParam.kind (Φ3),
    αλλιώς heuristic από το pkey (συμβατότητα με legacy params)."""
    if isinstance(p, str):
        pkey, kind = p, None
    else:
        pkey, kind = getattr(p, 'pkey', '') or '', getattr(p, 'kind', None)
    if kind in ('num', 'bool', 'text'):
        return kind
    if pkey == 'backwash_done':
        return 'bool'
    if pkey.startswith('location'):
        return 'text'
    return 'num'


# ── Φ3: Βιβλιοθήκη Μετρήσεων (A1 — μία κοινή, distinct ανά pkey) ──────────────
def _library(include_inactive=False):
    """Ενιαία βιβλιοθήκη μετρήσεων (dedup ανά pkey). Όλες οι εμφανίσεις ενός pkey
    κρατιούνται συγχρονισμένες από το lib_save (A1)."""
    seen, out = set(), []
    for p in (MonitorParam.query
              .order_by(MonitorParam.category, MonitorParam.sort, MonitorParam.id).all()):
        if p.pkey in seen:
            continue
        seen.add(p.pkey)
        active = (getattr(p, 'is_active', True) is not False)
        if not include_inactive and not active:
            continue
        out.append({'pkey': p.pkey, 'label': p.label, 'unit': p.unit or '',
                    'min_v': p.min_v, 'max_v': p.max_v,
                    'action_low': p.action_low or '', 'action_high': p.action_high or '',
                    'kind': _param_input_kind(p), 'category': (getattr(p, 'category', '') or ''),
                    'active': active})
    return out


def _lib_groups():
    """Βιβλιοθήκη (όλες, incl ανενεργές) ομαδοποιημένη ανά κατηγορία για το UI."""
    groups = {}
    for m in _library(include_inactive=True):
        groups.setdefault(m['category'] or 'Γενικό', []).append(m)
    order = ['Πισίνα', 'Νερό', 'ΖΝΧ', 'Γενικό']
    keys = sorted(groups.keys(), key=lambda k: (order.index(k) if k in order else len(order), k))
    return [{'name': k, 'rows': groups[k]} for k in keys]


def _slug_key(s):
    import re
    s = (s or '').strip().lower()
    s = re.sub(r'[^a-z0-9]+', '_', s).strip('_')
    return s[:40] or None


def point_params(area):
    """Φ1: οι παράμετροι ΕΝΟΣ σημείου. Αν έχει ανατεθεί συγκεκριμένες (AreaParam) τις
    επιστρέφει με τη σειρά τους· αλλιώς ΟΛΕΣ του template του (συμβατότητα — καμία αλλαγή)."""
    tpl = MonitorTemplate.query.get(area.template_key)
    tparams = list(tpl.params) if tpl else []
    rows = AreaParam.query.filter_by(area_id=area.id).order_by(AreaParam.sort, AreaParam.id).all()
    if not rows:
        return tparams
    bykey = {p.pkey: p for p in tparams}
    out = []
    for r in rows:
        p = bykey.get(r.pkey) or MonitorParam.query.filter_by(pkey=r.pkey).first()
        if p:
            out.append(p)
    return out


def _entry_points():
    """Σημεία για καταχώρηση: ενεργά engine σημεία ΕΚΤΟΣ coarse 'znx'."""
    return (Area.query.filter(Area.is_active == True, Area.engine_only.is_(True),
                              Area.template_key != 'znx')
            .order_by(Area.hotel_id, Area.template_key, Area.name).all())


# ── Φ-Β: Δομή δικτύων (δεντρική διαχείριση κόμβων + ανάθεση σημείων) ──────────
def _node_children_map():
    m = {}
    for n in MonitorNode.query.order_by(MonitorNode.sort, MonitorNode.name).all():
        m.setdefault(n.parent_id, []).append(n)
    return m


def _point_counts():
    rows = db.session.query(Area.node_id, db.func.count(Area.id)).group_by(Area.node_id).all()
    return {nid: c for nid, c in rows if nid is not None}


def _node_tree():
    """Επίπεδη λίστα κόμβων με depth (DFS) για indentation στο UI + πλήθος σημείων."""
    cmap = _node_children_map()
    pc = _point_counts()
    out = []
    def walk(parent_id, depth):
        for n in cmap.get(parent_id, []):
            out.append({'n': n, 'depth': depth, 'np': pc.get(n.id, 0)})
            walk(n.id, depth + 1)
    walk(None, 0)
    return out


def _node_options():
    return [{'id': r['n'].id, 'label': ('— ' * r['depth']) + r['n'].name,
             'parent_id': r['n'].parent_id} for r in _node_tree()]


def _would_cycle(node_id, new_parent):
    """True αν βάζοντας new_parent ως γονέα του node_id δημιουργείται κύκλος."""
    cur, seen = new_parent, 0
    while cur is not None and seen < 200:
        if cur == node_id:
            return True
        nd = MonitorNode.query.get(cur)
        cur = nd.parent_id if nd else None
        seen += 1
    return False


def _unique_node_key(name):
    base = _slug_key(name) or 'node'
    k, i = base[:40], 1
    while MonitorNode.query.filter_by(key=k).first():
        i += 1
        k = ('%s_%d' % (base, i))[:40]
    return k


def _node_pathmap():
    """Φ-Δ-2: node_id -> {'path': 'Δίκτυο Νερού › ΖΝΧ', 'order': tree_index, 'icon': ...}.
    Χρησιμοποιείται για ομαδοποίηση «Σήμερα»/«Καταγραφή» ανά δέντρο."""
    nodes = MonitorNode.query.all()
    nm = {n.id: n.name for n in nodes}
    pm = {n.id: n.parent_id for n in nodes}
    def path(nid):
        parts, cur, s = [], nid, 0
        while cur is not None and s < 50:
            parts.append(nm.get(cur, '?'))
            cur = pm.get(cur)
            s += 1
        return ' › '.join(reversed(parts))
    out = {}
    for i, row in enumerate(_node_tree()):
        out[row['n'].id] = {'path': path(row['n'].id), 'order': i,
                            'icon': (row['n'].icon or 'ti-sitemap')}
    return out


def _node_compliance(rows):
    """Φ-Δ: roll-up % συμμόρφωσης ανά κόμβο (κάθε κόμβος = δικά του + απογόνων).
    Επιστρέφει ordered λίστα (tree) μόνο κόμβων με δεδομένα στην περίοδο."""
    direct = {}
    for r in rows:
        nid = getattr(r['point'], 'node_id', None)
        if nid is None:
            continue
        n = sum(it['n'] for it in r['params'])
        o = sum(it['out'] for it in r['params'])
        d = direct.setdefault(nid, [0, 0])
        d[0] += n
        d[1] += o
    if not direct:
        return []
    cmap = _node_children_map()
    memo = {}
    def subtree(nid):
        if nid in memo:
            return memo[nid]
        N, O = direct.get(nid, [0, 0])[0], direct.get(nid, [0, 0])[1]
        for ch in cmap.get(nid, []):
            cn, co = subtree(ch.id)
            N += cn
            O += co
        memo[nid] = (N, O)
        return memo[nid]
    out = []
    for row in _node_tree():
        N, O = subtree(row['n'].id)
        if N == 0:
            continue
        out.append({'name': row['n'].name, 'depth': row['depth'], 'n': N, 'out': O,
                    'comp': round(100.0 * (N - O) / N) if N else 100})
    return out


@app.route('/dashboard/measurements/node/save', methods=['POST'])
def measurements_node_save():
    """Προσθήκη/επεξεργασία κόμβου (ομάδα/υποομάδα) + reparent με έλεγχο κύκλου."""
    if not is_admin():
        return redirect(url_for('login'))
    f = request.form
    nid = f.get('id')
    name = (f.get('name') or '').strip()[:80]
    pid = f.get('parent_id') or None
    pid = int(pid) if pid else None
    node_kind = (f.get('node_kind') or ('group' if not pid else 'subgroup')).strip()[:20]
    icon = (f.get('icon') or '').strip()[:40]
    try:
        sort = int(f.get('sort') or 0)
    except (TypeError, ValueError):
        sort = 0

    def go(m=''):
        return redirect(url_for('measurements_console') + '?tab=structure' + (('&msg=' + m) if m else ''))

    if not name:
        return go('Λείπει το όνομα.')
    if nid:
        n = MonitorNode.query.get(int(nid))
        if not n:
            return go('Δεν βρέθηκε ο κόμβος.')
        if pid and (pid == n.id or _would_cycle(n.id, pid)):
            return go('Μη έγκυρος γονέας (κύκλος).')
        n.name, n.parent_id, n.node_kind, n.icon, n.sort = name, pid, node_kind, icon, sort
        db.session.commit()
        log_activity('meas_node_save', 'edit %s' % name)
        return go('Αποθηκεύτηκε: ' + name)
    db.session.add(MonitorNode(key=_unique_node_key(name), parent_id=pid, name=name,
                               node_kind=node_kind, icon=icon, sort=sort, is_active=True))
    db.session.commit()
    log_activity('meas_node_save', 'add %s' % name)
    return go('Προστέθηκε: ' + name)


@app.route('/dashboard/measurements/node/<int:nid>/move', methods=['POST'])
def measurements_node_move(nid):
    """Φ-Β drag&drop: άλλαξε γονέα κόμβου (drop πάνω σε άλλον ή σε «κορυφή»).
    Έλεγχος κύκλου· σειρά = στο τέλος των νέων αδελφών."""
    if not is_admin():
        return jsonify(ok=False), 403
    n = MonitorNode.query.get(nid)
    if not n:
        return jsonify(ok=False, msg='Δεν βρέθηκε ο κόμβος'), 404
    data = request.get_json(silent=True) or {}
    pv = data.get('parent_id')
    pid = int(pv) if pv else None
    if pid == nid or (pid and _would_cycle(nid, pid)):
        return jsonify(ok=False, msg='Μη έγκυρος γονέας (κύκλος)')
    n.parent_id = pid
    mx = db.session.query(db.func.max(MonitorNode.sort)).filter(MonitorNode.parent_id == pid).scalar() or 0
    n.sort = mx + 1
    db.session.commit()
    log_activity('meas_node_move', '%s -> %s' % (nid, pid))
    return jsonify(ok=True)


@app.route('/dashboard/measurements/node/<int:nid>/toggle', methods=['POST'])
def measurements_node_toggle(nid):
    if not is_admin():
        return redirect(url_for('login'))
    n = MonitorNode.query.get(nid)
    if n:
        n.is_active = not bool(n.is_active)
        db.session.commit()
    return redirect(url_for('measurements_console') + '?tab=structure')


@app.route('/dashboard/measurements/node/<int:nid>/delete', methods=['POST'])
def measurements_node_delete(nid):
    """Διαγραφή μόνο αν ΔΕΝ έχει παιδιά ούτε σημεία (ασφάλεια — μηδέν απώλεια)."""
    if not is_admin():
        return redirect(url_for('login'))
    def go(m=''):
        return redirect(url_for('measurements_console') + '?tab=structure' + (('&msg=' + m) if m else ''))
    n = MonitorNode.query.get(nid)
    if not n:
        return go()
    if MonitorNode.query.filter_by(parent_id=nid).first():
        return go('Έχει υπο-κόμβους — μετακίνησέ τους πρώτα.')
    if Area.query.filter_by(node_id=nid).first():
        return go('Έχει σημεία — ανάθεσέ τα αλλού πρώτα.')
    db.session.delete(n)
    db.session.commit()
    log_activity('meas_node_delete', n.name)
    return go('Διαγράφηκε: ' + n.name)


@app.route('/dashboard/measurements/point/<int:pid>/node', methods=['POST'])
def measurements_point_node(pid):
    """Ανάθεση σημείου σε κόμβο (node_id). Κενό = καμία (σημερινή συμπεριφορά)."""
    if not is_admin():
        return redirect(url_for('login'))
    a = Area.query.get(pid)
    if a:
        v = request.form.get('node_id') or None
        a.node_id = int(v) if v else None
        db.session.commit()
        log_activity('meas_point_node', '%s -> %s' % (a.name, a.node_id))
    return redirect(url_for('measurements_console') + '?tab=structure')


# ── Φ-Γ: αυτόματη αντιστοίχιση υπαρχόντων σημείων σε κόμβους (conservative) ───
# Συντηρητικό: όλα τα ΖΝΧ μαζί (Ζεστό νερό), Όσμωση χωριστά (Επεξεργασία),
# Πισίνες στις Πισίνες. ΔΕΝ τρέχει στο boot (αλλάζει δεδομένα) — μόνο με κουμπί.
TEMPLATE_NODE_MAP = {
    'pool':        'pools',
    'znx':         'water_hot',
    'znx_tank':    'water_hot',
    'znx_kitchen': 'water_hot',
    'znx_remote':  'water_hot',
    'znx_dhw':     'water_hot',
    'znx_ro':      'water_ro',
}


def automap_points_to_nodes(overwrite=False):
    """Θέτει Area.node_id βάσει template_key. Idempotent· από προεπιλογή ΜΟΝΟ
    σε σημεία χωρίς κόμβο (διατηρεί χειροκίνητες αναθέσεις). Καμία διαγραφή."""
    cache = {}
    def node_for(key):
        if key not in cache:
            cache[key] = MonitorNode.query.filter_by(key=key).first()
        return cache[key]
    n = 0
    for a in Area.query.filter(Area.engine_only.is_(True)).all():
        if (not overwrite) and getattr(a, 'node_id', None):
            continue
        nk = TEMPLATE_NODE_MAP.get(a.template_key)
        node = node_for(nk) if nk else None
        if node and a.node_id != node.id:
            a.node_id = node.id
            n += 1
    if n:
        db.session.commit()
    return n


@app.route('/dashboard/measurements/node/point/add', methods=['POST'])
def measurements_node_point_add():
    """#5: δημιουργία σημείου μέτρησης κάτω από οποιονδήποτε κόμβο (γενικός τύπος).
    Μετρήσεις προστίθενται μετά με σύρσιμο από την καρτέλα «Σημεία»."""
    if not is_admin():
        return redirect(url_for('login'))
    f = request.form
    nid = f.get('node_id')
    hid = f.get('hotel_id')
    name = (f.get('name') or '').strip()[:120]
    def go(m=''):
        return redirect(url_for('measurements_console') + '?tab=structure' + (('&msg=' + m) if m else ''))
    if not (nid and hid and name):
        return go('Συμπλήρωσε κόμβο, ξενοδοχείο και όνομα.')
    db.session.add(Area(hotel_id=int(hid), template_key='generic', name=name, location='',
                        is_active=True, engine_only=True, node_id=int(nid)))
    db.session.commit()
    log_activity('meas_node_point_add', '%s @node %s' % (name, nid))
    return go('Προστέθηκε σημείο: ' + name)


@app.route('/dashboard/measurements/nodes/automap', methods=['POST'])
def measurements_nodes_automap():
    if not is_admin():
        return redirect(url_for('login'))
    n = automap_points_to_nodes(overwrite=False)
    log_activity('meas_nodes_automap', '%d points' % n)
    return redirect(url_for('measurements_console') + '?tab=structure&msg=' +
                    ('Αντιστοιχίστηκαν %d σημεία στο δέντρο.' % n if n else 'Όλα τα σημεία ήταν ήδη αντιστοιχισμένα.'))


# ── ΕΝΙΑΙΑ ΚΟΝΣΟΛΑ ΡΥΘΜΙΣΕΩΝ ─────────────────────────────────────────────────
@app.route('/dashboard/measurements')
def measurements_console():
    if not is_admin():
        return redirect(url_for('login'))
    tab = request.args.get('tab', 'points')
    hmap = {h.id: h.name for h in Hotel.query.all()}
    # points grouped by hotel
    pts = Area.query.filter(Area.engine_only.is_(True)).order_by(Area.hotel_id, Area.template_key, Area.name).all()
    by_hotel = {}
    for a in pts:
        by_hotel.setdefault(a.hotel_id, []).append(a)
    points_by_hotel = [{'hotel': hmap.get(hid, '—'), 'areas': items} for hid, items in by_hotel.items()]
    # Φ3 — ενιαία βιβλιοθήκη μετρήσεων (A1): παλέτα = ενεργές· διαχείριση = όλες (ομάδες)
    library = _library(include_inactive=False)
    lib_groups = _lib_groups()
    # Φ-Β — δομή δικτύων (υπολογισμός μόνο στο tab)
    node_tree, node_opts, assign_points = [], [], []
    if tab == 'structure':
        node_tree = _node_tree()
        node_opts = _node_options()
        for a in Area.query.filter(Area.engine_only.is_(True)).order_by(Area.hotel_id, Area.name).all():
            assign_points.append({'a': a, 'hotel': hmap.get(a.hotel_id, '—'), 'node_id': getattr(a, 'node_id', None)})
    area_chips = {}
    for a in pts:
        area_chips[a.id] = [{'pkey': pp.pkey, 'label': pp.label, 'unit': pp.unit or ''} for pp in point_params(a)]
    # periods
    tpl_periods = []
    for t in MonitorTemplate.query.filter_by(is_active=True).order_by(MonitorTemplate.sort, MonitorTemplate.name).all():
        tpl_periods.append({'tpl': t, 'periods': MonitorPeriod.query.filter_by(template_key=t.key)
                            .order_by(MonitorPeriod.sort, MonitorPeriod.id).all(),
                            'nparams': len(t.params or [])})
    return render_template('measurements_console.html', tab=tab,
                           points_by_hotel=points_by_hotel, tpl_periods=tpl_periods,
                           st=migration_status(), msg=request.args.get('msg'),
                           all_hotels=Hotel.query.order_by(Hotel.name).all(),
                           all_templates=MonitorTemplate.query.filter_by(is_active=True).order_by(MonitorTemplate.name).all(),
                           param_templates=MonitorTemplate.query.order_by(MonitorTemplate.sort).all(),
                           freq_label=FREQ_LABEL, library=library, area_chips=area_chips,
                           lib_groups=lib_groups,
                           node_tree=node_tree, node_opts=node_opts, assign_points=assign_points)


@app.route('/dashboard/measurements/point/<int:area_id>/params', methods=['POST'])
def measurements_point_params(area_id):
    """Φ2: αποθήκευση μετρήσεων ενός σημείου (drag&drop) → ξαναχτίζει AreaParam."""
    if not is_admin():
        return jsonify(ok=False), 403
    area = Area.query.get(area_id)
    if not area:
        return jsonify(ok=False), 404
    data = request.get_json(silent=True) or {}
    keys, seen = [], set()
    for k in (data.get('pkeys') or []):
        k = str(k)[:40]
        if k and k not in seen:
            seen.add(k)
            keys.append(k)
    AreaParam.query.filter_by(area_id=area_id).delete()
    for i, k in enumerate(keys):
        db.session.add(AreaParam(area_id=area_id, pkey=k, sort=i))
    db.session.commit()
    log_activity('meas_point_params', '%s: %d' % (area.name, len(keys)))
    return jsonify(ok=True, n=len(keys))


@app.route('/dashboard/measurements/autocreate', methods=['POST'])
def measurements_autocreate():
    if not is_admin():
        return redirect(url_for('login'))
    t, p = autocreate_granular_points()
    log_activity('meas_autocreate', f'{t} templates, {p} σημεία')
    return redirect(url_for('measurements_console') + '?tab=points&msg=' + ('Δημιουργήθηκαν %d τύποι περιοχής + %d σημεία.' % (t, p)))


@app.route('/dashboard/measurements/point/save', methods=['POST'])
def measurements_point_save():
    if not is_admin():
        return redirect(url_for('login'))
    f = request.form
    pid = f.get('point_id')
    name = (f.get('name') or '').strip()
    loc = (f.get('location') or '').strip()
    if pid:
        a = Area.query.get(int(pid))
        if a and name:
            a.name = name[:120]; a.location = loc[:120]
    else:
        hid = f.get('hotel_id'); tk = f.get('template_key')
        if hid and tk and name:
            db.session.add(Area(hotel_id=int(hid), template_key=tk, name=name[:120], location=loc[:120],
                                is_active=True, engine_only=True))
    db.session.commit()
    return redirect(url_for('measurements_console') + '?tab=points')


@app.route('/dashboard/measurements/point/<int:pid>/toggle', methods=['POST'])
def measurements_point_toggle(pid):
    if not is_admin():
        return redirect(url_for('login'))
    a = Area.query.get(pid)
    if a:
        a.is_active = not bool(a.is_active); db.session.commit()
    return redirect(url_for('measurements_console') + '?tab=points')


# ── Φ3: CRUD Βιβλιοθήκης Μετρήσεων (admin) ───────────────────────────────────
def _lib_num(x):
    x = (x or '').strip().replace(',', '.')
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


@app.route('/dashboard/measurements/lib/save', methods=['POST'])
def measurements_lib_save():
    """Προσθήκη/επεξεργασία μέτρησης. A1: edit ενημερώνει ΟΛΕΣ τις εμφανίσεις του pkey."""
    if not is_admin():
        return redirect(url_for('login'))
    f = request.form
    pkey  = (f.get('pkey') or '').strip()
    label = (f.get('label') or '').strip()
    unit  = (f.get('unit') or '').strip()[:20]
    mn, mx = _lib_num(f.get('min_v')), _lib_num(f.get('max_v'))
    low   = (f.get('action_low') or '').strip() or None
    high  = (f.get('action_high') or '').strip() or None
    kind  = (f.get('kind') or 'num').strip()
    kind  = kind if kind in ('num', 'bool', 'text') else 'num'
    category = (f.get('category') or '').strip()[:40]

    def _go(m):
        return redirect(url_for('measurements_console') + '?tab=library&msg=' + m)

    if not label:
        return _go('Λείπει το όνομα της μέτρησης.')

    if pkey:  # επεξεργασία υπάρχουσας (όλες οι εμφανίσεις)
        rows = MonitorParam.query.filter_by(pkey=pkey).all()
        if not rows:
            return _go('Δεν βρέθηκε η μέτρηση.')
        for p in rows:
            p.label, p.unit = label[:80], unit
            p.min_v, p.max_v = mn, mx
            p.action_low, p.action_high = low, high
            p.kind, p.category = kind, category
        db.session.commit()
        log_activity('meas_lib_save', 'edit %s' % pkey)
        return _go('Ενημερώθηκε: ' + label)

    # νέα μέτρηση
    newkey = (f.get('key') or '').strip() or _slug_key(label) or ''
    newkey = (_slug_key(newkey) or '')[:40]
    if not newkey:
        return _go('Δώσε ένα key (λατινικά) για τη μέτρηση.')
    if MonitorParam.query.filter_by(pkey=newkey).first():
        return _go('Υπάρχει ήδη μέτρηση με key «%s».' % newkey)
    nsort = (db.session.query(db.func.max(MonitorParam.sort)).scalar() or 0) + 1
    db.session.add(MonitorParam(template_key=None, pkey=newkey, label=label[:80], unit=unit,
                                min_v=mn, max_v=mx, action_low=low, action_high=high,
                                kind=kind, category=category, is_active=True, sort=nsort))
    db.session.commit()
    log_activity('meas_lib_save', 'add %s' % newkey)
    return _go('Προστέθηκε: ' + label)


@app.route('/dashboard/measurements/lib/<pkey>/toggle', methods=['POST'])
def measurements_lib_toggle(pkey):
    """Ενεργοποίηση/απενεργοποίηση μέτρησης (όλες οι εμφανίσεις). Δεν θίγει
    αναθέσεις/καταγραφές — απλώς κρύβεται από την παλέτα/νέες φόρμες."""
    if not is_admin():
        return redirect(url_for('login'))
    rows = MonitorParam.query.filter_by(pkey=pkey).all()
    if rows:
        cur = (getattr(rows[0], 'is_active', True) is not False)
        for p in rows:
            p.is_active = not cur
        db.session.commit()
        log_activity('meas_lib_toggle', '%s=%s' % (pkey, not cur))
    return redirect(url_for('measurements_console') + '?tab=library')


@app.route('/dashboard/measurements/migrate', methods=['POST'])
def measurements_migrate_run():
    if not is_admin():
        return redirect(url_for('login'))
    action = request.form.get('action')
    if action == 'points':
        made = ensure_measurement_points()
        msg = 'Δημιουργήθηκαν %d σημεία (coarse).' % made
    else:
        res = migrate_legacy_records()
        msg = ('Μεταφορά — Πισίνες: %d (%d ήδη)· Νερά: %d (%d ήδη)· ορφανά: %d.'
               % (res['pool'], res['pool_skip'], res['water'], res['water_skip'], res['orphan']))
    log_activity('meas_migrate', msg)
    return redirect(url_for('measurements_console') + '?tab=migrate&msg=' + msg)


def _next_period_key(template_key):
    keys = {p.key for p in MonitorPeriod.query.filter_by(template_key=template_key).all()}
    n = 1
    while f'p{n}' in keys:
        n += 1
    return f'p{n}'


@app.route('/dashboard/measurements/period/save', methods=['POST'])
def measurements_period_save():
    if not is_admin():
        return redirect(url_for('login'))
    f = request.form
    tk = (f.get('template_key') or '').strip()
    label = (f.get('label') or '').strip()
    tm = (f.get('time') or '').strip()
    try:
        sort = int(f.get('sort') or 0)
    except (ValueError, TypeError):
        sort = 0
    pid = f.get('period_id')
    if tk and label:
        if pid:
            p = MonitorPeriod.query.get(int(pid))
            if p:
                p.label = label[:40]; p.time = tm[:5]; p.sort = sort
        else:
            db.session.add(MonitorPeriod(template_key=tk, key=_next_period_key(tk), label=label[:40], time=tm[:5], sort=sort))
        db.session.commit()
    return redirect(url_for('measurements_console') + '?tab=periods')


@app.route('/dashboard/measurements/period/<int:period_id>/delete', methods=['POST'])
def measurements_period_delete(period_id):
    if not is_admin():
        return redirect(url_for('login'))
    p = MonitorPeriod.query.get(period_id)
    if p:
        db.session.delete(p); db.session.commit()
    return redirect(url_for('measurements_console') + '?tab=periods')


# ── ΦΟΡΜΑ ΚΑΤΑΧΩΡΗΣΗΣ (operational) ──────────────────────────────────────────
@app.route('/dashboard/measurements/entry')
def measurements_entry():
    if 'user_id' not in session or not can_log():
        return redirect(url_for('login'))
    user = current_user()
    points = _entry_points()
    if not is_admin():
        _hids = scoped_hotel_ids(user)
        points = [a for a in points if a.hotel_id in _hids]
    hmap = {h.id: h.name for h in Hotel.query.all()}
    hsel = request.args.get('hotel')
    try:
        hsel = int(hsel) if hsel else None
    except (ValueError, TypeError):
        hsel = None
    hotels_with_points = sorted({a.hotel_id for a in points})
    shown = [a for a in points if (hsel is None or a.hotel_id == hsel)]
    grouped = {}
    for a in shown:
        grouped.setdefault(a.hotel_id, []).append(a)
    _pmap = _node_pathmap()
    def _grp_e(a):
        nid = getattr(a, 'node_id', None)
        if nid and nid in _pmap:
            return (_pmap[nid]['order'], _pmap[nid]['path'])
        return (999, 'Χωρίς δίκτυο — προς αντιστοίχιση')
    points_by_hotel = []
    for hid, items in grouped.items():
        subs = {}
        for a in items:
            subs.setdefault(_grp_e(a), []).append(a)
        groups = [{'title': t, 'areas': sorted(ar, key=lambda x: x.name)}
                  for (o, t), ar in sorted(subs.items(), key=lambda kv: kv[0][0])]
        points_by_hotel.append({'hotel_id': hid, 'hotel': hmap.get(hid, '—'), 'groups': groups})

    sel = tpl = params = periods = None
    recent = []
    actions = []
    pid = request.args.get('point')
    if pid:
        sel = Area.query.get(int(pid))
        if sel:
            tpl = MonitorTemplate.query.get(sel.template_key)
            params = [{'pkey': p.pkey, 'label': p.label, 'unit': p.unit, 'min_v': p.min_v,
                       'max_v': p.max_v, 'low': p.action_low, 'high': p.action_high,
                       'kind': _param_input_kind(p)} for p in point_params(sel)]
            periods = MonitorPeriod.query.filter_by(template_key=sel.template_key).order_by(MonitorPeriod.sort, MonitorPeriod.id).all()
            recent = Reading.query.filter_by(area_id=sel.id).order_by(Reading.recorded_at.desc()).limit(10).all()
            if recent:
                try:
                    actions = area_actions(recent[0])
                except Exception:
                    actions = []
    return render_template('measurements_entry.html', points_by_hotel=points_by_hotel,
                           hotel_opts=[(hid, hmap.get(hid, '—')) for hid in hotels_with_points],
                           hsel=hsel, sel=sel, tpl=tpl, params=params, periods=periods,
                           recent=recent, actions=actions, today=date.today().isoformat())


@app.route('/dashboard/measurements/entry/save', methods=['POST'])
def measurements_entry_save():
    if 'user_id' not in session or not can_log():
        return redirect(url_for('login'))
    f = request.form
    area = Area.query.get(int(f.get('area_id'))) if f.get('area_id') else None
    if not area:
        return redirect(url_for('measurements_entry'))
    if not is_admin() and area.hotel_id not in scoped_hotel_ids(current_user()):
        return redirect(url_for('measurements_entry'))
    vals = {}
    for p in point_params(area):
        kind = _param_input_kind(p)
        raw = f.get(p.pkey)
        if kind == 'bool':
            if f.get(p.pkey):
                vals[p.pkey] = True
        elif kind == 'text':
            if raw:
                vals[p.pkey] = raw.strip()
        else:
            if raw not in (None, ''):
                try:
                    vals[p.pkey] = float(str(raw).replace(',', '.'))
                except (ValueError, TypeError):
                    pass
    period = (f.get('period') or 'day').strip()
    _rd = (f.get('record_date') or '').strip()
    try:
        rdate = date.fromisoformat(_rd) if _rd else date.today()
    except ValueError:
        rdate = date.today()
    if rdate > date.today():
        rdate = date.today()
    rec = Reading(area_id=area.id, template_key=area.template_key, user_id=current_user().id,
                  record_date=rdate, period=period, values=_json.dumps(vals),
                  notes=(f.get('notes') or '').strip())
    db.session.add(rec); db.session.commit()
    log_activity('meas_entry_save', f'{area.name}/{period}')
    return redirect(url_for('measurements_entry') + '?point=%d&ok=1' % area.id)


# ── Φ3c-2b: ΕΝΙΑΙΑ «Σήμερα» (engine) — σημεία ανά περιοχή + status ημέρας ─────
@app.route('/dashboard/measurements/today')
def measurements_today():
    if 'user_id' not in session or not can_log():
        return redirect(url_for('login'))
    user = current_user()
    points = _entry_points()
    if not is_admin():
        _hids = scoped_hotel_ids(user)
        points = [a for a in points if a.hotel_id in _hids]
    today = date.today()
    aids = [a.id for a in points]
    done = {}
    cnt = {}
    if aids:
        for r in Reading.query.filter(Reading.area_id.in_(aids), Reading.record_date == today).all():
            done.setdefault(r.area_id, {})[r.period] = r
            cnt[r.area_id] = cnt.get(r.area_id, 0) + 1
    hmap = {h.id: h.name for h in Hotel.query.all()}
    _pcache = {}

    def _periods(tk):
        if tk not in _pcache:
            _pcache[tk] = MonitorPeriod.query.filter_by(template_key=tk).order_by(
                MonitorPeriod.sort, MonitorPeriod.id).all()
        return _pcache[tk]

    _pmap = _node_pathmap()
    def _grp(a):
        nid = getattr(a, 'node_id', None)
        if nid and nid in _pmap:
            pm = _pmap[nid]
            return (pm['path'], pm['icon'], pm['order'])
        return ('Χωρίς δίκτυο — προς αντιστοίχιση', 'ti-help-circle', 900)

    by_hotel = {}
    alerts = []
    total = donen = 0
    for a in points:
        prs = _periods(a.template_key)
        slots = []
        for pr in prs:
            r = done.get(a.id, {}).get(pr.key)
            slots.append({'period': pr.key, 'label': pr.label, 'time': pr.time, 'done': bool(r)})
            total += 1
            if r:
                donen += 1
                try:
                    for act in area_actions(r):
                        alerts.append({'point': a.name, 'label': act.get('label'), 'action': act.get('action')})
                except Exception:
                    pass
        cat, icon, order = _grp(a)
        groups = by_hotel.setdefault(a.hotel_id, {})
        g = groups.setdefault(cat, {'title': cat, 'icon': icon, 'order': order, 'areas': []})
        g['areas'].append({'area': a, 'slots': slots, 'count': cnt.get(a.id, 0)})
    today_by_hotel = []
    for hid, groups in by_hotel.items():
        glist = sorted(groups.values(), key=lambda x: (x['order'], x['title']))
        today_by_hotel.append({'hotel': hmap.get(hid, '—'), 'groups': glist})
    return render_template('measurements_today.html', today_by_hotel=today_by_hotel,
                           alerts=alerts, total=total, donen=donen)


# ── Φ4b: ΣΤΑΤΙΣΤΙΚΑ (σημείο × παράμετρος, μέσος/μέγ/ελάχ/εκτός ορίων) ──────────
def _stats_compute(points, dfrom, dto):
    aids = [a.id for a in points]
    amap = {a.id: a for a in points}
    pmeta = {}   # template_key -> {pkey: (label, unit, min, max, sort)}
    agg = {}     # (area_id, pkey) -> stats
    if aids:
        q = Reading.query.filter(Reading.area_id.in_(aids),
                                 Reading.record_date >= dfrom, Reading.record_date <= dto)
        for r in q.all():
            try:
                vals = _json.loads(r.values or '{}')
            except Exception:
                vals = {}
            tk = r.template_key
            if tk not in pmeta:
                t = MonitorTemplate.query.get(tk)
                pmeta[tk] = {p.pkey: (p.label, p.unit, p.min_v, p.max_v, p.sort)
                             for p in (t.params if t else [])}
            for pk, v in vals.items():
                try:
                    fv = float(v)
                except (ValueError, TypeError):
                    continue
                d = agg.setdefault((r.area_id, pk), {'n': 0, 'sum': 0.0, 'min': None, 'max': None, 'out': 0})
                d['n'] += 1; d['sum'] += fv
                d['min'] = fv if d['min'] is None else min(d['min'], fv)
                d['max'] = fv if d['max'] is None else max(d['max'], fv)
                meta = pmeta.get(tk, {}).get(pk)
                if meta:
                    mn, mx = meta[2], meta[3]
                    if (mn is not None and fv < mn) or (mx is not None and fv > mx):
                        d['out'] += 1
    rows = []
    for a in points:
        prm = pmeta.get(a.template_key, {})
        items = []
        for (aid, pk), d in agg.items():
            if aid != a.id:
                continue
            meta = prm.get(pk, (pk, '', None, None, 99))
            avg = d['sum'] / d['n'] if d['n'] else 0
            items.append({'label': meta[0], 'unit': meta[1] or '', 'sort': meta[4],
                          'n': d['n'], 'avg': round(avg, 2),
                          'min': d['min'], 'max': d['max'], 'out': d['out'],
                          'comp': round(100.0 * (d['n'] - d['out']) / d['n']) if d['n'] else 100})
        if items:
            items.sort(key=lambda x: (x['sort'], x['label']))
            rows.append({'point': a, 'params': items})
    return rows


def _stats_range():
    today = date.today()
    rng = request.args.get('range')
    if rng == 'day':
        return today, today
    if rng == 'week':
        return today - timedelta(days=today.weekday()), today
    if rng == 'month':
        return today.replace(day=1), today
    if rng == 'year':
        return today.replace(month=1, day=1), today
    df = request.args.get('from') or today.replace(day=1).isoformat()
    dt = request.args.get('to') or today.isoformat()
    try:
        dfrom = date.fromisoformat(df)
    except Exception:
        dfrom = today.replace(day=1)
    try:
        dto = date.fromisoformat(dt)
    except Exception:
        dto = today
    return dfrom, dto


def _coverage(points, dfrom, dto):
    """Κάλυψη: αναμενόμενες (μέρες × περίοδοι/template) vs πραγματικές (distinct ημέρα/περίοδος)."""
    days = (dto - dfrom).days + 1
    if days < 1:
        days = 1
    out = []
    for a in points:
        nper = MonitorPeriod.query.filter_by(template_key=a.template_key).count() or 1
        expected = days * nper
        actual = (db.session.query(Reading.record_date, Reading.period)
                  .filter(Reading.area_id == a.id, Reading.record_date >= dfrom, Reading.record_date <= dto)
                  .distinct().count())
        missing = max(0, expected - actual)
        cov = round(100.0 * min(actual, expected) / expected) if expected else 100
        out.append({'point': a, 'expected': expected, 'actual': actual, 'missing': missing, 'cov': cov})
    return out


def _stats_xlsx(rows, cov, dfrom, dto, hotel_name):
    """Εκτυπώσιμο Excel: τίτλος + περίοδος/ξενοδοχείο + στατιστικά ανά σημείο + κάλυψη."""
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.properties import PageSetupProperties
    NAVY = '193847'
    wb = Workbook(); ws = wb.active; ws.title = 'Μετρήσεις'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', fgColor=NAVY)
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    r = 1
    ws.cell(r, 1, 'Εστία — CONDIAN HOTELS · Αναφορά Μετρήσεων').font = Font(bold=True, size=15, color=NAVY); r += 1
    ws.cell(r, 1, '%s · Περίοδος: %s έως %s' % (hotel_name, dfrom.strftime('%d/%m/%Y'), dto.strftime('%d/%m/%Y'))).font = Font(size=10, color='777777'); r += 2
    cols = ['Σημείο', 'Παράμετρος', 'Μονάδα', 'Πλήθος', 'Μέσος', 'Ελάχ', 'Μέγ', 'Εκτός ορίων', 'Συμμόρφωση %']
    for c, h in enumerate(cols, 1):
        cell = ws.cell(r, c, h); cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border
    r += 1
    for row in rows:
        for it in row['params']:
            vals = [row['point'].name, it['label'], it['unit'], it['n'], it['avg'], it['min'], it['max'], it['out'], it['comp']]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(r, c, v); cell.border = border
                if c >= 4:
                    cell.alignment = Alignment(horizontal='right')
                if c == 8 and it['out']:
                    cell.font = Font(color='B91C1C', bold=True)
            r += 1
    widths = [26, 22, 9, 9, 9, 9, 9, 12, 13]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, c).column_letter].width = w
    # Κάλυψη
    r += 2
    ws.cell(r, 1, 'Κάλυψη (τι έχει μετρηθεί / τι λείπει)').font = Font(bold=True, size=12, color=NAVY); r += 1
    for c, h in enumerate(['Σημείο', 'Αναμενόμενες', 'Έγιναν', 'Λείπουν', 'Κάλυψη %'], 1):
        cell = ws.cell(r, c, h); cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border
    r += 1
    for cc in cov:
        vals = [cc['point'].name, cc['expected'], cc['actual'], cc['missing'], cc['cov']]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v); cell.border = border
            if c >= 2:
                cell.alignment = Alignment(horizontal='right')
            if c == 4 and cc['missing']:
                cell.font = Font(color='B91C1C', bold=True)
        r += 1
    bio = _io.BytesIO(); wb.save(bio); return bio.getvalue()


@app.route('/dashboard/measurements/stats')
def measurements_stats():
    if 'user_id' not in session or not can_log():
        return redirect(url_for('login'))
    user = current_user()
    points = Area.query.filter(Area.engine_only.is_(True)).order_by(
        Area.hotel_id, Area.template_key, Area.name).all()
    if not is_admin():
        _hids = scoped_hotel_ids(user)
        points = [a for a in points if a.hotel_id in _hids]
    hmap = {h.id: h.name for h in Hotel.query.all()}
    hotel_ids = sorted({a.hotel_id for a in points})
    hsel = request.args.get('hotel')
    try:
        hsel = int(hsel) if hsel else None
    except (ValueError, TypeError):
        hsel = None
    if hsel:
        points = [a for a in points if a.hotel_id == hsel]
    dfrom, dto = _stats_range()
    rows = _stats_compute(points, dfrom, dto)
    cov = _coverage(points, dfrom, dto)
    hotel_name = hmap.get(hsel, 'Όλα τα ξενοδοχεία')
    _fmt = request.args.get('fmt')
    if _fmt == 'xlsx':
        data = _stats_xlsx(rows, cov, dfrom, dto, hotel_name)
        return Response(data, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition': 'attachment; filename=measurements-%s.xlsx' % dto.isoformat()})
    if _fmt == 'csv':
        lines = ['Σημείο,Παράμετρος,Μονάδα,Πλήθος,Μέσος,Ελάχ,Μέγ,Εκτός ορίων,Συμμόρφωση %']
        for r in rows:
            for it in r['params']:
                lines.append('%s,%s,%s,%d,%s,%s,%s,%d,%d' % (
                    (r['point'].name or '').replace(',', ' '), it['label'].replace(',', ' '),
                    it['unit'], it['n'], it['avg'], it['min'], it['max'], it['out'], it['comp']))
        csv = '﻿' + '\n'.join(lines)
        return Response(csv, mimetype='text/csv',
                        headers={'Content-Disposition': 'attachment; filename=measurements-stats.csv'})
    # KPIs + δεδομένα γραφημάτων
    total_n = sum(it['n'] for r in rows for it in r['params'])
    total_out = sum(it['out'] for r in rows for it in r['params'])
    overall = round(100.0 * (total_n - total_out) / total_n) if total_n else 100
    pt_labels = []; pt_comp = []
    for r in rows:
        n = sum(it['n'] for it in r['params']); o = sum(it['out'] for it in r['params'])
        pt_labels.append(r['point'].name or '—')
        pt_comp.append(round(100.0 * (n - o) / n) if n else 100)
    param_out = {}
    for r in rows:
        for it in r['params']:
            if it['out']:
                param_out[it['label']] = param_out.get(it['label'], 0) + it['out']
    po = sorted(param_out.items(), key=lambda kv: -kv[1])[:12]
    total_missing = sum(c['missing'] for c in cov)
    kpis = {'total_n': total_n, 'total_out': total_out, 'overall': overall,
            'npoints': len(rows), 'nparams_out': len(param_out), 'missing': total_missing}
    charts = {'pt_labels': pt_labels, 'pt_comp': pt_comp,
              'po_labels': [k for k, _ in po], 'po_vals': [v for _, v in po]}
    return render_template('measurements_stats.html', rows=rows, cov=cov, kpis=kpis, charts=charts,
                           node_comp=_node_compliance(rows),
                           dfrom=dfrom.isoformat(), dto=dto.isoformat(),
                           hotel_opts=[(hid, hmap.get(hid, '—')) for hid in hotel_ids],
                           hsel=hsel, hotel_name=hotel_name, cur_range=request.args.get('range', ''))


print('measurements module loaded (Φ1→Φ4 ενοποίηση μετρήσεων)')
