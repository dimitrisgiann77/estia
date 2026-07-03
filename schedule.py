# -*- coding: utf-8 -*-
"""
Εστία — Module «Πρόγραμμα Εργασίας» (Βάρδιες) — Φάση 1 (v12.40)
==============================================================
Plug-in: `import schedule` από το ΤΕΛΟΣ του app.py (αφού οριστούν app/db/helpers,
ΠΡΙΝ το init_db() ώστε το create_all να πιάσει τους νέους πίνακες).
Spec: 02_MODULES_ESTIA/ΠΡΟΓΡΑΜΜΑ_ΕΡΓΑΣΙΑΣ/00_SPEC.md

Αρχές:
 - Κάθε εργαζόμενος = User (login_enabled=false για το μαζικό προσωπικό).
 - Τμήματα οργανισμού-wide· ξενοδοχείο = ξεχωριστή διάσταση (home_hotel_id).
 - Κωδικολόγιο μισθοδοσίας (8 κωδικοί) + μηχανή ωρών (8.5 normal, έξτρα, σπαστή, νυχτερινή).
 - Workflow υποβολής: WeekPlan(τμήμα) -> ScheduleSubmission(ενοποιημένη/ξενοδοχείο),
   κανόνες (R1 ≥1 ρεπό/εβδ.), χρονικά κλειδώματα (Πέμπτη), versions + diff.
"""
import os, re, json, unicodedata, threading
from datetime import datetime, date, timedelta, time as dtime
from flask import request, redirect, url_for, render_template, session, jsonify, Response
from app import (app, db, current_user, is_admin, allowed_hotels, notify, notify_admins,
                 log_activity, Hotel, User, Setting, ROLE_RANK, role_rank, BASE_DIR,
                 send_email, EMAIL_TO_LIST, active_hotel_id)
from werkzeug.security import generate_password_hash

# ── Σταθερές μηχανής ──────────────────────────────────────────────────────────
NORMAL_HOURS = 8.5          # κανονικό ωράριο/μέρα (περιλαμβάνει 30' διάλειμμα)
HOTEL_CODES  = {'AST', 'CNT', 'SRG', 'PSV', 'PLM', 'IRO', 'CND', 'ΗΡΩ'}
HOTEL_NORM   = {'ΗΡΩ': 'IRO'}

# Κωδικολόγιο βαρδιών (από το σύστημα μισθοδοσίας — seed στο ShiftType)
SHIFT_CODES = [
    # code,  label,                 color,     counts_as_work, note, ergani
    ('ΕΡΓ',   'Εργασία',             '#16a34a', True,  'έξτρα = διάρκεια − 8,5', 'WORK'),
    ('ΔΡ',    'Δουλεμένο Ρεπό',      '#0d9488', True,  'δηλωμένο ρεπό που εργάστηκε — μετρά εργάσιμη', 'WORK'),
    ('ΑΝ',    'Ρεπό',                '#185FA5', False, 'ρεπό/εβδομάδα',          'OFF'),
    ('ΑΔ',    'Κανονική άδεια',      '#64748b', False, 'το βγάζει το λογιστήριο','LEAVE'),
    ('ΑΣΘ',   'Αναρρωτική',          '#0e9f6e', False, 'το βγάζει το λογιστήριο','SICK'),
    ('ΑΠ',    'Αδικαιολόγητος απών', '#dc2626', False, '—',                      'ABSENT'),
    ('ΑΡΓ',   'Αργία (ρεπό αργίας)', '#7e22ce', False, 'αν δουλέψει → ΕΡΓ',      'HOLIDAY'),
    ('ΑΝΕΥ',  'Άνευ αποδοχών',       '#ea580c', False, 'το βγάζει το λογιστήριο','UNPAID'),
    ('Ειδ.Α', 'Ειδική άδεια',        '#0891b2', False, 'το βγάζει το λογιστήριο','SPECIAL'),
]

# Κανονικά τμήματα (οργανισμού-wide) + aliases (καθαρισμός Ελλ/Αγγλ διπλών)
DEPARTMENTS = [
    ('Housekeeping',  'Housekeeping', '#0ea5e9', ['housekeeping', 'hk']),
    ('Reception',     'Reception',    '#6366f1', ['reception', 'ρεσεψιον', 'υποδοχη']),
    ('Service',       'Service',      '#f59e0b', ['service', 'σερβις']),
    ('Kitchen',       'Kitchen',      '#ef4444', ['kitchen', 'κουζινα']),
    ('Maintenance',   'Maintenance',  '#64748b', ['maintenance', 'συντηρηση']),
    ('Management',    'Management',   '#7e22ce', ['management', 'διοικηση', 'operations']),
    ('Pool Bar',      'Pool Bar',     '#06b6d4', ['poolbar', 'pool bar', 'pool']),
    ('Bar',           'Bar',          '#0891b2', ['bar', 'μπαρ', 'barman']),
    ('Restaurant',    'Restaurant',   '#d97706', ['restaurant', 'oliva', 'μπουφε']),
    ("Kid's Club",    "Kid's Club",   '#ec4899', ["kid's club", 'kids club', 'kidsclub']),
    ('Bellboy',       'Bellboy',      '#0d9488', ['bellboy', 'γκρουμ', 'groom']),
    ('Replacement',   'Replacement',  '#94a3b8', ['replacement', 'replacant', 'replacant']),
]

WEEKDAYS_EL = ['Δευτέρα', 'Τρίτη', 'Τετάρτη', 'Πέμπτη', 'Παρασκευή', 'Σάββατο', 'Κυριακή']
DOW_EL = {0: 'Δευτέρα', 1: 'Τρίτη', 2: 'Τετάρτη', 3: 'Πέμπτη', 4: 'Παρασκευή', 5: 'Σάββατο', 6: 'Κυριακή'}
MONTHS_EL = ['', 'Ιανουάριος', 'Φεβρουάριος', 'Μάρτιος', 'Απρίλιος', 'Μάιος', 'Ιούνιος',
             'Ιούλιος', 'Αύγουστος', 'Σεπτέμβριος', 'Οκτώβριος', 'Νοέμβριος', 'Δεκέμβριος']

WP_STATUS = ('draft', 'ready', 'submitted', 'locked')
WP_LABELS  = {'draft': 'Πρόχειρο', 'ready': 'Έτοιμο', 'submitted': 'Υποβλήθηκε', 'locked': 'Κλειδωμένο'}


# ── Normalization helpers ─────────────────────────────────────────────────────
def _acc(s):
    return ''.join(c for c in unicodedata.normalize('NFD', str(s)) if unicodedata.category(c) != 'Mn')

def _norm(s):
    return re.sub(r'[^a-zα-ω0-9]', '', _acc(s).strip().lower()) if s else ''

def monday_of(d):
    """Δευτέρα της εβδομάδας που περιέχει το d."""
    return d - timedelta(days=d.weekday())


# ── ΜΟΝΤΕΛΑ ───────────────────────────────────────────────────────────────────
class DepartmentGroup(db.Model):
    """v12.176 — master group τμημάτων (Κουζίνες/F&B/Όροφοι...). Single source· adjustable."""
    id      = db.Column(db.Integer, primary_key=True)
    name    = db.Column(db.String(60), nullable=False)   # v12.183 — όχι global unique (μοναδικό μεταξύ αδελφών)
    name_en = db.Column(db.String(60))
    color   = db.Column(db.String(9), default='#64748b')
    parent_id = db.Column(db.Integer)   # v12.181 — υποομάδες (self-FK, απεριόριστο βάθος)
    supervisor_user_id = db.Column(db.Integer)   # v12.182 — υπεύθυνος ομάδας/υποομάδας
    active  = db.Column(db.Boolean, default=True)
    sort    = db.Column(db.Integer, default=0)

class JobPosition(db.Model):
    """v12.178 — θέση εργασίας (Chef/Μάγειρας/Σερβιτόρος...). Seed από MgmtAssignment.position· adjustable."""
    id       = db.Column(db.Integer, primary_key=True)
    name     = db.Column(db.String(80), unique=True, nullable=False)
    color    = db.Column(db.String(9), default='#64748b')
    group_id = db.Column(db.Integer)   # προαιρετικά συνδεδεμένη σε ομάδα
    active   = db.Column(db.Boolean, default=True)
    sort     = db.Column(db.Integer, default=0)

class Department(db.Model):
    id        = db.Column(db.Integer, primary_key=True)
    name      = db.Column(db.String(60), unique=True, nullable=False)
    name_en   = db.Column(db.String(60))
    color     = db.Column(db.String(9), default='#64748b')
    aliases   = db.Column(db.Text)            # JSON list κανονικοποιημένων aliases
    active    = db.Column(db.Boolean, default=True)
    sort      = db.Column(db.Integer, default=0)
    is_leadership = db.Column(db.Boolean, default=False)   # v12.171 — λωρίδα Διεύθυνσης
    group_id  = db.Column(db.Integer)   # v12.176 soft FK -> department_group.id

class HotelDepartment(db.Model):
    """v12.166 — ποια (κοινά) τμήματα έχει κάθε ξενοδοχείο. Single source· προσθετικό.
    Κενό για ένα ξενοδοχείο = (fallback) όλα τα ενεργά τμήματα."""
    id            = db.Column(db.Integer, primary_key=True)
    hotel_id      = db.Column(db.Integer, db.ForeignKey('hotel.id'), nullable=False)
    department_id = db.Column(db.Integer, db.ForeignKey('department.id'), nullable=False)
    supervisor_user_id = db.Column(db.Integer)   # v12.167 soft FK -> user.id· owner-screen=οργανόγραμμα
    __table_args__ = (db.UniqueConstraint('hotel_id', 'department_id', name='uq_hotel_dept'),)

class ShiftType(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    code          = db.Column(db.String(12), unique=True, nullable=False)
    label         = db.Column(db.String(60))
    color         = db.Column(db.String(9), default='#64748b')
    default_start = db.Column(db.String(5))   # 'HH:MM'
    default_end   = db.Column(db.String(5))
    counts_as_work= db.Column(db.Boolean, default=False)
    count_as      = db.Column(db.String(10))   # v12.207: work|extra|repo|absence (παραμετροποιήσιμο)
    break_deduct  = db.Column(db.Boolean)       # v12.207: αφαίρεση 30' αν παρουσία ≥6ω
    payroll_note  = db.Column(db.String(120))
    ergani_type   = db.Column(db.String(16))
    active        = db.Column(db.Boolean, default=True)
    sort          = db.Column(db.Integer, default=0)

class EmploymentProfile(db.Model):
    id              = db.Column(db.Integer, primary_key=True)
    user_id         = db.Column(db.Integer, db.ForeignKey('user.id'), unique=True, nullable=False)
    agreement_amount= db.Column(db.Float)               # συμφωνημένο μηνιαίο ποσό
    days_per_month  = db.Column(db.Integer, default=26)
    hours_per_day   = db.Column(db.Float, default=8.0)
    agreement_type  = db.Column(db.String(20), default='Μηνιαίος')   # Μηνιαίος/Management/Ωρομίσθιος
    position        = db.Column(db.String(80))          # θέση/ειδικότητα
    hired_at        = db.Column(db.Date)
    left_at         = db.Column(db.Date)
    status          = db.Column(db.String(20), default='Ενεργός')

    @property
    def day_wage(self):
        try:
            return round(self.agreement_amount / (self.days_per_month or 26), 4) if self.agreement_amount else 0.0
        except Exception:
            return 0.0
    @property
    def hour_wage(self):
        dw = self.day_wage
        return round(dw / 8.0, 4) if dw else 0.0

class ShiftAssignment(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    user_id       = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, index=True)
    work_date     = db.Column(db.Date, nullable=False, index=True)
    shift_code    = db.Column(db.String(12), default='ΕΡΓ')     # κωδικός ShiftType
    segments      = db.Column(db.Text)            # JSON [{'start':'07:00','end':'15:30'}, ...]
    work_hotel_id = db.Column(db.Integer, db.ForeignKey('hotel.id'))   # != home αν δανεικός
    note          = db.Column(db.String(200))
    created_by    = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_at    = db.Column(db.DateTime, default=datetime.now)
    updated_at    = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    import_hash   = db.Column(db.String(40), index=True)
    # v12.131: επιτρέπονται ΠΟΛΛΕΣ βάρδιες/μέρα (αφαιρέθηκε το unique user_id+work_date)

class NameLink(db.Model):
    """Επιβεβαιωμενη συνδεση ονοματος προγραμματος -> master προφιλ.
    Αγκυρωνεται στο master (που κουβαλα τον αριθμο E0xxx + ΑΦΜ). Το φτιαχνει ΜΟΝΟ ο χρηστης."""
    id         = db.Column(db.Integer, primary_key=True)
    norm_name  = db.Column(db.String(120), unique=True, index=True, nullable=False)
    user_id    = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, index=True)
    raw_name   = db.Column(db.String(120))
    created_at = db.Column(db.DateTime, default=datetime.now)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))

class PendingShift(db.Model):
    """Προθαλαμος ταυτοποιησης: βαρδιες ονοματων ΧΩΡΙΣ επιβεβαιωμενη συνδεση.
    ΔΕΝ φτιαχνεται User -> δεν μπαινουν πουθενα στην πλατφορμα μεχρι να τα ταυτοποιησεις."""
    id            = db.Column(db.Integer, primary_key=True)
    norm_name     = db.Column(db.String(120), index=True, nullable=False)
    raw_name      = db.Column(db.String(120))
    epon          = db.Column(db.String(80))
    onoma         = db.Column(db.String(80))
    hotel_tag     = db.Column(db.String(20))    # upok raw
    dept_raw      = db.Column(db.String(120))
    employer      = db.Column(db.String(120))
    work_date     = db.Column(db.Date, index=True)
    shift_code    = db.Column(db.String(12))
    segments      = db.Column(db.Text)
    work_hotel_id = db.Column(db.Integer, db.ForeignKey('hotel.id'))
    import_hash   = db.Column(db.String(40), index=True)
    created_at    = db.Column(db.DateTime, default=datetime.now)
    created_by    = db.Column(db.Integer, db.ForeignKey('user.id'))
    # v12.132: επιτρέπονται πολλές βάρδιες/μέρα στον προθάλαμο (αφαιρέθηκε το unique norm_name+work_date)

class SchedulePeriodMark(db.Model):
    """v12.199 — Δήλωση manager στην ΕΠΙΣΤΡΟΦΗ εργαζομένου στο Πρόγραμμα (μετά από κενό ή αλλαγή ξεν.).
    Anchored στην ημερομηνία επιστροφής. kind: 'new'=νέα περίοδος (default) | 'continue'=συνέχεια (ενώνει το κενό ίδιου ξεν.)."""
    id         = db.Column(db.Integer, primary_key=True)
    user_id    = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, index=True)
    work_date  = db.Column(db.Date, nullable=False, index=True)
    hotel_id   = db.Column(db.Integer, db.ForeignKey('hotel.id'))
    kind       = db.Column(db.String(12), default='new')
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_at = db.Column(db.DateTime, default=datetime.now)

class Holiday(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    hol_date    = db.Column(db.Date, unique=True, nullable=False)
    description = db.Column(db.String(120))
    year        = db.Column(db.Integer, index=True)

class ScheduleRule(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    code        = db.Column(db.String(20), unique=True, nullable=False)
    description = db.Column(db.String(200))
    severity    = db.Column(db.String(8), default='block')   # block | warn
    params      = db.Column(db.Text)            # JSON
    active      = db.Column(db.Boolean, default=True)

class WeekPlan(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    hotel_id      = db.Column(db.Integer, db.ForeignKey('hotel.id'), nullable=False, index=True)
    department_id = db.Column(db.Integer, db.ForeignKey('department.id'), nullable=False, index=True)
    week_start    = db.Column(db.Date, nullable=False, index=True)   # Δευτέρα
    status        = db.Column(db.String(12), default='draft')
    updated_at    = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    updated_by    = db.Column(db.Integer, db.ForeignKey('user.id'))
    __table_args__ = (db.UniqueConstraint('hotel_id', 'department_id', 'week_start', name='uq_weekplan'),)

class ScheduleSubmission(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    hotel_id      = db.Column(db.Integer, db.ForeignKey('hotel.id'), nullable=False, index=True)
    week_start    = db.Column(db.Date, nullable=False, index=True)
    version       = db.Column(db.Integer, default=1)
    parent_version= db.Column(db.Integer)
    status        = db.Column(db.String(12), default='submitted')
    snapshot      = db.Column(db.Text)          # JSON: τι ακριβώς δηλώθηκε
    changes       = db.Column(db.Text)          # JSON diff από parent_version
    submitted_at  = db.Column(db.DateTime, default=datetime.now)
    submitted_by  = db.Column(db.Integer, db.ForeignKey('user.id'))


# ── MIGRATION (στήλες User) ───────────────────────────────────────────────────
def ensure_schedule_columns():
    """Auto-migration: νέες στήλες στον User (μη καταστροφικό, race-safe μέσω app._add_col)."""
    try:
        from app import _add_col
    except Exception:
        return
    with app.app_context():
        _add_col('user', 'department_id',     'department_id INTEGER')
        _add_col('user', 'employer',          'employer VARCHAR(120)')
        _add_col('user', 'subunit',           'subunit VARCHAR(20)')
        _add_col('user', 'home_hotel_id',     'home_hotel_id INTEGER')
        _add_col('user', 'login_enabled',     'login_enabled BOOLEAN')
        _add_col('user', 'employment_active', 'employment_active BOOLEAN')
        _add_col('hotel_department', 'supervisor_user_id', 'supervisor_user_id INTEGER')  # v12.167
        _add_col('department', 'is_leadership', 'is_leadership BOOLEAN')  # v12.171
        _add_col('department', 'group_id', 'group_id INTEGER')  # v12.176
        _add_col('department_group', 'parent_id', 'parent_id INTEGER')  # v12.181
        _add_col('department_group', 'supervisor_user_id', 'supervisor_user_id INTEGER')  # v12.182
        _add_col('shift_type', 'count_as', "count_as VARCHAR(10)")  # v12.207
        _add_col('shift_type', 'break_deduct', 'break_deduct BOOLEAN')  # v12.207
        try:  # backfill κατηγοριών (idempotent)
            for _st in ShiftType.query.all():
                if not _st.count_as:
                    _st.count_as = ('extra' if _st.code == 'ΕΩ' else 'repo' if _st.code == 'ΑΝ'
                                    else 'work' if _st.counts_as_work else 'absence')
                if _st.break_deduct is None:
                    _st.break_deduct = _st.count_as in ('work', 'extra')
            db.session.commit()
        except Exception:
            db.session.rollback()
        # v12.131 — επέτρεψε πολλές βάρδιες/μέρα: ρίξε το unique constraint (μόνο Postgres· αναστρέψιμο)
        try:
            from sqlalchemy import text as _text
            if db.engine.dialect.name == 'postgresql':
                db.session.execute(_text('ALTER TABLE shift_assignment DROP CONSTRAINT IF EXISTS uq_user_date'))
                db.session.execute(_text('ALTER TABLE pending_shift DROP CONSTRAINT IF EXISTS uq_pending_name_date'))
                db.session.execute(_text('ALTER TABLE department_group DROP CONSTRAINT IF EXISTS department_group_name_key'))  # v12.183
                db.session.commit()
        except Exception as _e:
            db.session.rollback()


# ── ΜΗΧΑΝΗ ΩΡΩΝ ───────────────────────────────────────────────────────────────
_RANGE_RE = re.compile(r'(\d{1,2}):(\d{2})\s*[-–—]\s*(\d{1,2}):(\d{2})')

def segments_hours(segments):
    """segments: list[{'start':'HH:MM','end':'HH:MM'}] -> συνολικές ώρες (νυχτερινή->ίδια μέρα)."""
    tot = 0.0
    for seg in (segments or []):
        try:
            sh, sm = [int(x) for x in str(seg['start']).split(':')]
            eh, em = [int(x) for x in str(seg['end']).split(':')]
        except Exception:
            continue
        a = sh * 60 + sm
        b = eh * 60 + em
        if b <= a:
            b += 1440           # περνά μεσάνυχτα -> μετρά στη μέρα έναρξης
        tot += (b - a) / 60.0
    return round(tot, 4)

def extra_hours(total):
    return round(max(0.0, total - NORMAL_HOURS), 4) if total and total > 0 else 0.0

# v12.118 — Διάλειμμα: 30' αφαιρούνται από τις πληρωτέες ώρες ΜΟΝΟ σε βάρδιες ≥6ω παρουσίας.
BREAK_HOURS = 0.5
BREAK_MIN_PRESENCE = 6.0
def shift_break(presence):
    return BREAK_HOURS if (presence and presence >= BREAK_MIN_PRESENCE) else 0.0

# v12.386 (P-060 δευτ.) — request-scoped cache: ShiftType-by-code + Holidays υπολογίζονται ΜΙΑ φορά
# ανά request (αντί query ανά βάρδια/aggregate). Εκτός request context → fallback (ίδια συμπεριφορά).
def _req_cache():
    try:
        from flask import g, has_request_context
        if not has_request_context():
            return None
        c = getattr(g, '_estia_c', None)
        if c is None:
            c = {}; g._estia_c = c
        return c
    except Exception:
        return None
def _st_map():
    c = _req_cache()
    if c is not None:
        m = c.get('stmap')
        if m is None:
            m = {s.code: s for s in ShiftType.query.all()}; c['stmap'] = m
        return m
    return {s.code: s for s in ShiftType.query.all()}
def _holidays():
    c = _req_cache()
    if c is not None:
        h = c.get('hol')
        if h is None:
            h = {x.hol_date for x in Holiday.query.all()}; c['hol'] = h
        return h
    return {x.hol_date for x in Holiday.query.all()}

def worked_hours(a):
    """Πληρωτέες ώρες = παρουσία − διάλειμμα. Το διάλειμμα (30' αν ≥6ω) εφαρμόζεται
    ΑΝΑ ΚΩΔΙΚΟ (ShiftType.break_deduct, παραμετροποιήσιμο στις Ρυθμίσεις)."""
    p = assignment_hours(a)
    _st = _st_map().get(a.shift_code)
    _bd = _st.break_deduct if (_st and _st.break_deduct is not None) else (a.shift_code in ('ΕΡΓ', 'ΕΩ'))
    brk = BREAK_HOURS if (_bd and p and p >= BREAK_MIN_PRESENCE) else 0.0
    return round(max(0.0, p - brk), 4)

# v12.119 — Επιτρεπτοί κωδικοί καταχώρησης (ρυθμιζόμενο από Πρόγραμμα·Ρυθμίσεις).
def _entry_codes_setting(role='admin'):
    key = 'sched_entry_codes_mgr' if role == 'manager' else 'sched_entry_codes'
    row = Setting.query.get(key)
    if row and row.value:
        try:
            vals = [c for c in json.loads(row.value) if c]
            return vals or None
        except Exception:
            pass
    return None  # None = όλοι οι ενεργοί κωδικοί
def _current_entry_role():
    return 'admin' if is_admin() else 'manager'
def entry_shift_types(role=None):
    if role is None:
        role = _current_entry_role()
    sts = ShiftType.query.filter_by(active=True).order_by(ShiftType.sort).all()
    allow = _entry_codes_setting(role)
    if allow is not None:
        sts = [s for s in sts if s.code in allow]
    return sts

def _validate_work_code(code, segs):
    """v12.120 — ΕΡΓ = ΑΚΡΙΒΩΣ 8ω + 30' (8,5 παρουσία)· διαφορετικά → ΕΩ (Έξτρα Ώρες)."""
    if code in ('ΕΡΓ', 'ΔΡ'):
        pres = segments_hours(segs)
        if abs(pres - NORMAL_HOURS) > 0.001:
            _lbl = 'ΔΡ' if code == 'ΔΡ' else 'ΕΡΓ'
            return (False, 'Ο κωδικός ' + _lbl + ' είναι μόνο για 8ω + 30΄ διάλειμμα (8,5 παρουσία). Για διαφορετικές ώρες χρησιμοποίησε τον κωδικό ΕΩ.')
    if code == 'ΕΩ' and not segs:
        return (False, 'Ο κωδικός ΕΩ χρειάζεται ώρες — όρισε βάρδια.')
    return (True, '')

def parse_cell(v):
    """Κελί Excel -> (shift_code, segments[list], work_hotel_tag).
    Πιστή μεταφορά λογικής ENGINE_v2: τμήματα ΕΡΓ, κωδικοί, cross-hotel tag."""
    if v is None:
        return (None, [], None)
    s = str(v).strip()
    if not s:
        return (None, [], None)
    tag = None
    m = re.match(r'^\s*(AST|CNT|SRG|PSV|PLM|IRO|CND|ΗΡΩ)\b\s*[-: ]*\s*(.*)$', s, re.I)
    if m and (m.group(2).strip() == '' or 'ΕΡΓ' in m.group(2).upper() or _RANGE_RE.search(m.group(2))):
        tag = HOTEL_NORM.get(m.group(1).upper(), m.group(1).upper())
        s = m.group(2).strip()
    up = _acc(s).upper()
    ranges = _RANGE_RE.findall(s)
    if ranges or 'ΕΡΓ' in up or (tag is not None and s):
        segs = [{'start': f'{int(h1):02d}:{m1}', 'end': f'{int(h2):02d}:{m2}'} for h1, m1, h2, m2 in ranges]
        if segs:
            _pres = segments_hours(segs)
            _code = 'ΕΡΓ' if abs(_pres - NORMAL_HOURS) < 0.001 else 'ΕΩ'
        else:
            _code = 'ΕΡΓ'
        return (_code, segs, tag)
    # κωδικοί χωρίς ώρες (σειρά: ειδικοί πριν τα γενικά)
    for c in ['ΑΣΘ', 'ΑΝΕΥ', 'ΑΡΓ', 'ΕΙΔ', 'ΑΔ', 'ΑΠ', 'ΑΑ', 'ΑΝ']:
        if up.startswith(_acc(c).upper()):
            return ({'ΕΙΔ': 'Ειδ.Α', 'ΑΑ': 'ΑΠ'}.get(c, c), [], None)
    return (None, [], None)   # άγνωστο -> αγνοείται

def assignment_hours(a):
    try:
        segs = json.loads(a.segments) if a.segments else []
    except Exception:
        segs = []
    return segments_hours(segs)

def _shift_count_as(code):
    """v12.207 — Κατηγορία μέτρησης κωδικού (data-driven από ShiftType.count_as).
    work=εργάσιμη · extra=έξτρα ώρες · repo=ρεπό · absence=άδεια/απουσία."""
    st = _st_map().get(code)
    if st and st.count_as:
        return st.count_as
    if code == 'ΕΡΓ': return 'work'
    if code == 'ΕΩ': return 'extra'
    if code == 'ΑΝ': return 'repo'
    if st and st.counts_as_work: return 'work'
    return 'absence'

def is_work_code(code):
    # «παράγει δουλεμένες ώρες» = εργάσιμη Ή έξτρα (διατηρεί την υπάρχουσα σημασία)
    return _shift_count_as(code) in ('work', 'extra')

def is_extra_code(code):
    return _shift_count_as(code) == 'extra'

def is_repo_code(code):
    return _shift_count_as(code) == 'repo'

def schedule_span(uid):
    """v12.197 — Πρώτη/τελευταία εμφάνιση εργαζομένου στο ΠΡΟΓΡΑΜΜΑ (ΟΛΕΣ οι βάρδιες,
    μαζί ΡΕΠΟ/άδεια = κάθε χρήση του από manager). Read-only, ζωντανός υπολογισμός.
    Επιστρέφει (first_date, last_date) ή (None, None) αν δεν υπάρχει καμία βάρδια."""
    from sqlalchemy import func as _func
    row = (db.session.query(_func.min(ShiftAssignment.work_date),
                            _func.max(ShiftAssignment.work_date))
           .filter(ShiftAssignment.user_id == uid,
                   ShiftAssignment.work_date.isnot(None))
           .one())
    return (row[0], row[1]) if row else (None, None)

# v12.198 — strict: κάθε ≥1 κενή μέρα Ή αλλαγή ξενοδοχείου ξεκινά νέα περίοδο χρήσης.
# (εύκολα ρυθμιζόμενο· μελλοντικά admin setting + manager pop-up στην επιστροφή)
SCHEDULE_PERIOD_MAX_GAP_DAYS = 1

def schedule_periods(uid):
    """v12.198 — Περίοδοι «χρήσης» εργαζομένου στο ΠΡΟΓΡΑΜΜΑ: συνεχόμενα διαστήματα
    παρουσίας (ΟΛΕΣ οι βάρδιες, μαζί ΡΕΠΟ/άδεια), σπασμένα ΑΝΑ ΞΕΝΟΔΟΧΕΙΟ και σε κάθε
    κενό > SCHEDULE_PERIOD_MAX_GAP_DAYS μερών. Read-only, ζωντανός υπολογισμός.
    Επιστρέφει λίστα dict {hotel_id, hotel, start, end, days} — νεότερη περίοδος πρώτη.
    days = πλήθος ΔΙΑΦΟΡΕΤΙΚΩΝ ημερομηνιών (πολλές βάρδιες/μέρα μετράνε 1)."""
    u = User.query.get(uid)
    home = getattr(u, 'home_hotel_id', None) if u else None
    hotels = {h.id: h.name for h in Hotel.query.all()}
    rows = (db.session.query(ShiftAssignment.work_date, ShiftAssignment.work_hotel_id)
            .filter(ShiftAssignment.user_id == uid,
                    ShiftAssignment.work_date.isnot(None))
            .all())
    seen = sorted({(d, (h or home)) for d, h in rows})   # (ημερομηνία, ξενοδοχείο) μοναδικά
    cont = {m.work_date for m in SchedulePeriodMark.query.filter_by(user_id=uid, kind='continue').all()}
    periods, cur = [], None
    for d, hid in seen:
        if cur and hid == cur['hotel_id'] and ((d - cur['end']).days <= SCHEDULE_PERIOD_MAX_GAP_DAYS or d in cont):
            cur['end'] = d; cur['days'] += 1
        else:
            cur = {'hotel_id': hid, 'hotel': hotels.get(hid, '—'),
                   'start': d, 'end': d, 'days': 1}
            periods.append(cur)
    periods.sort(key=lambda pr: pr['start'], reverse=True)
    return periods

def _schedule_return_context(uid, wd, this_hotel_id):
    """v12.199 — Αν η ημ. wd ξεκινά νέα περίοδο σε σχέση με την προηγούμενη βάρδια
    (κενό > MAX_GAP Ή αλλαγή ξενοδοχείου), επιστρέφει context για το pop-up· αλλιώς None."""
    u = User.query.get(uid)
    home = getattr(u, 'home_hotel_id', None) if u else None
    this_h = this_hotel_id or home
    prev = (ShiftAssignment.query
            .filter(ShiftAssignment.user_id == uid, ShiftAssignment.work_date < wd)
            .order_by(ShiftAssignment.work_date.desc()).first())
    if not prev:
        return None   # πρώτη εμφάνιση — όχι «επιστροφή»
    prev_h = prev.work_hotel_id or home
    gap = (wd - prev.work_date).days
    gap_break = gap > SCHEDULE_PERIOD_MAX_GAP_DAYS
    hotel_break = (prev_h != this_h)
    if not (gap_break or hotel_break):
        return None
    hotels = {h.id: h.name for h in Hotel.query.all()}
    dep = None
    if u and getattr(u, 'department_id', None):
        _d = Department.query.get(u.department_id); dep = _d.name if _d else None
    existing = SchedulePeriodMark.query.filter_by(user_id=uid, work_date=wd).first()
    reason = 'both' if (gap_break and hotel_break) else ('hotel' if hotel_break else 'gap')
    return {'user_id': uid, 'name': (u.full_name or u.username) if u else '',
            'date': wd.isoformat(), 'last_date': prev.work_date.isoformat(),
            'last_hotel': hotels.get(prev_h, '—'), 'this_hotel': hotels.get(this_h, '—'),
            'gap_days': gap, 'reason': reason, 'dept': dep, 'hotel_id': this_h,
            'marked': existing.kind if existing else None}

def aggregate(assignments, home_hotel_id=None):
    """Σύνολα από λίστα ShiftAssignment: work_days, repo, sundays, holidays_worked, extra, elsewhere."""
    hol = _holidays()
    work = repo = sundays = hol_worked = elsewhere = worked_repo = 0
    extra = 0.0
    for a in assignments:
        code = a.shift_code
        if is_extra_code(code):
            extra += worked_hours(a)
        elif is_work_code(code):
            work += 1
            if code == 'ΔΡ':          # v12.376 — δηλωμένο ρεπό που εργάστηκε (μετρά εργάσιμη ΚΑΙ ως δουλ. ρεπό)
                worked_repo += 1
            if a.work_date.weekday() == 6:
                sundays += 1
            if a.work_date in hol:
                hol_worked += 1
            if a.work_hotel_id and home_hotel_id and a.work_hotel_id != home_hotel_id:
                elsewhere += 1
        elif is_repo_code(code):
            repo += 1
    return {'work_days': work, 'repo': repo, 'sundays': sundays,
            'holidays_worked': hol_worked, 'extra_hours': round(extra, 2),
            'elsewhere_days': elsewhere, 'total_days': work, 'worked_repo': worked_repo}

def monthly_settlement(year, month, hotel_id=None, split=False):
    """Λίστα «ΠΡΟΣ ΛΟΓΙΣΤΗΡΙΟ» ανά εργαζόμενο για μήνα.
    v12.381 Φ3 — split=True: ΕΚ-ΠΕΡΙΤΡΟΠΗΣ βγαίνει σε μία γραμμή ΑΝΑ ξενοδοχείο-χρέωσης
    (μέρες/έξτρα/Κυρ./αργίες → ξεν. της ημέρας· ρεπό → έδρα· payable = μέρες×ημερομίσθιο).
    split=False (default) = ΠΑΛΙΑ συμπεριφορά ακριβώς (μία γραμμή/εργαζόμενο, όλα υπό home)
    — κρίσιμο για build_run που κάνει dedupe ανά user."""
    start = date(year, month, 1)
    end = date(year + (month // 12), (month % 12) + 1, 1)
    q = (db.session.query(ShiftAssignment)
         .filter(ShiftAssignment.work_date >= start, ShiftAssignment.work_date < end))
    rows = q.all()
    by_user = {}
    for a in rows:
        by_user.setdefault(a.user_id, []).append(a)
    try:
        import rotation as _ROT
        rot_ids = _ROT.rotational_user_ids()
    except Exception:
        rot_ids = set()
    _hotels = {h.id: h.name for h in Hotel.query.all()}
    out = []
    for uid, alist in by_user.items():
        u = User.query.get(uid)
        if not u:
            continue
        home = getattr(u, 'home_hotel_id', None)
        prof = EmploymentProfile.query.filter_by(user_id=uid).first()
        # Ομαδοποίηση ανά ξεν.-χρέωσης: work→work_hotel (ή home), ρεπό/άδεια→home. Split ΜΟΝΟ αν ζητηθεί + εκ-περιτροπής.
        if split and uid in rot_ids:
            groups = {}
            for a in alist:
                ch = a.work_hotel_id if (is_work_code(a.shift_code) and a.work_hotel_id) else home
                groups.setdefault(ch, []).append(a)
        else:
            groups = {home: alist}
        is_split = (split and uid in rot_ids and len(groups) > 1)
        for ch, sub in groups.items():
            if hotel_id and ch != hotel_id:
                continue
            agg = aggregate(sub, ch)
            payable = 0.0
            if prof and prof.agreement_amount:
                if prof.agreement_type == 'Management':
                    # fixed μηνιαίος: πλήρες ΜΟΝΟ στη γραμμή έδρας (αποφυγή πολλαπλασιασμού μισθού)
                    payable = round(prof.agreement_amount, 2) if ch == home else 0.0
                else:
                    payable = round(prof.day_wage * agg['total_days'] + extra_wage(prof, agg['extra_hours']), 2)
            out.append({'user': u, 'agg': agg, 'payable': payable, 'profile': prof,
                        'hotel_id': ch, 'hotel': _hotels.get(ch), 'split': is_split})
    out.sort(key=lambda r: (r['user'].full_name or '', r.get('hotel') or ''))
    return out

def extra_wage(prof, hours):
    try:
        return round((prof.hour_wage or 0) * (hours or 0), 2)
    except Exception:
        return 0.0


# ── POLICY / RULES helpers ────────────────────────────────────────────────────
POLICY_DEFAULTS = {
    'planning_horizon_weeks': 8,
    'cutoff_dow': 3,          # 0=Δευτ ... 3=Πέμπτη
    'cutoff_time': '18:00',
    'lead_days': 4,           # προθεσμία = lead_days πριν τη Δευτέρα της W (4 = Πέμπτη προηγ.)
    'allow_admin_override': 1,
    'manager_edit_locked': 0, # v12.230: αν 1, οι managers επεξεργάζονται και ΚΛΕΙΔΩΜΕΝΕΣ/ληγμένες εβδομάδες
}

def get_policy():
    pol = dict(POLICY_DEFAULTS)
    for row in Setting.query.filter(Setting.key.like('sched_%')).all():
        k = row.key[6:]
        if k in pol:
            try:
                pol[k] = type(POLICY_DEFAULTS[k])(row.value) if not isinstance(POLICY_DEFAULTS[k], str) else row.value
            except Exception:
                pass
    return pol

def set_policy(d):
    for k, v in d.items():
        if k in POLICY_DEFAULTS:
            row = Setting.query.get('sched_' + k)
            if not row:
                row = Setting(key='sched_' + k); db.session.add(row)
            row.value = str(v)
    db.session.commit()

def week_deadline(week_start):
    """datetime προθεσμίας οριστικοποίησης για εβδομάδα που ξεκινά (Δευτέρα) week_start."""
    pol = get_policy()
    try:
        hh, mm = [int(x) for x in str(pol['cutoff_time']).split(':')]
    except Exception:
        hh, mm = 18, 0
    d = week_start - timedelta(days=int(pol['lead_days']))
    return datetime.combine(d, dtime(hh, mm))

def week_editable(week_start, user=None):
    """True αν η εβδομάδα είναι ακόμη επεξεργάσιμη. v12.121: ο admin ΠΑΝΤΑ μπορεί
    (ιδιοκτήτης)· οι managers δεσμεύονται από την προθεσμία."""
    if user is None:
        user = current_user()
    if user is not None and role_rank(user.role) >= ROLE_RANK['admin']:
        return True
    if datetime.now() < week_deadline(week_start):
        return True
    # v12.230: ρυθμιζόμενη εξαίρεση — managers επεξεργάζονται κλειδωμένες/ληγμένες εβδομάδες
    if (user is not None and role_rank(user.role) >= ROLE_RANK['manager']
            and int(get_policy().get('manager_edit_locked', 0) or 0)):
        return True
    return False


# ── ROLE helpers ──────────────────────────────────────────────────────────────
def is_accountant():
    u = current_user()
    return u is not None and (u.role == 'accountant' or role_rank(u.role) >= ROLE_RANK['admin'])

def can_edit_schedule():
    return is_admin() or (current_user() and role_rank(current_user().role) >= ROLE_RANK['manager'])

def resolve_department(name):
    """Ταίριασμα ονόματος τμήματος -> Department (exact -> alias normalized)."""
    if not name:
        return None
    n = _norm(name)
    for d in Department.query.all():
        if _norm(d.name) == n or _norm(d.name_en or '') == n:
            return d
        try:
            for al in json.loads(d.aliases or '[]'):
                if _norm(al) == n:
                    return d
        except Exception:
            pass
    return None


# ── SEED (idempotent) ─────────────────────────────────────────────────────────
def seed_schedule():
  with app.app_context():
    try:
        # ShiftTypes
        if not ShiftType.query.first():
            for i, (code, label, color, cw, note, erg) in enumerate(SHIFT_CODES):
                db.session.add(ShiftType(code=code, label=label, color=color, counts_as_work=cw,
                                         payroll_note=note, ergani_type=erg, sort=i,
                                         default_start='07:00' if code in ('ΕΡΓ', 'ΔΡ') else None,
                                         default_end='15:30' if code in ('ΕΡΓ', 'ΔΡ') else None))
        # v12.120 — εξασφάλισε κωδικό ΕΩ (Έξτρα Ώρες) και σε υπάρχουσες βάσεις
        if not ShiftType.query.filter_by(code='ΕΩ').first():
            _erg = ShiftType.query.filter_by(code='ΕΡΓ').first()
            db.session.add(ShiftType(code='ΕΩ', label='Έξτρα Ώρες', color='#f59e0b',
                counts_as_work=True, payroll_note='ώρες ≠ 8,5 (πάνω/κάτω από ΕΡΓ)',
                ergani_type='WORK', sort=((_erg.sort if _erg else 0) + 1),
                default_start='07:00', default_end='17:00', active=True))
        # v12.375 — εξασφάλισε κωδικό ΔΡ (Δουλεμένο Ρεπό) και σε υπάρχουσες βάσεις
        if not ShiftType.query.filter_by(code='ΔΡ').first():
            _erg = ShiftType.query.filter_by(code='ΕΡΓ').first()
            db.session.add(ShiftType(code='ΔΡ', label='Δουλεμένο Ρεπό', color='#0d9488',
                counts_as_work=True, payroll_note='δηλωμένο ρεπό που εργάστηκε — μετρά εργάσιμη',
                ergani_type='WORK', sort=((_erg.sort if _erg else 0) + 1),
                count_as='work', break_deduct=True,
                default_start='07:00', default_end='15:30', active=True))
        # Departments
        if not Department.query.first():
            for i, (name, en, color, aliases) in enumerate(DEPARTMENTS):
                db.session.add(Department(name=name, name_en=en, color=color,
                                          aliases=json.dumps(aliases, ensure_ascii=False), sort=i))
        # Rules
        if not ScheduleRule.query.first():
            db.session.add(ScheduleRule(code='R1_repo', severity='block',
                description='Κάθε εργαζόμενος ≥1 ρεπό (ΑΝ) ανά εβδομάδα', params='{}'))
            db.session.add(ScheduleRule(code='R2_complete', severity='block',
                description='Καμία κενή ημέρα για assigned εργαζόμενο (πληρότητα 7/7)', params='{}'))
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f'seed_schedule skipped: {e}')


# ── IMPORT (workbook προγράμματος -> χρήστες/τμήματα/αναθέσεις) ────────────────
CODE_HOTELNAME = {
    'AST': 'Asterias', 'CNT': 'Central', 'SRG': 'Sergios',
    'PSV': 'Piskopiano', 'IRO': 'Iro', 'PLM': 'Palm',  # v12.52: PLM = Palm Island Suites
}
def _hotel_short(name):
    if not name:
        return ''
    nl = name.lower()
    for _code, _pref in CODE_HOTELNAME.items():
        if _pref.lower() in nl:
            return _code
    return name.split()[0][:4].upper()
_LABELS = {'τμημα': 'dept', 'ειδικοτητα': 'spec', 'εταιρεια': 'comp', 'υποκ': 'upok',
           'επωνυμο': 'epon', 'ονομα': 'onoma'}

def _resolve_hotel_by_code(code):
    if not code:
        return None
    code = HOTEL_NORM.get(str(code).strip().upper(), str(code).strip().upper())
    target = CODE_HOTELNAME.get(code)
    if not target:
        return None
    tn = _norm(target)
    for h in Hotel.query.all():
        if tn in _norm(h.name):
            return h
    return None

def _route_name(full):
    """Επιστρεφει master ΜΟΝΟ αν υπαρχει επιβεβαιωμενη συνδεση (NameLink, φτιαγμενη απο τον χρηστη).
    Αλλιως None -> ο καλων στελνει τη βαρδια στον προθαλαμο (PendingShift)."""
    nm = _norm(full)
    if not nm:
        return None
    link = NameLink.query.filter_by(norm_name=nm).first()
    if link:
        u = User.query.get(link.user_id)
        if u and u.is_active:
            return u
    return None

def _suggest_masters(full, limit=6):
    """Προτασεις (ΟΧΙ αυτοματη ενεργεια) πιθανων master για ενα ονομα — προτεραιοτητα locked Λογιστηριο.
    Επιστρεφει [(user, score)] ταξινομημενο. score: 100 ακριβες / 90 ιδιες λεξεις / 70 fuzzy / 40 ιδιο επωνυμο."""
    fn = _norm(full)
    if not fn:
        return []
    locked = _locked_uids()
    tgt = _tok_key(full)
    sur_t, _f = _name_parts(full)
    out = []
    for u in User.query.filter(User.is_active == True).all():
        un = u.full_name or ''
        sc = 0
        if _norm(un) == fn:
            sc = 100
        elif tgt and _tok_key(un) == tgt:
            sc = 90
        elif _names_likely_same(full, un):
            sc = 70
        else:
            sur_u, _ = _name_parts(un)
            if sur_t and sur_u == sur_t:
                sc = 40
        if sc:
            if u.id in locked:
                sc += 5
            out.append((u, sc))
    out.sort(key=lambda t: (-t[1], 0 if t[0].id in locked else 1, t[0].full_name or ''))
    return out[:limit]

def _find_or_create_user(epon, onoma, hotel, dept, employer, upok):
    full = (str(epon).strip() + ' ' + (str(onoma).strip() if onoma else '')).strip()
    # ΠΡΩΤΑ ταιριασμα με υπαρχον προφιλ (προτεραιοτητα master Λογιστηριο)
    user = _match_existing(full)
    created = False
    if user is None:
        base = _acc(full).lower().replace(' ', '.')
        uname = re.sub(r'[^a-z0-9.]', '', base) or ('emp' + str(ShiftAssignment.query.count() + 1))
        if User.query.filter_by(username=uname).first():
            uname = uname + '.' + str(User.query.count() + 1)
        user = User(username=uname[:50], password=generate_password_hash(os.urandom(8).hex()),
                    full_name=full[:100], role='staff', approved=True, is_active=True)
        db.session.add(user); db.session.flush(); created = True
    # μετα-στοιχεια ΜΟΝΟ σε νεο προφιλ - τα master κρατουν τα δικα τους assigned
    if created:
        if hotel: user.home_hotel_id = hotel.id
        if dept: user.department_id = dept.id
        if employer: user.employer = str(employer)[:120]
        if upok: user.subunit = str(upok)[:20]
        user.login_enabled = False
        user.employment_active = True
    return user

def import_schedule_workbook(source, only_year=None, created_by=None):
    """source = path ή bytes. Επιστρέφει στατιστικά. Idempotent (upsert ανά user/μέρα)."""
    import openpyxl, io
    if isinstance(source, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True)
    else:
        wb = openpyxl.load_workbook(source, data_only=True)
    stats = {'users_new': 0, 'users_seen': 0, 'assign_new': 0, 'assign_upd': 0,
             'no_hotel': 0, 'cells': 0, 'pending_new': 0, 'pending_upd': 0}
    before_users = User.query.count()
    seen_users = set()
    for ws in wb.worksheets:
        mr, mc = ws.max_row, min(ws.max_column, 40)
        r = 1
        while r <= mr:
            a = ws.cell(r, 1).value
            if a and _norm(a) == 'τμημα':
                colmap = {}; datecols = []
                for c in range(1, mc + 1):
                    v = ws.cell(r, c).value
                    nv = _norm(v)
                    if nv in _LABELS:
                        colmap[_LABELS[nv]] = c
                    d = None
                    if isinstance(v, datetime):
                        d = v.date()
                    elif isinstance(v, str):
                        mm = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', v)
                        if mm:
                            d = date(int(mm.group(3)), int(mm.group(2)), int(mm.group(1)))
                    if d:
                        datecols.append((c, d))
                ce, co = colmap.get('epon'), colmap.get('onoma')
                rr = r + 1
                while rr <= mr and not (ws.cell(rr, 1).value and _norm(ws.cell(rr, 1).value) == 'τμημα'):
                    epon = ws.cell(rr, ce).value if ce else None
                    if epon and str(epon).strip():
                        onoma = ws.cell(rr, co).value if co else ''
                        upok = ws.cell(rr, colmap['upok']).value if colmap.get('upok') else None
                        dept_v = ws.cell(rr, colmap['dept']).value if colmap.get('dept') else None
                        comp_v = ws.cell(rr, colmap['comp']).value if colmap.get('comp') else None
                        hotel = _resolve_hotel_by_code(upok)
                        dept = resolve_department(dept_v)
                        full = (str(epon).strip() + ' ' + (str(onoma).strip() if onoma else '')).strip()
                        nm = _norm(full)
                        master = _route_name(full)   # ΜΟΝΟ επιβεβαιωμενη συνδεση -> αλλιως προθαλαμος
                        if master and master.id not in seen_users:
                            seen_users.add(master.id)
                        if not hotel:
                            stats['no_hotel'] += 1
                        for c, dt in datecols:
                            if only_year and dt.year != only_year:
                                continue
                            cell = ws.cell(rr, c).value
                            code, segs, tag = parse_cell(cell)
                            if code is None:
                                continue
                            stats['cells'] += 1
                            whid = None
                            if tag:
                                th = _resolve_hotel_by_code(tag)
                                whid = th.id if th else None
                            elif hotel:
                                whid = hotel.id
                            segj = json.dumps(segs, ensure_ascii=False)
                            if master:
                                ex = ShiftAssignment.query.filter_by(user_id=master.id, work_date=dt, shift_code=code, segments=segj).first()
                                if ex:
                                    ex.work_hotel_id = whid
                                    stats['assign_upd'] += 1
                                else:
                                    db.session.add(ShiftAssignment(
                                        user_id=master.id, work_date=dt, shift_code=code,
                                        segments=segj, work_hotel_id=whid, created_by=created_by))
                                    stats['assign_new'] += 1
                            else:
                                pend = PendingShift.query.filter_by(norm_name=nm, work_date=dt, shift_code=code, segments=segj).first()
                                if pend:
                                    pend.work_hotel_id = whid
                                    pend.raw_name = full[:120]
                                    stats['pending_upd'] += 1
                                else:
                                    db.session.add(PendingShift(
                                        norm_name=nm, raw_name=full[:120],
                                        epon=str(epon)[:80], onoma=(str(onoma)[:80] if onoma else None),
                                        hotel_tag=(str(upok)[:20] if upok else None),
                                        dept_raw=(str(dept_v)[:120] if dept_v else None),
                                        employer=(str(comp_v)[:120] if comp_v else None),
                                        work_date=dt, shift_code=code, segments=segj,
                                        work_hotel_id=whid, created_by=created_by))
                                    stats['pending_new'] += 1
                    rr += 1
                db.session.commit()
                r = rr
            else:
                r += 1
    wb.close()
    stats['users_new'] = User.query.count() - before_users
    stats['users_seen'] = len(seen_users)
    return stats


# ── ROUTES helpers ────────────────────────────────────────────────────────────
def _auth():
    return 'user_id' in session

def _week_arg():
    w = request.args.get('week')
    if w:
        try:
            return monday_of(datetime.strptime(w, '%Y-%m-%d').date())
        except Exception:
            pass
    return monday_of(date.today())

def _dept_users(hotel_id, dept_id):
    q = User.query.filter(User.is_active == True)
    if hotel_id:
        q = q.filter(User.home_hotel_id == hotel_id)
    if dept_id:
        q = q.filter(User.department_id == dept_id)
    return q.order_by(User.full_name).all()

def week_grid(hotel_id, dept_id, week_start):
    return _grid_for_days(hotel_id, dept_id, [week_start + timedelta(days=i) for i in range(7)])

def _grid_for_days(hotel_id, dept_id, days, users=None):
    """v12.201 — Γενικό grid για ΟΠΟΙΕΣΔΗΠΟΤΕ μέρες (εβδομάδα ή μήνας). Ίδια δομή κελιών/totals.
    users=None → προσωπικό τμήματος· αλλιώς δοθείσα λίστα (π.χ. όλο το ξεν., ungrouped)."""
    borrowed_ids = set(); rot_ids = set()
    if users is None:
        users = _dept_users(hotel_id, dept_id)
        # v12.378 Φ2a — εκ-περιτροπής που μοιράζονται σε ΑΥΤΟ το ξεν. (έδρα αλλού) → read-only ορατότητα
        try:
            import rotation as _ROT
            rot_ids = _ROT.rotational_user_ids()
            if hotel_id:
                share_uids = _ROT.share_user_ids_for_hotel(hotel_id)
                have = {u.id for u in users}
                extra = [uu for uu in share_uids if uu not in have]
                if extra:
                    bq = User.query.filter(User.is_active == True, User.id.in_(extra),
                                           User.home_hotel_id != hotel_id)
                    if dept_id:
                        bq = bq.filter(User.department_id == dept_id)
                    for bu in bq.all():
                        users.append(bu); borrowed_ids.add(bu.id)
                    users = sorted(users, key=lambda x: x.full_name or '')
        except Exception:
            borrowed_ids = set(); rot_ids = set()
    uids = [u.id for u in users]
    amap = {}
    if uids:
        for a in (ShiftAssignment.query
                  .filter(ShiftAssignment.user_id.in_(uids))
                  .filter(ShiftAssignment.work_date >= days[0], ShiftAssignment.work_date <= days[-1]).all()):
            amap.setdefault((a.user_id, a.work_date.isoformat()), []).append(a)
    _colors = {st.code: st.color for st in ShiftType.query.all()}
    _hotels = {h.id: h.name for h in Hotel.query.all()}
    _meta = {}
    if uids:
        try:
            from payroll import EmployeePII as _PII
            for _p in _PII.query.filter(_PII.user_id.in_(uids)).all():
                _meta[_p.user_id] = {'emp_code': _p.emp_code, 'afm': _p.afm, 'locked': bool(_p.locked)}
        except Exception:
            _meta = {}
    rows = []
    for u in users:
        _is_rot = (u.id in rot_ids) or (u.id in borrowed_ids)   # v12.379 Φ2β
        _uhome = getattr(u, 'home_hotel_id', None)
        cells = []
        wk_hours = 0.0; wk_extra = 0.0; repo = 0; work_days = 0
        for d in days:
            alist = amap.get((u.id, d.isoformat())) or []
            if alist:
                day_hours = 0.0; day_extra = 0.0; day_work = False; day_repo = False
                labels = []; first = alist[0]
                for a in alist:
                    if is_extra_code(a.shift_code):
                        day_extra += worked_hours(a)
                    elif is_work_code(a.shift_code):
                        day_hours += worked_hours(a); day_work = True
                    elif is_repo_code(a.shift_code):
                        day_repo = True
                    try:
                        segs = json.loads(a.segments) if a.segments else []
                    except Exception:
                        segs = []
                    labels.append('\n'.join(f"{s['start']} - {s['end']}" for s in segs) if segs else a.shift_code)
                wk_extra += day_extra
                if day_work:
                    wk_hours += day_hours; work_days += 1
                if day_repo:
                    repo += 1
                try:
                    fsegs = json.loads(first.segments) if first.segments else []
                except Exception:
                    fsegs = []
                _entries = []
                for a in alist:
                    try:
                        _es = json.loads(a.segments) if a.segments else []
                    except Exception:
                        _es = []
                    _tm = ' & '.join("%s - %s" % (s.get('start'), s.get('end')) for s in _es) if _es else ''
                    _hn = _hotels.get(a.work_hotel_id) if (a.work_hotel_id and a.work_hotel_id != hotel_id) else None
                    _entries.append({'code': a.shift_code, 'segs': _es, 'wh': a.work_hotel_id,
                                     'times': _tm, 'hotel': _hn, 'hotel_short': _hotel_short(_hn) if _hn else '',
                                     'hours': (round(worked_hours(a), 1) if is_work_code(a.shift_code) else 0),
                                     'color': _colors.get(a.shift_code, '#64748b')})
                cells.append({'date': d.isoformat(), 'code': first.shift_code, 'segs': fsegs,
                              'label': '\n'.join(labels), 'hours': round(day_hours, 1),
                              'elsewhere': bool(first.work_hotel_id and first.work_hotel_id != hotel_id),
                              'wh': first.work_hotel_id, 'note': first.note or '',
                              'n': len(alist), 'entries': _entries,
                              # v12.379 Φ2β — κλείδωμα ανά μέρα: εκ-περιτροπής μέρα που ανήκει σε άλλο ξεν.
                              'locked': bool(_is_rot and ((first.work_hotel_id or _uhome) != hotel_id))})
            else:
                cells.append({'date': d.isoformat(), 'code': '', 'segs': [], 'label': '', 'hours': 0,
                              'elsewhere': False, 'wh': None, 'note': '', 'n': 0, 'entries': [],
                              'locked': False})
        _m = _meta.get(u.id) or {}
        _borrowed = u.id in borrowed_ids
        _rotational = (u.id in rot_ids) or _borrowed
        rows.append({'user': u, 'cells': cells, 'wk_hours': round(wk_hours, 1), 'wk_extra': round(wk_extra, 1), 'repo': repo, 'work_days': work_days,
                     'emp_code': _m.get('emp_code'), 'afm': _m.get('afm'), 'locked': _m.get('locked', False),
                     'rotational': _rotational, 'borrowed': _borrowed,  # v12.378 Φ2a
                     'home_hotel': (_hotels.get(getattr(u, 'home_hotel_id', None)) if _borrowed else None),
                     'readonly': _borrowed})
    return days, rows

def validate_hotel_week(hotel_id, week_start):
    """Εφαρμογή ScheduleRules σε όλο το ξενοδοχείο/εβδομάδα. Επιστρέφει issues."""
    days = [week_start + timedelta(days=i) for i in range(7)]
    issues = []
    rules = {r.code: r for r in ScheduleRule.query.filter_by(active=True).all()}
    users = User.query.filter(User.is_active == True, User.home_hotel_id == hotel_id).all()
    for u in users:
        alist = (ShiftAssignment.query.filter(ShiftAssignment.user_id == u.id)
                 .filter(ShiftAssignment.work_date >= days[0], ShiftAssignment.work_date <= days[6]).all())
        if not alist:
            continue
        codes = [a.shift_code for a in alist]
        if 'R1_repo' in rules and sum(1 for c in codes if is_repo_code(c)) < 1:
            issues.append({'user': u, 'rule': 'R1_repo', 'severity': rules['R1_repo'].severity,
                           'msg': f'{u.full_name}: κανένα ρεπό αυτή την εβδομάδα'})
        # v12.200 — μέτρημα ΔΙΑΚΡΙΤΩΝ ημερών (όχι βαρδιών: πολλές/μέρα δεν «γεμίζουν» κενά)
        days_set = {a.work_date for a in alist}
        if 'R2_complete' in rules and len(days_set) < 7:
            missing = [WEEKDAYS_EL[i] for i, dd in enumerate(days) if dd not in days_set]
            issues.append({'user': u, 'rule': 'R2_complete', 'severity': rules['R2_complete'].severity,
                           'msg': f'{u.full_name}: κενές ημέρες χωρίς κωδικό — ' + ', '.join(missing)})
    return issues


# ── ROUTE: Board (multi-week + πολλαπλά τμήματα δυναμικά) ──────────────────────
def _depts_present(hotel_id):
    """v12.173 (#2) — Τμήματα στο Πρόγραμμα = ΜΟΝΟ τα επιλεγμένα (HotelDepartment).
    Μη-επιλεγμένα ΔΕΝ εμφανίζονται (ακόμη κι αν έχουν άτομα). Μη ρυθμισμένο → όλα τα ενεργά."""
    if not hotel_id:
        return Department.query.filter_by(active=True).order_by(Department.sort).all()
    configured = {hd.department_id for hd in HotelDepartment.query.filter_by(hotel_id=hotel_id).all()}
    if configured:
        return Department.query.filter(Department.id.in_(configured), Department.active == True).order_by(Department.sort, Department.name).all()
    return Department.query.filter_by(active=True).order_by(Department.sort).all()

def _dept_supervisors(hotel_id):
    """v12.168 — {dept_id: όνομα υπευθύνου} ανά ξενοδοχείο, από HotelDepartment (owner=οργανόγραμμα)."""
    out = {}
    if not hotel_id:
        return out
    for hd in HotelDepartment.query.filter_by(hotel_id=hotel_id).all():
        if hd.supervisor_user_id:
            u = User.query.get(hd.supervisor_user_id)
            if u:
                out[hd.department_id] = u.full_name or u.username
    return out

def _build_block(hotel_id, dept_list, week_start, user):
    days = [week_start + timedelta(days=i) for i in range(7)]
    deptgrids = []
    sup = _dept_supervisors(hotel_id)
    for d in dept_list:
        _, rows = week_grid(hotel_id, d.id, week_start)
        deptgrids.append({'dept': d, 'rows': rows, 'supervisor': sup.get(d.id)})
    sub = None
    if hotel_id:
        sub = (ScheduleSubmission.query.filter_by(hotel_id=hotel_id, week_start=week_start)
               .order_by(ScheduleSubmission.version.desc()).first())
    return {
        'week_start': week_start, 'days': days, 'deptgrids': deptgrids,
        'editable': week_editable(week_start, user) and can_edit_schedule(),
        'issues': validate_hotel_week(hotel_id, week_start) if hotel_id else [],
        'deadline': week_deadline(week_start), 'sub': sub, 'iso': week_start.isoformat(),
        'label': f"{week_start.strftime('%d/%m')} – {(week_start + timedelta(days=6)).strftime('%d/%m/%Y')}",
    }

def _build_span_block(hotel_id, dept_list, days, user, label):
    """v12.388 — Πλάνο για ΟΠΟΙΟΔΗΠΟΤΕ εύρος ημερών (μήνας ή έτος): ΟΛΟ το προσωπικό σε ΜΙΑ λίστα,
    όλες οι μέρες ως στήλες, ίδια κελιά/editor. Row totals = για ΟΛΟ το εύρος. day_hdr με σήμανση μήνα."""
    ndays = len(days)
    hol = {h.hol_date for h in Holiday.query.all()}
    seen = set(); users = []
    for dep in dept_list:
        for u in _dept_users(hotel_id, dep.id):
            if u.id not in seen:
                seen.add(u.id); users.append(u)
    users.sort(key=lambda u: (u.full_name or u.username or '').lower())
    _wk = {}
    def _ed(d):
        ws = monday_of(d)
        if ws not in _wk:
            _wk[ws] = bool(week_editable(ws, user) and can_edit_schedule())
        return _wk[ws]
    _, rows = _grid_for_days(hotel_id, None, days, users=users)
    for r in rows:
        for c in r['cells']:
            c['edit'] = _ed(date.fromisoformat(c['date']))
    WD = ['Δε', 'Τρ', 'Τε', 'Πε', 'Πα', 'Σα', 'Κυ']
    today = date.today()
    day_work = [0] * ndays; day_repo = [0] * ndays
    for r in rows:
        for idx, c in enumerate(r['cells']):
            cds = [e['code'] for e in c['entries']]
            if any(is_work_code(x) and not is_extra_code(x) for x in cds):
                day_work[idx] += 1
            if any(is_repo_code(x) for x in cds):
                day_repo[idx] += 1
    day_hdr = [{'iso': d.isoformat(), 'd': d.day, 'wd': WD[d.weekday()],
                'we': d.weekday() >= 5, 'hol': d in hol, 'today': (d == today),
                'm': d.month, 'mstart': (d.day == 1), 'mlabel': MONTHS_EL[d.month][:3],
                'work': day_work[i], 'repo': day_repo[i]} for i, d in enumerate(days)]
    return {'rows': rows, 'day_hdr': day_hdr, 'ndays': ndays, 'label': label}

def _build_month_block(hotel_id, dept_list, year, month, user):
    """v12.204 — Μηνιαίο πλάνο (wrapper του _build_span_block για έναν μήνα)."""
    import calendar as _cal
    ndays = _cal.monthrange(year, month)[1]
    days = [date(year, month, dd) for dd in range(1, ndays + 1)]
    return _build_span_block(hotel_id, dept_list, days, user, '%s %d' % (MONTHS_EL[month], year))

def _build_year_block(hotel_id, dept_list, year, user):
    """v12.388 — Ετήσιο πλάνο: όλες οι μέρες του έτους ως στήλες (ίδιο grid/editor με το μηνιαίο)."""
    import calendar as _cal
    days = []
    for mo in range(1, 13):
        nd = _cal.monthrange(year, mo)[1]
        days += [date(year, mo, dd) for dd in range(1, nd + 1)]
    return _build_span_block(hotel_id, dept_list, days, user, str(year))

@app.route('/dashboard/schedule')
def schedule_board():
    if not _auth():
        return redirect(url_for('login'))
    user = current_user()
    hotels = allowed_hotels(user)
    hotel_id = request.args.get('hotel_id', type=int) or active_hotel_id() or (hotels[0].id if hotels else None)
    dept_list = _depts_present(hotel_id)
    week_start = _week_arg()
    pol = get_policy()
    horizon = max(1, int(pol.get('planning_horizon_weeks', 8)))
    weeks = max(1, min(request.args.get('weeks', type=int) or 1, horizon))
    # v12.55 — πλοήγηση/προβολή μήνα: αν δοθεί month, δείξε ΟΛΟΝ τον μήνα (5-6 εβδομάδες)
    import calendar as _cal
    mon_arg = request.args.get('month', type=int)
    yr_arg = request.args.get('year', type=int) or week_start.year
    sel_month = mon_arg or week_start.month
    sel_year = yr_arg
    if mon_arg:
        first = date(yr_arg, mon_arg, 1)
        week_start = monday_of(first)
        last = date(yr_arg, mon_arg, _cal.monthrange(yr_arg, mon_arg)[1])
        weeks = ((monday_of(last) - week_start).days // 7) + 1
    prev_m = (date(sel_year, sel_month, 1) - timedelta(days=1))
    nxt = date(sel_year, sel_month, 28) + timedelta(days=10)
    view_mode = request.args.get('view') or 'week'
    month_block = None; year_block = None
    if view_mode == 'month':
        month_block = _build_month_block(hotel_id, dept_list, sel_year, sel_month, user)
        blocks = []
    elif view_mode == 'year':
        year_block = _build_year_block(hotel_id, dept_list, sel_year, user)
        blocks = []
    else:
        blocks = [_build_block(hotel_id, dept_list, week_start + timedelta(days=7 * i), user) for i in range(weeks)]
    shift_types = ShiftType.query.filter_by(active=True).order_by(ShiftType.sort).all()
    shift_lookup = {st.code: st for st in shift_types}
    shift_types_json = json.dumps([{'code': st.code, 'color': st.color} for st in entry_shift_types()], ensure_ascii=False)
    cur_hotel = Hotel.query.get(hotel_id) if hotel_id else None
    return render_template('schedule_board.html',
        shift_lookup=shift_lookup, shift_types_json=shift_types_json,
        hotels=hotels, hotel_id=hotel_id, cur_hotel=cur_hotel, dept_list=dept_list,
        weekdays=WEEKDAYS_EL, shift_types=shift_types, blocks=blocks, weeks=weeks, horizon=horizon,
        week_start=week_start, week_start_iso=week_start.isoformat(),
        prev_week=(week_start - timedelta(days=7)).isoformat(),
        next_week=(week_start + timedelta(days=7)).isoformat(),
        month_el=MONTHS_EL, is_admin=is_admin(),
        sel_month=sel_month, sel_year=sel_year, view_mode=view_mode, month_block=month_block, year_block=year_block,
        prev_month=prev_m.month, prev_year=prev_m.year, next_month=nxt.month, next_year=nxt.year)


@app.route('/dashboard/schedule/submit_month', methods=['POST'])
def schedule_submit_month():
    """v12.213 — Υποβολή ΟΛΟΥ του μήνα: υποβάλλει κάθε εβδομάδα (με δεδομένα) που περνά τους κανόνες."""
    if not _auth() or not can_edit_schedule():
        return redirect(url_for('login'))
    user = current_user()
    hotel_id = request.form.get('hotel_id', type=int)
    year = request.form.get('year', type=int); month = request.form.get('month', type=int)
    import calendar as _cal
    if not (hotel_id and year and month):
        return redirect('/dashboard/schedule?embed=1')
    last = date(year, month, _cal.monthrange(year, month)[1])
    uids = [u.id for u in User.query.filter(User.is_active == True, User.home_hotel_id == hotel_id).all()]
    def _has(ws):
        if WeekPlan.query.filter_by(hotel_id=hotel_id, week_start=ws).first():
            return True
        if not uids:
            return False
        return db.session.query(ShiftAssignment.id).filter(
            ShiftAssignment.user_id.in_(uids),
            ShiftAssignment.work_date >= ws, ShiftAssignment.work_date <= ws + timedelta(days=6)).first() is not None
    submitted = 0; blocked = 0
    wk = monday_of(date(year, month, 1))
    while wk <= last:
        if not _has(wk):
            wk += timedelta(days=7); continue
        issues = validate_hotel_week(hotel_id, wk)
        if any(i['severity'] == 'block' for i in issues):
            blocked += 1; wk += timedelta(days=7); continue
        lastsub = (ScheduleSubmission.query.filter_by(hotel_id=hotel_id, week_start=wk)
                   .order_by(ScheduleSubmission.version.desc()).first())
        snap = _hotel_week_snapshot(hotel_id, wk)
        version = (lastsub.version + 1) if lastsub else 1
        changes = _diff_snapshots(json.loads(lastsub.snapshot) if lastsub and lastsub.snapshot else None, snap) if lastsub else []
        sub = ScheduleSubmission(hotel_id=hotel_id, week_start=wk, version=version,
                                 parent_version=(lastsub.version if lastsub else None),
                                 status='submitted', snapshot=json.dumps(snap, ensure_ascii=False),
                                 changes=json.dumps(changes, ensure_ascii=False), submitted_by=user.id)
        db.session.add(sub)
        for wp in WeekPlan.query.filter_by(hotel_id=hotel_id, week_start=wk).all():
            wp.status = 'submitted'
        _notify_accountants(hotel_id, wk, version, 'ΤΡΟΠΟΠΟΙΗΣΗ' if version > 1 else 'Νέα υποβολή', changes)
        submitted += 1
        wk += timedelta(days=7)
    db.session.commit()
    log_activity('schedule_submit_month', f'{month}/{year} ({submitted})', hotel_id=hotel_id)
    return redirect(f'/dashboard/schedule?view=month&hotel_id={hotel_id}&month={month}&year={year}&embed=1&ok=msub&n={submitted}&b={blocked}')


@app.route('/dashboard/schedule/month_export.xlsx')
def schedule_month_export():
    """v12.213 — Εξαγωγή μηνιαίου πλάνου (εργαζόμενοι × μέρες με κωδικούς) σε Excel."""
    if not _auth() or not can_edit_schedule():
        return redirect(url_for('login'))
    user = current_user()
    hotel_id = request.args.get('hotel_id', type=int)
    year = request.args.get('year', type=int) or date.today().year
    month = request.args.get('month', type=int) or date.today().month
    dept_list = _depts_present(hotel_id)
    mb = _build_month_block(hotel_id, dept_list, year, month, user)
    import openpyxl, io
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Μηνιαίο'
    navy = PatternFill('solid', fgColor='193847'); we = PatternFill('solid', fgColor='fde2e2')
    # v12.388 — στήλες HOTEL CODE + ΤΜΗΜΑ (home hotel / τμήμα ανά εργαζόμενο)
    _hnames = {h.id: h.name for h in Hotel.query.all()}
    _dnames = {d.id: d.name for d in Department.query.all()}
    NPRE = 3   # πλήθος στηλών πριν τις μέρες (Εργαζόμενος, HOTEL CODE, ΤΜΗΜΑ)
    hdr = ['Εργαζόμενος', 'HOTEL CODE', 'ΤΜΗΜΑ'] + ['%s %d' % (d['wd'], d['d']) for d in mb['day_hdr']] + ['Ώρες', 'Έξτρα', 'Ρεπό', 'Εργ.']
    ws.append(hdr)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF'); cell.fill = navy; cell.alignment = Alignment(horizontal='center')
    for r in mb['rows']:
        u = r['user']
        _hc = _hotel_short(_hnames.get(getattr(u, 'home_hotel_id', None), ''))
        _dep = _dnames.get(getattr(u, 'department_id', None), '')
        row = [u.full_name or u.username, _hc, _dep]
        for c in r['cells']:
            row.append(' / '.join((e['code'] + ((' ' + e['times']) if e.get('times') else '')) for e in c['entries']) if c['entries'] else '')
        row += [r['wk_hours'], r['wk_extra'], r['repo'], r['work_days']]
        ws.append(row)
    ws.freeze_panes = 'D2'
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18
    _ctr = Alignment(horizontal='center', vertical='center')
    for i, d in enumerate(mb['day_hdr']):
        col = openpyxl.utils.get_column_letter(NPRE + 1 + i); ws.column_dimensions[col].width = 13
        for rr in range(1, ws.max_row + 1):
            cell = ws.cell(row=rr, column=NPRE + 1 + i)
            cell.alignment = _ctr           # v12.388 — μέρες + περιεχόμενο κεντραρισμένα
            if d['we']:
                cell.fill = we
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    fn = 'monthly_%d_%02d.xlsx' % (year, month)
    return Response(bio.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=%s' % fn})


# ── API: autosave κελιού ──────────────────────────────────────────────────────
# ── v12.379 Φ2β — Guard καταχώρησης εκ-περιτροπής (κλείδωμα ανά μέρα + σκληρό όριο) ──
def _rotation_guard(uid, wd, code, board_hotel_id, pending=None):
    """Έλεγχος καταχώρησης βάρδιας για ΕΚ-ΠΕΡΙΤΡΟΠΗΣ εργαζόμενο.
    Επιστρέφει (ok, msg, work_hotel_override, is_rotational, acting_hotel_id).
    Για ΜΗ εκ-περιτροπής → (True,'',None,False,None): καμία επίδραση (ξεχωριστό κύκλωμα).
    pending: {(uid, acting): set(dates)} ήδη αποδεκτές σε ΑΥΤΟ το request (cumulative όριο)."""
    try:
        import rotation as _ROT
        if not _ROT.is_rotational(uid):
            return (True, '', None, False, None)
    except Exception:
        return (True, '', None, False, None)
    u = User.query.get(uid)
    home = getattr(u, 'home_hotel_id', None)
    acting = board_hotel_id or active_hotel_id() or home   # v12.380 F5 — fallback αν λείπει board_hotel_id
    nm = (u.full_name if u else '') or 'ο εργαζόμενος'
    wh = acting if (acting and acting != home) else None
    # 1) Ρεπό/άδειες ορίζονται ΜΟΝΟ από την έδρα
    if not is_work_code(code):
        if acting != home:
            return (False, '%s: τα ρεπό/άδειες ορίζονται από την έδρα του.' % nm, None, True, acting)
        return (True, '', None, True, acting)
    # 2) Αποκλειστικότητα/κλείδωμα ημέρας: υπάρχει ήδη βάρδια που ανήκει σε ΑΛΛΟ ξενοδοχείο
    for a in ShiftAssignment.query.filter_by(user_id=uid, work_date=wd).all():
        if (a.work_hotel_id or home) != acting:
            if is_work_code(a.shift_code):
                return (False, '%s έχει ήδη βάρδια εκείνη τη μέρα σε άλλο ξενοδοχείο.' % nm, None, True, acting)
            return (False, '%s: αυτή η μέρα ανήκει σε άλλο ξενοδοχείο (έδρα).' % nm, None, True, acting)
    # 3) Σκληρό όριο μεριδίου (μόνο όπου υπάρχει ορισμένο μερίδιο)
    quota = _ROT.days_quota(uid, acting)
    if quota is not None:
        wk = monday_of(wd)
        wdays = set()
        for a in ShiftAssignment.query.filter(
                ShiftAssignment.user_id == uid,
                ShiftAssignment.work_date >= wk,
                ShiftAssignment.work_date <= wk + timedelta(days=6)).all():
            if a.work_date == wd:
                continue   # η τρέχουσα μέρα (ξανα)γράφεται — δεν μετρά διπλά
            if is_work_code(a.shift_code) and (a.work_hotel_id or home) == acting:
                wdays.add(a.work_date)
        for pdate in (pending or {}).get((uid, acting), ()):
            if pdate != wd:
                wdays.add(pdate)
        if len(wdays) >= quota:
            return (False, 'Όριο εκ περιτροπής: %s δικαιούται %d μέρες/βδ σε αυτό το ξενοδοχείο.' % (nm, quota), None, True, acting)
    return (True, '', wh, True, acting)


def _rotation_can_modify(uid, a, board_hotel_id):
    """v12.379 Φ2β — Μπορεί το ξεν.-υποδοχής να ΔΙΑΓΡΑΨΕΙ/αντικαταστήσει υπάρχουσα βάρδια;
    Μπλοκάρει αν ο εκ-περιτροπής έχει βάρδια που ανήκει σε ΑΛΛΟ ξενοδοχείο (κλείδωμα ανά μέρα)."""
    try:
        import rotation as _ROT
        if not _ROT.is_rotational(uid):
            return (True, '')
    except Exception:
        return (True, '')
    u = User.query.get(uid)
    home = getattr(u, 'home_hotel_id', None)
    acting = board_hotel_id or active_hotel_id() or home   # v12.380 F5
    # F2 — έλεγξε ΟΛΕΣ τις βάρδιες της μέρας (όχι μόνο την πρώτη): μπλοκ αν ΚΑΠΟΙΑ ανήκει αλλού
    for _a in ShiftAssignment.query.filter_by(user_id=uid, work_date=a.work_date).all():
        if is_work_code(_a.shift_code) and (_a.work_hotel_id or home) != acting:
            return (False, 'Αυτή η μέρα ανήκει σε άλλο ξενοδοχείο (εκ περιτροπής).')
        if (not is_work_code(_a.shift_code)) and acting != home:
            return (False, 'Τα ρεπό/άδειες του εκ περιτροπής ορίζονται από την έδρα.')
    return (True, '')


@app.route('/dashboard/schedule/cell', methods=['POST'])
def schedule_cell():
    if not _auth():
        return ('', 401)
    if not can_edit_schedule():
        return jsonify(ok=False, err='forbidden'), 403
    user = current_user()
    d = request.json or {}
    try:
        uid = int(d['user_id'])
        wd = datetime.strptime(d['date'], '%Y-%m-%d').date()
    except Exception:
        return jsonify(ok=False, err='bad'), 400
    if not week_editable(monday_of(wd), user):
        return jsonify(ok=False, err='locked'), 423
    code = (d.get('code') or '').strip()
    segs = d.get('segments') or []
    note = (d.get('note') or '')[:200]
    whid = d.get('work_hotel_id'); whid = int(whid) if whid else None
    board_hid = d.get('board_hotel_id'); board_hid = int(board_hid) if board_hid else None
    a = ShiftAssignment.query.filter_by(user_id=uid, work_date=wd).first()
    if not code:
        if a:
            _dok, _dmsg = _rotation_can_modify(uid, a, board_hid)
            if not _dok:
                return jsonify(ok=False, err='rotation', msg=_dmsg), 400
            db.session.delete(a); db.session.commit()
        return jsonify(ok=True, deleted=True)
    # προεπιλεγμένες ώρες αν ΕΡΓ χωρίς segments
    if code in ('ΕΡΓ', 'ΔΡ') and not segs:
        st = ShiftType.query.filter_by(code=code).first()
        if st and st.default_start and st.default_end:
            segs = [{'start': st.default_start, 'end': st.default_end}]
    if code in ('ΕΡΓ', 'ΕΩ', 'ΔΡ'):
        _ok, _msg = _validate_work_code(code, segs)
        if not _ok:
            return jsonify(ok=False, err='invalid', msg=_msg), 400
    # v12.379 Φ2β — έλεγχος εκ-περιτροπής (κλείδωμα ημέρας + όριο) + auto work_hotel
    _gok, _gmsg, _gwh, _grot, _gact = _rotation_guard(uid, wd, code, board_hid)
    if not _gok:
        return jsonify(ok=False, err='rotation', msg=_gmsg), 400
    if _grot:
        whid = _gwh
    if not a:
        a = ShiftAssignment(user_id=uid, work_date=wd, created_by=user.id)
        db.session.add(a)
    a.shift_code = code
    a.segments = json.dumps(segs, ensure_ascii=False)
    a.work_hotel_id = whid
    a.note = note
    # WeekPlan -> draft (αν δεν υπάρχει)
    u = User.query.get(uid)
    if u and getattr(u, 'home_hotel_id', None) and getattr(u, 'department_id', None):
        wp = WeekPlan.query.filter_by(hotel_id=u.home_hotel_id, department_id=u.department_id,
                                      week_start=monday_of(wd)).first()
        if not wp:
            wp = WeekPlan(hotel_id=u.home_hotel_id, department_id=u.department_id,
                          week_start=monday_of(wd), status='draft')
            db.session.add(wp)
        if wp.status in ('submitted', 'locked'):
            wp.status = 'draft'
        wp.updated_by = user.id
    db.session.commit()
    hrs = worked_hours(a)
    return jsonify(ok=True, hours=round(hrs, 1), work=is_work_code(code))

@app.route('/dashboard/schedule/cells_bulk', methods=['POST'])
def schedule_cells_bulk():
    """v12.130 — μαζική εφαρμογή βάρδιας σε πολλά κελιά (drag-select).
    Δέχεται {user_id, dates[]} (γραμμή) ή {cells:[{user_id,date}]} (ορθογώνιο: πολλοί × μέρες)."""
    if not _auth():
        return ('', 401)
    if not can_edit_schedule():
        return jsonify(ok=False, err='forbidden'), 403
    user = current_user()
    d = request.json or {}
    pairs = []
    cells = d.get('cells')
    if cells:
        for c in cells:
            try:
                pairs.append((int(c['user_id']), datetime.strptime(c['date'], '%Y-%m-%d').date()))
            except Exception:
                continue
    else:
        try:
            uid0 = int(d['user_id'])
        except Exception:
            return jsonify(ok=False, err='bad'), 400
        for ds in (d.get('dates') or []):
            try:
                pairs.append((uid0, datetime.strptime(ds, '%Y-%m-%d').date()))
            except Exception:
                continue
    code = (d.get('code') or '').strip()
    segs = d.get('segments') or []
    _whid = d.get('work_hotel_id'); _whid = int(_whid) if _whid else None
    board_hid = d.get('board_hotel_id'); board_hid = int(board_hid) if board_hid else None
    if code in ('ΕΡΓ', 'ΔΡ') and not segs:
        st = ShiftType.query.filter_by(code=code).first()
        if st and st.default_start and st.default_end:
            segs = [{'start': st.default_start, 'end': st.default_end}]
    if code in ('ΕΡΓ', 'ΕΩ', 'ΔΡ'):
        _ok, _msg = _validate_work_code(code, segs)
        if not _ok:
            return jsonify(ok=False, err='invalid', msg=_msg), 400
    done = 0; locked = 0; blocked = 0; _last_msg = ''
    touched = set(); _pending = {}
    for uid, wd in pairs:
        if not week_editable(monday_of(wd), user):
            locked += 1; continue
        _wh_use = _whid
        if code:
            _gok, _gmsg, _gwh, _grot, _gact = _rotation_guard(uid, wd, code, board_hid, _pending)
            if _grot:
                if not _gok:
                    blocked += 1; _last_msg = _gmsg; continue
                _wh_use = _gwh
                if is_work_code(code):
                    _pending.setdefault((uid, _gact), set()).add(wd)
        else:
            _ex = ShiftAssignment.query.filter_by(user_id=uid, work_date=wd).first()
            if _ex:
                _dok, _dmsg = _rotation_can_modify(uid, _ex, board_hid)
                if not _dok:
                    blocked += 1; _last_msg = _dmsg; continue
        ShiftAssignment.query.filter_by(user_id=uid, work_date=wd).delete()
        if code:
            db.session.add(ShiftAssignment(user_id=uid, work_date=wd, shift_code=code,
                segments=json.dumps(segs, ensure_ascii=False), work_hotel_id=_wh_use, created_by=user.id))
        touched.add((uid, monday_of(wd)))
        done += 1
    for uid, ws in touched:
        u = User.query.get(uid)
        if u and getattr(u, 'home_hotel_id', None) and getattr(u, 'department_id', None):
            wp = WeekPlan.query.filter_by(hotel_id=u.home_hotel_id, department_id=u.department_id, week_start=ws).first()
            if not wp:
                wp = WeekPlan(hotel_id=u.home_hotel_id, department_id=u.department_id, week_start=ws, status='draft')
                db.session.add(wp)
            if wp.status in ('submitted', 'locked'):
                wp.status = 'draft'
            wp.updated_by = user.id
    db.session.commit()
    if done == 0 and blocked:
        return jsonify(ok=False, err='rotation', msg=_last_msg or 'Δεν επιτρέπεται (εκ περιτροπής).'), 400
    return jsonify(ok=True, done=done, locked=locked, blocked=blocked)

@app.route('/dashboard/schedule/day', methods=['POST'])
def schedule_day():
    """v12.131 — αποθήκευση ΟΛΗΣ της μέρας ενός εργαζομένου: αντικαθιστά τις εγγραφές
    με τη λίστα entries (καθεμία: κωδικός + ξενοδοχείο + ώρες). Επιτρέπει πολλές/μέρα."""
    if not _auth():
        return ('', 401)
    if not can_edit_schedule():
        return jsonify(ok=False, err='forbidden'), 403
    user = current_user()
    d = request.json or {}
    try:
        uid = int(d['user_id'])
        wd = datetime.strptime(d['date'], '%Y-%m-%d').date()
    except Exception:
        return jsonify(ok=False, err='bad'), 400
    if not week_editable(monday_of(wd), user):
        return jsonify(ok=False, err='locked', msg='Κλειδωμένη εβδομάδα.'), 423
    board_hid = d.get('board_hotel_id'); board_hid = int(board_hid) if board_hid else None
    norm = []
    for e in (d.get('entries') or []):
        code = (e.get('code') or '').strip()
        if not code:
            continue
        segs = e.get('segments') or []
        whid = e.get('work_hotel_id'); whid = int(whid) if whid else None
        if code in ('ΕΡΓ', 'ΔΡ') and not segs:
            st = ShiftType.query.filter_by(code='ΕΡΓ').first()
            if st and st.default_start and st.default_end:
                segs = [{'start': st.default_start, 'end': st.default_end}]
        if code in ('ΕΡΓ', 'ΕΩ', 'ΔΡ'):
            ok, msg = _validate_work_code(code, segs)
            if not ok:
                return jsonify(ok=False, err='invalid', msg=msg), 400
        else:
            segs = []; whid = None
        # v12.379 Φ2β — εκ-περιτροπής: κλείδωμα/όριο + auto work_hotel
        _gok, _gmsg, _gwh, _grot, _ = _rotation_guard(uid, wd, code, board_hid)
        if not _gok:
            return jsonify(ok=False, err='rotation', msg=_gmsg), 400
        if _grot:
            whid = _gwh
        norm.append((code, segs, whid))
    ShiftAssignment.query.filter_by(user_id=uid, work_date=wd).delete()
    for code, segs, whid in norm:
        db.session.add(ShiftAssignment(user_id=uid, work_date=wd, shift_code=code,
            segments=json.dumps(segs, ensure_ascii=False), work_hotel_id=whid, created_by=user.id))
    u = User.query.get(uid)
    if u and getattr(u, 'home_hotel_id', None) and getattr(u, 'department_id', None):
        ws = monday_of(wd)
        wp = WeekPlan.query.filter_by(hotel_id=u.home_hotel_id, department_id=u.department_id, week_start=ws).first()
        if not wp:
            wp = WeekPlan(hotel_id=u.home_hotel_id, department_id=u.department_id, week_start=ws, status='draft')
            db.session.add(wp)
        if wp.status in ('submitted', 'locked'):
            wp.status = 'draft'
        wp.updated_by = user.id
    db.session.commit()
    return jsonify(ok=True, n=len(norm))


@app.route('/dashboard/schedule/paste_cells', methods=['POST'])
def schedule_paste_cells():
    """v12.138 — Επικόλληση (Ctrl+V): γράφει την ίδια λίστα entries σε ΠΟΛΛΑ κελιά
    (user_id+date). Σέβεται κλειδωμένες εβδομάδες (τις προσπερνά)."""
    if not _auth():
        return ('', 401)
    if not can_edit_schedule():
        return jsonify(ok=False, err='forbidden'), 403
    user = current_user()
    d = request.json or {}
    board_hid = d.get('board_hotel_id'); board_hid = int(board_hid) if board_hid else None
    norm = []
    for e in (d.get('entries') or []):
        code = (e.get('code') or '').strip()
        if not code:
            continue
        segs = e.get('segments') or e.get('segs') or []
        whid = e.get('work_hotel_id') or e.get('wh'); whid = int(whid) if whid else None
        if code in ('ΕΡΓ', 'ΕΩ', 'ΔΡ'):
            ok, msg = _validate_work_code(code, segs)
            if not ok:
                return jsonify(ok=False, err='invalid', msg=msg), 400
        else:
            segs = []; whid = None
        norm.append((code, segs, whid))
    done = 0; locked = 0; blocked = 0; _last_msg = ''
    touched_wp = set(); _pending = {}
    for c in (d.get('cells') or []):
        try:
            uid = int(c['user_id']); wd = datetime.strptime(c['date'], '%Y-%m-%d').date()
        except Exception:
            continue
        if not week_editable(monday_of(wd), user):
            locked += 1; continue
        # v12.379 Φ2β — εκ-περιτροπής: guard ανά κελί (κλείδωμα/όριο + auto work_hotel)
        _entries_use = []; _blk = False
        for code, segs, whid in norm:
            _gok, _gmsg, _gwh, _grot, _gact = _rotation_guard(uid, wd, code, board_hid, _pending)
            if _grot:
                if not _gok:
                    _blk = True; _last_msg = _gmsg; break
                whid = _gwh
                if is_work_code(code):
                    _pending.setdefault((uid, _gact), set()).add(wd)
            _entries_use.append((code, segs, whid))
        if _blk:
            blocked += 1; continue
        ShiftAssignment.query.filter_by(user_id=uid, work_date=wd).delete()
        for code, segs, whid in _entries_use:
            db.session.add(ShiftAssignment(user_id=uid, work_date=wd, shift_code=code,
                segments=json.dumps(segs, ensure_ascii=False), work_hotel_id=whid, created_by=user.id))
        u = User.query.get(uid)
        if u and getattr(u, 'home_hotel_id', None) and getattr(u, 'department_id', None):
            ws = monday_of(wd)
            key = (u.home_hotel_id, u.department_id, ws)
            if key not in touched_wp:
                touched_wp.add(key)
                wp = WeekPlan.query.filter_by(hotel_id=u.home_hotel_id, department_id=u.department_id, week_start=ws).first()
                if not wp:
                    wp = WeekPlan(hotel_id=u.home_hotel_id, department_id=u.department_id, week_start=ws, status='draft')
                    db.session.add(wp)
                if wp.status in ('submitted', 'locked'):
                    wp.status = 'draft'
                wp.updated_by = user.id
        done += 1
    db.session.commit()
    prompt = None
    cells_in = d.get('cells') or []
    if norm and len(cells_in) == 1:
        try:
            _uid = int(cells_in[0]['user_id'])
            _wd = datetime.strptime(cells_in[0]['date'], '%Y-%m-%d').date()
            _wh = next((w for (_c, _s, w) in norm if w), None)
            ctx = _schedule_return_context(_uid, _wd, _wh)
            if ctx and not ctx['marked']:
                prompt = ctx
        except Exception:
            prompt = None
    if done == 0 and blocked:
        return jsonify(ok=False, err='rotation', msg=_last_msg or 'Δεν επιτρέπεται (εκ περιτροπής).'), 400
    return jsonify(ok=True, done=done, locked=locked, blocked=blocked, return_prompt=prompt)


@app.route('/dashboard/schedule/period_mark', methods=['POST'])
def schedule_period_mark():
    """v12.199 — Αποθηκεύει τη δήλωση του manager στην επιστροφή: νέα περίοδος ή συνέχεια."""
    if not _auth():
        return ('', 401)
    if not can_edit_schedule():
        return jsonify(ok=False, err='forbidden'), 403
    user = current_user()
    d = request.json or {}
    try:
        uid = int(d['user_id'])
        wd = datetime.strptime(d['date'], '%Y-%m-%d').date()
    except Exception:
        return jsonify(ok=False, err='bad'), 400
    kind = (d.get('kind') or 'new').strip()
    if kind not in ('new', 'continue'):
        kind = 'new'
    whid = d.get('hotel_id'); whid = int(whid) if whid else None
    m = SchedulePeriodMark.query.filter_by(user_id=uid, work_date=wd).first()
    if not m:
        m = SchedulePeriodMark(user_id=uid, work_date=wd); db.session.add(m)
    m.kind = kind; m.hotel_id = whid; m.created_by = user.id
    db.session.commit()
    return jsonify(ok=True)


@app.route('/dashboard/schedule/paste_excel', methods=['POST'])
def schedule_paste_excel():
    """v12.209/210 — Επικόλληση ΑΠΟ Excel: δέχεται items=[{user_id,date,text}] (block/grid),
    κάνει parse κάθε κελί (parse_cell) -> βάρδια και το εφαρμόζει. Άγνωστο κείμενο (π.χ. όνομα) αγνοείται."""
    if not _auth():
        return ('', 401)
    if not can_edit_schedule():
        return jsonify(ok=False, err='forbidden'), 403
    user = current_user()
    d = request.json or {}
    items = d.get('items') or []
    board_hid = d.get('board_hotel_id'); board_hid = int(board_hid) if board_hid else None
    short2id = {}
    for h in Hotel.query.all():
        sh = _hotel_short(h.name)
        if sh:
            short2id[sh.upper()] = h.id
    done = 0; locked = 0; skipped = 0; blocked = 0; touched = set(); _pending = {}
    unread = []   # v12.388 — δείγματα κειμένων που δεν αναγνωρίστηκαν (για ειδοποίηση)
    for it in items:
        try:
            uid = int(it['user_id']); wd = datetime.strptime(it['date'], '%Y-%m-%d').date()
        except Exception:
            continue
        code, segs, tag = parse_cell(it.get('text') or '')
        if not code:
            skipped += 1
            _t = (it.get('text') or '').strip()
            if _t and _t not in unread and len(unread) < 15:
                unread.append(_t)
            continue
        if not week_editable(monday_of(wd), user):
            locked += 1; continue
        whid = short2id.get((tag or '').upper()) if tag else None
        # v12.380 F1 — εκ-περιτροπής guard ΚΑΙ στο Excel-paste (κλείδωμα/όριο + auto work_hotel)
        _gok, _gmsg, _gwh, _grot, _gact = _rotation_guard(uid, wd, code, board_hid, _pending)
        if _grot:
            if not _gok:
                blocked += 1; continue
            whid = _gwh
            if is_work_code(code):
                _pending.setdefault((uid, _gact), set()).add(wd)
        ShiftAssignment.query.filter_by(user_id=uid, work_date=wd).delete()
        db.session.add(ShiftAssignment(user_id=uid, work_date=wd, shift_code=code,
            segments=json.dumps(segs, ensure_ascii=False), work_hotel_id=whid, created_by=user.id))
        touched.add((uid, monday_of(wd)))
        done += 1
    for uid, ws in touched:
        u = User.query.get(uid)
        if u and getattr(u, 'home_hotel_id', None) and getattr(u, 'department_id', None):
            wp = WeekPlan.query.filter_by(hotel_id=u.home_hotel_id, department_id=u.department_id, week_start=ws).first()
            if not wp:
                wp = WeekPlan(hotel_id=u.home_hotel_id, department_id=u.department_id, week_start=ws, status='draft')
                db.session.add(wp)
            if wp.status in ('submitted', 'locked'):
                wp.status = 'draft'
            wp.updated_by = user.id
    db.session.commit()
    if not done and blocked:
        return jsonify(ok=False, err='rotation', msg='Δεν επιτρέπεται (εκ περιτροπής — όριο/κλείδωμα ημέρας).')
    if not done and not locked:
        return jsonify(ok=False, msg='Δεν αναγνωρίστηκε βάρδια στο κείμενο του Excel.', skipped=skipped, unread=unread)
    return jsonify(ok=True, done=done, locked=locked, blocked=blocked, skipped=skipped, unread=unread)



# ── Αντιγραφή προηγούμενης εβδομάδας ──────────────────────────────────────────
@app.route('/dashboard/schedule/copyprev', methods=['POST'])
def schedule_copyprev():
    if not _auth() or not can_edit_schedule():
        return redirect(url_for('login'))
    user = current_user()
    hotel_id = request.form.get('hotel_id', type=int)
    dept_id = request.form.get('department_id', type=int)
    week_start = monday_of(datetime.strptime(request.form['week'], '%Y-%m-%d').date())
    if not week_editable(week_start, user):
        return redirect(f'/dashboard/schedule?hotel_id={hotel_id}&week={week_start}&embed=1&err=locked')
    prev = week_start - timedelta(days=7)
    users = _dept_users(hotel_id, dept_id) if dept_id else User.query.filter(User.is_active == True, User.home_hotel_id == hotel_id).all()
    n = 0
    for u in users:
        for i in range(7):
            sd, dd = prev + timedelta(days=i), week_start + timedelta(days=i)
            src = ShiftAssignment.query.filter_by(user_id=u.id, work_date=sd).first()
            if not src:
                continue
            dst = ShiftAssignment.query.filter_by(user_id=u.id, work_date=dd).first()
            if not dst:
                dst = ShiftAssignment(user_id=u.id, work_date=dd, created_by=user.id)
                db.session.add(dst)
            dst.shift_code = src.shift_code
            dst.segments = src.segments
            dst.work_hotel_id = src.work_hotel_id
            n += 1
    db.session.commit()
    log_activity('schedule_copyprev', f'{n} κελιά', hotel_id=hotel_id)
    return redirect(f'/dashboard/schedule?hotel_id={hotel_id}&week={week_start}&embed=1&ok=copied')


# ── ΥΠΟΒΟΛΗ (ενοποιημένη ανά ξενοδοχείο-εβδομάδα) ─────────────────────────────
def _hotel_week_snapshot(hotel_id, week_start):
    days = [week_start + timedelta(days=i) for i in range(7)]
    users = User.query.filter(User.is_active == True, User.home_hotel_id == hotel_id).all()
    snap = {}
    for u in users:
        alist = (ShiftAssignment.query.filter(ShiftAssignment.user_id == u.id)
                 .filter(ShiftAssignment.work_date >= days[0], ShiftAssignment.work_date <= days[6]).all())
        if not alist:
            continue
        dd = {}
        for a in alist:
            dd[a.work_date.isoformat()] = {'code': a.shift_code, 'segs': a.segments or '[]',
                                           'wh': a.work_hotel_id}
        snap[str(u.id)] = {'name': u.full_name, 'dept': getattr(u, 'department_id', None), 'days': dd}
    return snap

def _diff_snapshots(old, new):
    changes = []
    old = old or {}
    keys = set(old.keys()) | set(new.keys())
    for uid in keys:
        oname = (old.get(uid) or new.get(uid) or {}).get('name', uid)
        od = (old.get(uid) or {}).get('days', {})
        nd = (new.get(uid) or {}).get('days', {})
        for d in sorted(set(od.keys()) | set(nd.keys())):
            ov = od.get(d); nv = nd.get(d)
            if (ov or {}).get('code') != (nv or {}).get('code') or (ov or {}).get('segs') != (nv or {}).get('segs'):
                changes.append({'user_id': uid, 'name': oname, 'date': d,
                                'old': (ov or {}).get('code', '—'), 'new': (nv or {}).get('code', '—')})
    return changes

@app.route('/dashboard/schedule/submit', methods=['POST'])
def schedule_submit():
    if not _auth() or not can_edit_schedule():
        return redirect(url_for('login'))
    user = current_user()
    hotel_id = request.form.get('hotel_id', type=int)
    week_start = monday_of(datetime.strptime(request.form['week'], '%Y-%m-%d').date())
    issues = validate_hotel_week(hotel_id, week_start)
    blockers = [i for i in issues if i['severity'] == 'block']
    if blockers:
        return redirect(f'/dashboard/schedule?hotel_id={hotel_id}&week={week_start}&embed=1&err=rules')
    last = (ScheduleSubmission.query.filter_by(hotel_id=hotel_id, week_start=week_start)
            .order_by(ScheduleSubmission.version.desc()).first())
    snap = _hotel_week_snapshot(hotel_id, week_start)
    version = (last.version + 1) if last else 1
    changes = _diff_snapshots(json.loads(last.snapshot) if last and last.snapshot else None, snap) if last else []
    sub = ScheduleSubmission(hotel_id=hotel_id, week_start=week_start, version=version,
                             parent_version=(last.version if last else None),
                             status='submitted', snapshot=json.dumps(snap, ensure_ascii=False),
                             changes=json.dumps(changes, ensure_ascii=False),
                             submitted_by=user.id)
    db.session.add(sub)
    for wp in WeekPlan.query.filter_by(hotel_id=hotel_id, week_start=week_start).all():
        wp.status = 'submitted'
    db.session.commit()
    log_activity('schedule_submit', f'v{version}', hotel_id=hotel_id)
    # ειδοποίηση + email λογιστηρίου (background)
    hotel = Hotel.query.get(hotel_id)
    hn = hotel.name if hotel else ''
    label = 'ΤΡΟΠΟΠΟΙΗΣΗ' if version > 1 else 'Νέα υποβολή'
    _notify_accountants(hotel_id, week_start, version, label, changes)
    return redirect(f'/dashboard/schedule?hotel_id={hotel_id}&week={week_start}&embed=1&ok=submitted')

def _notify_accountants(hotel_id, week_start, version, label, changes):
    try:
        hotel = Hotel.query.get(hotel_id)
        hn = hotel.name if hotel else ''
        wk = week_start.strftime('%d/%m/%Y')
        accts = User.query.filter_by(role='accountant', is_active=True).all()
        for a in accts:
            notify(a.id, f'Πρόγραμμα {hn} — εβδ. {wk} ({label} v{version})', '/dashboard/schedule/submissions?embed=1')
        db.session.commit()
        recips = [a.email for a in accts if a.email] + list(EMAIL_TO_LIST)
        rows = ''.join(f"<tr><td>{c['name']}</td><td>{c['date']}</td><td>{c['old']}</td><td>{c['new']}</td></tr>" for c in changes)
        chtml = (f"<p><b>Αλλαγές (v{version}):</b></p><table border=1 cellpadding=4 style='border-collapse:collapse'>"
                 f"<tr><th>Εργαζόμενος</th><th>Ημ/νία</th><th>Πριν</th><th>Μετά</th></tr>{rows}</table>") if changes else ''
        html = (f"<h3>Πρόγραμμα Εργασίας — {hn}</h3>"
                f"<p>Εβδομάδα <b>{wk}</b> — <b>{label}</b> (έκδοση v{version}).</p>{chtml}"
                f"<p>Δες στο σύστημα: Εστία → Πρόγραμμα Εργασίας → Υποβολές.</p>")
        threading.Thread(target=lambda: send_email(f'[Εστία] Πρόγραμμα {hn} — εβδ. {wk} ({label})', html, recips)).start()
    except Exception as e:
        db.session.rollback()
        print('notify_accountants:', e)

@app.route('/dashboard/schedule/submissions')
def schedule_submissions():
    if not _auth():
        return redirect(url_for('login'))
    if not (is_admin() or is_accountant()):
        return ('Δεν έχετε πρόσβαση', 403)
    user = current_user()
    hids = {h.id for h in allowed_hotels(user)}
    q = ScheduleSubmission.query.order_by(ScheduleSubmission.submitted_at.desc())
    aid = active_hotel_id()
    if aid:
        q = q.filter(ScheduleSubmission.hotel_id == aid)
    subs = [s for s in q.limit(300).all() if (not hids) or s.hotel_id in hids]
    hotels = {h.id: h.name for h in Hotel.query.all()}
    return render_template('schedule_submissions.html', subs=subs, hotels=hotels,
                           month_el=MONTHS_EL)

@app.route('/dashboard/schedule/submission/<int:sid>')
def schedule_submission_view(sid):
    if not _auth():
        return redirect(url_for('login'))
    if not (is_admin() or is_accountant()):
        return ('Δεν έχετε πρόσβαση', 403)
    sub = ScheduleSubmission.query.get_or_404(sid)
    try:
        snap = json.loads(sub.snapshot or '{}')
        changes = json.loads(sub.changes or '[]')
    except Exception:
        snap, changes = {}, []
    days = [sub.week_start + timedelta(days=i) for i in range(7)]
    changed = {(c['user_id'], c['date']) for c in changes}
    hotel = Hotel.query.get(sub.hotel_id)
    return render_template('schedule_submission_view.html', sub=sub, snap=snap, changes=changes,
                           days=days, weekdays=WEEKDAYS_EL, changed=changed,
                           hotel=hotel, month_el=MONTHS_EL)


# ── EXPORT «ΠΡΟΣ ΛΟΓΙΣΤΗΡΙΟ» (μηνιαίο) ────────────────────────────────────────
@app.route('/dashboard/schedule/export.xlsx')
def schedule_export():
    if not _auth() or not (is_admin() or is_accountant()):
        return redirect(url_for('login'))
    import openpyxl, io
    year = request.args.get('year', type=int) or date.today().year
    month = request.args.get('month', type=int) or date.today().month
    hotel_id = request.args.get('hotel_id', type=int) or active_hotel_id()
    data = monthly_settlement(year, month, hotel_id, split=True)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'ΠΡΟΣ ΛΟΓΙΣΤΗΡΙΟ'
    ws.append(['Ονοματεπώνυμο', 'Ξενοδοχείο', 'Μήνας', 'Καθημερινές εργάσιμες', 'Κυριακές', 'Αργίες',
               'Έξτρα ώρες', 'Ρεπό', 'Δουλ. ρεπό', 'Μέρες αλλού', 'Σύνολο ημερών', 'Πληρωτέο συμφωνίας'])
    mname = MONTHS_EL[month]
    for r in data:
        a = r['agg']
        _nm = r['user'].full_name + (' (εκ περιτροπής)' if r.get('split') else '')
        ws.append([_nm, r.get('hotel') or '', mname, a['work_days'], a['sundays'], a['holidays_worked'],
                   a['extra_hours'], a['repo'], a.get('worked_repo', 0), a['elsewhere_days'], a['total_days'], r['payable']])
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    fn = f'PROS_LOGISTIRIO_{year}_{month:02d}.xlsx'
    return Response(bio.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={fn}'})


# ── v12.115 — ΜΗΝΙΑΙΑ ΣΥΓΚΕΝΤΡΩΤΙΚΗ ΚΑΤΑΣΤΑΣΗ (για μεροκάματα) ────────────────
def _monthly_rows(year, month, hotel_id=None, dept_id=None):
    """Ανά εργαζόμενο (home hotel, ΟΛΕΣ οι ώρες μαζί): ώρες/έξτρα/Κυριακές/εργάσιμες/ρεπό
    + ημερολόγιο μήνα (day->assignment). Group ανά home hotel + υποσύνολα + γενικό σύνολο."""
    import calendar as _cal
    ndays = _cal.monthrange(year, month)[1]
    start = date(year, month, 1)
    end = date(year + (month // 12), (month % 12) + 1, 1)
    hol = {h.hol_date for h in Holiday.query.all()}
    try:
        from payroll import EmployeePII as _PII
        piimap = {p.user_id: p for p in _PII.query.all()}
    except Exception:
        piimap = {}
    _colors = {st.code: st.color for st in ShiftType.query.all()}
    _hotels = {h.id: h.name for h in Hotel.query.all()}
    assigns = (ShiftAssignment.query
               .filter(ShiftAssignment.work_date >= start, ShiftAssignment.work_date < end).all())
    by_user = {}
    for a in assigns:
        by_user.setdefault(a.user_id, []).append(a)
    rows = []
    for uid, alist in by_user.items():
        u = User.query.get(uid)
        if not u or not u.is_active:
            continue
        hh = getattr(u, 'home_hotel_id', None)
        if hotel_id and hh != hotel_id:
            continue
        if dept_id and getattr(u, 'department_id', None) != dept_id:
            continue
        days = {}
        hours = extra = 0.0
        _wd = set(); _sun = set(); _hol = set(); _repo = set(); _wrepo = set()
        for a in alist:
            dd = days.setdefault(a.work_date.day, {'entries': [], 'code': '', 'wh': None, 'segs': '[]'})
            try:
                _es = json.loads(a.segments) if a.segments else []
            except Exception:
                _es = []
            _tm = ' & '.join("%s - %s" % (x.get('start'), x.get('end')) for x in _es) if _es else ''
            _hn = _hotels.get(a.work_hotel_id) if (a.work_hotel_id and a.work_hotel_id != hh) else None
            dd['entries'].append({'code': a.shift_code, 'segs': _es, 'wh': a.work_hotel_id,
                                  'times': _tm, 'hotel': _hn, 'hotel_short': _hotel_short(_hn) if _hn else '',
                                  'hours': (round(worked_hours(a), 1) if is_work_code(a.shift_code) else 0),
                                  'color': _colors.get(a.shift_code, '#64748b')})
            if not dd['code']:
                dd['code'] = a.shift_code; dd['wh'] = a.work_hotel_id; dd['segs'] = a.segments or '[]'
            if is_extra_code(a.shift_code):
                extra += worked_hours(a)
            elif is_work_code(a.shift_code):
                hours += worked_hours(a)
                _wd.add(a.work_date)
                if a.shift_code == 'ΔΡ':      # v12.376 — δουλεμένο ρεπό
                    _wrepo.add(a.work_date)
                if a.work_date.weekday() == 6:
                    _sun.add(a.work_date)
                if a.work_date in hol:
                    _hol.add(a.work_date)
            elif is_repo_code(a.shift_code):
                _repo.add(a.work_date)
        sundays = len(_sun); work_days = len(_wd); repo = len(_repo); hol_worked = len(_hol); worked_repo = len(_wrepo)
        prof = EmploymentProfile.query.filter_by(user_id=uid).first()
        payable = 0.0
        if prof and getattr(prof, 'agreement_amount', None):
            if getattr(prof, 'agreement_type', None) == 'Management':
                payable = round(prof.agreement_amount, 2)
            else:
                payable = round((getattr(prof, 'day_wage', 0) or 0) * work_days + extra_wage(prof, extra), 2)
        _pp = piimap.get(uid)
        rows.append({'user': u, 'hotel_id': hh, 'days': days,
                     'emp_code': (_pp.emp_code if _pp else None),
                     'afm': (_pp.afm if _pp else None),
                     'locked': (bool(_pp.locked) if _pp else False),
                     'hours': round(hours, 1), 'extra': round(extra, 1),
                     'sundays': sundays, 'holidays': hol_worked,
                     'work_days': work_days, 'repo': repo, 'worked_repo': worked_repo, 'payable': payable})
    rows.sort(key=lambda r: (r['user'].full_name or ''))
    hotels_by_id = {h.id: h for h in Hotel.query.all()}
    groups = {}
    for r in rows:
        groups.setdefault(r['hotel_id'], []).append(r)
    def _sub(rs):
        return {'hours': round(sum(x['hours'] for x in rs), 1),
                'extra': round(sum(x['extra'] for x in rs), 1),
                'sundays': sum(x['sundays'] for x in rs),
                'holidays': sum(x['holidays'] for x in rs),
                'work_days': sum(x['work_days'] for x in rs),
                'repo': sum(x['repo'] for x in rs),
                'worked_repo': sum(x['worked_repo'] for x in rs),
                'payable': round(sum(x['payable'] for x in rs), 2),
                'count': len(rs)}
    out = []
    for hid, rs in sorted(groups.items(),
                          key=lambda kv: (hotels_by_id.get(kv[0]).name if hotels_by_id.get(kv[0]) else 'zzz')):
        out.append({'hotel': hotels_by_id.get(hid), 'rows': rs, 'subtotal': _sub(rs)})
    return {'groups': out, 'grand': _sub(rows), 'ndays': ndays}


def _monthly_args():
    year = request.args.get('year', type=int) or date.today().year
    month = request.args.get('month', type=int) or date.today().month
    hotel_id = request.args.get('hotel_id', type=int) or 0
    dept_id = request.args.get('dept', type=int) or 0
    return year, month, hotel_id, dept_id


@app.route('/dashboard/schedule/monthly')
def schedule_monthly():
    if not _auth() or not (is_admin() or is_accountant()):
        return redirect(url_for('login'))
    user = current_user()
    year, month, hotel_id, dept_id = _monthly_args()
    view = request.args.get('view') or 'summary'
    data = _monthly_rows(year, month, hotel_id or None, dept_id or None)
    hotels = allowed_hotels(user) or Hotel.query.order_by(Hotel.name).all()
    depts = Department.query.order_by(Department.name).all()
    deptmap = {d.id: d.name for d in depts}
    hol = {h.hol_date for h in Holiday.query.all()}
    WD = ['Δε', 'Τρ', 'Τε', 'Πε', 'Πα', 'Σα', 'Κυ']
    day_hdr = []
    for d in range(1, data['ndays'] + 1):
        dt = date(year, month, d)
        day_hdr.append({'d': d, 'wd': WD[dt.weekday()], 'we': dt.weekday() >= 5,
                        'hol': dt in hol, 'iso': dt.isoformat()})
    shift_types = ShiftType.query.filter_by(active=True).order_by(ShiftType.sort).all()
    return render_template('schedule_monthly.html', data=data, year=year, month=month,
        hotel_id=hotel_id, dept_id=dept_id, view=view, hotels=hotels, depts=depts, deptmap=deptmap,
        months=MONTHS_EL, day_hdr=day_hdr, ndays=data['ndays'],
        can_edit=(can_edit_schedule() and is_admin()),
        shift_types=shift_types,
        shift_types_json=json.dumps([{'code': s.code, 'color': s.color} for s in entry_shift_types()], ensure_ascii=False),
        years=list(range(date.today().year - 2, date.today().year + 2)),
        is_admin=is_admin())


@app.route('/dashboard/schedule/monthly.xlsx')
def schedule_monthly_xlsx():
    if not _auth() or not (is_admin() or is_accountant()):
        return redirect(url_for('login'))
    import openpyxl, io
    from openpyxl.styles import Font, PatternFill
    year, month, hotel_id, dept_id = _monthly_args()
    data = _monthly_rows(year, month, hotel_id or None, dept_id or None)
    deptmap = {d.id: d.name for d in Department.query.all()}
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Μηνιαία'
    navy = PatternFill('solid', fgColor='193847')
    def boldrow(size=11):
        for c in range(1, 10):
            ws.cell(row=ws.max_row, column=c).font = Font(bold=True, size=size)
    ws.append(['Εστία — Μηνιαία συγκεντρωτική · %s %d' % (MONTHS_EL[month], year)])
    ws['A1'].font = Font(bold=True, size=14, color='193847')
    ws.append([])
    ws.append(['Ξενοδοχείο', 'Ονοματεπώνυμο', 'Τμήμα', 'Ώρες', 'Έξτρα ώρες',
               'Κυριακές', 'Εργάσιμες', 'Ρεπό', 'Πληρωτέο'])
    for c in range(1, 10):
        cc = ws.cell(row=ws.max_row, column=c); cc.font = Font(bold=True, color='FFFFFF'); cc.fill = navy
    for g in data['groups']:
        hn = g['hotel'].name if g['hotel'] else '— (χωρίς ξενοδοχείο)'
        for x in g['rows']:
            ws.append([hn, x['user'].full_name, deptmap.get(getattr(x['user'], 'department_id', None), ''),
                       x['hours'], x['extra'], x['sundays'], x['work_days'], x['repo'], x['payable']])
        s = g['subtotal']
        ws.append(['Σύνολο · %s' % hn, '', '', s['hours'], s['extra'], s['sundays'],
                   s['work_days'], s['repo'], s['payable']]); boldrow()
    gd = data['grand']
    ws.append(['ΓΕΝΙΚΟ ΣΥΝΟΛΟ', '', '', gd['hours'], gd['extra'], gd['sundays'],
               gd['work_days'], gd['repo'], gd['payable']]); boldrow(12)
    widths = [26, 30, 18, 9, 11, 10, 11, 8, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    fn = 'estia_monthly_%d_%02d.xlsx' % (year, month)
    return Response(bio.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=%s' % fn})


@app.route('/dashboard/schedule/monthly.pdf')
def schedule_monthly_pdf():
    if not _auth() or not (is_admin() or is_accountant()):
        return redirect(url_for('login'))
    from fpdf import FPDF
    year, month, hotel_id, dept_id = _monthly_args()
    data = _monthly_rows(year, month, hotel_id or None, dept_id or None)
    deptmap = {d.id: d.name for d in Department.query.all()}
    NAVY = (25, 56, 71); GREY = (120, 120, 120)
    pdf = FPDF(orientation='L', unit='mm', format='A4'); pdf.set_auto_page_break(True, margin=12)
    pdf.add_font('dv', '', os.path.join(BASE_DIR, 'assets', 'fonts', 'DejaVuSans.ttf'))
    pdf.add_font('dv', 'B', os.path.join(BASE_DIR, 'assets', 'fonts', 'DejaVuSans-Bold.ttf'))
    pdf.add_page()
    try:
        pdf.image(os.path.join(BASE_DIR, 'static', 'img', 'logo.png'), x=12, y=9, h=13)
    except Exception:
        pass
    pdf.set_xy(30, 10); pdf.set_font('dv', 'B', 15); pdf.set_text_color(*NAVY)
    pdf.cell(0, 8, 'Εστία — Μηνιαία συγκεντρωτική κατάσταση', ln=1)
    pdf.set_x(30); pdf.set_font('dv', '', 10); pdf.set_text_color(*GREY)
    pdf.cell(0, 6, '%s %d · Εκτύπωση: %s' % (MONTHS_EL[month], year, date.today().strftime('%d/%m/%Y')), ln=1)
    pdf.ln(5)
    headers = ['Ονοματεπώνυμο', 'Τμήμα', 'Ώρες', 'Έξτρα', 'Κυρ.', 'Εργάσ.', 'Ρεπό', 'Δ.ρεπό', 'Πληρωτέο']
    widths = [70, 48, 24, 24, 20, 24, 18, 19, 30]
    def hdr():
        pdf.set_font('dv', 'B', 9); pdf.set_text_color(255, 255, 255); pdf.set_fill_color(*NAVY)
        for h, w in zip(headers, widths):
            pdf.cell(w, 8, h, border=0, fill=True, align='L')
        pdf.ln(8)
    for g in data['groups']:
        hn = g['hotel'].name if g['hotel'] else '— (χωρίς ξενοδοχείο)'
        pdf.set_font('dv', 'B', 11); pdf.set_text_color(*NAVY)
        pdf.cell(0, 8, hn, ln=1)
        hdr()
        fill = False
        for x in g['rows']:
            pdf.set_fill_color(243, 247, 250) if fill else pdf.set_fill_color(255, 255, 255)
            pdf.set_font('dv', '', 8.5); pdf.set_text_color(40, 40, 40)
            vals = [x['user'].full_name, deptmap.get(getattr(x['user'], 'department_id', None), ''),
                    x['hours'], x['extra'], x['sundays'], x['work_days'], x['repo'],
                    (x['worked_repo'] or ''), x['payable']]
            for val, w in zip(vals, widths):
                s = str(val)
                while s and pdf.get_string_width(s) > w - 2 and len(s) > 3:
                    s = s[:-2]
                pdf.cell(w, 7, s, border=0, fill=True, align='L')
            pdf.ln(7); fill = not fill
        s = g['subtotal']
        pdf.set_font('dv', 'B', 9); pdf.set_text_color(*NAVY); pdf.set_fill_color(225, 235, 245)
        sub = ['Σύνολο · %s' % hn, '', s['hours'], s['extra'], s['sundays'], s['work_days'], s['repo'],
               (s['worked_repo'] or ''), s['payable']]
        for val, w in zip(sub, widths):
            ss = str(val)
            while ss and pdf.get_string_width(ss) > w - 2 and len(ss) > 3:
                ss = ss[:-2]
            pdf.cell(w, 7, ss, border=0, fill=True, align='L')
        pdf.ln(10)
    gd = data['grand']
    pdf.set_font('dv', 'B', 11); pdf.set_text_color(*NAVY)
    pdf.cell(0, 8, 'ΓΕΝΙΚΟ ΣΥΝΟΛΟ: %s ώρες · %s έξτρα · %s Κυρ. · %s εργάσιμες · %s ρεπό · %s δουλ. ρεπό · %s €'
             % (gd['hours'], gd['extra'], gd['sundays'], gd['work_days'], gd['repo'], gd['worked_repo'], gd['payable']), ln=1)
    out = bytes(pdf.output())
    fn = 'estia_monthly_%d_%02d.pdf' % (year, month)
    return Response(out, mimetype='application/pdf',
        headers={'Content-Disposition': 'attachment; filename=%s' % fn})

# ── v12.116 — ΕΠΟΠΤΕΙΑ ΠΡΟΓΡΑΜΜΑΤΟΣ (τι κάνουν οι managers, μηνιαία) ───────────
def _oversight_weeks(year, month):
    import calendar as _cal
    ndays = _cal.monthrange(year, month)[1]
    first = date(year, month, 1); last = date(year, month, ndays)
    w = monday_of(first); weeks = []
    while w <= last:
        weeks.append(w); w = w + timedelta(days=7)
    return weeks


def _oversight_data(year, month):
    weeks = _oversight_weeks(year, month)
    wkmin, wkmax = weeks[0], weeks[-1]
    umap = {u.id: u for u in User.query.all()}
    def uname(uid): 
        u = umap.get(uid); return (u.full_name if u else '—')
    # WeekPlans + Submissions στο εύρος
    wps = WeekPlan.query.filter(WeekPlan.week_start >= wkmin, WeekPlan.week_start <= wkmax).all()
    subs = (ScheduleSubmission.query
            .filter(ScheduleSubmission.week_start >= wkmin, ScheduleSubmission.week_start <= wkmax).all())
    wp_idx = {}
    for wp in wps:
        wp_idx[(wp.hotel_id, wp.department_id, wp.week_start)] = wp
    sub_idx = {}
    for s in subs:
        k = (s.hotel_id, s.week_start)
        if k not in sub_idx or (s.version or 0) > (sub_idx[k].version or 0):
            sub_idx[k] = s
    STAT = {'draft': ('Πρόχειρο', '#fde68a', '#92400e'),
            'submitted': ('Υποβλήθηκε', '#bbf7d0', '#166534'),
            'locked': ('Κλειδώθηκε', '#bfdbfe', '#1e40af')}
    # ── BY HOTEL ──
    hotels = Hotel.query.filter_by(is_active=True).order_by(Hotel.name).all()
    by_hotel = []
    for h in hotels:
        depts = _depts_present(h.id)
        rows = []
        done = total = 0
        for d in depts:
            cells = []
            for wk in weeks:
                wp = wp_idx.get((h.id, d.id, wk))
                total += 1
                st = wp.status if wp else None
                if st in ('submitted', 'locked'): done += 1
                cells.append({'wk': wk, 'status': st,
                              'label': (STAT.get(st, ('—', '#f1f5f9', '#94a3b8'))[0] if st else '—'),
                              'bg': (STAT.get(st, ('', '#f8fafc', ''))[1] if st else '#f8fafc'),
                              'fg': (STAT.get(st, ('', '', '#94a3b8'))[2] if st else '#94a3b8'),
                              'by': (uname(wp.updated_by) if wp and wp.updated_by else ''),
                              'when': (wp.updated_at.strftime('%d/%m %H:%M') if wp and wp.updated_at else '')})
            rows.append({'dept': d, 'cells': cells})
        sub_cells = []
        for wk in weeks:
            s = sub_idx.get((h.id, wk))
            sub_cells.append({'wk': wk, 'sub': s,
                              'by': (uname(s.submitted_by) if s and s.submitted_by else ''),
                              'when': (s.submitted_at.strftime('%d/%m %H:%M') if s and s.submitted_at else ''),
                              'ver': (s.version if s else None), 'sid': (s.id if s else None)})
        by_hotel.append({'hotel': h, 'depts': depts, 'rows': rows, 'subs': sub_cells,
                         'done': done, 'total': total})
    # ── BY MANAGER ── (όποιος έχει δραστηριότητα στην περίοδο)
    actor_ids = set([wp.updated_by for wp in wps if wp.updated_by] +
                    [s.submitted_by for s in subs if s.submitted_by])
    mgr = []
    for uid in actor_ids:
        u = umap.get(uid)
        if not u:
            continue
        my_wp = [wp for wp in wps if wp.updated_by == uid]
        my_sub = [s for s in subs if s.submitted_by == uid]
        hotel_ids = set([wp.hotel_id for wp in my_wp] + [s.hotel_id for s in my_sub])
        hotel_names = ', '.join(sorted({(Hotel.query.get(hid).name if Hotel.query.get(hid) else '—') for hid in hotel_ids}))
        times = [wp.updated_at for wp in my_wp if wp.updated_at] + [s.submitted_at for s in my_sub if s.submitted_at]
        last = max(times) if times else None
        sub_list = sorted(my_sub, key=lambda s: s.submitted_at or datetime.min, reverse=True)
        mgr.append({'user': u, 'role': u.role, 'n_wp': len(my_wp), 'n_sub': len(my_sub),
                    'hotels': hotel_names, 'last': (last.strftime('%d/%m/%Y %H:%M') if last else '—'),
                    'subs': sub_list})
    mgr.sort(key=lambda m: (-(m['n_sub'] + m['n_wp']), m['user'].full_name or ''))
    return {'weeks': weeks, 'by_hotel': by_hotel, 'by_manager': mgr}


@app.route('/dashboard/schedule/oversight')
def schedule_oversight():
    if not _auth() or not is_admin():
        return redirect(url_for('login'))
    year = request.args.get('year', type=int) or date.today().year
    month = request.args.get('month', type=int) or date.today().month
    view = request.args.get('view') or 'hotel'
    data = _oversight_data(year, month)
    return render_template('schedule_oversight.html', data=data, year=year, month=month,
        view=view, months=MONTHS_EL, years=list(range(date.today().year - 2, date.today().year + 2)),
        is_admin=is_admin())






# ── IMPORT page (upload workbook) ─────────────────────────────────────────────
@app.route('/dashboard/schedule/import', methods=['GET', 'POST'])
def schedule_import():
    if not _auth() or not is_admin():
        return redirect(url_for('login'))
    results = None
    if request.method == 'POST':
        files = request.files.getlist('files')
        if not files:
            f = request.files.get('file')
            files = [f] if f else []
        only_year = None if request.form.get('all_years') else request.form.get('year', type=int)
        results = []
        for f in files:
            if not f or not f.filename:
                continue
            try:
                data = f.read()
                kind, stats = import_any(data, f.filename, only_year=only_year, created_by=session.get('user_id'))
                results.append({'name': f.filename, 'kind': kind, 'stats': stats})
            except Exception as e:
                results.append({'name': f.filename, 'kind': 'error', 'stats': {'error': str(e)}})
        seed_schedule()
        log_activity('schedule_import_multi', f'{len(results)} αρχεία')
    try:
        from payroll import EmployeePII as _PII
        _locked = {p.user_id for p in _PII.query.filter_by(locked=True).all()}
    except Exception:
        _locked = set()
    purge_count = sum(1 for u in imported_staff_query().all() if u.id not in _locked)
    pending_total = PendingShift.query.count()
    return render_template('schedule_import.html', results=results, year=date.today().year, purge_count=purge_count, pending_total=pending_total)


# ── ΤΑΥΤΟΠΟΙΗΣΗ: προθαλαμος ονοματων χωρις επιβεβαιωμενο master ────────────────
def _confirm_link(norm_name, master_id, created_by=None):
    """Ο χρηστης επιβεβαιωνει: φτιαχνει NameLink + μεταφερει ΟΛΕΣ τις PendingShift στο master."""
    master = User.query.get(master_id)
    if not master:
        return 0
    pend = PendingShift.query.filter_by(norm_name=norm_name).all()
    rawn = pend[0].raw_name if pend else None
    link = NameLink.query.filter_by(norm_name=norm_name).first()
    if link:
        link.user_id = master_id
    else:
        db.session.add(NameLink(norm_name=norm_name, user_id=master_id,
                                raw_name=rawn, created_by=created_by))
    moved = 0
    for ps in pend:
        dup = ShiftAssignment.query.filter_by(user_id=master_id, work_date=ps.work_date,
            shift_code=ps.shift_code, segments=ps.segments).first()
        if not dup:
            db.session.add(ShiftAssignment(user_id=master_id, work_date=ps.work_date,
                shift_code=ps.shift_code, segments=ps.segments,
                work_hotel_id=ps.work_hotel_id, created_by=created_by))
            moved += 1
        db.session.delete(ps)
    db.session.commit()
    return moved

def _pending_pii_map():
    try:
        from payroll import EmployeePII as _PII
        return {pp.user_id: pp for pp in _PII.query.all()}
    except Exception:
        return {}

@app.context_processor
def _inject_pending_identify():
    def pending_identify_count():
        try:
            return db.session.query(PendingShift.norm_name).distinct().count()
        except Exception:
            return 0
    return {'pending_identify_count': pending_identify_count}

@app.route('/dashboard/schedule/identify')
def schedule_identify():
    if not _auth() or not is_admin():
        return redirect(url_for('login'))
    # v12.374 (P-051 cleanup) — αποσύρθηκε· ανακατεύθυνση στην ενοποιημένη «Ταυτοποίηση εισαγωγών»
    return redirect(url_for('schedule_import_hub') + ('?embed=1' if request.args.get('embed') else ''))

@app.route('/dashboard/schedule/identify/link', methods=['POST'])
def schedule_identify_link():
    if not _auth() or not is_admin():
        return redirect(url_for('login'))
    nm = (request.form.get('norm_name') or '').strip()
    mid = request.form.get('master_id', type=int)
    if nm and mid:
        moved = _confirm_link(nm, mid, created_by=session.get('user_id'))
        u = User.query.get(mid)
        log_activity('schedule_identify_link', '%s -> %s (%d βάρδιες)' % (nm, (u.full_name if u else mid), moved))
    _bk = request.form.get('back')
    return redirect((_bk + ('&embed=1' if '?' in _bk else '?embed=1')) if _bk else url_for('schedule_identify'))

@app.route('/dashboard/schedule/identify/dismiss', methods=['POST'])
def schedule_identify_dismiss():
    if not _auth() or not is_admin():
        return redirect(url_for('login'))
    nm = (request.form.get('norm_name') or '').strip()
    if nm:
        n = PendingShift.query.filter_by(norm_name=nm).delete()
        db.session.commit()
        log_activity('schedule_identify_dismiss', '%s (%s)' % (nm, n))
    _bk = request.form.get('back')
    return redirect((_bk + ('&embed=1' if '?' in _bk else '?embed=1')) if _bk else url_for('schedule_identify'))

@app.route('/dashboard/schedule/identify/search')
def schedule_identify_search():
    if not _auth() or not is_admin():
        return jsonify([])
    qraw = (request.args.get('q') or '').strip()
    q = _norm(qraw)
    if not qraw:
        return jsonify([])
    piimap = _pending_pii_map()
    locked = _locked_uids()
    res = []
    for u in User.query.filter(User.is_active == True).all():
        pp = piimap.get(u.id)
        afm = (pp.afm if pp else '') or ''
        emp = (pp.emp_code if pp else '') or ''
        if (q and q in _norm(u.full_name or '')) or (qraw in afm) or (qraw.upper() in emp.upper()):
            res.append({'id': u.id, 'name': u.full_name, 'emp_code': emp, 'afm': afm,
                        'locked': u.id in locked})
        if len(res) >= 20:
            break
    return jsonify(res)

def _exact_master_map():
    """norm_name -> λιστα ενεργων users με ΑΚΡΙΒΕΣ ιδιο ονομα."""
    m = {}
    for u in User.query.filter(User.is_active == True).all():
        nm = _norm(u.full_name or '')
        if nm:
            m.setdefault(nm, []).append(u)
    return m

def _seed_locked_links(created_by=None):
    """Ταυτοτητα (ΟΧΙ merge): NameLink καθε locked master -> στο ιδιο του το ονομα.
    Skip: ασαφη (ιδιο norm_name σε >1 locked) + οσα εχουν ηδη link."""
    locked = _locked_uids()
    by_nm = {}
    for u in User.query.filter(User.is_active == True).all():
        if u.id in locked:
            nm = _norm(u.full_name or '')
            if nm:
                by_nm.setdefault(nm, []).append(u)
    n = 0; skipped = 0
    for nm, us in by_nm.items():
        if len(us) != 1:
            skipped += 1; continue
        if NameLink.query.filter_by(norm_name=nm).first():
            continue
        db.session.add(NameLink(norm_name=nm, user_id=us[0].id,
                                raw_name=us[0].full_name, created_by=created_by))
        n += 1
    db.session.commit()
    return n, skipped

@app.route('/dashboard/schedule/identify/seed_locked', methods=['POST'])
def schedule_identify_seed_locked():
    if not _auth() or not is_admin():
        return redirect(url_for('login'))
    n, skipped = _seed_locked_links(created_by=session.get('user_id'))
    log_activity('schedule_identify_seed_locked', '%d links, %d ασαφη' % (n, skipped))
    # μετα το seed, τρεξε και αυτοματη ταυτοποιηση ακριβων για οσα ηδη ειναι στον προθαλαμο
    _bk = request.form.get('back')
    return redirect((_bk + ('&embed=1' if '?' in _bk else '?embed=1')) if _bk else url_for('schedule_identify'))

@app.route('/dashboard/schedule/identify/auto', methods=['POST'])
def schedule_identify_auto():
    if not _auth() or not is_admin():
        return redirect(url_for('login'))
    locked = _locked_uids()
    exact = _exact_master_map()
    norms = [r[0] for r in db.session.query(PendingShift.norm_name).distinct().all()]
    linked = 0
    for nm in norms:
        cands = exact.get(nm) or []
        target = None
        if len(cands) == 1:
            target = cands[0]
        elif len(cands) > 1:
            lk = [u for u in cands if u.id in locked]
            if len(lk) == 1:
                target = lk[0]
        if target:
            _confirm_link(nm, target.id, created_by=session.get('user_id'))
            linked += 1
    log_activity('schedule_identify_auto', '%d ακριβη ταιριασματα' % linked)
    _bk = request.form.get('back')
    return redirect((_bk + ('&embed=1' if '?' in _bk else '?embed=1')) if _bk else url_for('schedule_identify'))


@app.route('/dashboard/schedule/identify/clear_all', methods=['POST'])
def schedule_identify_clear_all():
    # v12.114 — καθαρισμος ΟΛΟΥ του προθαλαμου (1 κλικ αντι N «Αγνόησε»)· ιδανικο
    # για reset πριν σωστο re-import (π.χ. οταν τα ξενοδοχεια ηταν λαθος στην πηγη).
    if not _auth() or not is_admin():
        return redirect(url_for('login'))
    n = PendingShift.query.delete()
    db.session.commit()
    log_activity('schedule_identify_clear_all', '%d βαρδιες προθαλαμου' % (n or 0))
    _bk = request.form.get('back')
    if _bk:
        return redirect(_bk + ('&embed=1' if '?' in _bk else '?embed=1'))
    return redirect(url_for('schedule_identify') + ('?embed=1' if request.args.get('embed') else ''))


@app.route('/dashboard/schedule/imported')
def schedule_imported():
    # v12.114 — ενιαια οθονη: εισηγμενα (keyless) προφιλ με badge + προταση 🔒 master
    # + κουμπια «Συγχωνευση με Master» / «Διαγραφη». Ξεχωριζει «αληθινος→merge» απο «σκουπιδι→delete».
    if not _auth() or not is_admin():
        return redirect(url_for('login'))
    # v12.374 (P-051 cleanup) — αποσύρθηκε· ανακατεύθυνση στην ενοποιημένη «Ταυτοποίηση εισαγωγών»
    return redirect(url_for('schedule_import_hub') + ('?embed=1' if request.args.get('embed') else ''))


@app.route('/dashboard/schedule/imported/merge', methods=['POST'])
def schedule_imported_merge():
    if not _auth() or not is_admin():
        return redirect(url_for('login'))
    drop_id = int(request.form.get('drop_id') or 0)
    master_id = int(request.form.get('master_id') or 0)
    drop = User.query.get(drop_id); master = User.query.get(master_id)
    locked = _locked_uids()
    if drop and master and drop.id != master.id and drop.id not in locked:
        # μεταφορα ΜΟΝΟ βαρδιων στο master (διπλη μερα: κραταμε του master)
        keep_dates = {a.work_date for a in ShiftAssignment.query.filter_by(user_id=master_id).all()}
        for a in ShiftAssignment.query.filter_by(user_id=drop_id).all():
            if a.work_date in keep_dates:
                db.session.delete(a)
            else:
                a.user_id = master_id; keep_dates.add(a.work_date)
        # NameLink ωστε μελλοντικα imports αυτου του ονοματος να πανε στο master
        nm = _norm(drop.full_name or '')
        if nm:
            link = NameLink.query.filter_by(norm_name=nm).first()
            if link:
                link.user_id = master_id
            else:
                db.session.add(NameLink(norm_name=nm, user_id=master_id,
                                        raw_name=drop.full_name, created_by=session.get('user_id')))
        db.session.commit()
        # τωρα ο drop ειναι χωρις βαρδιες -> πληρης διαγραφη (cascade)
        try:
            from payroll import _hard_delete_user as _hdel
            _hdel(drop_id)
        except Exception:
            db.session.rollback()
        log_activity('schedule_imported_merge', '%d -> %d' % (drop_id, master_id))
    _bk = request.form.get('back')
    if _bk:
        return redirect(_bk + ('&embed=1' if '?' in _bk else '?embed=1'))
    return redirect(url_for('schedule_imported') + ('?embed=1' if request.args.get('embed') else ''))


@app.route('/dashboard/schedule/imported/delete', methods=['POST'])
def schedule_imported_delete():
    if not _auth() or not is_admin():
        return redirect(url_for('login'))
    uid = int(request.form.get('uid') or 0)
    try:
        from payroll import _hard_delete_user as _hdel
        ok, _r = _hdel(uid)
        log_activity('schedule_imported_delete', '%d (%s)' % (uid, 'ok' if ok else _r))
    except Exception:
        db.session.rollback()
    _bk = request.form.get('back')
    if _bk:
        return redirect(_bk + ('&embed=1' if '?' in _bk else '?embed=1'))
    return redirect(url_for('schedule_imported') + ('?embed=1' if request.args.get('embed') else ''))


# ── v12.372 (P-051 Φ3) — ΕΝΙΑΙΑ «Ταυτοποίηση εισαγωγών»: προθάλαμος + εισηγμένα σε 2 tabs ──
@app.route('/dashboard/schedule/import_hub')
def schedule_import_hub():
    if not _auth() or not is_admin():
        return redirect(url_for('login'))
    piimap = _pending_pii_map(); locked = _locked_uids()
    # Tab 1 — εκκρεμείς βάρδιες (PendingShift), ίδια λογική με schedule_identify
    groups = {}
    for ps in PendingShift.query.order_by(PendingShift.work_date).all():
        g = groups.get(ps.norm_name)
        if not g:
            g = {'norm': ps.norm_name, 'raw': ps.raw_name, 'count': 0, 'hotel_tag': ps.hotel_tag,
                 'dept': ps.dept_raw, 'employer': ps.employer, 'dmin': ps.work_date, 'dmax': ps.work_date}
            groups[ps.norm_name] = g
        g['count'] += 1
        if ps.work_date:
            if not g['dmin'] or ps.work_date < g['dmin']: g['dmin'] = ps.work_date
            if not g['dmax'] or ps.work_date > g['dmax']: g['dmax'] = ps.work_date
    pending = []
    for g in groups.values():
        g['suggestions'] = [{'id': u.id, 'name': u.full_name,
                             'emp_code': (piimap.get(u.id).emp_code if piimap.get(u.id) else None),
                             'afm': (piimap.get(u.id).afm if piimap.get(u.id) else None),
                             'locked': u.id in locked, 'score': sc}
                            for u, sc in _suggest_masters(g['raw'] or g['norm'], limit=5)]
        pending.append(g)
    pending.sort(key=lambda x: (-x['count'], x['raw'] or ''))
    # Tab 2 — εισηγμένα προφίλ (keyless), ίδια λογική με schedule_imported
    imp = []
    for u in imported_staff_query().filter(User.is_active == True).all():
        if u.id in locked:
            continue
        sugg = None
        for cand, sc in _suggest_masters(u.full_name or '', limit=8):
            if cand.id == u.id or cand.id not in locked:
                continue
            pp = piimap.get(cand.id)
            sugg = {'id': cand.id, 'name': cand.full_name,
                    'emp_code': (pp.emp_code if pp else None), 'afm': (pp.afm if pp else None), 'score': sc}
            break
        imp.append({'id': u.id, 'name': u.full_name,
                    'shifts': ShiftAssignment.query.filter_by(user_id=u.id).count(), 'suggestion': sugg})
    imp.sort(key=lambda r: (0 if r['suggestion'] else 1, -r['shifts'], r['name'] or ''))
    log_activity('schedule_import_hub', '%d pending / %d imported' % (len(pending), len(imp)))
    return render_template('schedule_import_hub.html', pending=pending, n_pending=len(pending),
                           imported=imp, n_imported=len(imp))


# ── PASTE: μαζική επικόλληση προγράμματος από Excel (όνομα + Δ–Κ) ──────────────
def _name_index():
    ix, ixs = {}, {}
    for u in User.query.filter(User.is_active == True).all():
        if getattr(u, 'employment_active', None) is False:
            continue
        fn = u.full_name or ''
        k = _norm(fn)
        if k:
            ix.setdefault(k, []).append(u)
        toks = sorted([t for t in re.sub(r'[^a-zα-ω0-9 ]', '', _acc(fn).lower()).split() if t])
        ks = ''.join(toks)
        if ks:
            ixs.setdefault(ks, []).append(u)
    return ix, ixs

def _match_name(name, ix, ixs):
    k = _norm(name)
    if k in ix:
        return (ix[k][0], None) if len(ix[k]) == 1 else (None, 'διπλό όνομα στο Μητρώο')
    toks = sorted([t for t in re.sub(r'[^a-zα-ω0-9 ]', '', _acc(name).lower()).split() if t])
    ks = ''.join(toks)
    if ks in ixs:
        return (ixs[ks][0], None) if len(ixs[ks]) == 1 else (None, 'αμφίσημο όνομα')
    return (None, 'δεν βρέθηκε στο Μητρώο')

def _cell_label(code, segs):
    if code is None:
        return '—'
    if code == 'ΕΡΓ':
        if segs:
            return 'ΕΡΓ ' + ', '.join('%s-%s' % (x.get('start'), x.get('end')) for x in segs)
        return 'ΕΡΓ'
    return code

def _parse_paste(raw, monday):
    ix, ixs = _name_index()
    days = [monday + timedelta(days=i) for i in range(7)]
    rows = []
    for line in (raw or '').splitlines():
        if not line.strip():
            continue
        cells = line.split('	')
        name = cells[0].strip()
        if not name or _norm(name) in ('ονοματεπωνυμο', 'ονομα', 'επωνυμο', 'τμημα'):
            continue
        user = _route_name(name); warn = None if user else "προς ταυτοποίηση (μη συνδεδεμένο)"
        dayvals = cells[1:8]
        parsed = []
        for i, dt in enumerate(days):
            raw_v = dayvals[i].strip() if i < len(dayvals) else ''
            code, segs, tag = parse_cell(raw_v) if raw_v else (None, [], None)
            whid = None
            if tag:
                th = _resolve_hotel_by_code(tag); whid = th.id if th else None
            elif user is not None:
                whid = getattr(user, 'home_hotel_id', None)
            unknown = bool(raw_v) and code is None
            lbl = _cell_label(code, segs)
            if unknown and (_norm(raw_v) in ('ρεπο', 'ρεπ', 'ρ', 'off', 'repo', 'dayoff') or raw_v in ('-', '—', '/')):
                unknown = False; lbl = 'ΡΕΠΟ'   # ρεπό = κενή ημέρα (καμία ανάθεση)
            parsed.append({'date': dt, 'raw': raw_v, 'code': code, 'segs': segs,
                           'whid': whid, 'label': lbl, 'unknown': unknown})
        rows.append({'name': name, 'user': user, 'warn': warn, 'days': parsed,
                     'matched': user is not None})
    return rows, days

@app.route('/dashboard/schedule/paste', methods=['GET', 'POST'])
def schedule_paste():
    if not _auth() or not is_admin():
        return redirect(url_for('login'))
    monday = _week_arg()
    rows = None; saved = None
    if request.method == 'POST':
        raw = request.form.get('data', '')
        ws = request.form.get('week_start', '')
        try:
            monday = monday_of(datetime.strptime(ws, '%Y-%m-%d').date())
        except Exception:
            monday = _week_arg()
        rows, days = _parse_paste(raw, monday)
        if request.form.get('action') == 'commit':
            user = current_user(); n_new = n_upd = 0
            for r in rows:
                if not r['user']:
                    continue
                for d in r['days']:
                    if d['code'] is None:
                        continue
                    a = ShiftAssignment.query.filter_by(user_id=r['user'].id, work_date=d['date']).first()
                    if a:
                        a.shift_code = d['code']; a.segments = json.dumps(d['segs'], ensure_ascii=False)
                        a.work_hotel_id = d['whid']; n_upd += 1
                    else:
                        db.session.add(ShiftAssignment(user_id=r['user'].id, work_date=d['date'],
                            shift_code=d['code'], segments=json.dumps(d['segs'], ensure_ascii=False),
                            work_hotel_id=d['whid'], created_by=user.id)); n_new += 1
                    u = r['user']
                    if getattr(u, 'home_hotel_id', None) and getattr(u, 'department_id', None):
                        wp = WeekPlan.query.filter_by(hotel_id=u.home_hotel_id, department_id=u.department_id, week_start=monday).first()
                        if not wp:
                            wp = WeekPlan(hotel_id=u.home_hotel_id, department_id=u.department_id, week_start=monday, status='draft'); db.session.add(wp)
                        elif wp.status in ('submitted', 'locked'):
                            wp.status = 'draft'
            db.session.commit()
            seed_schedule()
            log_activity('schedule_paste', 'νέες %d / ενημ. %d' % (n_new, n_upd))
            saved = {'new': n_new, 'upd': n_upd,
                     'unmatched': sum(1 for r in rows if not r['user'])}
    else:
        days = [monday + timedelta(days=i) for i in range(7)]
    return render_template('schedule_paste.html', rows=rows, days=days, monday=monday,
                           week_start=monday.strftime('%Y-%m-%d'), saved=saved,
                           raw=request.form.get('data', ''))



# ── ADMIN settings (κωδικοί / τμήματα / αργίες / πολιτική / κανόνες) ───────────
@app.route('/dashboard/schedule/settings', methods=['GET', 'POST'])
def schedule_settings():
    if not _auth() or not is_admin():
        return redirect(url_for('login'))
    if request.method == 'POST':
        act = request.form.get('action')
        if act == 'policy':
            set_policy({k: request.form.get(k) for k in POLICY_DEFAULTS if request.form.get(k) is not None})
        elif act == 'holiday':
            try:
                hd = datetime.strptime(request.form['hol_date'], '%Y-%m-%d').date()
                if not Holiday.query.filter_by(hol_date=hd).first():
                    db.session.add(Holiday(hol_date=hd, description=request.form.get('description', '')[:120], year=hd.year))
                    db.session.commit()
            except Exception:
                db.session.rollback()
        elif act == 'holiday_del':
            h = Holiday.query.get(request.form.get('id', type=int))
            if h:
                db.session.delete(h); db.session.commit()
        elif act == 'rule_toggle':
            r = ScheduleRule.query.get(request.form.get('id', type=int))
            if r:
                r.active = not r.active; db.session.commit()
        elif act == 'shift_toggle':
            s = ShiftType.query.get(request.form.get('id', type=int))
            if s:
                s.active = not s.active; db.session.commit()
        elif act == 'shift_colors':
            import re as _re
            for st in ShiftType.query.all():
                _c = request.form.get('color_%d' % st.id)
                if _c and _re.match(r'^#[0-9a-fA-F]{6}$', _c):
                    st.color = _c
            db.session.commit()
        elif act == 'shift_meta':
            for st in ShiftType.query.all():
                _cat = request.form.get('cat_%d' % st.id)
                if _cat in ('work', 'extra', 'repo', 'absence'):
                    st.count_as = _cat
                    st.counts_as_work = _cat in ('work', 'extra')  # συγχρονισμός legacy
                st.break_deduct = (request.form.get('brk_%d' % st.id) is not None)
            db.session.commit()
        elif act == 'entry_codes':
            for _key, _field in (('sched_entry_codes', 'entry_code'), ('sched_entry_codes_mgr', 'entry_code_mgr')):
                codes = request.form.getlist(_field)
                row = Setting.query.get(_key)
                if not row:
                    row = Setting(key=_key); db.session.add(row)
                row.value = json.dumps(codes, ensure_ascii=False)
            db.session.commit()
        return redirect('/dashboard/schedule/settings?embed=1&ok=1')
    _allow = _entry_codes_setting('admin'); _allow_m = _entry_codes_setting('manager')
    return render_template('schedule_settings.html',
        policy=get_policy(), shift_types=ShiftType.query.order_by(ShiftType.sort).all(),
        entry_codes=_allow, entry_all=(_allow is None),
        entry_codes_mgr=_allow_m, entry_all_mgr=(_allow_m is None),
        depts=Department.query.order_by(Department.sort).all(),
        holidays=Holiday.query.order_by(Holiday.hol_date).all(),
        rules=ScheduleRule.query.all(), dow_el=DOW_EL)


# ── ROUTE: Διαχείριση Προσωπικού (οργανόγραμμα: τμήμα/ξενοδοχείο/login) ────────
@app.route('/dashboard/schedule/staff', methods=['GET', 'POST'])
def schedule_staff():
    if not _auth() or not is_admin():
        return redirect(url_for('login'))
    if request.method == 'POST':
        act = request.form.get('action')
        if act == 'add':
            full = (request.form.get('full_name') or '').strip()
            if full:
                base = re.sub(r'[^a-z0-9.]', '', _acc(full).lower().replace(' ', '.')) or 'emp'
                uname = base
                i = 1
                while User.query.filter_by(username=uname).first():
                    i += 1; uname = f'{base}.{i}'
                u = User(username=uname[:50], password=generate_password_hash(os.urandom(8).hex()),
                         full_name=full[:100], role='staff', approved=True, is_active=True,
                         department_id=request.form.get('department_id', type=int) or None,
                         home_hotel_id=request.form.get('home_hotel_id', type=int) or None,
                         employer=(request.form.get('employer') or '')[:120] or None,
                         login_enabled=bool(request.form.get('login_enabled')),
                         employment_active=True)
                db.session.add(u); db.session.commit()
                log_activity('staff_add', full)
        elif act == 'edit':
            u = User.query.get(request.form.get('user_id', type=int))
            if u:
                import people  # v12.169 — μέσω helper (ένα write-path + ιστορικό)
                people.assign_user_org(u, request.form.get('home_hotel_id', type=int) or None,
                                       request.form.get('department_id', type=int) or None)
                u.employer = (request.form.get('employer') or '')[:120] or None
                u.login_enabled = bool(request.form.get('login_enabled'))
                db.session.commit()
        elif act == 'toggle_active':
            u = User.query.get(request.form.get('user_id', type=int))
            if u:
                u.employment_active = not bool(getattr(u, 'employment_active', True))
                db.session.commit()
        qs = f"?embed=1&hotel_id={request.form.get('f_hotel','')}&department_id={request.form.get('f_dept','')}"
        return redirect('/dashboard/schedule/staff' + qs)
    f_hotel = request.args.get('hotel_id', type=int)
    f_dept = request.args.get('department_id', type=int)
    q = User.query.filter(User.is_active == True)
    if f_hotel:
        q = q.filter(User.home_hotel_id == f_hotel)
    if f_dept:
        q = q.filter(User.department_id == f_dept)
    users = q.order_by(User.full_name).limit(800).all()
    dup_groups = find_dup_groups()
    return render_template('schedule_staff.html',
        users=users, depts=Department.query.order_by(Department.sort).all(),
        hotels=Hotel.query.filter_by(is_active=True).order_by(Hotel.name).all(),
        f_hotel=f_hotel, f_dept=f_dept, dup_groups=dup_groups,
        dept_map={d.id: d.name for d in Department.query.all()},
        hotel_map={h.id: h.name for h in Hotel.query.all()})


# ── ΕΚΚΑΘΑΡΙΣΗ / ΣΥΓΧΩΝΕΥΣΗ ΔΙΠΛΩΝ ΕΡΓΑΖΟΜΕΝΩΝ ──────────────────────────────
def _lev(a, b):
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cur.append(min(prev[j] + 1, cur[j - 1] + 1, prev[j - 1] + (ca != cb)))
        prev = cur
    return prev[-1]

def _name_parts(full):
    toks = _acc(full or '').lower().split()
    sur = toks[0] if toks else ''
    first = ' '.join(toks[1:]) if len(toks) > 1 else ''
    return _norm(sur), _norm(first)

def _likely_dup(u1, u2):
    s1, f1 = _name_parts(u1.full_name)
    s2, f2 = _name_parts(u2.full_name)
    if not s1 or s1 != s2:
        return False                       # διαφορετικό επώνυμο -> όχι
    if f1 == f2:
        return True                        # ίδιο πλήρες όνομα
    if not f1 or not f2:
        return True                        # ο ένας χωρίς μικρό (π.χ. "FARO" vs "FARO ANNA")
    if f1.startswith(f2) or f2.startswith(f1):
        return True
    if _lev(f1, f2) <= 2 and min(len(f1), len(f2)) >= 3:
        return True                        # JOEY vs JOY, typos
    return False

def _names_likely_same(full_a, full_b):
    s1, f1 = _name_parts(full_a); s2, f2 = _name_parts(full_b)
    if not s1 or s1 != s2:
        return False
    if f1 == f2 or not f1 or not f2:
        return True
    if f1.startswith(f2) or f2.startswith(f1):
        return True
    pl = 0
    for a, b in zip(f1, f2):
        if a == b: pl += 1
        else: break
    if pl >= 4:                            # ΝΙΚΟΣ <-> ΝΙΚΟΛΑΟΣ (κοινο προθεμα >=4)
        return True
    if _lev(f1, f2) <= 2 and min(len(f1), len(f2)) >= 3:
        return True                        # NATALIA <-> NATALIIA
    return False

def _locked_uids():
    try:
        from payroll import EmployeePII as _PII
        return {row[0] for row in db.session.query(_PII.user_id).filter_by(locked=True).all()}
    except Exception:
        return set()

def _tok_key(name):
    return ''.join(sorted([t for t in re.sub(r'[^a-z\u03b1-\u03c90-9 ]', '', _acc(name or '').lower()).split() if t]))

def _match_existing(full):
    """Ταιριαζει το ονομα με υπαρχον προφιλ - ΠΡΟΤΕΡΑΙΟΤΗΤΑ στο master (Λογιστηριο).
    Ετσι οι βαρδιες κουμπωνουν στα master προφιλ αντι να φτιαχνονται διπλα."""
    fn = _norm(full)
    if not fn:
        return None
    users = User.query.filter(User.is_active == True).all()
    locked = _locked_uids()
    def prefer(cs):
        for u in cs:
            if u.id in locked: return u
        return cs[0] if cs else None
    ex = [u for u in users if _norm(u.full_name) == fn]
    if ex: return prefer(ex)
    tgt = _tok_key(full)
    ts = [u for u in users if tgt and _tok_key(u.full_name) == tgt]
    if ts: return prefer(ts)
    fz = [u for u in users if u.id in locked and _names_likely_same(full, u.full_name or '')]
    if fz: return fz[0]
    return None

def find_dup_groups():
    users = User.query.filter(User.is_active == True).order_by(User.full_name).all()
    parent = {u.id: u.id for u in users}
    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]; x = parent[x]
        return x
    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra
    by_sur = {}
    for u in users:
        s, _ = _name_parts(u.full_name)
        by_sur.setdefault(s, []).append(u)
    for s, lst in by_sur.items():
        if not s or len(lst) < 2:
            continue
        for i in range(len(lst)):
            for j in range(i + 1, len(lst)):
                if _likely_dup(lst[i], lst[j]):
                    union(lst[i].id, lst[j].id)
    groups = {}
    umap = {u.id: u for u in users}
    for u in users:
        groups.setdefault(find(u.id), []).append(u)
    out = []
    for gid, members in groups.items():
        if len(members) > 1:
            # εμπλουτισμός: πλήθος βαρδιών + αν είναι κλειδωμένο (Λογιστήριο)
            try:
                from payroll import EmployeePII as _PII
                _locked = {row[0] for row in db.session.query(_PII.user_id).filter_by(locked=True).all()}
            except Exception:
                _locked = set()
            for m in members:
                m._assign_n = ShiftAssignment.query.filter_by(user_id=m.id).count()
                m._locked = m.id in _locked
            # default keep = πρώτα το κλειδωμένο (Λογιστήριο), μετά όποιος έχει τις περισσότερες βάρδιες
            members.sort(key=lambda m: (not m._locked, -m._assign_n))
            out.append(members)
    out.sort(key=lambda g: g[0].full_name or '')
    return out

def merge_users(keep_id, drop_ids):
    """v12.103 — ΜΟΝΟ βάρδιες μεταφέρονται στον keep. Θέση/τμήμα/ξενοδοχείο/προφίλ
    του keep ΜΕΝΟΥΝ ως έχουν (τα ήδη assigned στη βάση). Οι διπλοί αρχειοθετούνται."""
    keep = User.query.get(keep_id)
    if not keep:
        return 0
    moved = 0
    keep_dates = {a.work_date for a in ShiftAssignment.query.filter_by(user_id=keep_id).all()}
    for d in drop_ids:
        u = User.query.get(d)
        if not u or u.id == keep_id:
            continue
        # ΜΟΝΟ βάρδιες -> keep (διπλή μέρα: κρατάμε του keep)
        for a in ShiftAssignment.query.filter_by(user_id=d).all():
            if a.work_date in keep_dates:
                db.session.delete(a)
            else:
                a.user_id = keep_id; keep_dates.add(a.work_date); moved += 1
        # ΔΕΝ μεταφέρουμε θέση/τμήμα/ξενοδοχείο/προφίλ — καθαρίζουμε του διπλού
        dp = EmploymentProfile.query.filter_by(user_id=d).first()
        if dp:
            db.session.delete(dp)
        try:
            import faults as _flt
            for us in _flt.UserSpecialty.query.filter_by(user_id=d).all():
                db.session.delete(us)
        except Exception:
            pass
        db.session.delete(u)
    db.session.commit()
    return moved

@app.route('/dashboard/schedule/staff/merge', methods=['POST'])
def schedule_staff_merge():
    if not _auth() or not is_admin():
        return redirect(url_for('login'))
    keep_id = request.form.get('keep_id', type=int)
    drop_ids = [int(x) for x in request.form.getlist('drop_ids') if x.isdigit() and int(x) != keep_id]
    if keep_id and drop_ids:
        n = merge_users(keep_id, drop_ids)
        log_activity('staff_merge', f'keep={keep_id} drop={drop_ids} moved={n}')
    return redirect('/dashboard/schedule/staff?embed=1&ok=merged')


# ── Καθαρισμός εισαγμένου προσωπικού (login ανενεργό) ─────────────────────────
def imported_staff_query():
    return User.query.filter(User.login_enabled == False)

@app.route('/dashboard/schedule/staff/purge', methods=['POST'])
def schedule_staff_purge():
    if not _auth() or not is_admin():
        return redirect(url_for('login'))
    # v12.114 — full cascade ανά χρηστη μεσω _hard_delete_user (καθαριζει ολους τους
    # σχετιζομενους πινακες + προστατευει τα κλειδωμενα)· per-user try ωστε ενα FK
    # σφαλμα να ΜΗΝ μπλοκαρει τους υπολοιπους (πριν: μερικη διαγραφη -> FK violation).
    staff = imported_staff_query().all()
    try:
        from payroll import EmployeePII as _PII, _hard_delete_user as _hdel
        locked_ids = {p.user_id for p in _PII.query.filter_by(locked=True).all()}
    except Exception:
        locked_ids = set(); _hdel = None
    staff = [u for u in staff if u.id not in locked_ids]   # v12.56: μην σβήνεις κλειδωμένους (Epsilon)
    done = 0; failed = 0
    for u in staff:
        ok = False
        if _hdel:
            try:
                ok, _r = _hdel(u.id)
            except Exception:
                db.session.rollback(); ok = False
        if ok: done += 1
        else:  failed += 1
    log_activity('schedule_staff_purge', 'v12.114: %d διαγραφηκαν, %d απετυχαν' % (done, failed))
    return redirect('/dashboard/schedule/import?embed=1&purged=' + str(done))


# ── ΕΙΣΑΓΩΓΗ ΜΗΤΡΩΟΥ ΕΡΓΑΖΟΜΕΝΩΝ (payroll «Εργαζόμενοι» — καθαρή πηγή) ─────────
REG_NAME_KEYS = {'ονοματεπωνυμο', 'επωνυμο'}
def _parse_date_cell(v):
    if v is None or v == '':
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', str(v).strip())
    if m:
        y = int(m.group(3)); y = y + 2000 if y < 100 else y
        try:
            return date(y, int(m.group(2)), int(m.group(1)))
        except Exception:
            return None
    return None

def _to_float(v):
    try:
        return float(str(v).replace(',', '.'))
    except Exception:
        return None

def _registry_header(ws, maxr=12):
    """Βρες γραμμή-κεφαλίδα μητρώου: έχει τμημα + (επωνυμο|ονοματεπωνυμο) και ΟΧΙ ημερομηνίες."""
    REG = {'id': ['id'], 'active': ['ενεργος', 'κατασταση'], 'dept': ['τμημα'],
           'position': ['θεση', 'ειδικοτητα'], 'epon': ['επωνυμο'], 'onoma': ['ονομα'],
           'fullname': ['ονοματεπωνυμο'], 'employer': ['εταιρεια'], 'upok': ['υποκ'],
           'amount': ['συμφωνια', 'συμφωνημενοποσο'], 'days': ['ημερεσμηνα', 'ημερες'],
           'hours': ['ωρεςμερα', 'ωρες'], 'hired': ['ημπροσληψης', 'ημερομηνιαπροσληψης'],
           'left': ['ημαποχωρησης', 'ημερομηνιααποχωρησης'], 'email': ['email'], 'phone': ['τηλεφωνο']}
    for r in range(1, min(maxr, ws.max_row) + 1):
        cmap = {}; has_date = False
        for c in range(1, min(ws.max_column, 30) + 1):
            v = ws.cell(r, c).value
            nv = _norm(v)
            if isinstance(v, datetime) or (isinstance(v, str) and re.match(r'\d{1,2}/\d{1,2}/\d{4}', v.strip())):
                has_date = True
            for key, alts in REG.items():
                if nv in alts and key not in cmap:
                    cmap[key] = c
        if not has_date and 'dept' in cmap and ('epon' in cmap or 'fullname' in cmap):
            return r, cmap
    return None, None

def detect_kind(wb):
    for ws in wb.worksheets:
        # schedule = έχει ΤΜΗΜΑ + στήλες ημερομηνιών
        for r in range(1, min(6, ws.max_row) + 1):
            if _norm(ws.cell(r, 1).value) == 'τμημα':
                for c in range(1, min(ws.max_column, 30) + 1):
                    v = ws.cell(r, c).value
                    if isinstance(v, datetime) or (isinstance(v, str) and re.match(r'\d{1,2}/\d{1,2}/\d{4}', str(v).strip())):
                        return 'schedule'
    for ws in wb.worksheets:
        hr, _ = _registry_header(ws)
        if hr:
            return 'registry'
    return 'unknown'

def import_staff_registry(source, hotel_hint=None):
    import openpyxl, io
    wb = (openpyxl.load_workbook(io.BytesIO(source), data_only=True)
          if isinstance(source, (bytes, bytearray)) else openpyxl.load_workbook(source, data_only=True))
    hint_hotel = _resolve_hotel_by_code(hotel_hint) if hotel_hint else None
    stats = {'users_new': 0, 'users_upd': 0, 'profiles': 0, 'inactive': 0, 'rows': 0}
    # διάλεξε ΕΝΑ φύλλο μητρώου: προτίμησε «Εργαζόμενοι», αλλιώς το 1ο με header μητρώου
    target = None
    for ws in wb.worksheets:
        if 'εργαζομεν' in _norm(ws.title):
            if _registry_header(ws)[0]:
                target = ws; break
    if target is None:
        for ws in wb.worksheets:
            if _registry_header(ws)[0]:
                target = ws; break
    for ws in ([target] if target else []):
        hr, cm = _registry_header(ws)
        for r in range(hr + 1, ws.max_row + 1):
            def g(k):
                return ws.cell(r, cm[k]).value if cm.get(k) else None
            if cm.get('fullname'):
                full = str(g('fullname') or '').strip()
            else:
                full = (str(g('epon') or '').strip() + ' ' + str(g('onoma') or '').strip()).strip()
            if not full or _norm(full) in ('', 'συνολο', 'ονοματεπωνυμο'):
                continue
            stats['rows'] += 1
            dept = resolve_department(g('dept'))
            upok = g('upok')
            hotel = _resolve_hotel_by_code(upok) if upok else hint_hotel
            # active
            av = _norm(g('active'))
            active = True
            if av in ('οχι', 'ανενεργος', 'αποχωρησε', 'inactive', 'no'):
                active = False
            fn = _norm(full)
            user = None
            for u in User.query.all():
                if _norm(u.full_name) == fn:
                    user = u; break
            if not user:
                base = re.sub(r'[^a-z0-9.]', '', _acc(full).lower().replace(' ', '.')) or 'emp'
                un = base; i = 1
                while User.query.filter_by(username=un).first():
                    i += 1; un = f'{base}.{i}'
                user = User(username=un[:50], password=generate_password_hash(os.urandom(8).hex()),
                            full_name=full[:100], role='staff', approved=True, is_active=True,
                            login_enabled=False)
                db.session.add(user); db.session.flush()
                stats['users_new'] += 1
            else:
                stats['users_upd'] += 1
            if dept and not user.department_id:
                user.department_id = dept.id
            if hotel and not user.home_hotel_id:
                user.home_hotel_id = hotel.id
            if g('employer') and not user.employer:
                user.employer = str(g('employer'))[:120]
            if upok and not user.subunit:
                user.subunit = str(upok)[:20]
            if user.login_enabled is None:
                user.login_enabled = False
            user.employment_active = active
            if not active:
                stats['inactive'] += 1
            # EmploymentProfile
            prof = EmploymentProfile.query.filter_by(user_id=user.id).first()
            amount = _to_float(g('amount'))
            if amount or g('position') or g('hired'):
                if not prof:
                    prof = EmploymentProfile(user_id=user.id); db.session.add(prof); stats['profiles'] += 1
                if amount:
                    prof.agreement_amount = amount
                dd = _to_float(g('days'));  hh = _to_float(g('hours'))
                if dd:
                    prof.days_per_month = int(dd)
                if hh:
                    prof.hours_per_day = hh
                if g('position'):
                    prof.position = str(g('position'))[:80]
                if g('hired'):
                    prof.hired_at = _parse_date_cell(g('hired'))
                if g('left'):
                    prof.left_at = _parse_date_cell(g('left'))
                prof.status = 'Ενεργός' if active else 'Ανενεργός'
        db.session.commit()
    wb.close()
    return stats

def import_any(source, filename='', only_year=None, created_by=None):
    """Auto-detect: εισάγει ΚΑΙ μητρώο ΚΑΙ πρόγραμμα αν υπάρχουν στο ίδιο αρχείο."""
    import openpyxl, io
    wb = openpyxl.load_workbook(io.BytesIO(source) if isinstance(source, (bytes, bytearray)) else source, data_only=True)
    has_reg = any(_registry_header(ws)[0] for ws in wb.worksheets)
    has_sch = False
    for ws in wb.worksheets:
        for r in range(1, min(6, ws.max_row) + 1):
            if _norm(ws.cell(r, 1).value) == 'τμημα':
                for c in range(1, min(ws.max_column, 30) + 1):
                    v = ws.cell(r, c).value
                    if isinstance(v, datetime) or (isinstance(v, str) and re.match(r'\d{1,2}/\d{1,2}/\d{4}', str(v).strip())):
                        has_sch = True; break
            if has_sch:
                break
        if has_sch:
            break
    wb.close()
    code = None
    m = re.match(r'^\s*([A-Za-zΑ-Ωα-ω]{2,4})\b', filename or '')
    if m:
        code = m.group(1).upper()
    out = {'registry': None, 'schedule': None}
    if has_reg:
        out['registry'] = import_staff_registry(source, hotel_hint=code)
    if has_sch:
        out['schedule'] = import_schedule_workbook(source, only_year=only_year, created_by=created_by)
    kind = '+'.join([k for k in ('registry', 'schedule') if out[k] is not None]) or 'unknown'
    return kind, out


# ══════════════════════════════════════════════════════════════════════════════
# v12.359 — Κονσόλα «Κατάσταση Προσωπικού ανά Ξενοδοχείο» (read-only grid + Excel)
# Δυναμικό: grid ΚΑΙ excel διαβάζουν από PAYROLL_GRID_COLS (single source of truth).
# Πηγές (read-only): EmploymentProfile · schedule.aggregate() · payroll.LegalNetImport · _hotel_short.
# ══════════════════════════════════════════════════════════════════════════════
STAFF_STATUS_YEAR   = 2026
STAFF_STATUS_MONTHS = [1, 2, 3, 4, 5, 6]   # Ιαν→Ιουν

# Ενιαίο μητρώο στηλών — άλλαξε label/src ΕΔΩ και ενημερώνονται grid + excel μαζί.
# src = κλειδί μέσα στο dict του μηνιαίου cell (βλ. _staff_status_rows).
# grp: 'mgmt'=Management (αχνό κόκκινο) · 'acct'=Λογιστήριο (αχνό πράσινο έως Μάιο) · 'bal'=Υπόλοιπο
PAYROLL_GRID_COLS = [
    {'k': 'hotel',  'label': 'HOTEL CODE',      'src': 'hotel_code', 'kind': 'text',  'grp': ''},
    {'k': 'agree',  'label': 'ΠΟΣΟ ΣΥΜΦΩΝΙΑΣ',   'src': 'agreement',  'kind': 'money', 'grp': 'mgmt', 'cap': 'MANAGEMENT'},
    {'k': 'day',    'label': 'ΗΜΕΡΟΜΙΣΘΙΟ',      'src': 'day_wage',   'kind': 'money', 'grp': 'mgmt', 'cap': 'MANAGEMENT'},
    {'k': 'hour',   'label': 'ΩΡΟΜΙΣΘΙΟ',        'src': 'hour_wage',  'kind': 'money', 'grp': 'mgmt', 'cap': 'MANAGEMENT'},
    {'k': 'paymgmt','label': 'ΠΛΗΡΩΤΕΟ MANAGEMENT','src': 'pay_mgmt', 'kind': 'money', 'grp': ''},
    {'k': 'extra',  'label': 'ΕΞΤΡΑ ΩΡΕΣ',       'src': 'extra',      'kind': 'num',   'grp': ''},
    {'k': 'work',   'label': 'ΕΡΓΑΣΙΜΕΣ',        'src': 'work',       'kind': 'int',   'grp': ''},
    {'k': 'repo',   'label': 'ΡΕΠΟ',             'src': 'repo',       'kind': 'int',   'grp': ''},
    {'k': 'net',    'label': 'ΚΑΘΑΡΟ ΠΛΗΡΩΤΕΟ',   'src': 'net',        'kind': 'money', 'grp': 'acct', 'cap': 'ΛΟΓΙΣΤΗΡΙΟ'},
    {'k': 'dx',     'label': 'ΔΧ',               'src': 'gift_xmas',   'kind': 'money', 'grp': 'acct'},
    {'k': 'aa',     'label': 'ΑΑ',               'src': 'comp_leave',  'kind': 'money', 'grp': 'acct'},
    {'k': 'dp',     'label': 'ΔΠ',               'src': 'gift_easter', 'kind': 'money', 'grp': 'acct'},
    {'k': 'ea',     'label': 'ΕΑ',               'src': 'leave_allow', 'kind': 'money', 'grp': 'acct'},
    {'k': 'apol',   'label': 'Αποζ. Απολ.',      'src': 'comp_dismiss','kind': 'money', 'grp': 'acct'},
    {'k': 'bal',    'label': 'ΥΠΟΛΟΙΠΟ',         'src': 'balance',    'kind': 'money', 'grp': 'bal'},
]

def _staff_status_rows(year=None, months=None):
    """Read-only γραμμές ανά (εργαζόμενος × ξενοδοχείο). Home hotel = Οργανόγραμμα (User.home_hotel_id).
    Κάθε μήνας cell: hotel_code/agreement/day_wage/hour_wage/extra/work/repo/net.
    «Καθαρό πληρωτέο» (LegalNetImport.net_legal) μπαίνει ΜΟΝΟ στη γραμμή του home hotel (δεν διπλομετριέται)."""
    year = year or STAFF_STATUS_YEAR
    months = months or STAFF_STATUS_MONTHS
    try:
        from payroll import _employees, LegalNetImport
    except Exception:
        _employees, LegalNetImport = None, None
    if _employees:
        emp_rows = _employees('active')
    else:
        emp_rows = [{'user': u, 'profile': EmploymentProfile.query.filter_by(user_id=u.id).first(),
                     'hotel_id': getattr(u, 'home_hotel_id', None)}
                    for u in User.query.filter((User.employment_active == True) | (User.employment_active.is_(None))).all()]
    hotel_name = {h.id: h.name for h in Hotel.query.all()}
    start = date(year, months[0], 1)
    endm = months[-1]
    end = date(year + (endm // 12), (endm % 12) + 1, 1)
    uids = [r['user'].id for r in emp_rows]
    shifts = []
    if uids:
        shifts = (ShiftAssignment.query
                  .filter(ShiftAssignment.user_id.in_(uids),
                          ShiftAssignment.work_date >= start,
                          ShiftAssignment.work_date < end).all())
    # (uid, work_hotel_id|None, month) -> [assignments]
    by_uhm = {}
    for a in shifts:
        by_uhm.setdefault((a.user_id, a.work_hotel_id, a.work_date.month), []).append(a)
    # (uid, μήνας) -> {net, gift_xmas, comp_leave, gift_easter, leave_allow, comp_dismiss}
    # Καθαρά Λογιστηρίου (LegalNetImport.net_legal) ανά κατηγορία period_kind.
    SPECIAL_MATCH = [
        ('gift_xmas',    lambda n: 'δωρο' in n and 'χριστ' in n and 'προηγ' not in n),
        ('gift_easter',  lambda n: 'δωρο' in n and 'πασχα' in n and 'προηγ' not in n),
        ('leave_allow',  lambda n: 'επιδομα' in n and 'αδει' in n and 'προηγ' not in n),
        ('comp_leave',   lambda n: 'αποζημιωσ' in n and 'αδει' in n and 'προηγ' not in n),
        ('comp_dismiss', lambda n: 'αποζημιωσ' in n and 'απολυσ' in n),
    ]
    pay_by_um = {}   # (uid, μήνας) -> dict κατηγοριών
    if LegalNetImport and uids:
        for li in (LegalNetImport.query
                   .filter(LegalNetImport.year == year,
                           LegalNetImport.month.in_(months),
                           LegalNetImport.user_id.in_(uids)).all()):
            k = (li.user_id, li.month); d = pay_by_um.setdefault(k, {})
            pk = (li.period_kind or 'monthly'); val = li.net_legal or 0.0
            if pk == 'monthly':
                d['net'] = round(d.get('net', 0.0) + val, 2)
            else:
                nn = _norm(pk)
                for key, match in SPECIAL_MATCH:
                    if match(nn):
                        d[key] = round(d.get(key, 0.0) + val, 2); break
    rows = []
    for er in emp_rows:
        u = er['user']; prof = er.get('profile'); pii = er.get('pii')
        afm = (pii.afm if (pii and pii.afm) else '')
        home = er.get('hotel_id') or getattr(u, 'home_hotel_id', None)
        agreement = round(prof.agreement_amount, 2) if (prof and prof.agreement_amount) else None
        day_wage  = round(prof.day_wage, 2) if prof else None
        hour_wage = round(prof.hour_wage, 2) if prof else None
        hotels_worked = set()
        for (uid, whid, mo) in by_uhm.keys():
            if uid == u.id:
                hotels_worked.add(whid if whid else home)
        if home:
            hotels_worked.add(home)
        if not hotels_worked:
            hotels_worked = {home}
        ordered = ([home] if home in hotels_worked else []) + \
                  sorted([h for h in hotels_worked if h != home], key=lambda x: (x or 0))
        for hk in ordered:
            is_home = (hk == home)
            mcells = {}
            any_data = False
            for mo in months:
                alist = []
                for (uid, whid, m2), lst in by_uhm.items():
                    if uid == u.id and m2 == mo and ((whid or home) == hk):
                        alist += lst
                agg = aggregate(alist, home) if alist else {'work_days': 0, 'repo': 0, 'extra_hours': 0.0}
                pay = pay_by_um.get((u.id, mo), {}) if is_home else {}
                net = pay.get('net') if is_home else None
                sp = {k: (pay.get(k) if is_home else None) for k in
                      ('gift_xmas', 'comp_leave', 'gift_easter', 'leave_allow', 'comp_dismiss')}
                active = bool(alist) or (net is not None) or any(v for v in sp.values())
                _dw = day_wage if active else None
                _work = agg['work_days'] or 0
                _paymgmt = round((_dw or 0) * _work, 2) if (active and _dw and _work) else None
                _bal = round(_paymgmt - net, 2) if (_paymgmt is not None and net is not None) else None
                mcells[mo] = {
                    'active': active,
                    'hotel_code': (_hotel_short(hotel_name.get(hk, '')) if hk else '') if active else '',
                    'agreement': agreement if active else None,
                    'day_wage': _dw,
                    'hour_wage': hour_wage if active else None,
                    'pay_mgmt': _paymgmt,
                    'extra': (round(agg['extra_hours'], 2) or None) if active else None,
                    'work': (agg['work_days'] or None) if active else None,
                    'repo': (agg['repo'] or None) if active else None,
                    'net': net,
                    'gift_xmas': sp['gift_xmas'], 'comp_leave': sp['comp_leave'],
                    'gift_easter': sp['gift_easter'], 'leave_allow': sp['leave_allow'],
                    'comp_dismiss': sp['comp_dismiss'],
                    'balance': _bal,
                }
                if active:
                    any_data = True
            if is_home or any_data:
                rows.append({'user': u, 'hotel_id': hk, 'is_home': is_home, 'afm': afm,
                             'hotel_code': _hotel_short(hotel_name.get(hk, '')) if hk else '',
                             'name': u.full_name or u.username, 'months': mcells})
    rows.sort(key=lambda r: (r['name'] or '', 0 if r['is_home'] else 1, r['hotel_code'] or ''))
    return rows


@app.route('/dashboard/schedule/staff_status')
def schedule_staff_status():
    if not _auth():
        return redirect(url_for('login'))
    if not is_admin():
        return redirect(url_for('schedule_board'))
    rows = _staff_status_rows()
    # σύνολα ανά (μήνας, στήλη) για το footer
    totals = {}
    for mo in STAFF_STATUS_MONTHS:
        tt = {}
        for col in PAYROLL_GRID_COLS:
            if col['kind'] in ('money', 'num', 'int'):
                tot = 0.0
                for r in rows:
                    v = r['months'][mo].get(col['src'])
                    if isinstance(v, (int, float)):
                        tot += v
                tt[col['src']] = round(tot, 2) if tot else None
        totals[mo] = tt
    return render_template('schedule_staff_status.html',
        rows=rows, cols=PAYROLL_GRID_COLS, months=STAFF_STATUS_MONTHS, totals=totals,
        month_el=MONTHS_EL, year=STAFF_STATUS_YEAR, is_admin=is_admin(), n_emp=len(rows))


@app.route('/dashboard/schedule/staff_status.xlsx')
def schedule_staff_status_xlsx():
    if not _auth() or not is_admin():
        return redirect(url_for('login'))
    rows = _staff_status_rows()
    import openpyxl, io
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    cols = PAYROLL_GRID_COLS; ncol = len(cols)
    nmonths = len(STAFF_STATUS_MONTHS)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Προσωπικό'
    navy = PatternFill('solid', fgColor='153847'); navy2 = PatternFill('solid', fgColor='0F2A36')
    grey = PatternFill('solid', fgColor='334155'); grey2 = PatternFill('solid', fgColor='3E4C5E')
    gold = PatternFill('solid', fgColor='CAA64A'); totfill = PatternFill('solid', fgColor='FBF3DE')
    zebra = PatternFill('solid', fgColor='F7F9FB')
    # group fills
    mgmt_hdr = PatternFill('solid', fgColor='7A2F2F'); mgmt_cell = PatternFill('solid', fgColor='FCE6E6')
    acct_hdr = PatternFill('solid', fgColor='2F6B3A'); acct_cell = PatternFill('solid', fgColor='E4F4E6')
    bal_hdr  = PatternFill('solid', fgColor='8A6D1C'); bal_cell  = PatternFill('solid', fgColor='FFF4D6')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='E3E9EF'); gside = Side(style='medium', color='CAA64A')
    money_fmt = '#,##0.00'; num_fmt = '0.00'; int_fmt = '0'
    def fmt_of(kind):
        return money_fmt if kind == 'money' else (num_fmt if kind == 'num' else (int_fmt if kind == 'int' else None))
    def hdr_style(grp):
        if grp == 'mgmt': return mgmt_hdr, 'F7DEDE'
        if grp == 'acct': return acct_hdr, 'DFF2E3'
        if grp == 'bal':  return bal_hdr, 'FBEEC4'
        return None, 'FFFFFF'
    def cell_fill(grp, mo, even):
        if grp == 'mgmt': return mgmt_cell
        if grp == 'acct' and mo <= 5: return acct_cell
        if grp == 'bal': return bal_cell
        return zebra if even else None
    # Row1 identity (ΑΦΜ · ΞΕΝ · Ονοματεπώνυμο) + month headers
    for col_i, lbl in ((1, 'ΑΦΜ'), (2, 'ΞΕΝ.'), (3, 'Ονοματεπώνυμο')):
        ws.merge_cells(start_row=1, start_column=col_i, end_row=2, end_column=col_i)
        a1 = ws.cell(row=1, column=col_i, value=lbl)
        a1.font = Font(bold=True, color='FFFFFF'); a1.fill = navy2; a1.alignment = center
    c = 4
    for mi, mo in enumerate(STAFF_STATUS_MONTHS):
        ws.merge_cells(start_row=1, start_column=c, end_row=1, end_column=c + ncol - 1)
        mc = ws.cell(row=1, column=c, value=MONTHS_EL[mo].upper())
        mc.font = Font(bold=True, color='FFFFFF', size=12); mc.fill = (navy if mi % 2 == 0 else navy2); mc.alignment = center
        for j, col in enumerate(cols):
            lbl = col['label'] + (('\n' + col['cap']) if col.get('cap') else '')
            hc = ws.cell(row=2, column=c + j, value=lbl)
            gf, fcolor = hdr_style(col.get('grp'))
            # ΛΟΓΙΣΤΗΡΙΟ πράσινο μόνο έως Μάιο· αλλιώς γκρι
            if col.get('grp') == 'acct' and mo > 5:
                gf, fcolor = None, 'FFFFFF'
            hc.fill = gf if gf else (grey if mi % 2 == 0 else grey2)
            hc.font = Font(bold=True, color=fcolor, size=9); hc.alignment = center
            hc.border = Border(left=(gside if j == 0 else thin), right=thin, top=thin, bottom=thin)
        c += ncol
    # Data
    r = 3
    for ri, row in enumerate(rows):
        even = (ri % 2 == 1)
        ac = ws.cell(row=r, column=1, value=(row.get('afm') or ''))
        ac.font = Font(color='475569'); ac.alignment = Alignment(vertical='center')
        hc2 = ws.cell(row=r, column=2, value=(row.get('hotel_code') or ''))
        hc2.font = Font(bold=True, color='0F2A36'); hc2.alignment = Alignment(horizontal='center', vertical='center')
        nc = ws.cell(row=r, column=3, value=row['name'])
        nc.font = Font(bold=True, color='153847'); nc.alignment = Alignment(vertical='center')
        if even:
            ac.fill = zebra; hc2.fill = zebra; nc.fill = zebra
        for _idc in (1, 2, 3):
            ws.cell(row=r, column=_idc).border = Border(left=thin, right=(gside if _idc == 3 else thin), top=thin, bottom=thin)
        c = 4
        for mi, mo in enumerate(STAFF_STATUS_MONTHS):
            cell = row['months'][mo]
            for j, col in enumerate(cols):
                v = cell.get(col['src'])
                x = ws.cell(row=r, column=c + j, value=v)
                x.border = Border(left=(gside if j == 0 else thin), right=thin, top=thin, bottom=thin)
                f = fmt_of(col['kind'])
                if f and v is not None: x.number_format = f
                if col['kind'] in ('money', 'num', 'int'): x.alignment = Alignment(horizontal='right')
                fill = cell_fill(col.get('grp'), mo, even)
                if fill: x.fill = fill
            c += ncol
        r += 1
    # Totals
    tr = r
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=3)
    tc = ws.cell(row=tr, column=1, value='ΣΥΝΟΛΑ')
    tc.font = Font(bold=True, color='153847'); tc.alignment = Alignment(vertical='center')
    for _idc in (1, 2, 3):
        tcell = ws.cell(row=tr, column=_idc); tcell.fill = gold
        tcell.border = Border(left=thin, right=(gside if _idc == 3 else thin), top=Side(style='medium', color='CAA64A'), bottom=thin)
    c = 4
    for mi, mo in enumerate(STAFF_STATUS_MONTHS):
        for j, col in enumerate(cols):
            if col['kind'] in ('money', 'num', 'int'):
                tot = 0.0
                for row in rows:
                    v = row['months'][mo].get(col['src'])
                    if isinstance(v, (int, float)): tot += v
                x = ws.cell(row=tr, column=c + j, value=round(tot, 2) if tot else None)
                x.number_format = int_fmt if col['kind'] == 'int' else money_fmt
                x.font = Font(bold=True, color='153847'); x.alignment = Alignment(horizontal='right')
            else:
                x = ws.cell(row=tr, column=c + j, value=None)
            x.fill = totfill
            x.border = Border(left=(gside if j == 0 else thin), right=thin, top=Side(style='medium', color='CAA64A'), bottom=thin)
        c += ncol
    ws.freeze_panes = 'D3'
    ws.row_dimensions[1].height = 22; ws.row_dimensions[2].height = 34
    ws.column_dimensions['A'].width = 13; ws.column_dimensions['B'].width = 7; ws.column_dimensions['C'].width = 28
    for i in range(4, 4 + ncol * nmonths):
        ws.column_dimensions[get_column_letter(i)].width = 11
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return Response(bio.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=staff_status_%d.xlsx' % STAFF_STATUS_YEAR})
